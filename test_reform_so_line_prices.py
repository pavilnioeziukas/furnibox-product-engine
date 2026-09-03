from pathlib import Path
import json
import sys
import tempfile
import types
import unittest

from openpyxl import Workbook, load_workbook

if "dotenv" not in sys.modules:
    dotenv_stub = types.ModuleType("dotenv")
    dotenv_stub.load_dotenv = lambda *_args, **_kwargs: None
    sys.modules["dotenv"] = dotenv_stub

from reform_so_line_prices import (
    Item,
    add_generated_boms_to_graph,
    apply_target_business_category_rules,
    build_from_application_config,
    build_reform_so_line_prices,
    calculate_boms,
    classify_missing_pricing_bom,
    component_cost_only_manufacture_products,
    exclude_bom_products_from_non_bom,
    inherit_generated_apack_rules,
    inherit_unambiguous_analog_rules,
    key,
    fpack_labour_cost,
    load_target_dataset_graph,
    load_tamara_pricing_reference,
    load_prices,
    write_component_cost_breakdown,
)
from manifest.manifest_writer import calculate_file_hash
from so_pricing_rules import (
    PricingRule,
    compose_bom_category_rule,
    empty_config,
    load_config,
    migrate_legacy_workbook,
    save_config,
)


class ReformSoLinePriceTests(unittest.TestCase):
    @staticmethod
    def pricing_rule(sku, assembly=0, storage=0, packaging=0):
        return PricingRule(
            sku,
            "CATEGORY",
            "Category",
            "Odoo / Category",
            assembly,
            storage,
            packaging,
            0,
            0,
            0,
        )

    def test_component_cost_breakdown_formats_level_ii_and_top_bom_groups(self):
        workbook = Workbook()
        workbook.remove(workbook.active)
        bom_rows = [
            {
                "component_details": [
                    {
                        "top": "TOP-1", "level_ii": "HRD-1", "level_ii_qty": 1,
                        "component": "PART-1", "component_qty": 2, "total_qty": 2,
                        "unit_price": 1, "line_cost": 2, "status": "OK",
                        "cost_source": "REFORM PURCHASE PRICE",
                    },
                    {
                        "top": "TOP-1", "level_ii": "PACK-1", "level_ii_qty": 1,
                        "component": "BOX-1", "component_qty": 1, "total_qty": 1,
                        "unit_price": 2, "line_cost": 2, "status": "OK",
                        "cost_source": "REFORM PURCHASE PRICE",
                    },
                    {
                        "top": "TOP-2", "level_ii": "HRD-2", "level_ii_qty": 1,
                        "component": "PART-2", "component_qty": 1, "total_qty": 1,
                        "unit_price": 3, "line_cost": 3, "status": "OK",
                        "cost_source": "REFORM PURCHASE PRICE",
                    },
                ]
            }
        ]

        write_component_cost_breakdown(workbook, bom_rows)
        sheet = workbook["BOM COMPONENT COSTS"]

        self.assertEqual(sheet["A2"].fill.fgColor.rgb, "00DCE6F1")
        self.assertEqual(sheet["A3"].fill.fgColor.rgb, "00FDE9D9")
        self.assertEqual(sheet["A4"].fill.fgColor.rgb, "00DCE6F1")
        self.assertEqual(sheet["A2"].border.top.style, "medium")
        self.assertEqual(sheet["J3"].border.bottom.style, "medium")
        self.assertEqual(sheet["A4"].border.top.style, "medium")
        self.assertEqual(sheet["J4"].border.bottom.style, "medium")

    def test_fpack_labour_cost_uses_tamara_floor_rate_and_cap(self):
        self.assertEqual(fpack_labour_cost(20), 4.0)
        self.assertEqual(fpack_labour_cost(39.2), 4.0)
        self.assertEqual(fpack_labour_cost(49), 5.0)
        self.assertEqual(fpack_labour_cost(98), 10.0)
        self.assertEqual(fpack_labour_cost(150), 10.0)

    def test_fpack_price_replaces_static_assembly_with_tamara_labour(self):
        top = "FPACK-EU-CAB01-BAS001"
        part = "CABINET-PART"
        packaging = "EU FP PACK"
        rules = {
            key(top): self.pricing_rule(top, assembly=7, storage=0.1),
        }

        rows, _ = calculate_boms(
            {top: ("FPACK", [Item(part, 1), Item(packaging, 1)])},
            {
                key(part): ("Part", 49.0, "CABINET PART CALCULATION"),
                key(packaging): ("Packaging", 2.0, "LAST PURCHASE PRICE"),
            },
            rules,
            adjustment=0,
            graph={key(top): [(part, 1), (packaging, 1)]},
            authoritative_rule_tops={top},
        )

        self.assertEqual(rows[0]["status"], "COMPLETE")
        self.assertEqual(rows[0]["addons"][0], 5.0)
        self.assertEqual(rows[0]["addons"][1], 0.1)
        self.assertAlmostEqual(rows[0]["cost"], 51.0)
        self.assertAlmostEqual(rows[0]["final"], 56.1)

    def test_nested_fpack_replaces_static_tariff_with_tamara_labour(self):
        top = "EUB-C-CAB01-BAS003"
        fpack = "FPACK-EU-CAB01-BAS003"
        part = "CABINET-PART"
        rules = {
            key(top): self.pricing_rule(top),
            key(fpack): self.pricing_rule(fpack, assembly=0.1),
        }
        rows, details = calculate_boms(
            {top: ("CABINET", [Item(fpack, 1)])},
            {key(part): ("Part", 49.0, "CABINET PART CALCULATION")},
            rules,
            adjustment=-0.07,
            graph={key(top): [(fpack, 1)], key(fpack): [(part, 1)]},
        )
        self.assertEqual(rows[0]["status"], "COMPLETE")
        self.assertAlmostEqual(rows[0]["addons"][0], 5.0)
        self.assertAlmostEqual(rows[0]["adjustment"], -0.35)
        fpack_detail = next(row for row in details if row["rule"].sku == fpack)
        self.assertAlmostEqual(fpack_detail["addons"][0], 5.0)
        self.assertIn("MIN(10, MAX(4", fpack_detail["calculation"])

    def test_generated_manufacture_children_are_component_cost_only(self):
        apack = "APACK-EU-C-CAB01-BAS001-A"
        part = "CON7X50"
        rules = {
            key(apack): PricingRule(
                apack, "", "APACK", "", 5, 0, 0, 0, 0, 0
            ),
        }
        boms = {
            apack: ("APACK", [Item(part, 4)]),
        }
        prices = {
            key(part): ("Connector", 0.25, "DIRECT PRICE"),
        }

        rows, details = calculate_boms(
            boms,
            prices,
            rules,
            adjustment=0,
            graph={key(apack): [(part, 4)]},
            component_cost_only_tops={apack},
        )

        self.assertEqual(rows[0]["status"], "COMPLETE")
        self.assertEqual(rows[0]["issues"], "")
        self.assertAlmostEqual(rows[0]["cost"], 1.0)
        self.assertAlmostEqual(rows[0]["final"], 6.0)
        self.assertEqual([row["level"] for row in details], ["LEVEL I BOM"])

    def test_tamara_category_expression_is_composed_from_editable_rates(self):
        rule = compose_bom_category_rule(
            "SHELF-PP",
            "8.2+25.1",
            empty_config(),
        )

        self.assertEqual(rule.category_id, "8.2+25.1")
        for actual, expected in zip(
            rule.addons,
            (7.16, 0.8, 1.0, 0.0, 0.05, 0.0),
        ):
            self.assertAlmostEqual(actual, expected)
        self.assertAlmostEqual(sum(rule.addons), 9.01)

    def test_target_business_categories_cover_shelf_and_assembled_cabinet(self):
        shelf = "EUB-C-CAB01-SLF001"
        shelf_pp = "EU-SREW-SHELF-163x564-WW-PP"
        cabinet = "EUB-C-CAB01-BAS001-A"
        apack = "APACK-EU-CAB01-BAS001-A"
        dataset = {
            "products": [
                {
                    "sku": shelf,
                    "product_type": "CABINET SHELF",
                    "bom_type": "KIT",
                    "components": [
                        {"sku": shelf_pp, "quantity": 1},
                        {"sku": "SLF-PINS-HRD-6", "quantity": 1},
                    ],
                },
                {
                    "sku": shelf_pp,
                    "product_type": "SHELF PREPACK",
                    "bom_type": "MANUFACTURE",
                    "components": [
                        {"sku": "SHELF-PART", "quantity": 1},
                        {"sku": "SHELF-PACK", "quantity": 0.4},
                        {"sku": "TERMO 90X48", "quantity": 1},
                    ],
                },
                {
                    "sku": cabinet,
                    "product_type": "CABINETS",
                    "bom_type": "KIT",
                    "components": [
                        {"sku": apack, "quantity": 1},
                        {"sku": "UNI-P-ACC01-HRD206D-A", "quantity": 1},
                    ],
                },
                {
                    "sku": apack,
                    "product_type": "PREPACK CABINETS",
                    "bom_type": "MANUFACTURE",
                    "components": [
                        {"sku": "CABINET-PART", "quantity": 1},
                        {"sku": "N PACK EU", "quantity": 1},
                    ],
                },
            ],
        }

        rules, authoritative = apply_target_business_category_rules(
            {}, dataset, empty_config(), reference={}
        )

        self.assertAlmostEqual(sum(rules[key(shelf_pp)].addons), 1.15)
        self.assertAlmostEqual(sum(rules[key(shelf)].addons), 1.82)
        self.assertAlmostEqual(sum(rules[key(apack)].addons), 63.27)
        self.assertAlmostEqual(sum(rules[key(cabinet)].addons), 64.87)
        self.assertEqual(
            authoritative,
            {key(shelf), key(shelf_pp), key(cabinet), key(apack)},
        )

    def test_target_business_categories_cover_flat_pack_usb_cabinet(self):
        sku = "USB-C-CAB01-UPP010"
        dataset = {"products": [{
            "sku": sku,
            "product_type": "CABINETS",
            "bom_type": "KIT",
            "components": [
                {"sku": "FPACK-US-CAB01-UPP010", "quantity": 1},
                {"sku": "UNI-P-ACC01-HRD202D", "quantity": 1},
            ],
        }]}
        rules, authoritative = apply_target_business_category_rules(
            {}, dataset, empty_config(), reference={}
        )
        self.assertIn(key(sku), rules)
        self.assertEqual(rules[key(sku)].addons, (0, 0, 0, 0, 0, 0))
        self.assertNotIn(key(sku), authoritative)

    def test_authoritative_product_category_does_not_double_child_addons(self):
        top = "SHELF"
        child = "SHELF-PP"
        rules = {
            key(top): self.pricing_rule(top, assembly=10),
            key(child): self.pricing_rule(child, assembly=4),
        }
        rows, details = calculate_boms(
            {top: ("CABINET SHELF", [Item(child, 1)])},
            {key("PART"): ("Part", 2.0, "DIRECT PRICE")},
            rules,
            adjustment=0,
            graph={key(top): [(child, 1)], key(child): [("PART", 1)]},
            authoritative_rule_tops={top},
        )

        self.assertEqual(rows[0]["status"], "COMPLETE")
        self.assertEqual(rows[0]["cost"], 2.0)
        self.assertEqual(sum(rows[0]["addons"]), 10.0)
        self.assertEqual([row["level"] for row in details], ["LEVEL I BOM"])

    def test_target_shelf_subtypes_and_us_pack_categories(self):
        products = []
        expectations = {
            "US-SREW-SHELF-420x339-WW-PP": ("8+26.1", 1.15),
            "EU-SREW-SHELF-ROD-563x340-WW-PP": ("8.1+25.1", 6.65),
            "EU-SREW-SHELF-LEDROD-563x340-WW-PP": ("8.2+25.1", 9.01),
        }
        for sku in expectations:
            components = [{"sku": "PART", "quantity": 1}]
            if sku.startswith("US-"):
                components.append({"sku": "L0377", "quantity": 0.4})
            products.append({
                "sku": sku,
                "product_type": "SHELF PREPACK",
                "bom_type": "MANUFACTURE",
                "components": components,
            })

        rules, _ = apply_target_business_category_rules(
            {}, {"products": products}, empty_config(), reference={}
        )

        for sku, (expression, total) in expectations.items():
            self.assertEqual(rules[key(sku)].category_id, expression)
            self.assertAlmostEqual(sum(rules[key(sku)].addons), total)

    def test_fpack_market_category_follows_sku_not_stale_reference(self):
        products = [
            {"sku": "FPACK-EU-CAB01-BAS001", "product_type": "PREPACK CABINETS"},
            {"sku": "FPACK-US-CAB01-BAS001", "product_type": "PREPACK CABINETS"},
        ]
        reference = {
            key("FPACK-EU-CAB01-BAS001"): "3+21.1",
            key("FPACK-US-CAB01-BAS001"): "3+20.1",
        }

        rules, authoritative = apply_target_business_category_rules(
            {}, {"products": products}, empty_config(), reference=reference
        )

        self.assertEqual(rules[key("FPACK-EU-CAB01-BAS001")].category_id, "3+20.1")
        self.assertEqual(rules[key("FPACK-US-CAB01-BAS001")].category_id, "3+21.1")
        self.assertEqual(authoritative, {key(product["sku"]) for product in products})

    def test_bom_material_cost_uses_reform_purchase_price(self):
        with tempfile.TemporaryDirectory() as directory:
            price_path = Path(directory) / "Reform_Final_Prices.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "REFORM PRICE LIST"
            sheet.append([
                "Internal Reference",
                "Name",
                "Adjusted Furnibox Purchase Price",
                "Reform Markup Factor",
                "Reform Purchase Price",
                "Price Source",
            ])
            sheet.append([
                "MATERIAL-1",
                "Material",
                10.0,
                1.05,
                10.5,
                "LAST PURCHASE PRICE",
            ])
            workbook.save(price_path)

            prices = load_prices(price_path)

            self.assertEqual(prices[key("MATERIAL-1")][1], 10.5)

    def test_versioned_tamara_reference_wins_over_market_heuristic(self):
        sku = "EUB-C-CAB01-BNF001-A"
        apack = "APACK-EU-CAB01-BNF001-A"
        reference = load_tamara_pricing_reference()
        self.assertEqual(reference[key(sku)], "12+9+23.1+24.1")
        rules, authoritative = apply_target_business_category_rules(
            {},
            {"products": [{
                "sku": sku,
                "product_type": "CABINETS",
                "bom_type": "KIT",
                "components": [{"sku": apack, "quantity": 1}],
            }]},
            empty_config(),
        )
        self.assertEqual(rules[key(sku)].category_id, "12+9+23.1+24.1")
        self.assertIn(key(sku), authoritative)

    def test_every_versioned_tamara_expression_has_configured_rates(self):
        reference = load_tamara_pricing_reference()
        self.assertEqual(len(reference), 3218)
        for sku, expression in reference.items():
            compose_bom_category_rule(sku, expression, empty_config())

    def test_generated_cost_only_scope_does_not_hide_parent_rule_gap(self):
        cabinet = "EUB-C-CAB01-BAS001-A"
        apack = "APACK-EU-C-CAB01-BAS001-A"
        part = "PART"
        rules = {
            key(cabinet): PricingRule(
                cabinet, "", "CABINET-A", "", 1, 0, 0, 0, 0, 0
            ),
        }
        boms = {
            cabinet: ("CABINETS (Assembled)", [Item(apack, 1)]),
        }
        prices = {
            key(part): ("Part", 2.0, "DIRECT PRICE"),
        }

        rows, _ = calculate_boms(
            boms,
            prices,
            rules,
            adjustment=0,
            graph={
                key(cabinet): [(apack, 1)],
                key(apack): [(part, 1)],
            },
            component_cost_only_tops={apack},
        )

        self.assertEqual(rows[0]["status"], "BLOCKED")
        self.assertIn(
            f"Missing LEVEL II BOM pricing rule: {apack}",
            rows[0]["issues"],
        )

    def test_apack_rule_inherits_only_from_exact_generated_fpack(self):
        fpack = "FPACK-EU-CAB01-BNF002"
        apack = "APACK-EU-C-CAB01-BNF002-A"
        source = PricingRule(
            fpack, "8", "FPACK", "", 5, 0.2, 0, 0, 0, 0
        )
        shelf_pp = "EUB-PACK-CAB01-SLF301-PP"
        dataset = {
            "products": [{
                "sku": apack,
                "generated_from": fpack,
                "bom_type": "MANUFACTURE",
            }, {
                "sku": shelf_pp,
                "generated_from": "",
                "bom_type": "MANUFACTURE",
            }, {
                "sku": "FPACK-EU-CAB01-BNF002",
                "generated_from": "",
                "bom_type": "MANUFACTURE",
            }, {
                "sku": "EUB-C-CAB01-BNF002-A",
                "generated_from": "EUB-C-CAB01-BNF002",
                "bom_type": "KIT",
            }],
        }

        rules = inherit_generated_apack_rules(
            {key(fpack): source},
            dataset,
        )

        self.assertEqual(rules[key(apack)].sku, apack)
        self.assertEqual(rules[key(apack)].addons, source.addons)
        self.assertEqual(
            component_cost_only_manufacture_products(dataset),
            {key(fpack), key(apack), key(shelf_pp)},
        )

    def test_internal_manufacture_cost_uses_bom_instead_of_direct_history(self):
        fpack = "FPACK-EU-CAB01-BAS001"
        part = "EU-SIDE-SREW-800X590-WW"
        rules = {
            key(fpack): self.pricing_rule(fpack),
        }
        rows, _ = calculate_boms(
            {fpack: ("FPACK", [Item(part, 2)])},
            {
                key(fpack): ("Old direct FPACK", 999.0, "LAST PURCHASE PRICE"),
                key(part): ("Cabinet part", 5.0, "CABINET PART CALCULATION"),
            },
            rules,
            adjustment=0,
            graph={key(fpack): [(part, 2)]},
            component_cost_only_tops={fpack},
        )

        self.assertEqual(rows[0]["status"], "COMPLETE")
        self.assertAlmostEqual(rows[0]["cost"], 10.0)
        self.assertAlmostEqual(rows[0]["addons"][0], 4.0)
        self.assertAlmostEqual(rows[0]["final"], 14.0)

    def test_non_positive_direct_component_price_blocks_bom(self):
        top = "FPACK-EU-CAB01-BAS001"
        part = "PART-ZERO"
        rows, _ = calculate_boms(
            {top: ("FPACK", [Item(part, 1)])},
            {key(part): ("Part", 0.0, "LAST PURCHASE PRICE")},
            {key(top): self.pricing_rule(top)},
            adjustment=0,
            graph={key(top): [(part, 1)]},
        )

        self.assertEqual(rows[0]["status"], "BLOCKED")
        self.assertIsNone(rows[0]["final"])
        self.assertIn(
            "Non-positive component price: PART-ZERO (0)",
            rows[0]["issues"],
        )

    def test_configured_bom_without_target_components_is_blocked(self):
        top = "EMPTY-BOM"
        rows, _ = calculate_boms(
            {top: ("CABINETS", [])},
            {},
            {key(top): self.pricing_rule(top)},
            adjustment=0,
            graph={},
        )

        self.assertEqual(rows[0]["status"], "BLOCKED")
        self.assertIsNone(rows[0]["final"])
        self.assertIn(
            "Target BOM has no components: EMPTY-BOM",
            rows[0]["issues"],
        )

    def test_missing_pricing_bom_is_attributed_to_reform_source(self):
        issue = classify_missing_pricing_bom(
            "SOURCE-BOM",
            {key("SOURCE-BOM"): []},
        )
        self.assertIn("REFORM_BOM_MISSING_COMPONENTS", issue)

    def test_production_only_bom_points_to_lifecycle_audit(self):
        issue = classify_missing_pricing_bom(
            "ODOO-ONLY",
            {},
            {"product_catalog": []},
            {key("ODOO-ONLY")},
        )
        self.assertIn("PRODUCTION_ODOO_BOM_NOT_IN_TARGET_DATASET", issue)
        self.assertIn("Product Lifecycle Audit", issue)

    def test_missing_generated_bom_is_attributed_to_product_engine(self):
        issue = classify_missing_pricing_bom(
            "GENERATED-BOM",
            {},
            {
                "product_catalog": [
                    {
                        "sku": "GENERATED-BOM",
                        "generated_from": "SOURCE-BOM",
                    }
                ]
            },
        )
        self.assertIn("PRODUCT_ENGINE_GENERATED_BOM_MISSING", issue)
        self.assertIn("SOURCE-BOM", issue)

    def test_missing_reform_bom_is_distinct_from_scope_mismatch(self):
        dataset_issue = classify_missing_pricing_bom(
            "KNOWN-PRODUCT",
            {},
            {"product_catalog": [{"sku": "KNOWN-PRODUCT"}]},
        )
        scope_issue = classify_missing_pricing_bom(
            "CONFIG-ONLY",
            {},
            {"product_catalog": []},
        )
        self.assertIn("REFORM_BOM_NOT_PROVIDED", dataset_issue)
        self.assertIn("PRICING_BOM_SCOPE_MISMATCH", scope_issue)

    def test_missing_rule_inherits_from_unanimous_exact_analogs(self):
        target = "FPACK-EU-CAB03-WAL015"
        analogs = [
            "FPACK-EU-CAB01-WAL015",
            "FPACK-EU-CAB02-WAL015",
            "FPACK-US-CAB03-WAL015",
        ]
        dataset = {
            "product_catalog": [
                {
                    "sku": sku,
                    "has_bom": True,
                    "product_type": "PREPACK CABINETS",
                    "name_2": "WALL Cabinet - W30 H80 D37",
                }
                for sku in [target, *analogs]
            ],
        }
        rules = {
            key(sku): self.pricing_rule(sku, assembly=4)
            for sku in analogs
        }

        result = inherit_unambiguous_analog_rules(rules, dataset)

        self.assertEqual(result[key(target)].sku, target)
        self.assertEqual(result[key(target)].addons, (4, 0, 0, 0, 0, 0))

    def test_analog_category_labels_do_not_block_identical_pricing(self):
        target = "FPACK-EU-CAB03-WAL015"
        first = self.pricing_rule("FPACK-EU-CAB01-WAL015", assembly=4)
        second = PricingRule(
            "FPACK-EU-CAB02-WAL015",
            "DIFFERENT-ID",
            "Different category label",
            "Different Odoo category",
            4,
            0,
            0,
            0,
            0,
            0,
        )
        dataset = {"product_catalog": [{
            "sku": sku,
            "has_bom": True,
            "product_type": "PREPACK CABINETS",
            "name_2": "WALL Cabinet - W30 H80 D37",
        } for sku in (target, first.sku, second.sku)]}

        result = inherit_unambiguous_analog_rules(
            {key(first.sku): first, key(second.sku): second},
            dataset,
        )

        self.assertEqual(result[key(target)].addons, (4, 0, 0, 0, 0, 0))

    def test_shelf_pp_rule_may_inherit_across_width_only(self):
        target = "EUB-PACK-CAB03-SLF902-PP"
        analogs = [
            ("EUB-PACK-CAB03-SLF901-PP", "W20"),
            ("EUB-PACK-CAB03-SLF903-PP", "W60"),
            ("EUB-PACK-CAB03-SLF904-PP", "W80"),
        ]
        catalog = [{
            "sku": target,
            "has_bom": True,
            "product_type": "SHELF PREPACK",
            "name_2": "Shelf - W40 D60 - Shelf with integrated light (silver)",
        }]
        catalog.extend({
            "sku": sku,
            "has_bom": True,
            "product_type": "SHELF PREPACK",
            "name_2": f"Shelf - {width} D60 - Shelf with integrated light (silver)",
        } for sku, width in analogs)
        clothing_rod = "EUB-PACK-CAB03-SLF909-PP"
        catalog.append({
            "sku": clothing_rod,
            "has_bom": True,
            "product_type": "SHELF PREPACK",
            "name_2": "Shelf - W40 D60 - Shelf with integrated light (silver) & clothing rod",
        })
        rules = {
            key(sku): self.pricing_rule(
                sku,
                assembly=7.16,
                storage=0.5,
                packaging=1,
            )
            for sku, _ in analogs
        }
        rules[key(clothing_rod)] = self.pricing_rule(
            clothing_rod,
            assembly=7.16,
            storage=0.7,
            packaging=1,
        )

        result = inherit_unambiguous_analog_rules(
            rules,
            {"product_catalog": catalog},
        )

        self.assertEqual(result[key(target)].sku, target)
        self.assertEqual(
            result[key(target)].addons,
            (7.16, 0.5, 1, 0, 0, 0),
        )

    def test_ambiguous_analog_profiles_remain_without_rule(self):
        target = "TARGET-PP"
        dataset = {"product_catalog": [{
            "sku": sku,
            "has_bom": True,
            "product_type": "SHELF PREPACK",
            "name_2": name,
        } for sku, name in (
            (target, "Shelf - W40 D60 - Standard"),
            ("ANALOG-1-PP", "Shelf - W20 D60 - Standard"),
            ("ANALOG-2-PP", "Shelf - W60 D60 - Standard"),
        )]}
        rules = {
            key("ANALOG-1-PP"): self.pricing_rule("ANALOG-1-PP", storage=0.5),
            key("ANALOG-2-PP"): self.pricing_rule("ANALOG-2-PP", storage=0.7),
        }

        result = inherit_unambiguous_analog_rules(rules, dataset)

        self.assertNotIn(key(target), result)

    def test_bom_product_is_excluded_from_non_bom_pricing(self):
        bom_item = ("BOM", "BOM", "Category", "Pricing", 0, 0, 0, 0)
        non_bom_item = (
            "NON-BOM", "Non-BOM", "Category", "Pricing", 0, 0, 0, 0
        )

        result = exclude_bom_products_from_non_bom(
            [bom_item, non_bom_item],
            {key("BOM"): [("PART", 1)]},
        )

        self.assertEqual(result, [non_bom_item])

    def test_pricing_uses_full_target_dataset_for_shelf_pp(self):
        with tempfile.TemporaryDirectory() as directory:
            base = Path(directory)
            bom = base / "Reform BOM.xlsx"
            bom.write_bytes(b"reform")
            dataset_path = base / "target.json"
            shelf = "SHELF"
            shelf_pp = "SHELF-PART-PP"
            dataset_path.write_text(json.dumps({
                "environment": "production",
                "source": {"file_hash": calculate_file_hash(bom)},
                "products": [
                    {
                        "sku": shelf,
                        "components": [
                            {"sku": shelf_pp, "quantity": 1},
                            {"sku": "PINS", "quantity": 1},
                        ],
                    },
                    {
                        "sku": shelf_pp,
                        "components": [
                            {"sku": "SHELF-PART", "quantity": 1},
                            {"sku": "PACKAGE", "quantity": 0.4},
                            {"sku": "STICKER", "quantity": 2},
                        ],
                    },
                ],
            }), encoding="utf-8")
            _, graph = load_target_dataset_graph(dataset_path, bom)
            prices = {
                key("SHELF-PART"): ("Part", 10.0, "DIRECT"),
                key("PACKAGE"): ("Package", 2.0, "DIRECT"),
                key("STICKER"): ("Sticker", 0.1, "DIRECT"),
                key("PINS"): ("Pins", 1.0, "DIRECT"),
            }
            from reform_so_line_prices import resolve_component_cost
            result = resolve_component_cost(shelf, prices, graph)
            self.assertAlmostEqual(result["cost"], 12.0)
            self.assertEqual(result["issues"], [])

    def test_stale_target_dataset_is_blocked(self):
        with tempfile.TemporaryDirectory() as directory:
            base = Path(directory)
            bom = base / "Reform BOM.xlsx"
            bom.write_bytes(b"current")
            dataset = base / "target.json"
            dataset.write_text(json.dumps({
                "environment": "production",
                "source": {"file_hash": "old"},
                "products": [{
                    "sku": "TOP",
                    "components": [{"sku": "PART", "quantity": 1}],
                }],
            }), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "ne iš pateikto Reform"):
                load_target_dataset_graph(dataset, bom)

    def test_untransformed_apack_dataset_is_blocked(self):
        with tempfile.TemporaryDirectory() as directory:
            base = Path(directory)
            bom = base / "Reform BOM.xlsx"
            bom.write_bytes(b"current")
            dataset = base / "target.json"
            dataset.write_text(json.dumps({
                "environment": "production",
                "source": {"file_hash": calculate_file_hash(bom)},
                "products": [{
                    "sku": "APACK-EU-C-CAB01-BAS001-A",
                    "components": [{"sku": "PART", "quantity": 1}],
                }],
            }), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "seno Dataset"):
                load_target_dataset_graph(dataset, bom)

    def test_assembled_cabinet_uses_generated_apack_and_hrd_a_boms(self):
        cabinet = "EUB-C-CAB01-BAS001"
        assembled = f"{cabinet}-A"
        fpack = "FPACK-EU-CAB01-BAS001"
        apack = "APACK-EU-C-CAB01-BAS001-A"
        hrd = "UNI-P-ACC01-HRD206D"
        hrd_a = f"{hrd}-A"
        graph = {
            key(cabinet): [(hrd, 1), (fpack, 1)],
            key(hrd): [("HRD-PART", 2)],
            key(fpack): [("CABINET-PART", 3)],
        }
        products = [
            {"sku": cabinet, "product_category": "All / CABINETS"},
            {"sku": assembled, "product_category": "All / CABINETS (Assembled)"},
        ]

        generated = add_generated_boms_to_graph(graph, products)

        self.assertEqual(generated[key(assembled)], [(hrd_a, 1.0), (apack, 1.0)])
        self.assertEqual(generated[key(hrd_a)], [("HRD-PART", 2.0)])
        self.assertEqual(generated[key(apack)], [
            ("CABINET-PART", 3.0),
            ("N PACK EU", 1.0),
            ("STICKER UP", 2.0),
        ])

        def rule(sku, assembly=0):
            return PricingRule(sku, "", "", "", assembly, 0, 0, 0, 0, 0)

        rules = {
            key(sku): rule(sku, assembly=5.5 if sku == assembled else 0)
            for sku in (cabinet, assembled, fpack, apack, hrd, hrd_a)
        }
        boms = {
            cabinet: ("CABINETS", [Item(hrd, 1), Item(fpack, 1)]),
            assembled: ("CABINETS (Assembled)", [Item(hrd_a, 1), Item(apack, 1)]),
        }
        prices = {
            key("HRD-PART"): ("Hardware", 4.0, "DIRECT PRICE"),
            key("CABINET-PART"): ("Cabinet part", 10.0, "DIRECT PRICE"),
            key("TERMO 90X48"): ("Label", 0.1, "DIRECT PRICE"),
            key("EU FP PACK"): ("Flat pack", 1.0, "DIRECT PRICE"),
            key("N PACK EU"): ("Assembled pack", 2.0, "DIRECT PRICE"),
            key("STICKER UP"): ("Sticker", 0.05, "DIRECT PRICE"),
        }

        rows, _ = calculate_boms(boms, prices, rules, adjustment=0, graph=generated)
        by_sku = {row["sku"]: row for row in rows}
        self.assertAlmostEqual(by_sku[cabinet]["cost"], 39.1)
        self.assertAlmostEqual(by_sku[assembled]["cost"], 40.1)
        self.assertGreater(by_sku[assembled]["cost"], by_sku[cabinet]["cost"])
        self.assertEqual(sum(by_sku[assembled]["addons"]), 5.5)

    def test_generated_fpack_and_apack_use_tamara_market_packaging(self):
        fpack_eu = "FPACK-EU-CAB03-BNF002"
        fpack_us = "FPACK-US-CAB03-BNF002"
        graph = {
            key(fpack_eu): [("EU-PART", 1)],
            key(fpack_us): [("US-PART", 1)],
        }
        products = [
            {"sku": fpack_eu, "product_category": "PREPACK CABINETS"},
            {"sku": fpack_us, "product_category": "PREPACK CABINETS"},
        ]

        generated = add_generated_boms_to_graph(graph, products)

        self.assertEqual(generated[key(fpack_eu)], [
            ("EU-PART", 1.0),
            ("TERMO 90X48", 1.0),
            ("EU FP PACK", 1.0),
        ])
        self.assertEqual(generated[key("APACK-EU-C-CAB03-BNF002-A")], [
            ("EU-PART", 1.0),
            ("N PACK EU", 1.0),
            ("STICKER UP", 2.0),
        ])
        self.assertEqual(generated[key(fpack_us)], [
            ("US-PART", 1.0),
            ("TERMO 90X48", 1.0),
            ("L0377", 1.0),
            ("US FP PACK", 1.0),
        ])
        self.assertEqual(generated[key("APACK-US-C-CAB03-BNF002-A")], [
            ("US-PART", 1.0),
            ("N PACK US", 1.0),
            ("L0377", 1.0),
            ("STICKER UP", 2.0),
        ])

    def test_bom_category_breakdown_and_non_bom_logic(self):
        with tempfile.TemporaryDirectory() as directory:
            base = Path(directory)
            model = base / "model.xlsx"
            prices = base / "Reform_Final_Prices.xlsx"
            output = base / "Reform_SO_Line_Prices.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "bomai"
            ws.append([None] * 24)
            ws.append(["lv1", None, None, "lv2", None, "qty2", "lv3", "qty3"])
            ws.append(["TOP-1", "TOP-1", "CABINET", "SUB-1", "PACK", 2, "PART-1", 3])
            ws.append([None, "TOP-1", "CABINET", None, "PACK", None, "PART-2", 1])
            ws.append(["TOP-1", "TOP-1", "CABINET", "DIRECT-1", "HARDWARE", 4, None, None])
            rules = wb.create_sheet("Kainodaros kategorijos")
            rules.append(["SKU", "ID", "Name", "Odoo", "Assembly", "Storage", "Packaging", "Pallet", "Other", "Markup", "Total"])
            rules.append(["SUB-1", 8, "PACK", "", 1, 2, 3, 4, 5, 6, 21])
            rules.append(["DIRECT-1", 7, "HARDWARE", "", 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 2.1])
            rules.append(["TOP-1", 1, "CABINET", "", 10, 20, 30, 40, 50, 60, 210])
            non_bom = wb.create_sheet("Ne BOM pozicijos")
            non_bom.append(["SKU", "Description", "Group", "Category", "Cost", "New price", "Preparation", "Storage", "Bag", "Sticker", "Total"])
            non_bom.append(["ACC-1", "Accessory", "ACC", 11, 0, 0, 0.1, 0.2, 0.03, 0.02, 0])
            wb.save(model)

            wb = Workbook()
            ws = wb.active
            ws.title = "REFORM PRICE LIST"
            ws.append(["Internal Reference", "Name", "Adjusted Furnibox Purchase Price", "Reform Markup Factor", "Reform Purchase Price"])
            ws.append(["PART-1", "Part 1", 2, 1, 2])
            ws.append(["PART-2", "Part 2", 5, 1, 5])
            ws.append(["DIRECT-1", "Direct", 7, 1, 7])
            ws.append(["ACC-1", "Accessory", 10, 1, 10])
            wb.save(prices)

            counts = build_reform_so_line_prices(model, prices, output)
            self.assertEqual(counts, (1, 1, 0))
            result = load_workbook(output, data_only=True)
            self.assertEqual(result.sheetnames, [
                "SO LINE PRICES", "BOM COMPONENT COSTS", "BOM CATEGORY BREAKDOWN", "CATEGORY RULES",
                "NON-BOM RULES", "DIAGNOSTICS", "INFO",
            ])
            sheet = result["SO LINE PRICES"]
            headers = {cell.value: cell.column for cell in sheet[1]}
            rows = {sheet.cell(r, 1).value: r for r in range(2, sheet.max_row + 1)}
            bom = rows["TOP-1"]
            # Components: (2*3 + 5*1)*2 + 7*4 = 50
            self.assertAlmostEqual(sheet.cell(bom, headers["Component / Purchase Cost"]).value, 50)
            # Add-ons: SUB-1*2 + DIRECT-1 once + TOP-1 once = 254.1
            self.assertAlmostEqual(sheet.cell(bom, headers["Pricing Add-ons Total"]).value, 254.1)
            self.assertAlmostEqual(sheet.cell(bom, headers["Adjustment Amount"]).value, -17.787)
            self.assertAlmostEqual(sheet.cell(bom, headers["Final Reform SO Unit Price"]).value, 286.313)
            non = rows["ACC-1"]
            self.assertAlmostEqual(sheet.cell(non, headers["Final Reform SO Unit Price"]).value, 10.35)
            self.assertEqual(sheet.cell(non, headers["Status"]).value, "COMPLETE")

            components = result["BOM COMPONENT COSTS"]
            component_headers = {cell.value: cell.column for cell in components[1]}
            component_rows = {
                components.cell(row, component_headers["Purchased Component SKU"]).value: row
                for row in range(2, components.max_row + 1)
            }
            part_1 = component_rows["PART-1"]
            self.assertEqual(components.cell(part_1, component_headers["Level II SKU"]).value, "SUB-1")
            self.assertAlmostEqual(components.cell(part_1, component_headers["Total Qty in Top BOM"]).value, 6)
            self.assertAlmostEqual(components.cell(part_1, component_headers["Purchase Unit Price"]).value, 2)
            self.assertAlmostEqual(components.cell(part_1, component_headers["Component Cost"]).value, 12)
            direct = component_rows["DIRECT-1"]
            self.assertAlmostEqual(components.cell(direct, component_headers["Component Cost"]).value, 28)

    def test_application_config_replaces_legacy_workbook_at_runtime(self):
        with tempfile.TemporaryDirectory() as directory:
            base = Path(directory)
            legacy = base / "legacy.xlsx"
            bom_input = base / "Reform_BOM_Input.xlsx"
            config = base / "so_pricing_rules.json"
            prices = base / "prices.xlsx"
            output = base / "result.xlsx"

            wb = Workbook()
            ws = wb.active; ws.title = "bomai"
            ws.append([]); ws.append([]); ws.append([None, "TOP", "CABINET", "SUB", None, 2, "PART", 3])
            ws = wb.create_sheet("Kainodaros kategorijos")
            ws.append(["SKU", "ID", "Name", "Odoo", "Assembly", "Storage", "Packaging", "Pallet", "Other", "Markup"])
            ws.append(["SUB", "2", "Pack", "", 1, 0, 0, 0, 0, 0])
            ws.append(["TOP", "1", "Cabinet", "", 10, 0, 0, 0, 0, 0])
            ws = wb.create_sheet("Ne BOM pozicijos")
            ws.append(["SKU", "Name", "Group", "Category", "", "", "Preparation", "Storage", "Bag", "Sticker"])
            wb.save(legacy)
            migrated = migrate_legacy_workbook(legacy)
            self.assertEqual(migrated["schema_version"], 3)
            self.assertEqual(len(migrated["bom_categories"]), 2)
            self.assertEqual(len(migrated["bom_skus"]), 2)
            self.assertEqual(migrated["bom_skus"][0]["category_id"], "BOM-001")
            save_config(config, migrated)
            self.assertEqual(load_config(config)["schema_version"], 3)
            legacy.unlink()

            wb = Workbook(); ws = wb.active; ws.title = "BOM - Input"
            ws.append(["BOM SKU Code", "Part 1 Code", "Part 1 Qty"])
            ws.append(["TOP", "SUB", 2]); ws.append(["SUB", "PART", 3]); wb.save(bom_input)
            wb = Workbook(); ws = wb.active; ws.title = "REFORM PRICE LIST"
            ws.append(["Internal Reference", "Name", "Adjusted Furnibox Purchase Price", "Reform Markup Factor", "Reform Purchase Price"])
            ws.append(["PART", "Part", 5, 1, 5]); wb.save(prices)

            self.assertEqual(build_from_application_config(bom_input, prices, config, output), (1, 0, 0))
            result = load_workbook(output, data_only=True)["SO LINE PRICES"]
            headers = {cell.value: cell.column for cell in result[1]}
            self.assertAlmostEqual(result.cell(2, headers["Final Reform SO Unit Price"]).value, 41.16)


if __name__ == "__main__":
    unittest.main()
