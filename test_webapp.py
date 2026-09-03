from __future__ import annotations

import importlib
import hashlib
import io
from pathlib import Path
from openpyxl import Workbook


def load_webapp(monkeypatch, tmp_path):
    for name in (
        "PRODUCT_ENGINE_APP_NAME",
        "PRODUCT_ENGINE_BRAND",
        "PRODUCT_ENGINE_BRAND_MARK",
        "PRODUCT_ENGINE_HERO_EYEBROW",
        "PRODUCT_ENGINE_HERO_TITLE",
        "PRODUCT_ENGINE_HERO_DESCRIPTION",
        "PRODUCT_ENGINE_STATE_DIR",
        "PRODUCT_ENGINE_WEB_SECRET",
        "PRODUCT_ENGINE_WEB_PASSWORD",
        "PRODUCT_ENGINE_ENABLED_ACTIONS",
        "PRODUCT_ENGINE_ACTION_MODULES",
        "PRODUCT_ENGINE_SHOW_BOM_WORKSPACE",
        "PRODUCT_ENGINE_SHOW_PRICING_NAV",
    ):
        monkeypatch.delenv(name, raising=False)
    monkeypatch.setenv("FURNIBOX_WEB_STATE_DIR", str(tmp_path / "state"))
    monkeypatch.setenv("FURNIBOX_WEB_SECRET", "test-secret")
    monkeypatch.delenv("FURNIBOX_WEB_PASSWORD", raising=False)
    monkeypatch.delenv("FURNIBOX_SHARED_DATA", raising=False)
    monkeypatch.delenv("FURNIBOX_SHARED_DATA_DIR", raising=False)

    import webapp.app as webapp

    return importlib.reload(webapp)


def test_health_and_index(monkeypatch, tmp_path):
    webapp = load_webapp(monkeypatch, tmp_path)
    client = webapp.app.test_client()

    assert client.get("/health").get_json() == {"status": "ok"}
    assert client.get("/").status_code == 200
    assert "Atnaujinti Reform kainodarą" in client.get("/").get_data(as_text=True)
    assert "Audituoti nebeaktualius Odoo produktus ir BOM" in client.get("/").get_data(as_text=True)
    assert "upload-progress" in client.get("/").get_data(as_text=True)
    assert "chunked-upload.js" in client.get("/").get_data(as_text=True)


def test_pricing_control_explains_inputs_and_source_priority(monkeypatch, tmp_path):
    webapp = load_webapp(monkeypatch, tmp_path)
    page = webapp.app.test_client().get("/pricing-control")
    text = page.get_data(as_text=True)

    assert page.status_code == 200
    assert "Rankiniu būdu valdome" in text
    assert "Kainos šaltinių pasirinkimo tvarka" in text
    assert "Odoo Standard Price nėra atsarginis šaltinis" in text
    assert "Odoo nekeičiamas" in text
    assert "Nustatomas produkto kainodaros tipas" in text
    assert "Perkamas komponentas" in text
    assert "Cabinet / Shelf Part" in text
    assert "APACK / HRD-A / Shelf-PP" in text
    assert "MIN(10 €, MAX(4 €, Cabinet Parts savikaina ÷ 9,8))" in text
    assert "Ne BOM produktas" in text


def test_pricing_control_saves_manual_values_together(monkeypatch, tmp_path):
    webapp = load_webapp(monkeypatch, tmp_path)
    client = webapp.app.test_client()
    response = client.post("/pricing-control/manual-values", data={
        "adjustment_percent": "-5",
        "back_rate_per_m2": "12",
        "processing_rate_per_m2": "18",
        "ww_material_rate_per_m2": "9",
        "bb_material_rate_per_m2": "8",
        "no_material_rate_per_m2": "9",
        "small_part_threshold_m2": "0.4",
        "small_part_surcharge": "1.2",
        "furnix_markup_percent": "2",
        "output_decimals": "3",
    })

    assert response.status_code == 302
    assert webapp.load_config(webapp.SO_PRICING_CONFIG_PATH)["adjustment_rate"] == -0.05
    saved = webapp.load_cabinet_parts_parameters(webapp.CABINET_PARTS_PARAMETERS_PATH)
    assert saved.processing_rate_per_m2 == 18
    assert saved.output_decimals == 3


def test_pricing_rules_explains_internal_and_legacy_category_ids(monkeypatch, tmp_path):
    webapp = load_webapp(monkeypatch, tmp_path)
    page = webapp.app.test_client().get("/pricing-rules").get_data(as_text=True)

    assert "Priskirta SKU" in page
    assert "Techninė informacija" in page
    assert "Vidinis ID:" in page
    assert "Seno MAP kodas:" in page
    assert "kasdieniam valdymui jie nereikalingi" in page


def test_pricing_control_searches_latest_run_trace_and_blockers(monkeypatch, tmp_path):
    webapp = load_webapp(monkeypatch, tmp_path)
    job_dir = webapp.RUN_DIR / "pricing123"
    files_dir = job_dir / "files"
    files_dir.mkdir(parents=True)

    prices = Workbook()
    sheet = prices.active
    sheet.title = "PRICE RESULTS"
    sheet.append(["SKU", "Name", "Position Type", "Product Category", "Component / Purchase Cost", "Pricing Add-ons Total", "Adjustment Amount", "Final Reform SO Unit Price", "Control Status", "Applied Rule IDs", "Issues / Review Reason"])
    sheet.append(["SKU-OK", "Test product", "BOM", "Cabinet", 10, 2, -0.1, 11.9, "CALCULATED", "R002, R003", ""])
    trace = prices.create_sheet("PRICE TRACE")
    trace.append(["SKU", "Step #", "Step Type", "Rule ID", "Input / Component / Rule", "Qty / Multiplier", "Unit Price", "Amount", "Source", "Step Status", "Explanation"])
    trace.append(["SKU-OK", 1, "MATERIAL", "R001", "COMP-1", 2, 5, 10, "Last Purchase Price", "CALCULATED", "Test"])
    trace.append(["SKU-OK", 2, "PRICING ADD-ON", "R003/R004", "SKU-OK | C01 | Cabinet", 1, None, 1.9, "LEVEL I BOM", "CALCULATED", "Test tariff"])
    price_path = files_dir / "Reform_SO_Line_Prices_COMPLETE_ONLY.xlsx"
    prices.save(price_path)

    blocked = Workbook()
    blockers = blocked.active
    blockers.title = "BLOCKERS"
    blockers.append(["SKU", "Position Type", "Status", "Issues"])
    blockers.append(["SKU-BLOCKED", "BOM", "BLOCKED", "Missing component price"])
    blocker_path = files_dir / "Reform_Pricing_BLOCKED.xlsx"
    blocked.save(blocker_path)

    webapp.write_job(job_dir, {"id": "pricing123", "action": "refresh_reform_pricing", "title": "Pricing", "status": "BLOCKED", "created_at": "2026-09-03T00:00:00+00:00", "files": [{"name": price_path.name, "path": str(price_path)}, {"name": blocker_path.name, "path": str(blocker_path)}]})
    client = webapp.app.test_client()

    ok_page = client.get("/pricing-control?sku=SKU-OK").get_data(as_text=True)
    blocked_page = client.get("/pricing-control?sku=SKU-BLOCKED").get_data(as_text=True)

    assert "11.9000 €" in ok_page
    assert "COMP-1" in ok_page
    assert "Paskutinė pirkimo kaina" in ok_page
    assert "BOM struktūra ir komponentų kainos" in ok_page
    assert "Iš kur gauta kaina" in ok_page
    assert "Galutinės kainos formulė" in ok_page
    assert "Komponentų savikaina pagal BOM" in ok_page
    assert "Taikoma tik priedams. Komponentų savikaina nemažinama." in ok_page
    assert "Techniniai audito kodai" in ok_page
    assert "Ką reiškia R001–R007" in ok_page
    assert "Kategorijos darbų ir aptarnavimo tarifai" in ok_page
    assert 'href="#rule-R003"' in ok_page
    assert 'href="#rule-R004"' in ok_page
    assert "Atidaryti kategorijų konfigūraciją" in ok_page
    assert "FPACK, HRD, APACK ir Shelf tarifai" in ok_page
    assert '/pricing-rules#bom-categories' in ok_page
    assert "Missing component price" in blocked_page
    assert "R006" in blocked_page


def test_pricing_search_stops_after_matching_trace_block(monkeypatch, tmp_path):
    webapp = load_webapp(monkeypatch, tmp_path)
    path = tmp_path / "prices.xlsx"
    workbook = Workbook()
    results = workbook.active
    results.title = "PRICE RESULTS"
    results.append(["SKU", "Name", "Position Type", "Product Category", "Component / Purchase Cost", "Pricing Add-ons Total", "Adjustment Amount", "Final Reform SO Unit Price", "Control Status", "Applied Rule IDs", "Issues / Review Reason"])
    results.append(["A", "A", "BOM", "C", 1, 0, 0, 1, "CALCULATED", "R002", ""])
    trace = workbook.create_sheet("PRICE TRACE")
    trace.append(["SKU", "Step Type", "Input / Component / Rule"])
    trace.append(["A", "MATERIAL", "PART-A"])
    trace.append(["B", "MATERIAL", "PART-B"])
    workbook.save(path)

    match, rows, _ = webapp._read_pricing_workbook_match(path, "a")

    assert match["sku"] == "A"
    assert [row["Input / Component / Rule"] for row in rows] == ["PART-A"]


def test_furnix_profile_can_expose_only_selected_addon(monkeypatch, tmp_path):
    webapp = load_webapp(monkeypatch, tmp_path)
    monkeypatch.setenv("PRODUCT_ENGINE_BRAND", "Furnix")
    monkeypatch.setenv("PRODUCT_ENGINE_APP_NAME", "Furnix Product Engine")
    monkeypatch.setenv("PRODUCT_ENGINE_HERO_TITLE", "Furnix Odoo ataskaitos")
    monkeypatch.setenv("PRODUCT_ENGINE_ENABLED_ACTIONS", "stock_by_location")
    monkeypatch.setenv("PRODUCT_ENGINE_SHOW_BOM_WORKSPACE", "false")
    monkeypatch.setenv("PRODUCT_ENGINE_SHOW_PRICING_NAV", "false")
    webapp = importlib.reload(webapp)

    page = webapp.app.test_client().get("/").get_data(as_text=True)

    assert "Furnix <strong>Product Engine</strong>" in page
    assert "Furnix Odoo ataskaitos" in page
    assert "Generuoti SKU likučius pagal lokaciją" in page
    assert "Atnaujinti Reform kainodarą" not in page
    assert "Reform BOM įvestis" not in page
    assert "Pirkimo kainodara" not in page


def test_supply_result_separates_mo_and_catalog_statuses():
    template = (Path(__file__).parent / "webapp" / "templates" / "job.html").read_text(
        encoding="utf-8"
    )
    assert "MO–PO:" in template
    assert "Catalog–PO:" in template


def test_upload_accepts_xlsx_and_rejects_other_files(monkeypatch, tmp_path):
    webapp = load_webapp(monkeypatch, tmp_path)
    client = webapp.app.test_client()

    accepted = client.post(
        "/upload",
        data={"file": (io.BytesIO(b"xlsx"), "Reform BOM.xlsx")},
        content_type="multipart/form-data",
    )
    rejected = client.post(
        "/upload",
        data={"file": (io.BytesIO(b"csv"), "Reform BOM.csv")},
        content_type="multipart/form-data",
    )

    assert accepted.status_code == 302
    assert rejected.status_code == 400


def test_chunked_upload_persists_and_completes_atomically(monkeypatch, tmp_path):
    webapp = load_webapp(monkeypatch, tmp_path)
    client = webapp.app.test_client()
    content = b"large Reform workbook contents"

    started = client.post(
        "/upload/chunked", json={"filename": "Reform BOM.xlsx", "size": len(content)}
    )
    upload_id = started.get_json()["upload_id"]
    first = client.put(
        f"/upload/chunked/{upload_id}",
        data=content[:9],
        headers={"Upload-Offset": "0", "Content-Type": "application/octet-stream"},
    )
    second = client.put(
        f"/upload/chunked/{upload_id}",
        data=content[9:],
        headers={"Upload-Offset": "9", "Content-Type": "application/octet-stream"},
    )
    completed = client.post(
        f"/upload/chunked/{upload_id}/complete",
        json={"sha256": hashlib.sha256(content).hexdigest()},
    )

    assert started.status_code == 200
    assert first.get_json()["offset"] == 9
    assert second.get_json()["offset"] == len(content)
    assert completed.status_code == 200
    saved = webapp.UPLOAD_DIR / completed.get_json()["filename"]
    assert saved.read_bytes() == content
    assert saved.parent == tmp_path / "state" / "uploads"
    assert not list(webapp.CHUNK_UPLOAD_DIR.iterdir())


def test_chunked_upload_retries_are_idempotent_and_offsets_are_checked(monkeypatch, tmp_path):
    webapp = load_webapp(monkeypatch, tmp_path)
    client = webapp.app.test_client()
    upload_id = client.post(
        "/upload/chunked", json={"filename": "bom.xlsx", "size": 6}
    ).get_json()["upload_id"]
    headers = {"Upload-Offset": "0", "Content-Type": "application/octet-stream"}

    assert client.put(f"/upload/chunked/{upload_id}", data=b"abc", headers=headers).status_code == 200
    retry = client.put(f"/upload/chunked/{upload_id}", data=b"abc", headers=headers)
    gap = client.put(
        f"/upload/chunked/{upload_id}",
        data=b"z",
        headers={"Upload-Offset": "4", "Content-Type": "application/octet-stream"},
    )

    assert retry.get_json()["offset"] == 3
    assert gap.status_code == 409
    assert (webapp.CHUNK_UPLOAD_DIR / f"{upload_id}.part").read_bytes() == b"abc"


def test_chunked_upload_validates_file_and_completion(monkeypatch, tmp_path):
    webapp = load_webapp(monkeypatch, tmp_path)
    client = webapp.app.test_client()

    assert client.post("/upload/chunked", json={"filename": "bom.csv", "size": 3}).status_code == 400
    upload_id = client.post(
        "/upload/chunked", json={"filename": "bom.xlsx", "size": 3}
    ).get_json()["upload_id"]
    assert client.post(f"/upload/chunked/{upload_id}/complete", json={}).status_code == 409


def test_release_without_dataset_does_not_lock_queue(monkeypatch, tmp_path):
    webapp = load_webapp(monkeypatch, tmp_path)
    job_id = "missingdataset"
    job_dir = webapp.RUN_DIR / job_id
    job_dir.mkdir()
    webapp.write_job(
        job_dir,
        {
            "id": job_id,
            "action": "release_files",
            "title": "Generuoti BOM importo failus",
            "status": "QUEUED",
            "created_at": webapp.utc_now(),
            "upload": None,
            "files": [],
        },
    )
    webapp._reserved_jobs.add(job_id)

    webapp.run_job(job_id, "release_files", None)

    assert webapp.read_job(job_dir)["status"] == "ERROR"
    assert job_id not in webapp._reserved_jobs


def test_job_retention_keeps_latest_ten_per_action(monkeypatch, tmp_path):
    webapp = load_webapp(monkeypatch, tmp_path)
    for index in range(12):
        job_dir = webapp.RUN_DIR / f"job{index:02d}"
        job_dir.mkdir()
        webapp.write_job(job_dir, {
            "id": job_dir.name,
            "action": "audit",
            "title": "Audit",
            "status": "PASS",
            "created_at": f"2026-09-{index + 1:02d}T00:00:00+00:00",
            "files": [],
        })
    removed = webapp.prune_completed_jobs(keep_per_action=10)
    assert removed == ["job01", "job00"]
    assert len(list(webapp.RUN_DIR.iterdir())) == 10


def test_job_retention_never_deletes_running_or_latest_target(monkeypatch, tmp_path):
    webapp = load_webapp(monkeypatch, tmp_path)
    target_job = webapp.RUN_DIR / "target"
    target_job.mkdir()
    (target_job / "files").mkdir()
    (target_job / "files" / "Furnibox_Target_Dataset.json").write_text("{}")
    webapp.write_job(target_job, {
        "id": "target", "action": "pricing", "title": "Pricing",
        "status": "BLOCKED", "created_at": "2026-01-01T00:00:00+00:00", "files": [],
    })
    running = webapp.RUN_DIR / "running"
    running.mkdir()
    webapp.write_job(running, {
        "id": "running", "action": "pricing", "title": "Pricing",
        "status": "RUNNING", "created_at": "2025-01-01T00:00:00+00:00", "files": [],
    })
    for index in range(2):
        old = webapp.RUN_DIR / f"old{index}"
        old.mkdir()
        webapp.write_job(old, {
            "id": old.name, "action": "pricing", "title": "Pricing",
            "status": "FAIL", "created_at": f"2024-01-0{index + 1}T00:00:00+00:00", "files": [],
        })
    webapp.prune_completed_jobs(keep_per_action=0)
    assert target_job.exists()
    assert running.exists()
