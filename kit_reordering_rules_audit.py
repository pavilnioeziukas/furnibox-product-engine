"""Read-only Production audit of Reordering Rules for KIT BOM parent products."""

from __future__ import annotations

import argparse
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
import json
from pathlib import Path
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from odoo_client import OdooClient


SKU_COLUMN = "Product/Internal Reference"


class AuditError(RuntimeError):
    """Raised when the audit cannot be performed safely."""


@dataclass(frozen=True)
class AuditSettings:
    url: str
    db: str
    login: str
    api_key: str
    output_dir: Path
    log_dir: Path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Tik skaitymo režimu patikrina, kurie KIT BOM parent produktai "
            "Production aplinkoje turi Reordering Rules."
        )
    )
    parser.add_argument("--source", required=True, type=Path, help="lv1_KIT.xlsx failas")
    parser.add_argument("--output", required=True, type=Path, help="Excel ataskaita")
    parser.add_argument(
        "--audit",
        type=Path,
        help="JSON auditas (numatyta: toks pats kelias kaip --output, su .json)",
    )
    parser.add_argument(
        "--env-file",
        type=Path,
        default=Path(".env"),
        help="Production konfigūracijos failas (numatyta: .env)",
    )
    return parser.parse_args()


def load_settings_from_file(env_file: Path) -> AuditSettings:
    if not env_file.exists():
        raise AuditError(f"Nerastas konfigūracijos failas: {env_file}")
    values: dict[str, str] = {}
    for raw_line in env_file.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip("\"'")
    required = {
        "url": str(values.get("ODOO_URL") or "").strip().rstrip("/"),
        "db": str(values.get("ODOO_DB") or "").strip(),
        "login": str(values.get("ODOO_LOGIN") or "").strip(),
        "api_key": str(values.get("ODOO_API_KEY") or "").strip(),
    }
    missing = [name for name, value in required.items() if not value]
    if missing:
        raise AuditError("Neužpildyti .env laukai: " + ", ".join(missing))
    url = required["url"].lower()
    if "stage" in url or url != "https://odoo.furnibox.lt":
        raise AuditError(
            "Diagnostika skirta tik Production https://odoo.furnibox.lt; "
            f"gauta: {required['url']}"
        )
    output_dir = Path.cwd() / "output" / "production"
    return AuditSettings(
        **required,
        output_dir=output_dir,
        log_dir=Path.cwd() / "logs",
    )


def read_kit_skus(path: Path) -> list[str]:
    if not path.exists():
        raise AuditError(f"Nerastas KIT failas: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header_row = next(sheet.iter_rows(values_only=True), None)
        if not header_row:
            raise AuditError("KIT failas tuščias.")
        headers = list(header_row)
        try:
            sku_index = headers.index(SKU_COLUMN)
        except ValueError as exc:
            raise AuditError(
                f"Faile nėra stulpelio „{SKU_COLUMN}“. Rasti: {headers}"
            ) from exc

        skus: list[str] = []
        seen: set[str] = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            raw = row[sku_index] if sku_index < len(row) else None
            sku = str(raw).strip() if raw is not None else ""
            if sku and sku not in seen:
                seen.add(sku)
                skus.append(sku)
        if not skus:
            raise AuditError("KIT faile nerasta nė vieno parent Internal Reference.")
        return skus
    finally:
        workbook.close()


def chunks(values: list, size: int = 100):
    for start in range(0, len(values), size):
        yield values[start : start + size]


def relation_name(value) -> str:
    return str(value[1]) if isinstance(value, (list, tuple)) and len(value) > 1 else ""


def relation_id(value):
    return value[0] if isinstance(value, (list, tuple)) and value else None


def fetch_products(client: OdooClient, skus: list[str]) -> list[dict]:
    rows: list[dict] = []
    for batch in chunks(skus):
        rows.extend(
            client.search_read_all(
                "product.product",
                [["default_code", "in", batch]],
                ["id", "default_code", "name", "active", "product_tmpl_id", "categ_id"],
                context={"active_test": False},
            )
        )
    return rows


def fetch_orderpoints(client: OdooClient, product_ids: list[int]) -> list[dict]:
    fields = [
        "id",
        "active",
        "product_id",
        "warehouse_id",
        "location_id",
        "company_id",
        "product_min_qty",
        "product_max_qty",
        "qty_multiple",
        "trigger",
        "route_id",
    ]
    rows: list[dict] = []
    for batch in chunks(product_ids):
        rows.extend(
            client.search_read_all(
                "stock.warehouse.orderpoint",
                [["product_id", "in", batch]],
                fields,
                context={"active_test": False},
            )
        )
    return rows


def style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 55)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = width


def analyze(skus: list[str], products: list[dict], orderpoints: list[dict]) -> dict:
    products_by_sku: dict[str, list[dict]] = {}
    products_by_id: dict[int, dict] = {}
    for product in products:
        sku = str(product.get("default_code") or "").strip()
        products_by_sku.setdefault(sku, []).append(product)
        products_by_id[product["id"]] = product

    active_rules = [rule for rule in orderpoints if rule.get("active", True)]
    active_counts = Counter(relation_id(rule.get("product_id")) for rule in active_rules)
    all_counts = Counter(relation_id(rule.get("product_id")) for rule in orderpoints)
    missing_skus = [sku for sku in skus if sku not in products_by_sku]
    duplicate_skus = {
        sku: rows for sku, rows in products_by_sku.items() if len(rows) > 1
    }
    blocker_skus = sorted(
        {
            product.get("default_code")
            for product in products
            if active_counts.get(product["id"], 0)
        }
    )
    return {
        "products_by_sku": products_by_sku,
        "products_by_id": products_by_id,
        "active_counts": active_counts,
        "all_counts": all_counts,
        "active_rules": active_rules,
        "missing_skus": missing_skus,
        "duplicate_skus": duplicate_skus,
        "blocker_skus": blocker_skus,
    }


def rule_row(rule: dict, products_by_id: dict[int, dict]) -> list:
    product_id = relation_id(rule.get("product_id"))
    product = products_by_id.get(product_id, {})
    return [
        product.get("default_code", ""),
        product.get("name", relation_name(rule.get("product_id"))),
        product_id,
        product.get("active", ""),
        relation_name(product.get("categ_id")),
        rule.get("id"),
        rule.get("active", ""),
        relation_name(rule.get("warehouse_id")),
        relation_name(rule.get("location_id")),
        rule.get("product_min_qty"),
        rule.get("product_max_qty"),
        rule.get("qty_multiple"),
        rule.get("trigger"),
        relation_name(rule.get("route_id")),
        relation_name(rule.get("company_id")),
    ]


RULE_HEADERS = [
    "SKU",
    "Produkto pavadinimas",
    "Produkto ID",
    "Produkto aktyvumas",
    "Kategorija",
    "Rule ID",
    "Rule aktyvumas",
    "Sandėlis",
    "Lokacija",
    "Min",
    "Max",
    "Multiple",
    "Trigger",
    "Route",
    "Įmonė",
]


def build_report(
    report_path: Path,
    source_path: Path,
    production_url: str,
    skus: list[str],
    products: list[dict],
    orderpoints: list[dict],
) -> dict:
    result = analyze(skus, products, orderpoints)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "SUMMARY"
    summary.append(["Rodiklis", "Reikšmė"])
    for row in [
        ("Statusas", "BLOCKED" if result["blocker_skus"] else "PASS"),
        ("Patikros laikas", datetime.now().isoformat(timespec="seconds")),
        ("Aplinka", production_url),
        ("Šaltinio failas", str(source_path)),
        ("Unikalūs KIT parent SKU", len(skus)),
        ("Odoo rasti produktų variantai", len(products)),
        ("Importą blokuojantys SKU", len(result["blocker_skus"])),
        ("Aktyvios Reordering Rules", len(result["active_rules"])),
        ("Visos Rules, įskaitant archyvuotas", len(orderpoints)),
        ("Odoo nerasti KIT SKU", len(result["missing_skus"])),
        ("Dubliuoti SKU Odoo", len(result["duplicate_skus"])),
        ("Pastaba", "Tik skaitymo diagnostika. Odoo pakeitimai: 0."),
    ]:
        summary.append(row)

    blockers = workbook.create_sheet("IMPORT BLOCKERS")
    blockers.append(RULE_HEADERS)
    for rule in sorted(
        result["active_rules"],
        key=lambda row: (relation_name(row.get("product_id")), row.get("id", 0)),
    ):
        blockers.append(rule_row(rule, result["products_by_id"]))

    all_rules = workbook.create_sheet("ALL RULES")
    all_rules.append(RULE_HEADERS)
    for rule in sorted(
        orderpoints,
        key=lambda row: (relation_name(row.get("product_id")), row.get("id", 0)),
    ):
        all_rules.append(rule_row(rule, result["products_by_id"]))

    no_rules = workbook.create_sheet("KIT WITHOUT ACTIVE RULE")
    no_rules.append(["SKU", "Produkto ID", "Produkto pavadinimas", "Aktyvus"])
    for sku in skus:
        rows = result["products_by_sku"].get(sku, [])
        if not rows:
            no_rules.append([sku, "", "NERASTAS ODOO", ""])
        elif not any(result["active_counts"].get(product["id"], 0) for product in rows):
            for product in rows:
                no_rules.append([sku, product["id"], product["name"], product["active"]])

    duplicates = workbook.create_sheet("DUPLICATE SKU")
    duplicates.append(["SKU", "Produkto ID", "Pavadinimas", "Aktyvus", "Template ID"])
    for sku, rows in sorted(result["duplicate_skus"].items()):
        for product in rows:
            duplicates.append(
                [
                    sku,
                    product["id"],
                    product["name"],
                    product["active"],
                    relation_id(product["product_tmpl_id"]),
                ]
            )

    for sheet in workbook.worksheets:
        style_sheet(sheet)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(report_path)

    return {
        "status": "BLOCKED" if result["blocker_skus"] else "PASS",
        "production_url": production_url,
        "source": str(source_path),
        "kit_parent_skus": len(skus),
        "products_found": len(products),
        "blocker_skus": result["blocker_skus"],
        "active_reordering_rules": len(result["active_rules"]),
        "all_reordering_rules": len(orderpoints),
        "missing_skus": result["missing_skus"],
        "duplicate_skus": sorted(result["duplicate_skus"]),
        "odoo_changes": 0,
    }


def main() -> int:
    try:
        args = parse_args()
        settings = load_settings_from_file(args.env_file)
        skus = read_kit_skus(args.source)

        client = OdooClient(settings)
        client.authenticate()
        products = fetch_products(client, skus)
        orderpoints = fetch_orderpoints(client, [row["id"] for row in products])

        result = build_report(
            args.output,
            args.source,
            settings.url,
            skus,
            products,
            orderpoints,
        )
        audit_path = args.audit or args.output.with_suffix(".json")
        audit_path.parent.mkdir(parents=True, exist_ok=True)
        audit_path.write_text(
            json.dumps(result, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        print("=" * 80)
        print("KIT REORDERING RULES AUDIT")
        print("=" * 80)
        print(f"Statusas: {result['status']}")
        print(f"Production URL: {settings.url}")
        print(f"KIT parent SKU: {result['kit_parent_skus']}")
        print(f"Importą blokuojantys SKU: {len(result['blocker_skus'])}")
        print(f"Aktyvios Reordering Rules: {result['active_reordering_rules']}")
        print(f"Visos Rules: {result['all_reordering_rules']}")
        print(f"Odoo nerasti SKU: {len(result['missing_skus'])}")
        print(f"Dubliuoti SKU: {len(result['duplicate_skus'])}")
        print(f"Ataskaita: {args.output}")
        print(f"Auditas: {audit_path}")
        print("Odoo pakeitimai: 0")
        return 0
    except Exception as exc:
        print(f"KLAIDA: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
