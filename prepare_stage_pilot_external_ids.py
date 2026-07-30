from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from bom_release.generator import IMPORT_HEADERS, canon


class StageExternalIdError(RuntimeError):
    """Piloto External ID negalima saugiai persieti su Stage."""


def _external_id_name(row: dict[str, Any]) -> str:
    module = str(row.get("module") or "").strip()
    name = str(row.get("name") or "").strip()
    return f"{module}.{name}" if module and name else ""


def _preferred_external_ids(
    client: Any,
    model: str,
    record_ids: set[int],
) -> dict[int, str]:
    if not record_ids:
        return {}
    rows = client.search_read_all(
        "ir.model.data",
        [["model", "=", model], ["res_id", "in", sorted(record_ids)]],
        ["module", "name", "res_id"],
    )
    grouped: dict[int, list[str]] = defaultdict(list)
    for row in rows:
        external_id = _external_id_name(row)
        if external_id:
            grouped[int(row["res_id"])].append(external_id)
    result: dict[int, str] = {}
    for record_id, values in grouped.items():
        result[record_id] = sorted(
            set(values),
            key=lambda value: (not value.startswith("__export__."), value),
        )[0]
    return result


def _read_pilot_skus(path: Path) -> tuple[list[str], list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if workbook.sheetnames != ["BOM import"]:
            raise StageExternalIdError(
                f"{path.name}: tikėtasi vieno lapo 'BOM import'."
            )
        rows = workbook["BOM import"].iter_rows(values_only=True)
        header = list(next(rows, ()))
        if header != IMPORT_HEADERS:
            raise StageExternalIdError(
                f"{path.name}: neteisingi importo stulpeliai."
            )
        parents: list[str] = []
        components: list[str] = []
        for row_number, raw in enumerate(rows, start=2):
            row = list(raw[: len(IMPORT_HEADERS)])
            if not any(value not in (None, "") for value in row):
                continue
            if any(
                isinstance(value, str) and value.startswith("=")
                for value in row
            ):
                raise StageExternalIdError(
                    f"{path.name}:{row_number}: formulės neleidžiamos."
                )
            parent = canon(row[0])
            component = canon(row[3])
            if parent:
                parents.append(parent)
            if component:
                components.append(component)
        if not parents:
            raise StageExternalIdError("Piloto faile nėra BOM produktų.")
        if not components:
            raise StageExternalIdError("Piloto faile nėra komponentų.")
        duplicates = sorted(
            sku for sku in set(parents) if parents.count(sku) > 1
        )
        if duplicates:
            raise StageExternalIdError(
                "Piloto faile dubliuojasi parent SKU: "
                + ", ".join(duplicates[:10])
            )
        return sorted(set(parents)), sorted(set(components))
    finally:
        workbook.close()


def _resolve_stage_ids(
    client: Any,
    parent_skus: list[str],
    component_skus: list[str],
) -> tuple[dict[str, str], dict[str, str]]:
    wanted = set(parent_skus) | set(component_skus)
    rows = client.search_read_all(
        "product.product",
        [["default_code", "in", sorted(wanted)]],
        ["id", "default_code", "product_tmpl_id", "active"],
        context={"active_test": False},
    )
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        sku = canon(row.get("default_code"))
        if sku in wanted:
            grouped[sku].append(row)

    missing = sorted(wanted - set(grouped))
    ambiguous = sorted(
        sku
        for sku, matches in grouped.items()
        if len({int(row["id"]) for row in matches}) != 1
        or len(
            {
                int(row["product_tmpl_id"][0])
                for row in matches
                if row.get("product_tmpl_id")
            }
        )
        != 1
    )
    if missing or ambiguous:
        details = []
        if missing:
            details.append("Stage nerasti SKU: " + ", ".join(missing[:10]))
        if ambiguous:
            details.append(
                "Stage neunikalūs SKU: " + ", ".join(ambiguous[:10])
            )
        raise StageExternalIdError("; ".join(details))

    unique = {sku: matches[0] for sku, matches in grouped.items()}
    template_ids = {
        int(unique[sku]["product_tmpl_id"][0]) for sku in parent_skus
    }
    product_ids = {int(unique[sku]["id"]) for sku in component_skus}
    template_external = _preferred_external_ids(
        client, "product.template", template_ids
    )
    product_external = _preferred_external_ids(
        client, "product.product", product_ids
    )
    parent_map = {
        sku: template_external.get(
            int(unique[sku]["product_tmpl_id"][0]), ""
        )
        for sku in parent_skus
    }
    component_map = {
        sku: product_external.get(int(unique[sku]["id"]), "")
        for sku in component_skus
    }
    missing_parent_ids = sorted(sku for sku, value in parent_map.items() if not value)
    missing_component_ids = sorted(
        sku for sku, value in component_map.items() if not value
    )
    if missing_parent_ids or missing_component_ids:
        details = []
        if missing_parent_ids:
            details.append(
                "Stage parent neturi product.template External ID: "
                + ", ".join(missing_parent_ids[:10])
            )
        if missing_component_ids:
            details.append(
                "Stage komponentas neturi product.product External ID: "
                + ", ".join(missing_component_ids[:10])
            )
        raise StageExternalIdError("; ".join(details))
    return parent_map, component_map


def prepare_stage_pilot(
    client: Any,
    source: Path,
    output: Path,
    audit_path: Path | None = None,
) -> dict[str, Any]:
    parent_skus, component_skus = _read_pilot_skus(source)
    parent_map, component_map = _resolve_stage_ids(
        client, parent_skus, component_skus
    )
    workbook = load_workbook(source, data_only=False)
    try:
        sheet = workbook["BOM import"]
        parent_changes = 0
        component_changes = 0
        operation_type_values: set[str] = set()
        for row_number in range(2, sheet.max_row + 1):
            parent = canon(sheet.cell(row_number, 1).value)
            component = canon(sheet.cell(row_number, 4).value)
            if parent:
                cell = sheet.cell(row_number, 2)
                parent_changes += int(cell.value != parent_map[parent])
                cell.value = parent_map[parent]
            if component:
                cell = sheet.cell(row_number, 5)
                component_changes += int(
                    cell.value != component_map[component]
                )
                cell.value = component_map[component]
            operation_type = str(sheet.cell(row_number, 9).value or "").strip()
            if operation_type:
                operation_type_values.add(operation_type)
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output)
    finally:
        workbook.close()

    audit = {
        "status": "PASS",
        "source": str(source.resolve()),
        "output": str(output.resolve()),
        "parent_skus": parent_skus,
        "component_skus": component_skus,
        "parent_external_ids_changed": parent_changes,
        "component_external_ids_changed": component_changes,
        "operation_type_external_ids_unchanged": sorted(operation_type_values),
        "odoo_changes": 0,
    }
    if audit_path:
        audit_path.parent.mkdir(parents=True, exist_ok=True)
        audit_path.write_text(
            json.dumps(audit, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    return audit


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Persieja BOM piloto produktų External ID su Stage pagal unikalų "
            "Internal Reference. Odoo tik skaito ir nekeičia."
        )
    )
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--audit", type=Path)
    parser.add_argument("--env-file", type=Path, default=Path(".env.stage"))
    return parser.parse_args()


def main() -> None:
    from dotenv import load_dotenv
    from config import load_settings
    from odoo_client import OdooClient

    args = parse_args()
    env_file = args.env_file.resolve()
    if not env_file.is_file():
        raise FileNotFoundError(f"Nerastas Stage env failas: {env_file}")
    load_dotenv(env_file, override=True)
    settings = load_settings()
    if "stage" not in settings.url.lower():
        raise PermissionError(
            f"BLOCKED: URL nėra Stage aplinka: {settings.url}"
        )
    client = OdooClient(settings)
    client.authenticate()
    audit = prepare_stage_pilot(
        client, args.source, args.output, args.audit
    )
    print("=" * 80)
    print("BOM RELEASE STAGE EXTERNAL ID REMAP")
    print("=" * 80)
    print("Statusas:", audit["status"])
    print("Stage URL:", settings.url)
    print("BOM produktai:", len(audit["parent_skus"]))
    print("Komponentai:", len(audit["component_skus"]))
    print(
        "Pakeisti parent External ID:",
        audit["parent_external_ids_changed"],
    )
    print(
        "Pakeisti komponentų External ID:",
        audit["component_external_ids_changed"],
    )
    print("Failas:", args.output.resolve())
    if args.audit:
        print("Auditas:", args.audit.resolve())
    print("Odoo pakeitimai: 0")


if __name__ == "__main__":
    main()
