from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from bom_release.generator import generate_release_files
from bom_release.import_validator import validate_release_imports
from test_bom_release_generator import FakeClient, dataset, plan


class BomReleaseImportValidatorTests(unittest.TestCase):
    def test_generated_release_round_trip_matches_dataset(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory)
            generate_release_files(
                dataset=dataset(),
                plan=plan(),
                client=FakeClient(),
                output_dir=output,
            )
            result = validate_release_imports(
                dataset=dataset(),
                release_id="RELEASE-1",
                release_reference="RELEASE-1",
                import_dir=output,
            )
            self.assertTrue(result.passed, result.errors)
            self.assertEqual(result.actual_boms, 1)
            self.assertEqual(result.component_rows, 1)
            self.assertEqual(result.operation_rows, 1)

    def test_changed_component_quantity_fails(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory)
            files = generate_release_files(
                dataset=dataset(),
                plan=plan(),
                client=FakeClient(),
                output_dir=output,
            )
            release_file = next(item.path for item in files if item.level > 0)
            workbook = load_workbook(release_file)
            workbook["BOM import"]["F2"] = 99
            workbook.save(release_file)
            workbook.close()

            result = validate_release_imports(
                dataset=dataset(),
                release_id="RELEASE-1",
                release_reference="RELEASE-1",
                import_dir=output,
            )
            self.assertFalse(result.passed)
            self.assertTrue(
                any("kiekis 99" in error for error in result.errors),
                result.errors,
            )

    def test_changed_new_bom_sequence_fails(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory)
            files = generate_release_files(
                dataset=dataset(),
                plan=plan(),
                client=FakeClient(),
                output_dir=output,
            )
            release_file = next(item.path for item in files if item.level > 0)
            workbook = load_workbook(release_file)
            workbook["BOM import"]["H2"] = 0
            workbook.save(release_file)
            workbook.close()

            result = validate_release_imports(
                dataset=dataset(),
                release_id="RELEASE-1",
                release_reference="RELEASE-1",
                import_dir=output,
            )
            self.assertFalse(result.passed)
            self.assertTrue(
                any("naujo BOM sequence 0" in error for error in result.errors),
                result.errors,
            )

    def test_changed_operation_type_external_id_fails(self):
        from openpyxl import load_workbook

        source = dataset()
        source["products"][0]["product_type"] = "CABINET HARDWARE"
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory)
            files = generate_release_files(
                dataset=source,
                plan=plan(),
                client=FakeClient(),
                output_dir=output,
            )
            workbook = load_workbook(files[0].path)
            workbook["BOM import"]["I2"] = "wrong.operation_type"
            workbook.save(files[0].path)
            workbook.close()

            result = validate_release_imports(
                dataset=source,
                release_id="RELEASE-1",
                release_reference="RELEASE-1",
                import_dir=output,
            )
            self.assertFalse(result.passed)
            self.assertTrue(
                any(
                    "Operation Type/External ID 'wrong.operation_type'"
                    in error
                    for error in result.errors
                ),
                result.errors,
            )

    def test_manufacture_dataset_without_operations_fails(self):
        invalid = dataset()
        invalid["products"][0]["operations"] = []
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory)
            generate_release_files(
                dataset=dataset(),
                plan=plan(),
                client=FakeClient(),
                output_dir=output,
            )
            result = validate_release_imports(
                dataset=invalid,
                release_id="RELEASE-1",
                release_reference="RELEASE-1",
                import_dir=output,
            )
            self.assertFalse(result.passed)
            self.assertTrue(
                any(
                    "Manufacture BOM neturi operacijų" in error
                    for error in result.errors
                ),
                result.errors,
            )

    def test_legacy_manufacture_cabinet_shelf_without_operations_is_blocked(self):
        cabinet_shelf = dataset()
        cabinet_shelf["products"][0]["sku"] = "EUB-C-CAB03-SLF901"
        cabinet_shelf["products"][0]["reform_category"] = "CABINET SHELF"
        cabinet_shelf["products"][0]["operations"] = []
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory)
            result = validate_release_imports(
                dataset=cabinet_shelf,
                release_id="RELEASE-1",
                release_reference="RELEASE-1",
                import_dir=output,
            )
            self.assertTrue(
                any(
                    "Manufacture BOM neturi operacijų" in error
                    for error in result.errors
                ),
                result.errors,
            )


if __name__ == "__main__":
    unittest.main()
