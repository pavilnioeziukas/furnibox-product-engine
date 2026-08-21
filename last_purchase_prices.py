"""Eksportuoja paskutinę patvirtintą kiekvieno produkto pirkimo kainą iš Odoo."""

from __future__ import annotations

import logging
import os
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from config import load_settings
from excel_writer import HEADER_FILL, HEADER_FONT
from odoo_client import OdooClient
from purchase_price_adjustments import load_adjustments


BASE_DIR = Path(__file__).resolve().parent
SHARED_DATA_DIR = Path(
    os.getenv(
        "FURNIBOX_SHARED_DATA",
        BASE_DIR / "web_state" / "shared_data",
    )
).resolve()
PURCHASE_PRICE_ADJUSTMENTS_PATH = SHARED_DATA_DIR / "purchase_price_adjustments.json"
REFORM_VENDOR_NAME = "Reform Supply & Logistics, UAB"
REFORM_VENDOR_MARKUP_FACTOR = 1.05

EXPORT_COLUMNS = [
    "Internal Reference",
    "Name",
    "Product Category/Name",
    "Purchase Order",
    "Vendor",
    "Ordered Quantity",
    "Last Purchase Price",
    "Order Date",
]

CALCULATOR_COLUMNS = [
    "Adjusted Purchase Price",
    "Markup Factor",
    "Reform Price",
]

ADJUSTMENT_SHEET = "PURCHASE PRICE ADJUSTMENTS"

OUTPUT_COLUMNS = [
    "Real Purchase Price" if column == "Last Purchase Price" else column
    for column in EXPORT_COLUMNS
] + CALCULATOR_COLUMNS


def relation_id(value):
    return value[0] if isinstance(value, list) and value else None


def relation_name(value):
    return value[1] if isinstance(value, list) and len(value) >= 2 else ""


def reform_markup_factor(vendor):
    if str(vendor or "").strip().casefold() == REFORM_VENDOR_NAME.casefold():
        return REFORM_VENDOR_MARKUP_FACTOR
    return 1.0


def build_last_purchase_prices(purchase_lines, products):
    """Sujungia naujausią PO eilutę su tikrais product.product laukais."""
    products_by_id = {row["id"]: row for row in products}
    latest_lines = {}

    for line in purchase_lines:
        product_id = relation_id(line.get("product_id"))
        if product_id is None:
            continue

        candidate_key = (line.get("date_order") or "", line.get("id") or 0)
        current = latest_lines.get(product_id)
        if current is not None:
            current_key = (current.get("date_order") or "", current.get("id") or 0)
            if candidate_key <= current_key:
                continue
        latest_lines[product_id] = line

    missing_product_ids = sorted(set(latest_lines) - set(products_by_id))
    if missing_product_ids:
        preview = ", ".join(map(str, missing_product_ids[:10]))
        suffix = "..." if len(missing_product_ids) > 10 else ""
        raise RuntimeError(
            "Odoo negrąžino product.product įrašų šiems ID: "
            f"{preview}{suffix}"
        )

    rows = []
    for product_id in sorted(latest_lines):
        line = latest_lines[product_id]
        product = products_by_id[product_id]
        rows.append({
            "Internal Reference": product.get("default_code") or "",
            "Name": product.get("name") or "",
            "Product Category/Name": relation_name(product.get("categ_id")),
            "Purchase Order": relation_name(line.get("order_id")),
            "Vendor": relation_name(line.get("partner_id")),
            "Ordered Quantity": line.get("product_qty"),
            "Last Purchase Price": line.get("price_unit"),
            "Order Date": line.get("date_order") or "",
        })

    return rows


def load_last_purchase_prices(client):
    """Nuskaito PO eilutes ir atskirus produktų laukus iš Odoo."""
    purchase_lines = client.purchase_order_lines()
    product_ids = {
        relation_id(line.get("product_id"))
        for line in purchase_lines
        if relation_id(line.get("product_id")) is not None
    }
    products = client.search_read_all(
        "product.product",
        [["id", "in", sorted(product_ids)]],
        ["id", "default_code", "name", "categ_id"],
        context={"active_test": False},
    ) if product_ids else []
    return build_last_purchase_prices(purchase_lines, products)


def write_purchase_prices(
    path,
    purchase_prices,
    metadata,
    purchase_price_adjustments=None,
):
    """Sukuria komponentų pirkimo ir Reform kainų kontrolinį failą."""
    if purchase_price_adjustments is None:
        purchase_price_adjustments = load_adjustments(PURCHASE_PRICE_ADJUSTMENTS_PATH)

    workbook = Workbook()

    info = workbook.active
    info.title = "INFO"
    for row in [
        ("Generated", datetime.now().isoformat(sep=" ", timespec="seconds")),
        ("Odoo URL", metadata["url"]),
        ("Database", metadata["db"]),
        ("User", metadata["login"]),
        ("Odoo UID", metadata["uid"]),
        ("Products with Last Purchase Price", len(purchase_prices)),
        ("Purchase price adjustments source", str(PURCHASE_PRICE_ADJUSTMENTS_PATH)),
        ("Purchase price adjustments loaded", len(purchase_price_adjustments)),
    ]:
        info.append(row)
    info.column_dimensions["A"].width = 36
    info.column_dimensions["B"].width = 80

    adjustments = workbook.create_sheet(ADJUSTMENT_SHEET)
    adjustments.append([
        "Internal Reference",
        "Adjusted Purchase Price",
        "Real Purchase Price (reference)",
        "Comment",
    ])

    applied_adjustments = 0

    for row in purchase_prices:
        sku = str(row.get("Internal Reference") or "").strip()
        real_price = row.get("Last Purchase Price")
        adjustment = purchase_price_adjustments.get(sku)

        if adjustment:
            adjusted_price = adjustment["adjusted_purchase_price"]
            comment = adjustment.get("comment", "")
            applied_adjustments += 1
        else:
            adjusted_price = real_price
            comment = ""

        adjustments.append([
            sku,
            adjusted_price,
            real_price,
            comment,
        ])

    for cell in adjustments[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    adjustments.freeze_panes = "A2"
    adjustments.auto_filter.ref = adjustments.dimensions
    adjustments.column_dimensions["A"].width = 24
    adjustments.column_dimensions["B"].width = 25
    adjustments.column_dimensions["C"].width = 29
    adjustments.column_dimensions["D"].width = 55

    review_fill = PatternFill("solid", fgColor="E2F0D9")
    for row_number in range(2, adjustments.max_row + 1):
        adjustments.cell(row_number, 2).number_format = '0.0000 [$€-x-euro2]'
        adjustments.cell(row_number, 3).number_format = '0.0000 [$€-x-euro2]'
        adjustments.cell(row_number, 2).fill = review_fill

    prices = workbook.create_sheet("COMPONENT PRICES")
    invalid_columns = [
        key
        for row in purchase_prices
        for key in row
        if key not in EXPORT_COLUMNS
    ]
    if invalid_columns:
        raise RuntimeError(
            "Eksporte aptikti neleistini stulpeliai: "
            + ", ".join(sorted(set(invalid_columns)))
        )

    prices.append(OUTPUT_COLUMNS)
    for row_number, row in enumerate(purchase_prices, start=2):
        prices.append(
            [row.get(column, "") for column in EXPORT_COLUMNS]
            + [None, None, None]
        )
        prices.cell(row_number, 9).value = f"='{ADJUSTMENT_SHEET}'!B{row_number}"
        prices.cell(row_number, 10).value = reform_markup_factor(
            row.get("Vendor")
        )
        prices.cell(row_number, 11).value = (
            f'=IF(I{row_number}="",G{row_number},I{row_number})*J{row_number}'
        )

    for cell in prices[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        alignment = copy(cell.alignment)
        alignment.horizontal = "center"
        alignment.vertical = "center"
        alignment.wrap_text = True
        cell.alignment = alignment

    prices.freeze_panes = "A2"
    prices.auto_filter.ref = prices.dimensions

    column_widths = [22, 45, 38, 20, 35, 18, 19, 21, 25, 16, 18]
    for index, width in enumerate(column_widths, start=1):
        column_letter = prices.cell(row=1, column=index).column_letter
        prices.column_dimensions[column_letter].width = width

    formula_fill = PatternFill("solid", fgColor="E2F0D9")
    for row_number in range(2, prices.max_row + 1):
        prices.cell(row_number, 7).number_format = '0.0000 [$€-x-euro2]'
        prices.cell(row_number, 9).number_format = '0.0000 [$€-x-euro2]'
        prices.cell(row_number, 10).number_format = "0.00"
        prices.cell(row_number, 11).number_format = '0.0000 [$€-x-euro2]'
        prices.cell(row_number, 9).fill = formula_fill
        prices.cell(row_number, 10).fill = formula_fill
        prices.cell(row_number, 11).fill = formula_fill

    note = workbook.create_sheet("PRICING RULES", 1)
    note.append(["Field", "Rule"])
    note.append([
        "Real Purchase Price",
        "Last confirmed purchase price from Odoo.",
    ])
    note.append([
        "Adjusted Purchase Price",
        "Product Engine purchase price adjustment when configured; otherwise Real Purchase Price.",
    ])
    note.append([
        "Purchase price adjustment source",
        str(PURCHASE_PRICE_ADJUSTMENTS_PATH),
    ])
    note.append([
        "Markup Factor",
        f"{REFORM_VENDOR_MARKUP_FACTOR:.2f} only when Vendor is exactly "
        f"'{REFORM_VENDOR_NAME}'; otherwise 1.00.",
    ])
    note.append([
        "Reform Price",
        "Adjusted Purchase Price multiplied by Markup Factor.",
    ])
    note.append([
        "Non-BOM pricing",
        "Preparation, storage, bag and sticker are applied later by SO pricing rules.",
    ])

    for cell in note[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    note.column_dimensions["A"].width = 28
    note.column_dimensions["B"].width = 100
    note.freeze_panes = "A2"

    info.append(("Purchase price adjustments applied", applied_adjustments))
    info.append(("Reform vendor", REFORM_VENDOR_NAME))
    info.append(("Reform vendor markup factor", REFORM_VENDOR_MARKUP_FACTOR))

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    workbook.save(path)


def main():
    settings = load_settings()
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[
            logging.FileHandler(
                settings.log_dir / "last_purchase_prices.log",
                encoding="utf-8",
            ),
            logging.StreamHandler(),
        ],
    )

    client = OdooClient(settings)

    logging.info("Jungiamasi prie Odoo...")
    uid = client.authenticate()
    logging.info("Prisijungta. UID=%s", uid)

    logging.info("Nuskaitomos patvirtintų ir užbaigtų pirkimų eilutės...")
    purchase_prices = load_last_purchase_prices(client)
    logging.info("Produktų su paskutine pirkimo kaina: %s", len(purchase_prices))

    purchase_price_adjustments = load_adjustments(PURCHASE_PRICE_ADJUSTMENTS_PATH)
    logging.info(
        "Tamaros korekcijų saugykloje: %s",
        len(purchase_price_adjustments),
    )

    output_path = settings.output_dir / "Last_Purchase_Prices.xlsx"
    write_purchase_prices(
        output_path,
        purchase_prices,
        {
            "url": settings.url,
            "db": settings.db,
            "login": settings.login,
            "uid": uid,
        },
        purchase_price_adjustments=purchase_price_adjustments,
    )

    print()
    print("KOMPONENTŲ PIRKIMO IR REFORM KAINŲ FAILAS SUKURTAS")
    print("Failas:", output_path)
    print("Produktai su kaina:", len(purchase_prices))
    print("Tamaros korekcijų saugykloje:", len(purchase_price_adjustments))


if __name__ == "__main__":
    main()
