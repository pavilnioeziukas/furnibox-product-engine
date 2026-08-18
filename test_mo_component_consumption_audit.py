from dataclasses import asdict

from openpyxl import load_workbook

from mo_component_consumption_audit import ConsumptionGap, _write_xlsx, detect_gaps


PRODUCTS = {
    10: {"default_code": "FIN", "display_name": "Finished"},
    20: {"default_code": "COMP-A", "display_name": "Component A"},
    21: {"default_code": "COMP-B", "display_name": "Component B"},
}


def production(move_ids=(100, 101)):
    return {
        "id": 1, "name": "WH/MO/34218", "state": "done",
        "product_id": [10, "Finished"], "qty_produced": 5,
        "bom_id": [30, "BOM"], "date_finished": "2025-07-18 11:20:17",
        "company_id": [1, "Company"], "move_raw_ids": list(move_ids),
    }


def test_detects_zero_and_partial_consumption_but_not_fully_consumed():
    moves = [
        {"id": 100, "product_id": [20, "Component A"], "product_uom_qty": 10, "quantity": 0, "state": "cancel", "product_uom": [1, "Units"]},
        {"id": 101, "product_id": [21, "Component B"], "product_uom_qty": 5, "quantity": 3, "state": "done", "product_uom": [1, "Units"]},
        {"id": 102, "product_id": [20, "Component A"], "product_uom_qty": 7, "quantity": 7, "state": "done", "product_uom": [1, "Units"]},
    ]
    rows = detect_gaps([production((100, 101, 102))], moves, PRODUCTS)
    assert [(row.component_sku, row.missing_qty) for row in rows] == [("COMP-A", 10), ("COMP-B", 2)]
    assert all(row.mo == "WH/MO/34218" for row in rows)


def test_ignores_fully_consumed_and_zero_plan_moves():
    moves = [
        {"id": 100, "product_id": [20, "Component A"], "product_uom_qty": 10, "quantity": 10, "state": "done"},
        {"id": 101, "product_id": [21, "Component B"], "product_uom_qty": 0, "quantity": 0, "state": "cancel"},
    ]
    assert detect_gaps([production()], moves, PRODUCTS) == []


def test_writes_formatted_excel_result(tmp_path):
    row = ConsumptionGap(
        mo_id=34218, mo="WH/MO/34218", finished_product_sku="UNI-P-ACC02-MIS005",
        finished_product="ACCESSORIES - Interior", produced_qty=5, bom="BOM",
        completion_date="2025-07-18 08:20:17", company="FurniBox LT, UAB",
        component_sku="2601189846", component="ARENA BASKET", planned_qty=10,
        consumed_qty=0, missing_qty=10, uom="Units", move_state="cancel",
    )
    summary = {
        "generated_at": "2026-08-18T12:00:00+00:00", "days_checked": 550,
        "completion_date_from": "2025-02-14 12:00:00", "completed_mos_checked": 1,
        "raw_component_moves_checked": 2, "mos_with_short_consumption": 1,
        "component_gaps": 1,
    }
    path = tmp_path / "audit.xlsx"
    _write_xlsx(path, [asdict(row)], summary)
    workbook = load_workbook(path, data_only=False)
    assert workbook.sheetnames == ["Santrauka", "Neatitikimai"]
    sheet = workbook["Neatitikimai"]
    assert sheet.freeze_panes == "A2"
    assert sheet["A2"].value == "WH/MO/34218"
    assert sheet["H2"].value == "2601189846"
    assert sheet["L2"].value == 10
    assert list(sheet.tables) == ["MOComponentConsumptionGaps"]
