"""Nuskaito Production BOM operacijų etalonus naujų BOM importui.

SAUGA: scenarijus naudoja tik fields_get ir search_read metodus. Jis nieko
nekuria, neredaguoja ir netrina Odoo sistemoje.
"""

from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
from statistics import median

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import load_settings
from odoo_client import OdooClient


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_PATH = BASE_DIR / "output" / "production" / "BOM_Operations_Reference.xlsx"


def m2o_id(value):
    return int(value[0]) if isinstance(value, (list, tuple)) and value else None


def m2o_name(value):
    return str(value[1]) if isinstance(value, (list, tuple)) and len(value) > 1 else ""


def fields_get(client: OdooClient, model: str) -> dict:
    """Pasiima techninius laukų vardus, etiketes, tipus ir pasirinkimus."""
    return client.models.execute_kw(
        client.settings.db,
        client.uid,
        client.settings.api_key,
        model,
        "fields_get",
        [],
        {"attributes": ["string", "type", "relation", "selection", "readonly"]},
    )


def selected_boms(boms: list[dict]) -> list[dict]:
    """Vienam produktui palieka aktyvų BOM su mažiausia sequence ir naujausią."""
    grouped = defaultdict(list)
    for bom in boms:
        template_id = m2o_id(bom.get("product_tmpl_id"))
        if template_id:
            grouped[template_id].append(bom)
    result = []
    for rows in grouped.values():
        rows.sort(key=lambda row: (
            int(row.get("sequence") or 0),
            str(row.get("write_date") or ""),
            int(row["id"]),
        ), reverse=False)
        lowest_sequence = int(rows[0].get("sequence") or 0)
        same_sequence = [
            row for row in rows if int(row.get("sequence") or 0) == lowest_sequence
        ]
        same_sequence.sort(
            key=lambda row: (str(row.get("write_date") or ""), int(row["id"])),
            reverse=True,
        )
        result.append(same_sequence[0])
    return result


def add_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    for index, header in enumerate(headers, start=1):
        values = [len(str(ws.cell(row, index).value or "")) for row in range(1, min(ws.max_row, 300) + 1)]
        ws.column_dimensions[get_column_letter(index)].width = min(max(values, default=len(header)) + 2, 55)


def main() -> None:
    settings = load_settings()
    if "stage" in settings.url.lower():
        raise PermissionError("Etalonai turi būti nuskaitomi iš Production, ne Stage.")

    client = OdooClient(settings)
    uid = client.authenticate()

    # 1 ŽINGSNIS: nustatome realius Production techninius laukus. Taip
    # neatspėjame Furnibox individualių Auto Finish ir Autoplan laukų vardų.
    bom_fields = fields_get(client, "mrp.bom")
    operation_model = "mrp.routing.workcenter"
    operation_fields = fields_get(client, operation_model)

    # 2 ŽINGSNIS: pasiimame aktyvius Manufacture BOM ir pagal sutartą taisyklę
    # kiekvienam produktui parenkame vieną galiojantį etaloną.
    wanted_bom_fields = [
        name for name in [
            "id", "code", "active", "sequence", "product_tmpl_id", "product_id",
            "type", "write_date", "operation_ids",
        ] if name in bom_fields
    ]
    boms = client.search_read_all(
        "mrp.bom", [["active", "=", True], ["type", "=", "normal"]],
        wanted_bom_fields, context={"active_test": False},
    )
    chosen = selected_boms(boms)
    chosen_ids = {int(row["id"]) for row in chosen}

    # 3 ŽINGSNIS: prie BOM prijungiame produkto Internal Reference ir pilną
    # kategorijos kelią. Jo paskutinė dalis naudojama kaip pakategorė.
    template_ids = {m2o_id(row.get("product_tmpl_id")) for row in chosen} - {None}
    templates = client.search_read_all(
        "product.template", [["id", "in", sorted(template_ids)]],
        ["id", "default_code", "categ_id"], context={"active_test": False},
    )
    template_by_id = {int(row["id"]): row for row in templates}
    category_ids = {m2o_id(row.get("categ_id")) for row in templates} - {None}
    categories = client.search_read_all(
        "product.category", [["id", "in", sorted(category_ids)]],
        ["id", "name", "complete_name"],
    )
    category_by_id = {int(row["id"]): row for row in categories}

    # 4 ŽINGSNIS: nuskaitome operacijų pavadinimus, darbo centrus ir laikus.
    useful_operation_fields = [
        name for name in [
            "id", "bom_id", "name", "sequence", "workcenter_id", "time_mode",
            "time_cycle_manual", "time_cycle", "time_mode_batch", "batch_size",
            "worksheet_type", "worksheet", "company_id", "write_date",
        ] if name in operation_fields
    ]
    operations = client.search_read_all(
        operation_model, [["bom_id", "in", sorted(chosen_ids)]],
        useful_operation_fields, order="bom_id asc, sequence asc, id asc",
    )
    bom_by_id = {int(row["id"]): row for row in chosen}

    raw_rows = []
    summary_values = defaultdict(list)
    signature_counts = defaultdict(Counter)
    for operation in operations:
        bom_id = m2o_id(operation.get("bom_id"))
        bom = bom_by_id.get(bom_id)
        if not bom:
            continue
        template = template_by_id.get(m2o_id(bom.get("product_tmpl_id")), {})
        category = category_by_id.get(m2o_id(template.get("categ_id")), {})
        category_path = str(category.get("complete_name") or category.get("name") or "")
        subcategory = category_path.rsplit("/", 1)[-1].strip()
        sku = str(template.get("default_code") or "").strip()
        operation_name = str(operation.get("name") or "").strip()
        workcenter = m2o_name(operation.get("workcenter_id"))
        manual_time = operation.get("time_cycle_manual")
        calculated_time = operation.get("time_cycle")
        time_value = manual_time if isinstance(manual_time, (int, float)) else calculated_time
        raw_rows.append([
            sku, category_path, subcategory, bom_id, bom.get("code"),
            operation.get("sequence"), operation_name, workcenter,
            operation.get("time_mode"), manual_time, calculated_time,
            operation.get("time_mode_batch"), operation.get("batch_size"),
            operation.get("worksheet_type"), operation.get("worksheet"),
        ])
        key = (subcategory, operation_name, workcenter)
        if isinstance(time_value, (int, float)):
            summary_values[key].append(float(time_value))
        signature_counts[subcategory][(operation_name, workcenter)] += 1

    # 5 ŽINGSNIS: suvestinė rodo dažniausias operacijas ir medianinius laikus.
    # Medianą naudojame analizei, tačiau prieš generatorių Edgaras dar patvirtins
    # galutinę pakategorės taisyklę.
    summary_rows = []
    for subcategory, counter in sorted(signature_counts.items()):
        for (operation_name, workcenter), count in counter.most_common():
            times = summary_values.get((subcategory, operation_name, workcenter), [])
            summary_rows.append([
                subcategory, operation_name, workcenter, count,
                median(times) if times else "", min(times) if times else "",
                max(times) if times else "",
            ])

    # 6 ŽINGSNIS: atskirame lape pateikiame kandidatus Auto Finish / Autoplan
    # laukams su techniniu vardu ir Odoo etikete.
    keywords = ("auto", "finish", "plan", "operation", "work order")
    candidate_rows = []
    for model, metadata in (("mrp.bom", bom_fields), (operation_model, operation_fields)):
        for technical_name, info in sorted(metadata.items()):
            searchable = f"{technical_name} {info.get('string', '')}".lower()
            if any(keyword in searchable for keyword in keywords):
                candidate_rows.append([
                    model, technical_name, info.get("string"), info.get("type"),
                    info.get("relation"), info.get("readonly"), str(info.get("selection") or ""),
                ])

    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "SUMMARY BY SUBCATEGORY", [
        "Subcategory", "Operation", "Work Center", "BOM Count",
        "Median Time", "Minimum Time", "Maximum Time",
    ], summary_rows)
    add_sheet(wb, "BOM OPERATIONS", [
        "Parent SKU", "Category Path", "Subcategory", "BOM ID", "BOM Reference",
        "Operation Sequence", "Operation", "Work Center", "Time Mode",
        "Manual Time", "Calculated Time", "Batch Time Mode", "Batch Size",
        "Worksheet Type", "Worksheet",
    ], raw_rows)
    add_sheet(wb, "FIELD CANDIDATES", [
        "Model", "Technical Field", "Odoo Label", "Type", "Relation",
        "Readonly", "Selection",
    ], candidate_rows)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)

    print("Prisijungta prie Production Odoo. UID=", uid)
    print("\nBOM OPERACIJŲ ETALONAI NUSKAITYTI")
    print("Failas:", OUTPUT_PATH)
    print("Pasirinkti Manufacture BOM:", len(chosen))
    print("Operacijų eilutės:", len(raw_rows))
    print("Pakategorių operacijų kombinacijos:", len(summary_rows))
    print("Techninių laukų kandidatai:", len(candidate_rows))
    print("Odoo pakeitimų neatlikta.")


if __name__ == "__main__":
    main()
