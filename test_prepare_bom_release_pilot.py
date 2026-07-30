from __future__ import annotations

import tempfile
import unittest
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from bom_release.generator import (
    APACK_OPERATION_TYPE_EXTERNAL_ID,
    IMPORT_HEADERS,
    MANUFACTURE,
    NEW_BOM_SEQUENCE,
)
from prepare_bom_release_pilot import PilotPreparationError, prepare_pilot


def parent_row(
    sku: str,
    operation_type_external_id: str = APACK_OPERATION_TYPE_EXTERNAL_ID,
) -> list[object]:
    return [
        sku,
        "__export__.parent",
        1,
        "PART-1",
        "__export__.part",
        2,
        MANUFACTURE,
        NEW_BOM_SEQUENCE,
        operation_type_external_id,
        "REFORM_v08_20260730",
        True,
        True,
        "Surinkimas",
        "Surinkimas",
        "manual",
        3,
        100,
    ]


def write_source(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOM import"
    sheet.append(IMPORT_HEADERS)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


def write_audit(path: Path, links, review=()):
    rows = []
    for apack, hrd in links:
        rows.append(
            {
                "apack_sku": apack,
                "hrd_a_sku": hrd,
                "component_sku": "HINGE",
                "quantity": 1,
                "decision": "TRANSFER_TO_APACK",
            }
        )
    for apack in review:
        rows.append(
            {
                "apack_sku": apack,
                "hrd_a_sku": "",
                "decision": "DEFAULT_HRD_REVIEW",
            }
        )
    path.write_text(json.dumps({"status": "PASS", "rows": rows}), encoding="utf-8")


class PrepareBomReleasePilotTests(unittest.TestCase):
    def test_extracts_one_complete_apack_group(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "release.xlsx"
            audit = root / "audit.json"
            output = root / "pilot.xlsx"
            continuation = [None] * len(IMPORT_HEADERS)
            continuation[3:6] = ["PART-2", "__export__.part2", 1]
            continuation[12:17] = [
                "Pakavimas",
                "Pakavimas",
                "manual",
                2,
                101,
            ]
            write_source(
                source,
                [
                    parent_row("APACK-EU-C-CAB01-AAA001-A"),
                    continuation,
                    parent_row("APACK-EU-C-CAB01-BBB001-A"),
                    parent_row("HRD-EU-C-CAB01-AAA001-A"),
                ],
            )
            write_audit(
                audit,
                [(
                    "APACK-EU-C-CAB01-AAA001-A",
                    "HRD-EU-C-CAB01-AAA001-A",
                )],
            )

            sku, hrd, apacks, components, operations = prepare_pilot(
                source, audit, output
            )

            self.assertEqual(sku, "APACK-EU-C-CAB01-AAA001-A")
            self.assertEqual(hrd, "HRD-EU-C-CAB01-AAA001-A")
            self.assertEqual(apacks, [sku])
            self.assertEqual(components, 3)
            self.assertEqual(operations, 3)
            workbook = load_workbook(output, read_only=True)
            try:
                rows = list(workbook["BOM import"].iter_rows(values_only=True))
                self.assertEqual(len(rows), 4)
                self.assertEqual(rows[1][0], sku)
                self.assertIsNone(rows[2][0])
                self.assertEqual(rows[3][0], hrd)
            finally:
                workbook.close()

    def test_requested_sku_is_used(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "release.xlsx"
            audit = root / "audit.json"
            output = root / "pilot.xlsx"
            wanted = "APACK-EU-C-CAB01-BBB001-A"
            hrd = "HRD-EU-C-CAB01-BBB001-A"
            write_source(
                source,
                [
                    parent_row("APACK-EU-C-CAB01-AAA001-A"),
                    parent_row(wanted),
                    parent_row(hrd),
                ],
            )
            write_audit(audit, [(wanted, hrd)])
            sku, _, _, _, _ = prepare_pilot(
                source, audit, output, wanted.lower()
            )
            self.assertEqual(sku, wanted)

    def test_wrong_routing_is_rejected(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "release.xlsx"
            audit = root / "audit.json"
            write_source(
                source,
                [
                    parent_row("APACK-EU-C-CAB01-AAA001-A", "Wrong"),
                    parent_row("HRD-EU-C-CAB01-AAA001-A"),
                ],
            )
            write_audit(
                audit,
                [(
                    "APACK-EU-C-CAB01-AAA001-A",
                    "HRD-EU-C-CAB01-AAA001-A",
                )],
            )
            with self.assertRaisesRegex(
                PilotPreparationError,
                "Operation Type/External ID",
            ):
                prepare_pilot(source, audit, root / "pilot.xlsx")

    def test_usb_legacy_exception_is_not_auto_selected(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "release.xlsx"
            audit = root / "audit.json"
            write_source(
                source,
                [
                    parent_row("APACK-USB-C-CAB01-WAL045-A"),
                    parent_row("HRD-USB-C-CAB01-WAL045-A"),
                ],
            )
            write_audit(
                audit,
                [(
                    "APACK-USB-C-CAB01-WAL045-A",
                    "HRD-USB-C-CAB01-WAL045-A",
                )],
            )
            with self.assertRaisesRegex(PilotPreparationError, "saugus APACK"):
                prepare_pilot(source, audit, root / "pilot.xlsx")

    def test_shared_hrd_includes_complete_apack_cohort(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "release.xlsx"
            audit = root / "audit.json"
            output = root / "pilot.xlsx"
            first = "APACK-EU-C-CAB01-BAS001-A"
            second = "APACK-EU-C-CAB01-BAS002-A"
            hrd = "HRD-EU-C-CAB01-BAS001-A"
            write_source(
                source,
                [parent_row(first), parent_row(second), parent_row(hrd)],
            )
            write_audit(audit, [(first, hrd), (second, hrd)])

            sku, linked_hrd, apacks, _, _ = prepare_pilot(
                source, audit, output, second
            )

            self.assertEqual(sku, second)
            self.assertEqual(linked_hrd, hrd)
            self.assertEqual(apacks, [first, second])
            workbook = load_workbook(output, read_only=True)
            try:
                parents = [
                    row[0]
                    for row in list(
                        workbook["BOM import"].iter_rows(
                            min_row=2, values_only=True
                        )
                    )
                    if row[0]
                ]
                self.assertEqual(parents, [first, second, hrd])
            finally:
                workbook.close()

    def test_default_review_cannot_be_requested(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "release.xlsx"
            audit = root / "audit.json"
            apack = "APACK-EU-C-CAB01-BNF001-A"
            write_source(source, [parent_row(apack)])
            write_audit(audit, [], review=[apack])
            with self.assertRaisesRegex(
                PilotPreparationError, "nėra patvirtintas transformuotas"
            ):
                prepare_pilot(
                    source, audit, root / "pilot.xlsx", apack
                )


if __name__ == "__main__":
    unittest.main()
