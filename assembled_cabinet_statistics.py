#!/usr/bin/env python3
"""Surenkamų spintelių kiekio pagal Sales Order ataskaita.

Atranka:
- produkto Internal Reference (`default_code`) baigiasi `-A`;
- sumuojamas Sales Order eilutės užsakytas kiekis (`product_uom_qty`).

Įvestis:
- `orders.txt` projekto kataloge, po vieną SO numerį kiekvienoje eilutėje.

Rezultatas:
- `output/Assembled_Cabinets_By_SO.xlsx`
"""

from __future__ import annotations

import sys
import xmlrpc.client
from collections import defaultdict
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from config import load_settings


BASE_DIR = Path(__file__).resolve().parent
ORDERS_FILE = BASE_DIR / "orders.txt"
OUTPUT_FILENAME = "Assembled_Cabinets_By_SO.xlsx"


def load_orders(path: Path) -> list[str]:
    if not path.exists():
        raise FileNotFoundError(
            f"Nerastas užsakymų failas: {path}\n"
            "Sukurkite orders.txt projekto kataloge ir įrašykite po vieną SO numerį eilutėje."
        )

    orders: list[str] = []
    seen: set[str] = set()

    for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
        order = raw_line.strip()
        if not order or order.startswith("#"):
            continue
        if order not in seen:
            seen.add(order)
            orders.append(order)

    if not orders:
        raise ValueError("orders.txt yra tuščias.")

    return orders


def connect() -> tuple[Any, int, Any]:
    settings = load_settings()

    common = xmlrpc.client.ServerProxy(
        f"{settings.url}/xmlrpc/2/common",
        allow_none=True,
    )
    uid = common.authenticate(
        settings.db,
        settings.login,
        settings.api_key,
        {},
    )
    if not uid:
        raise RuntimeError(
            "Nepavyko prisijungti prie Odoo. Patikrinkite pasirinktos aplinkos prisijungimo duomenis."
        )

    models = xmlrpc.client.ServerProxy(
        f"{settings.url}/xmlrpc/2/object",
        allow_none=True,
    )
    return settings, uid, models


def execute_kw(
    settings: Any,
    uid: int,
    models: Any,
    model: str,
    method: str,
    args: list[Any],
    kwargs: dict[str, Any] | None = None,
) -> Any:
    return models.execute_kw(
        settings.db,
        uid,
        settings.api_key,
        model,
        method,
        args,
        kwargs or {},
    )


def fetch_report_rows(
    settings: Any,
    uid: int,
    models: Any,
    order_names: list[str],
) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]], list[str]]:
    print("Ieškoma Sales Order...")

    orders = execute_kw(
        settings,
        uid,
        models,
        "sale.order",
        "search_read",
        [[("name", "in", order_names)]],
        {
            "fields": ["id", "name", "state", "partner_id"],
            "limit": 0,
        },
    )

    orders_by_id = {order["id"]: order for order in orders}
    orders_by_name = {order["name"]: order for order in orders}
    not_found = [name for name in order_names if name not in orders_by_name]

    print(f"Rasta SO: {len(orders)}")
    print(f"Nerasta SO: {len(not_found)}")

    if not orders:
        return [], orders_by_name, not_found

    print("Nuskaitomos Sales Order eilutės...")

    lines = execute_kw(
        settings,
        uid,
        models,
        "sale.order.line",
        "search_read",
        [[
            ("order_id", "in", list(orders_by_id)),
            ("display_type", "=", False),
        ]],
        {
            "fields": [
                "order_id",
                "product_id",
                "product_uom_qty",
                "product_uom",
                "name",
            ],
            "limit": 0,
        },
    )

    product_ids = sorted(
        {
            line["product_id"][0]
            for line in lines
            if line.get("product_id")
        }
    )

    products: list[dict[str, Any]] = []
    if product_ids:
        print(f"Nuskaitomos produktų kortelės: {len(product_ids)}")
        products = execute_kw(
            settings,
            uid,
            models,
            "product.product",
            "read",
            [product_ids],
            {"fields": ["default_code", "name"]},
        )

    products_by_id = {product["id"]: product for product in products}
    details: list[dict[str, Any]] = []

    for line in lines:
        product_value = line.get("product_id")
        if not product_value:
            continue

        product_id = product_value[0]
        product = products_by_id.get(product_id, {})
        internal_reference = str(product.get("default_code") or "").strip()

        # Odoo sąrašuose Internal Reference gali būti rodomas laužtiniuose skliaustuose.
        # Atrenkame tik tuos produktus, kurių tikrasis kodas baigiasi "-A".
        code = internal_reference.strip("[]").strip()

        if not code.upper().endswith("-A"):
            continue

        order_id = line["order_id"][0]
        order = orders_by_id[order_id]

        details.append(
            {
                "so": order["name"],
                "state": order.get("state", ""),
                "customer": order["partner_id"][1] if order.get("partner_id") else "",
                "internal_reference": internal_reference,
                "product_name": product.get("name") or product_value[1],
                "quantity": line.get("product_uom_qty") or 0,
                "uom": line["product_uom"][1] if line.get("product_uom") else "",
                "line_description": line.get("name") or "",
            }
        )

    return details, orders_by_name, not_found


def autosize(ws: Any) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)


def save_report(
    output_path: Path,
    order_names: list[str],
    details: list[dict[str, Any]],
    orders_by_name: dict[str, dict[str, Any]],
    not_found: list[str],
) -> None:
    totals: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    line_counts: dict[str, int] = defaultdict(int)

    for row in details:
        so = row["so"]
        totals[so] += Decimal(str(row["quantity"]))
        line_counts[so] += 1

    wb = Workbook()

    summary = wb.active
    summary.title = "SUMMARY"
    summary.append(
        [
            "Sales Order",
            "Status",
            "Customer",
            "-A product lines",
            "Assembled cabinets qty",
            "Result",
        ]
    )

    not_found_set = set(not_found)

    for so in order_names:
        order = orders_by_name.get(so)

        if so in not_found_set or not order:
            summary.append([so, "", "", 0, 0, "SO NOT FOUND"])
            continue

        qty = totals[so]
        summary.append(
            [
                so,
                order.get("state", ""),
                order["partner_id"][1] if order.get("partner_id") else "",
                line_counts.get(so, 0),
                float(qty),
                "FOUND -A PRODUCTS" if qty else "NO -A PRODUCTS",
            ]
        )

    details_ws = wb.create_sheet("DETAILS")
    details_ws.append(
        [
            "Sales Order",
            "Status",
            "Customer",
            "Internal Reference",
            "Product",
            "Quantity",
            "UoM",
            "SO line description",
        ]
    )

    order_position = {name: index for index, name in enumerate(order_names)}
    for row in sorted(
        details,
        key=lambda item: (
            order_position.get(item["so"], 999999),
            item["internal_reference"],
        ),
    ):
        details_ws.append(
            [
                row["so"],
                row["state"],
                row["customer"],
                row["internal_reference"],
                row["product_name"],
                float(Decimal(str(row["quantity"]))),
                row["uom"],
                row["line_description"],
            ]
        )

    not_found_ws = wb.create_sheet("NOT_FOUND")
    not_found_ws.append(["Sales Order"])
    for so in not_found:
        not_found_ws.append([so])

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        autosize(ws)

    wb.save(output_path)


def main() -> int:
    try:
        order_names = load_orders(ORDERS_FILE)
        print(f"Nuskaityta SO numerių: {len(order_names)}")

        settings, uid, models = connect()
        print(f"Prisijungta prie Odoo: {settings.url}")

        details, orders_by_name, not_found = fetch_report_rows(
            settings,
            uid,
            models,
            order_names,
        )

        output_path = settings.output_dir / OUTPUT_FILENAME
        save_report(
            output_path,
            order_names,
            details,
            orders_by_name,
            not_found,
        )

        total_qty = sum(
            (Decimal(str(row["quantity"])) for row in details),
            Decimal("0"),
        )

        print(f"Rasta '-A' produkto eilučių: {len(details)}")
        print(f"Bendras surenkamų spintelių kiekis: {total_qty}")
        print(f"Rezultatas išsaugotas: {output_path}")
        return 0

    except Exception as exc:
        print(f"KLAIDA: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
