"""Detect all Reform products in Odoo: BOM parents and every component SKU."""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import load_settings
from odoo_client import OdooClient


def canon(value: object) -> str:
    return str(value or "").strip().upper()


def find_bom_input(base: Path) -> Path:
    candidates = []
    for folder in (base / "data", base):
        for pattern in ("Reform_BOM_Input*.xlsx", "BOM_for Furnibox*.xlsx"):
            candidates.extend(folder.glob(pattern))
    candidates = [path for path in candidates if not path.name.startswith("~$")]
    if not candidates:
        raise FileNotFoundError("Nerastas Reform BOM Input .xlsx failas aplanke data.")
    return max(candidates, key=lambda path: path.stat().st_mtime)


def load_reform_universe(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        header_row = None
        headers = None
        for row_no, row in enumerate(
            ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True), start=1
        ):
            values = [str(value).strip() if value is not None else "" for value in row]
            if "BOM SKU Code" in values:
                header_row, headers = row_no, values
                break
        if header_row is None:
            continue

        index = {value: i for i, value in enumerate(headers) if value}
        part_columns = []
        for header, code_index in index.items():
            match = re.fullmatch(r"Part\s+(\d+)\s+Code", header, flags=re.IGNORECASE)
            if not match:
                continue
            number = int(match.group(1))
            group_index = index.get(f"Part {number} Group")
            part_columns.append((number, code_index, group_index))
        part_columns.sort()

        products = {}
        used_by = defaultdict(set)
        source_rows = defaultdict(set)
        groups = defaultdict(Counter)

        for excel_row, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
        ):
            parent_raw = row[index["BOM SKU Code"]]
            parent = canon(parent_raw)
            if not parent:
                continue
            products.setdefault(parent, {
                "sku": str(parent_raw).strip(),
                "is_parent": False,
                "is_component": False,
                "category": "",
                "name_1": "",
                "name_2": "",
            })
            products[parent]["is_parent"] = True
            products[parent]["category"] = str(row[index.get("Product Catagory", -1)] or "").strip()
            products[parent]["name_1"] = str(row[index.get("SKU name 1", -1)] or "").strip()
            products[parent]["name_2"] = str(row[index.get("SKU name 2", -1)] or "").strip()
            source_rows[parent].add(excel_row)

            for number, code_index, group_index in part_columns:
                component_raw = row[code_index]
                component = canon(component_raw)
                if not component:
                    continue
                products.setdefault(component, {
                    "sku": str(component_raw).strip(),
                    "is_parent": False,
                    "is_component": False,
                    "category": "",
                    "name_1": "",
                    "name_2": "",
                })
                products[component]["is_component"] = True
                used_by[component].add(parent)
                source_rows[component].add(excel_row)
                if group_index is not None:
                    group = str(row[group_index] or "").strip()
                    if group:
                        groups[component][group] += 1

        for sku, product in products.items():
            product["part_group"] = groups[sku].most_common(1)[0][0] if groups[sku] else ""
            product["part_group_options"] = "; ".join(
                f"{group} ({count})" for group, count in groups[sku].most_common()
            )
            product["used_by_count"] = len(used_by[sku])
            product["source_rows"] = ", ".join(map(str, sorted(source_rows[sku])))
        wb.close()
        return products, ws.title, header_row

    wb.close()
    raise ValueError("Nerasta BOM SKU Code antraštė.")


def role(product: dict) -> str:
    if product["is_parent"] and product["is_component"]:
        return "BOM PARENT + COMPONENT"
    if product["is_parent"]:
        return "BOM PARENT"
    return "COMPONENT ONLY"


def read_odoo_products(client: OdooClient):
    rows = client.search_read_all(
        "product.product",
        [["default_code", "!=", False]],
        ["id", "default_code", "name", "categ_id", "active", "product_tmpl_id"],
        context={"active_test": False},
    )
    result = {}
    duplicates = defaultdict(list)
    for row in rows:
        sku = canon(row.get("default_code"))
        if not sku:
            continue
        duplicates[sku].append(row)
        if sku not in result or (row.get("active") and not result[sku].get("active")):
            result[sku] = row
    return result, duplicates


def style_sheet(ws):
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        values = [str(ws.cell(row, col).value or "") for row in range(1, min(ws.max_row, 500) + 1)]
        ws.column_dimensions[get_column_letter(col)].width = min(max(map(len, values)) + 2, 55)


def write_output(path: Path, reform_products, odoo_products, duplicates, info):
    wb = Workbook()
    ws = wb.active
    ws.title = "PRODUCT DETECTION"
    headers = [
        "Reform SKU", "Reform Role", "Reform Product Category", "Reform Part Group",
        "Part Group Options", "Reform Name 1", "Reform Name 2", "Used By BOM Count",
        "Odoo SKU", "Odoo Product ID", "Odoo Template ID", "Odoo Product Name",
        "Odoo Category", "Odoo Active", "Exists in Odoo", "Next Step", "Source Rows",
    ]
    ws.append(headers)
    summary = Counter()
    missing_by_role = Counter()
    missing_by_group = Counter()
    for sku, product in sorted(reform_products.items()):
        odoo = odoo_products.get(sku)
        exists = odoo is not None
        product_role = role(product)
        summary["TOTAL"] += 1
        summary["EXISTS" if exists else "MISSING"] += 1
        summary[product_role] += 1
        if not exists:
            missing_by_role[product_role] += 1
            missing_by_group[product["category"] or product["part_group"] or "UNCLASSIFIED"] += 1
        category = odoo.get("categ_id") if odoo else None
        template = odoo.get("product_tmpl_id") if odoo else None
        ws.append([
            product["sku"], product_role, product["category"], product["part_group"],
            product["part_group_options"], product["name_1"], product["name_2"],
            product["used_by_count"], odoo.get("default_code") if odoo else "",
            odoo.get("id") if odoo else "", template[0] if template else "",
            odoo.get("name") if odoo else "", category[1] if category else "",
            odoo.get("active") if odoo else "", "YES" if exists else "NO",
            "NONE" if exists else "CREATE PRODUCT", product["source_rows"],
        ])
    style_sheet(ws)

    summary_ws = wb.create_sheet("SUMMARY")
    summary_ws.append(["Metric", "Count"])
    for key in ("TOTAL", "EXISTS", "MISSING", "BOM PARENT", "BOM PARENT + COMPONENT", "COMPONENT ONLY"):
        summary_ws.append([key, summary[key]])
    summary_ws.append([])
    summary_ws.append(["Missing by role", "Count"])
    for key, value in sorted(missing_by_role.items()):
        summary_ws.append([key, value])
    summary_ws.append([])
    summary_ws.append(["Missing by category / part group", "Count"])
    for key, value in missing_by_group.most_common():
        summary_ws.append([key, value])
    style_sheet(summary_ws)

    diagnostics = wb.create_sheet("DIAGNOSTICS")
    diagnostics.append(["Type", "SKU", "Count", "Message"])
    for sku, rows in sorted(duplicates.items()):
        if len(rows) > 1:
            diagnostics.append(["DUPLICATE ODOO SKU", sku, len(rows), "Keli Odoo produktai turi tą patį Internal Reference"])
    style_sheet(diagnostics)

    info_ws = wb.create_sheet("INFO")
    info_ws.append(["Parameter", "Value"])
    for key, value in info.items():
        info_ws.append([key, value])
    style_sheet(info_ws)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return summary, missing_by_role


def main():
    base = Path(__file__).resolve().parent
    input_path = find_bom_input(base)
    output_path = base / "output" / "Product_Detection_All.xlsx"
    print("Nuskaitomi visi Reform produktai ir komponentai...")
    reform_products, sheet_name, header_row = load_reform_universe(input_path)
    print("Unikalių Reform SKU:", len(reform_products))

    settings = load_settings()
    client = OdooClient(settings)
    uid = client.authenticate()
    print(f"Prisijungta prie Odoo. UID={uid}")
    odoo_products, duplicates = read_odoo_products(client)
    print("Odoo produktų su Internal Reference:", len(odoo_products))
    summary, missing_by_role = write_output(output_path, reform_products, odoo_products, duplicates, {
        "Odoo URL": settings.url,
        "Odoo DB": settings.db,
        "Reform source": str(input_path),
        "Reform sheet": sheet_name,
        "Header row": header_row,
        "Comparison": "TRIM + UPPERCASE",
    })
    print("\nVISŲ PRODUKTŲ PALYGINIMAS BAIGTAS")
    print("Visi Reform SKU:", summary["TOTAL"])
    print("Yra Odoo:", summary["EXISTS"])
    print("Nėra Odoo:", summary["MISSING"])
    for product_role, count in sorted(missing_by_role.items()):
        print(f"Nėra Odoo – {product_role}: {count}")
    print("Rezultatas:", output_path)


if __name__ == "__main__":
    main()
