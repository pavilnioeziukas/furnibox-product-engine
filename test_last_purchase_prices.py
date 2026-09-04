from pathlib import Path
import sys
import tempfile
import types
import unittest

from openpyxl import load_workbook


config_stub = types.ModuleType("config")
config_stub.load_settings = lambda: None
sys.modules.setdefault("config", config_stub)

from last_purchase_prices import resolve_shared_data_dir, write_purchase_prices


class ComponentPriceWorkbookTests(unittest.TestCase):
    def test_product_engine_shared_data_path_has_priority(self):
        from unittest.mock import patch

        with patch.dict(
            "os.environ",
            {
                "PRODUCT_ENGINE_SHARED_DATA_DIR": "/tmp/canonical-shared",
                "FURNIBOX_SHARED_DATA": "/tmp/stale-shared",
                "FURNIBOX_SHARED_DATA_DIR": "/tmp/other-stale-shared",
            },
            clear=False,
        ):
            self.assertEqual(
                resolve_shared_data_dir(),
                Path("/tmp/canonical-shared"),
            )

    def test_adjustment_without_purchase_history_is_included_as_reform_price(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            output = Path(temporary_directory) / "Last_Purchase_Prices.xlsx"

            write_purchase_prices(
                output,
                [],
                {
                    "url": "https://example.test",
                    "db": "test",
                    "login": "test@example.test",
                    "uid": 1,
                },
                purchase_price_adjustments={
                    "EU FP PACK": {
                        "adjusted_purchase_price": 2.09,
                        "comment": "Tamara approved price",
                    }
                },
            )

            workbook = load_workbook(output, data_only=False)
            adjustments = workbook["PURCHASE PRICE ADJUSTMENTS"]
            prices = workbook["COMPONENT PRICES"]

            self.assertEqual(adjustments["A2"].value, "EU FP PACK")
            self.assertEqual(adjustments["B2"].value, 2.09)
            self.assertIsNone(adjustments["C2"].value)
            self.assertEqual(prices["A2"].value, "EU FP PACK")
            self.assertEqual(prices["I2"].value, "='PURCHASE PRICE ADJUSTMENTS'!B2")
            self.assertEqual(prices["J2"].value, 1.0)
            self.assertEqual(prices["K2"].value, '=IF(I2="",G2,I2)*J2')

    def test_tamara_master_prices_override_zero_odoo_prices_for_hrd_components(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            output = Path(temporary_directory) / "Last_Purchase_Prices.xlsx"
            rows = []
            for sku in ("7001730", "3284903"):
                row = {column: "" for column in [
                    "Internal Reference", "Name", "Product Category/Name",
                    "Purchase Order", "Vendor", "Ordered Quantity",
                    "Last Purchase Price", "Order Date",
                ]}
                row["Internal Reference"] = sku
                row["Vendor"] = "Reform Supply & Logistics, UAB"
                row["Last Purchase Price"] = 0
                rows.append(row)

            write_purchase_prices(
                output,
                rows,
                {"url": "x", "db": "x", "login": "x", "uid": 1},
                purchase_price_adjustments={
                    "7001730": {"adjusted_purchase_price": 1.48},
                    "3284903": {"adjusted_purchase_price": 0.04},
                },
            )

            workbook = load_workbook(output, data_only=False)
            adjustments = workbook["PURCHASE PRICE ADJUSTMENTS"]
            self.assertEqual(adjustments["B2"].value, 1.48)
            self.assertEqual(adjustments["B3"].value, 0.04)
            self.assertEqual(adjustments["C2"].value, 0)
            self.assertEqual(adjustments["C3"].value, 0)

    def test_adjustment_matching_purchase_sku_is_case_insensitive(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            output = Path(temporary_directory) / "Last_Purchase_Prices.xlsx"
            row = {column: "" for column in [
                "Internal Reference", "Name", "Product Category/Name",
                "Purchase Order", "Vendor", "Ordered Quantity",
                "Last Purchase Price", "Order Date",
            ]}
            row["Internal Reference"] = "Eu Fp Pack"
            row["Last Purchase Price"] = 1.5

            write_purchase_prices(
                output,
                [row],
                {"url": "x", "db": "x", "login": "x", "uid": 1},
                purchase_price_adjustments={
                    "EU FP PACK": {"adjusted_purchase_price": 2.09}
                },
            )

            workbook = load_workbook(output, data_only=False)
            adjustments = workbook["PURCHASE PRICE ADJUSTMENTS"]
            self.assertEqual(adjustments.max_row, 2)
            self.assertEqual(adjustments["A2"].value, "Eu Fp Pack")
            self.assertEqual(adjustments["B2"].value, 2.09)

    def test_workbook_contains_editable_component_pricing_inputs_and_reform_formula(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            output = Path(temporary_directory) / "Last_Purchase_Prices.xlsx"

            write_purchase_prices(
                output,
                [
                    {
                        "Internal Reference": "ACCS-TEST",
                        "Name": "Test component",
                        "Product Category/Name": "All / Components / CABINET HARDWARE",
                        "Purchase Order": "P00001",
                        "Vendor": "Test vendor",
                        "Ordered Quantity": 100,
                        "Last Purchase Price": 2.5,
                        "Order Date": "2026-08-10 12:00:00",
                    }
                ],
                {
                    "url": "https://example.test",
                    "db": "test",
                    "login": "test@example.test",
                    "uid": 1,
                },
            )

            workbook = load_workbook(output, data_only=False)

            self.assertEqual(
                workbook.sheetnames,
                [
                    "INFO",
                    "PRICING RULES",
                    "PURCHASE PRICE ADJUSTMENTS",
                    "COMPONENT PRICES",
                ],
            )

            prices = workbook["COMPONENT PRICES"]
            headers = {cell.value: cell.column for cell in prices[1]}

            self.assertIn("Real Purchase Price", headers)
            self.assertNotIn("Last Purchase Price", headers)

            self.assertEqual(
                prices.cell(
                    2,
                    headers["Adjusted Purchase Price"],
                ).value,
                "='PURCHASE PRICE ADJUSTMENTS'!B2",
            )

            self.assertEqual(
                prices.cell(
                    2,
                    headers["Markup Factor"],
                ).value,
                1.0,
            )

            self.assertEqual(
                prices.cell(
                    2,
                    headers["Reform Price"],
                ).value,
                '=IF(I2="",G2,I2)*J2',
            )

            adjustments = workbook["PURCHASE PRICE ADJUSTMENTS"]

            self.assertEqual(
                adjustments["A2"].value,
                "ACCS-TEST",
            )
            self.assertEqual(
                adjustments["B2"].value,
                2.5,
            )
            self.assertEqual(
                workbook.calculation.calcMode,
                "auto",
            )

            write_purchase_prices(
                output,
                [
                    {
                        "Internal Reference": "ACCS-TEST",
                        "Name": "Test component",
                        "Product Category/Name": "All / Components / CABINET HARDWARE",
                        "Purchase Order": "P00002",
                        "Vendor": "Reform Supply & Logistics, UAB",
                        "Ordered Quantity": 50,
                        "Last Purchase Price": 2.75,
                        "Order Date": "2026-08-11 12:00:00",
                    }
                ],
                {
                    "url": "https://example.test",
                    "db": "test",
                    "login": "test@example.test",
                    "uid": 1,
                },
                purchase_price_adjustments={
                    "ACCS-TEST": {
                        "adjusted_purchase_price": 3.25,
                        "comment": "Test adjustment",
                    }
                },
            )

            regenerated = load_workbook(
                output,
                data_only=False,
            )

            self.assertEqual(
                regenerated["PURCHASE PRICE ADJUSTMENTS"]["B2"].value,
                3.25,
            )
            self.assertEqual(
                regenerated["PURCHASE PRICE ADJUSTMENTS"]["C2"].value,
                2.75,
            )
            self.assertEqual(
                regenerated["PURCHASE PRICE ADJUSTMENTS"]["D2"].value,
                "Test adjustment",
            )
            self.assertEqual(
                regenerated["COMPONENT PRICES"]["J2"].value,
                1.05,
            )
            self.assertEqual(
                regenerated["COMPONENT PRICES"]["K2"].value,
                '=IF(I2="",G2,I2)*J2',
            )


if __name__ == "__main__":
    unittest.main()
