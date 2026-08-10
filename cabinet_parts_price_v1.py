"""Atlieka esamų Odoo ir naujų Reform FPACK Cabinet Parts kainų auditą.

Šaltiniai – aktyvūs Odoo FPACK BOM ir 6 veiksmo MAP_Comparison.xlsx. NEW BOM LINES pateikia FPACK,
detalės SKU ir kiekį. Cabinet Parts atpažįstamos pagal SKU esančius matmenis;
jos neprivalo būti NEW PRODUCTS lape, nes kainą gali reikėti perskaičiuoti ir
jau Odoo egzistuojančiai detalei.

Skaičiavimas atkartoja failo „Detaliu kainos perskaiciavimas“ lapo
„Cabinet part kainso“ logiką. Scenarijus tik sukuria kontrolinį Excel failą;
Odoo duomenų nekeičia.
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
import os
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from output_paths import environment_output_dir
from odoo_client import OdooClient


BASE_DIR = Path(__file__).resolve().parent
SCRIPT_VERSION = "7A-FURNIX-TRANSFER-PRICE-20260810-05"
SOURCE_NAME = "MAP_Comparison.xlsx"
OUTPUT_NAME = "Existing_and_New_Cabinet_Parts_Prices.xlsx"
PARAMETERS_NAME = "Cabinet_Parts_Price_Parameters_v03.xlsx"
DIMENSION_RE = re.compile(
    r"(?<!\d)(\d+(?:[.,]\d+)?)\s*[xX]\s*(\d+(?:[.,]\d+)?)(?!\d)"
)


@dataclass(frozen=True)
class PriceParameters:
    back_rate_per_m2: float = 11.0
    processing_rate_per_m2: float = 17.04
    ww_material_rate_per_m2: float = 49.706 / 5.7
    bb_material_rate_per_m2: float = 40.774 / 5.7
    no_material_rate_per_m2: float = 49.706 / 5.7
    small_part_threshold_m2: float = 0.5
    small_part_surcharge: float = 1.0
    furnix_markup_percent: float = 0.0
    output_decimals: int = 4

    def material_rate(self, color: str) -> float | None:
        return {
            "WW": self.ww_material_rate_per_m2,
            "BB": self.bb_material_rate_per_m2,
            "NO": self.no_material_rate_per_m2,
        }.get(color)


@dataclass(frozen=True)
class PriceCalculation:
    sku: str
    dimension_1_mm: int | float
    dimension_2_mm: int | float
    area_m2: float
    color: str
    part_type: str
    back_rate_per_m2: float
    processing_rate_per_m2: float
    material_rate_per_m2: float
    small_part_surcharge: float
    unit_price: float


@dataclass(frozen=True)
class OdooProductPrice:
    product_id: int
    sku: str
    standard_price: float
    active: bool


@dataclass(frozen=True)
class CombinedBomLine:
    parent: str
    component: str
    existing_quantity: float | None
    new_quantity: float | None
    effective_quantity: float
    fpack_source: str
    component_source: str
    change_status: str


DEFAULT_PARAMETERS = PriceParameters()

PARAMETER_FIELDS = {
    "BACK rate, EUR/m²": "back_rate_per_m2",
    "Processing rate, EUR/m²": "processing_rate_per_m2",
    "WW material rate, EUR/m²": "ww_material_rate_per_m2",
    "BB material rate, EUR/m²": "bb_material_rate_per_m2",
    "NO material rate, EUR/m²": "no_material_rate_per_m2",
    "Small part threshold, m²": "small_part_threshold_m2",
    "Small part surcharge, EUR/unit": "small_part_surcharge",
    "Furnix markup, %": "furnix_markup_percent",
    "Output decimals": "output_decimals",
}


def furnix_transfer_price(
    unit_cost: float,
    parameters: PriceParameters = DEFAULT_PARAMETERS,
) -> tuple[float, float]:
    """Grąžina Furnix antkainį EUR ir pardavimo kainą Furnibox."""
    markup_eur = unit_cost * parameters.furnix_markup_percent / 100
    return markup_eur, unit_cost + markup_eur


def load_parameters(path: Path) -> PriceParameters:
    if not path.exists():
        raise FileNotFoundError(
            f"Nerastas kainodaros parametrų failas:\n{path}\n"
            "Įkelkite Cabinet_Parts_Price_Parameters_v03.xlsx į projekto katalogą."
        )
    workbook = load_workbook(path, data_only=True, read_only=False)
    if "PARAMETERS" not in workbook.sheetnames:
        raise ValueError(f"Parametrų faile '{path.name}' nerastas lapas PARAMETERS.")
    ws = workbook["PARAMETERS"]
    values: dict[str, float | int] = {}
    seen: set[str] = set()
    for row in range(2, ws.max_row + 1):
        name = str(ws.cell(row, 1).value or "").strip()
        if not name:
            continue
        if name not in PARAMETER_FIELDS:
            raise ValueError(f"Nežinomas parametras lape PARAMETERS: '{name}'.")
        if name in seen:
            raise ValueError(f"Parametras įrašytas kelis kartus: '{name}'.")
        seen.add(name)
        raw_value = ws.cell(row, 2).value
        try:
            number = float(str(raw_value).replace(",", "."))
        except (TypeError, ValueError):
            raise ValueError(f"Parametro '{name}' reikšmė nėra skaičius: '{raw_value}'.")
        field = PARAMETER_FIELDS[name]
        if field == "output_decimals":
            if not number.is_integer() or not 0 <= number <= 8:
                raise ValueError("Output decimals turi būti sveikas skaičius nuo 0 iki 8.")
            values[field] = int(number)
        else:
            if number < 0:
                raise ValueError(f"Parametro '{name}' reikšmė negali būti neigiama.")
            values[field] = number
    missing = sorted(set(PARAMETER_FIELDS) - seen)
    if missing:
        raise ValueError("Parametrų faile trūksta: " + ", ".join(missing))
    return PriceParameters(**values)


def canon(value: object) -> str:
    return str(value or "").strip().upper()


def m2o_id(value: object) -> int | None:
    if value is False or value is None:
        return None
    if isinstance(value, (list, tuple)) and value:
        return int(value[0])
    if isinstance(value, int):
        return value
    return None


def parse_number(value: str) -> int | float:
    number = float(value.replace(",", "."))
    return int(number) if number.is_integer() else number


def parse_dimensions(sku: str) -> tuple[int | float, int | float] | None:
    match = DIMENSION_RE.search(sku)
    if not match:
        return None
    return parse_number(match.group(1)), parse_number(match.group(2))


def parse_color(sku: str) -> str:
    return canon(sku).rsplit("-", 1)[-1]


def calculate_unit_price(
    sku: str,
    dimensions: tuple[int | float, int | float],
    parameters: PriceParameters = DEFAULT_PARAMETERS,
) -> PriceCalculation:
    dimension_1, dimension_2 = dimensions
    if dimension_1 <= 0 or dimension_2 <= 0:
        raise ValueError("matmenys turi būti didesni už nulį")

    area_m2 = float(dimension_1) * float(dimension_2) / 1_000_000
    surcharge = (
        parameters.small_part_surcharge
        if area_m2 < parameters.small_part_threshold_m2
        else 0.0
    )
    color = parse_color(sku)
    is_back = "BACK" in canon(sku)

    if is_back:
        unit_price = area_m2 * parameters.back_rate_per_m2 + surcharge
        return PriceCalculation(
            sku=sku,
            dimension_1_mm=dimension_1,
            dimension_2_mm=dimension_2,
            area_m2=area_m2,
            color=color,
            part_type="BACK",
            back_rate_per_m2=parameters.back_rate_per_m2,
            processing_rate_per_m2=0.0,
            material_rate_per_m2=0.0,
            small_part_surcharge=surcharge,
            unit_price=unit_price,
        )

    material_rate = parameters.material_rate(color)
    if material_rate is None:
        raise ValueError(
            f"nežinomas spalvos kodas '{color}' (leidžiama: WW, BB, NO)"
        )
    unit_price = (
        area_m2 * parameters.processing_rate_per_m2
        + area_m2 * material_rate
        + surcharge
    )
    return PriceCalculation(
        sku=sku,
        dimension_1_mm=dimension_1,
        dimension_2_mm=dimension_2,
        area_m2=area_m2,
        color=color,
        part_type="STANDARD",
        back_rate_per_m2=0.0,
        processing_rate_per_m2=parameters.processing_rate_per_m2,
        material_rate_per_m2=material_rate,
        small_part_surcharge=surcharge,
        unit_price=unit_price,
    )


def header_map(ws) -> dict[str, int]:
    result: dict[str, int] = {}
    for cell in ws[1]:
        name = str(cell.value or "").strip()
        if name:
            result[name] = cell.column
    return result


def required_column(headers: dict[str, int], name: str, sheet_name: str) -> int:
    if name not in headers:
        raise ValueError(f"Lape '{sheet_name}' nerastas privalomas stulpelis: {name}")
    return headers[name]


def load_new_products(ws) -> tuple[dict[str, str], dict[str, tuple[int | float, int | float]]]:
    headers = header_map(ws)
    sku_col = required_column(headers, "SKU", ws.title)
    action_col = required_column(headers, "Required Action", ws.title)
    category_col = required_column(headers, "Category", ws.title)

    display_skus: dict[str, str] = {}
    dimensions: dict[str, tuple[int | float, int | float]] = {}
    for row in range(2, ws.max_row + 1):
        sku = str(ws.cell(row, sku_col).value or "").strip()
        action = canon(ws.cell(row, action_col).value)
        category = canon(ws.cell(row, category_col).value)
        parsed = parse_dimensions(sku)
        if (
            sku
            and action == "CREATE PRODUCT"
            and "CABINET PART" in category
            and parsed
        ):
            key = canon(sku)
            display_skus[key] = sku
            dimensions[key] = parsed
    return display_skus, dimensions


def load_fpack_lines(
    ws,
    explicit_cabinet_parts: set[str] | None = None,
) -> tuple[
    list[tuple[str, str, float]],
    dict[str, str],
    dict[str, tuple[int | float, int | float]],
    list[str],
]:
    headers = header_map(ws)
    parent_col = required_column(headers, "Parent SKU", ws.title)
    component_col = required_column(headers, "Component SKU", ws.title)
    quantity_col = required_column(headers, "Quantity", ws.title)

    grouped: dict[tuple[str, str], dict[str, object]] = {}
    display_skus: dict[str, str] = {}
    dimensions: dict[str, tuple[int | float, int | float]] = {}
    invalid_quantities: list[str] = []
    for row in range(2, ws.max_row + 1):
        parent = str(ws.cell(row, parent_col).value or "").strip()
        component = str(ws.cell(row, component_col).value or "").strip()
        parent_key = canon(parent)
        component_key = canon(component)
        parsed = parse_dimensions(component)
        color = parse_color(component)
        looks_like_cabinet_part = (
            "BACK" in component_key
            or color in {"WW", "BB", "NO"}
            or component_key in (explicit_cabinet_parts or set())
        )
        if (
            not parent_key.startswith("FPACK-")
            or not component_key
            or not parsed
            or not looks_like_cabinet_part
        ):
            continue
        display_skus[component_key] = component
        dimensions[component_key] = parsed
        raw_quantity = ws.cell(row, quantity_col).value
        try:
            quantity = float(str(raw_quantity).replace(",", "."))
        except (TypeError, ValueError):
            invalid_quantities.append(
                f"{parent} / {component}: netinkamas kiekis '{raw_quantity}'"
            )
            continue
        key = (parent_key, component_key)
        if key not in grouped:
            grouped[key] = {
                "parent": parent,
                "component": display_skus[component_key],
                "quantity": 0.0,
            }
        grouped[key]["quantity"] = float(grouped[key]["quantity"]) + quantity

    rows = [
        (
            str(value["parent"]),
            str(value["component"]),
            float(value["quantity"]),
        )
        for value in grouped.values()
    ]
    rows.sort(key=lambda item: (canon(item[0]), canon(item[1])))
    return rows, display_skus, dimensions, invalid_quantities


def load_odoo_cabinet_parts(
    client: OdooClient,
) -> tuple[dict[str, OdooProductPrice], dict[str, list[int]], list[tuple[str, str]]]:
    """Read all active and archived products in a CABINET PART category."""
    categories = client.search_read_all(
        "product.category", [], ["id", "name", "complete_name"]
    )
    category_ids = sorted({
        int(row["id"])
        for row in categories
        if "CABINET PART" in {
            canon(part)
            for part in str(
                row.get("complete_name") or row.get("name") or ""
            ).split("/")
        }
    })
    if not category_ids:
        raise ValueError("Odoo nerasta produktų kategorija pavadinimu 'CABINET PART'.")

    rows = client.search_read_all(
        "product.product",
        [["categ_id", "in", category_ids]],
        ["id", "default_code", "standard_price", "active"],
        context={"active_test": False},
    )

    grouped: dict[str, list[dict]] = defaultdict(list)
    diagnostics: list[tuple[str, str]] = []
    for row in rows:
        key = canon(row.get("default_code"))
        if not key:
            diagnostics.append((f"Odoo ID {row['id']}", "trūksta Internal Reference"))
            continue
        grouped[key].append(row)

    prices: dict[str, OdooProductPrice] = {}
    duplicates: dict[str, list[int]] = {}
    for key, matches in grouped.items():
        if len(matches) != 1:
            duplicates[key] = sorted(int(row["id"]) for row in matches)
            continue
        row = matches[0]
        prices[key] = OdooProductPrice(
            product_id=int(row["id"]),
            sku=str(row.get("default_code") or "").strip(),
            standard_price=float(row.get("standard_price") or 0.0),
            active=bool(row.get("active")),
        )
    return prices, duplicates, diagnostics


def load_odoo_fpack_lines(
    client: OdooClient,
) -> tuple[list[tuple[str, str, float]], list[tuple[str, str]]]:
    """Load the effective active BOM of every active Odoo FPACK product."""
    categories = client.search_read_all(
        "product.category", [], ["id", "name", "complete_name"]
    )
    cabinet_category_ids = {
        int(row["id"])
        for row in categories
        if "CABINET PART" in {
            canon(part)
            for part in str(row.get("complete_name") or row.get("name") or "").split("/")
        }
    }
    products = client.search_read_all(
        "product.product",
        [],
        ["id", "default_code", "active", "product_tmpl_id", "categ_id"],
        context={"active_test": False},
    )
    fpack_products = [
        row for row in products
        if bool(row.get("active")) and canon(row.get("default_code")).startswith("FPACK-")
    ]
    part_by_id = {
        int(row["id"]): str(row.get("default_code") or "").strip()
        for row in products
        if bool(row.get("active"))
        and m2o_id(row.get("categ_id")) in cabinet_category_ids
        and str(row.get("default_code") or "").strip()
    }
    template_ids = sorted({m2o_id(row.get("product_tmpl_id")) for row in fpack_products} - {None})
    if not template_ids:
        return [], []
    boms = client.search_read_all(
        "mrp.bom",
        [["active", "=", True], ["product_tmpl_id", "in", template_ids]],
        ["id", "active", "sequence", "product_tmpl_id", "product_id", "write_date"],
        order="product_tmpl_id asc, sequence asc, write_date desc, id desc",
    )
    selected_by_product: dict[int, dict] = {}
    diagnostics: list[tuple[str, str]] = []
    for product in fpack_products:
        product_id = int(product["id"])
        template_id = m2o_id(product.get("product_tmpl_id"))
        candidates = [
            bom for bom in boms
            if m2o_id(bom.get("product_tmpl_id")) == template_id
            and m2o_id(bom.get("product_id")) in (None, product_id)
        ]
        if not candidates:
            diagnostics.append((str(product.get("default_code") or product_id), "aktyvus FPACK neturi aktyvaus BOM"))
            continue
        lowest_sequence = min(int(bom.get("sequence") or 0) for bom in candidates)
        effective = [bom for bom in candidates if int(bom.get("sequence") or 0) == lowest_sequence]
        selected_by_product[product_id] = max(
            effective,
            key=lambda bom: (str(bom.get("write_date") or ""), int(bom["id"])),
        )

    selected_bom_ids = sorted({int(bom["id"]) for bom in selected_by_product.values()})
    if not selected_bom_ids:
        return [], diagnostics
    bom_lines = client.search_read_all(
        "mrp.bom.line",
        [["bom_id", "in", selected_bom_ids]],
        ["id", "bom_id", "product_id", "product_qty"],
        order="bom_id asc, id asc",
    )
    lines_by_bom: dict[int, list[dict]] = defaultdict(list)
    for line in bom_lines:
        bom_id = m2o_id(line.get("bom_id"))
        if bom_id is not None:
            lines_by_bom[bom_id].append(line)

    grouped: dict[tuple[str, str], float] = defaultdict(float)
    for product in fpack_products:
        selected = selected_by_product.get(int(product["id"]))
        if selected is None:
            continue
        parent = str(product.get("default_code") or "").strip()
        for line in lines_by_bom.get(int(selected["id"]), []):
            component = part_by_id.get(m2o_id(line.get("product_id")) or -1)
            if not component:
                continue
            try:
                quantity = float(line.get("product_qty") or 0.0)
            except (TypeError, ValueError):
                diagnostics.append((f"{parent} / {component}", f"netinkamas Odoo BOM kiekis '{line.get('product_qty')}'"))
                continue
            grouped[(parent, component)] += quantity
    rows = [(parent, component, quantity) for (parent, component), quantity in grouped.items()]
    rows.sort(key=lambda item: (canon(item[0]), canon(item[1])))
    return rows, diagnostics


def combine_bom_lines(
    existing_lines: list[tuple[str, str, float]],
    new_lines: list[tuple[str, str, float]],
) -> list[CombinedBomLine]:
    existing = {(canon(parent), canon(component)): (parent, component, quantity) for parent, component, quantity in existing_lines}
    new = {(canon(parent), canon(component)): (parent, component, quantity) for parent, component, quantity in new_lines}
    existing_parents = {parent for parent, _ in existing}
    new_parents = {parent for parent, _ in new}
    result: list[CombinedBomLine] = []
    for key in sorted(set(existing) | set(new)):
        old = existing.get(key)
        fresh = new.get(key)
        parent = (fresh or old)[0]
        component = (fresh or old)[1]
        old_quantity = old[2] if old else None
        new_quantity = fresh[2] if fresh else None
        parent_key = key[0]
        fpack_source = (
            "EXISTING + NEW" if parent_key in existing_parents and parent_key in new_parents
            else "NEW" if parent_key in new_parents
            else "EXISTING"
        )
        if old and fresh:
            component_source = "EXISTING + NEW"
            change_status = "UNCHANGED" if abs(old_quantity - new_quantity) < 1e-9 else "QUANTITY CHANGED"
        elif fresh:
            component_source = "NEW"
            change_status = "ADDED" if parent_key in existing_parents else "NEW FPACK"
        else:
            component_source = "EXISTING"
            change_status = "REMOVED" if parent_key in new_parents else "EXISTING ONLY"
        effective_quantity = new_quantity if parent_key in new_parents else old_quantity
        result.append(CombinedBomLine(
            parent=parent,
            component=component,
            existing_quantity=old_quantity,
            new_quantity=new_quantity,
            effective_quantity=float(effective_quantity or 0.0),
            fpack_source=fpack_source,
            component_source=component_source,
            change_status=change_status,
        ))
    return result


def style_output(ws, last_row: int) -> None:
    dark_fill = PatternFill("solid", fgColor="1F4E78")
    light_fill = PatternFill("solid", fgColor="D9EAF7")
    group_fill = PatternFill("solid", fgColor="E9ECEF")
    white_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E2F3")
    medium_blue = Side(style="medium", color="5B9BD5")

    for cell in ws[1]:
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:U{last_row}"
    ws.sheet_view.showGridLines = False

    widths = {
        "A": 30, "B": 30, "C": 18, "D": 42, "E": 20, "F": 20,
        "G": 16, "H": 14, "I": 17, "J": 14, "K": 10, "L": 15,
        "M": 16, "N": 18, "O": 21, "P": 19, "Q": 25, "R": 17,
        "S": 17, "T": 20, "U": 22,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for row in range(2, last_row + 1):
        for col in (7, 8, 9):
            ws.cell(row, col).number_format = "0.####"
        for col in (12, 13):
            ws.cell(row, col).number_format = "0.000000"
        for col in range(14, 18):
            ws.cell(row, col).number_format = '0.0000 [$€-x-euro2]'
        ws.cell(row, 21).number_format = '0.0000 [$€-x-euro2]'
        for col in range(1, 22):
            ws.cell(row, col).alignment = Alignment(vertical="center")
            ws.cell(row, col).border = Border(bottom=thin_gray)
        if ws.cell(row, 1).value:
            for col in range(1, 22):
                ws.cell(row, col).fill = group_fill
                ws.cell(row, col).border = Border(top=medium_blue, bottom=thin_gray)
            ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 21).fill = light_fill


def add_parameters_sheet(wb: Workbook, parameters: PriceParameters) -> None:
    ws = wb.create_sheet("PARAMETERS")
    ws.append(["Parameter", "Value", "Unit / rule"])
    rows = [
        ("BACK rate", parameters.back_rate_per_m2, "EUR/m²"),
        ("Processing rate", parameters.processing_rate_per_m2, "EUR/m²"),
        ("WW material rate", parameters.ww_material_rate_per_m2, "EUR/m²"),
        ("BB material rate", parameters.bb_material_rate_per_m2, "EUR/m²"),
        ("NO material rate", parameters.no_material_rate_per_m2, "EUR/m²"),
        ("Small part threshold", parameters.small_part_threshold_m2, "m²; surcharge when area < threshold"),
        ("Small part surcharge", parameters.small_part_surcharge, "EUR/unit"),
        ("Furnix markup", parameters.furnix_markup_percent, "% of Furnix unit cost"),
        ("Output decimals", parameters.output_decimals, "digits after decimal point"),
    ]
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{ws.max_row}"
    ws["B2"].number_format = '0.0000 [$€-x-euro2]'
    ws["B3"].number_format = '0.0000 [$€-x-euro2]'
    ws["B4"].number_format = '0.0000 [$€-x-euro2]'
    ws["B5"].number_format = '0.0000 [$€-x-euro2]'
    ws["B6"].number_format = '0.0000 [$€-x-euro2]'
    ws["B7"].number_format = "0.000000"
    ws["B8"].number_format = '0.0000 [$€-x-euro2]'
    ws["B9"].number_format = '0.00"%"'
    ws["B10"].number_format = "0"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 42
    ws.sheet_view.showGridLines = False


def add_fpack_breakdown_sheet(
    wb: Workbook,
    lines: list[CombinedBomLine],
    calculations: dict[str, PriceCalculation],
    parameters: PriceParameters,
) -> None:
    ws = wb.create_sheet("FPACK PRICE BREAKDOWN", 1)
    ws.append([
        "FPACK SKU", "FPACK Source", "Component SKU", "Component Source", "Change Status",
        "Existing Quantity", "New Quantity", "Effective Quantity", "Part Type", "Color",
        "Dimension 1, mm", "Dimension 2, mm", "Unit Area, m²", "Total Area, m²",
        "Material Cost, EUR", "BACK Cost, EUR", "Processing Cost, EUR",
        "Small Part Surcharge, EUR", "Calculated Unit Cost, EUR",
        "Furnix Markup, %", "Furnix Markup, EUR/unit",
        "Furnix Sales Price to Furnibox, EUR/unit",
        "Component Total Purchase Price, EUR", "FPACK Cabinet Parts Purchase Price, EUR",
    ])
    parent_totals: dict[str, float] = defaultdict(float)
    prepared: list[tuple[CombinedBomLine, PriceCalculation]] = []
    for line in lines:
        calculation = calculations.get(canon(line.component))
        if calculation is None:
            continue
        prepared.append((line, calculation))
        rounded_unit_cost = round(calculation.unit_price, parameters.output_decimals)
        _, transfer_price = furnix_transfer_price(rounded_unit_cost, parameters)
        parent_totals[canon(line.parent)] += (
            round(transfer_price, parameters.output_decimals) * line.effective_quantity
        )
    for line, item in prepared:
        total_area = item.area_m2 * line.effective_quantity
        rounded_unit_cost = round(item.unit_price, parameters.output_decimals)
        markup_eur, transfer_price = furnix_transfer_price(rounded_unit_cost, parameters)
        rounded_transfer_price = round(transfer_price, parameters.output_decimals)
        ws.append([
            line.parent, line.fpack_source, line.component, line.component_source, line.change_status,
            line.existing_quantity, line.new_quantity, line.effective_quantity, item.part_type, item.color,
            item.dimension_1_mm, item.dimension_2_mm, item.area_m2, total_area,
            total_area * item.material_rate_per_m2,
            total_area * item.back_rate_per_m2,
            total_area * item.processing_rate_per_m2,
            item.small_part_surcharge * line.effective_quantity,
            rounded_unit_cost,
            parameters.furnix_markup_percent / 100,
            round(markup_eur, parameters.output_decimals),
            rounded_transfer_price,
            rounded_transfer_price * line.effective_quantity,
            parent_totals[canon(line.parent)],
        ])
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:X{max(ws.max_row, 1)}"
    ws.sheet_view.showGridLines = False
    widths = [30, 18, 44, 20, 20, 16, 14, 17, 14, 10, 17, 17, 15, 16, 19, 17, 21, 25, 23, 17, 25, 32, 32, 34]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in range(2, ws.max_row + 1):
        for column in (6, 7, 8):
            ws.cell(row, column).number_format = "0.####"
        for column in (13, 14):
            ws.cell(row, column).number_format = "0.000000"
        ws.cell(row, 20).number_format = "0.00%"
        for column in list(range(15, 20)) + list(range(21, 25)):
            decimals = "0" * parameters.output_decimals
            ws.cell(row, column).number_format = (
                f'0.{decimals} [$€-x-euro2]' if decimals else '0 [$€-x-euro2]'
            )


def add_unique_prices_sheet(
    wb: Workbook,
    calculations: dict[str, PriceCalculation],
    odoo_prices: dict[str, OdooProductPrice],
    duplicate_odoo_skus: dict[str, list[int]],
    odoo_checked: bool,
    parameters: PriceParameters,
    existing_component_skus: set[str],
    new_component_skus: set[str],
) -> None:
    ws = wb.create_sheet("CABINET PART PRICES", 0)
    headers = [
        "Internal Reference", "Odoo Product ID", "Odoo Active",
        "Current Furnibox Purchase Cost", "Furnix Unit Cost",
        "Furnix Markup, %", "Furnix Markup, EUR",
        "Furnix Sales Price to Furnibox", "Change EUR", "Change %",
        "Product Status", "BOM Source", "Proposed Action", "Part Type", "Color",
        "Dimension 1, mm", "Dimension 2, mm", "Area, m²",
    ]
    ws.append(headers)
    for key in sorted(calculations):
        item = calculations[key]
        furnix_cost = round(item.unit_price, parameters.output_decimals)
        markup_eur, transfer_price = furnix_transfer_price(furnix_cost, parameters)
        transfer_price = round(transfer_price, parameters.output_decimals)
        current = odoo_prices.get(key)
        duplicate_ids = duplicate_odoo_skus.get(key)
        if duplicate_ids:
            status = "DUPLICATE IN ODOO"
            action = "REVIEW DUPLICATE"
            current_price = change = change_percent = None
            product_id = ", ".join(map(str, duplicate_ids))
            active = None
        elif current:
            status = "EXISTING" if current.active else "ARCHIVED"
            current_price = current.standard_price
            change = transfer_price - current_price
            change_percent = change / current_price if current_price else None
            if not current.active:
                action = "REVIEW ARCHIVED"
            else:
                action = "NO CHANGE" if abs(change) < 0.00005 else "UPDATE COST"
            product_id = current.product_id
            active = "YES" if current.active else "NO"
        elif odoo_checked:
            status = "NEW"
            action = "CREATE PRODUCT / SET COST"
            current_price = change = change_percent = None
            product_id = active = None
        else:
            status = "ODOO NOT CHECKED"
            action = "CHECK ODOO"
            current_price = change = change_percent = None
            product_id = active = None
        if key in existing_component_skus and key in new_component_skus:
            bom_source = "EXISTING + NEW"
        elif key in new_component_skus:
            bom_source = "NEW"
        elif key in existing_component_skus:
            bom_source = "EXISTING"
        elif current and not current.active:
            bom_source = "ARCHIVED"
        else:
            bom_source = "NOT IN ACTIVE FPACK BOM"
        ws.append([
            item.sku, product_id, active, current_price,
            furnix_cost, parameters.furnix_markup_percent / 100,
            round(markup_eur, parameters.output_decimals), transfer_price,
            change, change_percent, status, bom_source, action, item.part_type, item.color,
            item.dimension_1_mm, item.dimension_2_mm, item.area_m2,
        ])
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:R{max(ws.max_row, 1)}"
    for column, width in {
        "A": 44, "B": 18, "C": 13, "D": 28, "E": 20, "F": 17,
        "G": 20, "H": 30, "I": 16, "J": 14, "K": 20, "L": 22,
        "M": 28, "N": 15, "O": 10, "P": 17, "Q": 17, "R": 14,
    }.items():
        ws.column_dimensions[column].width = width
    for row in range(2, ws.max_row + 1):
        for column in (4, 5, 7, 8, 9):
            ws.cell(row, column).number_format = '0.0000 [$€-x-euro2]'
        ws.cell(row, 6).number_format = "0.00%"
        ws.cell(row, 10).number_format = "0.00%"
        ws.cell(row, 18).number_format = "0.000000"
    ws.sheet_view.showGridLines = False


def build_workbook(
    source_path: Path,
    output_path: Path,
    parameters: PriceParameters = DEFAULT_PARAMETERS,
    odoo_client: OdooClient | None = None,
) -> tuple[int, int, int, int]:
    source_wb = load_workbook(source_path, read_only=False, data_only=False)
    required_sheets = {"NEW BOM LINES"}
    missing_sheets = sorted(required_sheets - set(source_wb.sheetnames))
    if missing_sheets:
        raise ValueError("Trūksta lapų: " + ", ".join(missing_sheets))

    explicit_skus: dict[str, str] = {}
    if "NEW PRODUCTS" in source_wb.sheetnames:
        explicit_skus, _ = load_new_products(source_wb["NEW PRODUCTS"])
    new_lines, cabinet_part_skus, dimensions, invalid_quantities = load_fpack_lines(
        source_wb["NEW BOM LINES"], set(explicit_skus)
    )

    new_component_skus = {canon(component) for _, component, _ in new_lines}
    existing_lines: list[tuple[str, str, float]] = []
    odoo_bom_errors: list[tuple[str, str]] = []
    odoo_prices: dict[str, OdooProductPrice] = {}
    duplicate_odoo_skus: dict[str, list[int]] = {}
    odoo_product_errors: list[tuple[str, str]] = []
    if odoo_client is not None:
        odoo_prices, duplicate_odoo_skus, odoo_product_errors = (
            load_odoo_cabinet_parts(odoo_client)
        )
        existing_lines, odoo_bom_errors = load_odoo_fpack_lines(odoo_client)
        for key, item in odoo_prices.items():
            cabinet_part_skus[key] = item.sku
            parsed = parse_dimensions(item.sku)
            if parsed:
                dimensions[key] = parsed
            else:
                odoo_product_errors.append((item.sku, "SKU nerasti matmenys"))
        for key in duplicate_odoo_skus:
            cabinet_part_skus.setdefault(key, key)
            parsed = parse_dimensions(cabinet_part_skus[key])
            if parsed:
                dimensions[key] = parsed
            else:
                odoo_product_errors.append((cabinet_part_skus[key], "SKU nerasti matmenys"))

    lines = combine_bom_lines(existing_lines, new_lines)
    existing_component_skus = {canon(component) for _, component, _ in existing_lines}
    bom_skus = existing_component_skus | new_component_skus

    audit_skus = set(dimensions)
    if not audit_skus:
        raise ValueError("Faile ir Odoo nerasta nė vienos Cabinet Part detalės su matmenimis.")

    calculations: dict[str, PriceCalculation] = {}
    calculation_errors: list[tuple[str, str]] = []
    for component_key in sorted(audit_skus):
        try:
            calculations[component_key] = calculate_unit_price(
                cabinet_part_skus[component_key], dimensions[component_key], parameters
            )
        except ValueError as exc:
            calculation_errors.append((cabinet_part_skus[component_key], str(exc)))

    wb = Workbook()
    ws = wb.active
    ws.title = "CALCULATION DETAILS"
    headers = [
        "Product/Internal Reference",
        "FPACK SKU",
        "FPACK Source",
        "BoM Lines/Component/Internal Reference",
        "Component Source", "Change Status",
        "Existing Quantity", "New Quantity", "Effective Quantity",
        "Part Type", "Color", "Unit Area, m²", "BOM Area, m²",
        "BACK Rate, EUR/m²", "Processing Rate, EUR/m²",
        "Material Rate, EUR/m²", "Small Part Surcharge, EUR",
        "Dimension 1, mm", "Dimension 2, mm", "Calculation Status",
        "Calculated Unit Price, EUR",
    ]
    ws.append(headers)

    previous_parent = None
    for excel_row, line in enumerate(lines, start=2):
        component_key = canon(line.component)
        width, height = dimensions[component_key]
        quantity = line.effective_quantity
        quantity_value: int | float = int(quantity) if quantity.is_integer() else quantity
        first_in_group = canon(line.parent) != canon(previous_parent)
        calculation = calculations.get(component_key)
        ws.append([
            line.parent if first_in_group else None,
            line.parent, line.fpack_source,
            line.component, line.component_source, line.change_status,
            line.existing_quantity, line.new_quantity, quantity_value,
            calculation.part_type if calculation else None,
            calculation.color if calculation else parse_color(line.component),
            calculation.area_m2 if calculation else width * height / 1_000_000,
            width * height * quantity / 1_000_000,
            calculation.back_rate_per_m2 if calculation else None,
            calculation.processing_rate_per_m2 if calculation else None,
            calculation.material_rate_per_m2 if calculation else None,
            calculation.small_part_surcharge if calculation else None,
            width, height,
            "CALCULATED" if calculation else "ERROR",
            calculation.unit_price if calculation else None,
        ])
        previous_parent = line.parent

    last_row = max(ws.max_row, 1)
    style_output(ws, last_row)

    diagnostics = wb.create_sheet("DIAGNOSTICS")
    diagnostics.append(["Type", "SKU / ryšys", "Message"])
    for message in invalid_quantities:
        diagnostics.append(["INVALID QUANTITY", message.split(":", 1)[0], message])
    for sku, message in calculation_errors:
        diagnostics.append(["PRICE CALCULATION ERROR", sku, message])
    for sku, message in odoo_product_errors:
        diagnostics.append(["ODOO PRODUCT ERROR", sku, message])
    for sku, message in odoo_bom_errors:
        diagnostics.append(["ODOO FPACK BOM ERROR", sku, message])
    for sku, product_ids in sorted(duplicate_odoo_skus.items()):
        diagnostics.append([
            "DUPLICATE ODOO SKU", sku,
            "Tas pats Internal Reference rastas keliose Odoo kortelėse: "
            + ", ".join(map(str, product_ids)),
        ])
    for cell in diagnostics[1]:
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.font = Font(color="FFFFFF", bold=True)
    diagnostics.freeze_panes = "A2"
    diagnostics.auto_filter.ref = f"A1:C{max(diagnostics.max_row, 1)}"
    diagnostics.column_dimensions["A"].width = 42
    diagnostics.column_dimensions["B"].width = 45
    diagnostics.column_dimensions["C"].width = 75
    diagnostics.sheet_view.showGridLines = False

    info = wb.create_sheet("INFO")
    info.append(["Parameter", "Value"])
    info.append(["Script version", SCRIPT_VERSION])
    info.append(["Source", str(source_path)])
    info.append(["Dimensional FPACK components", len(cabinet_part_skus)])
    info.append(["Combined audit rows", len(lines)])
    info.append(["Unique Cabinet Parts from BOM", len(bom_skus)])
    info.append(["Existing Odoo FPACK count", len({canon(parent) for parent, _, _ in existing_lines})])
    info.append(["New MAP FPACK count", len({canon(parent) for parent, _, _ in new_lines})])
    info.append(["Combined FPACK count", len({canon(line.parent) for line in lines})])
    info.append(["Invalid quantities", len(invalid_quantities)])
    info.append(["Calculated unique Cabinet Parts", len(calculations)])
    info.append(["Price calculation errors", len(calculation_errors)])
    info.append(["Odoo prices checked", "YES" if odoo_client is not None else "NO"])
    info.append(["Existing unique products in Odoo", len(odoo_prices)])
    info.append(["New products not found in Odoo", len(set(calculations) - set(odoo_prices) - set(duplicate_odoo_skus)) if odoo_client is not None else "N/A"])
    info.append(["Duplicate Odoo SKU", len(duplicate_odoo_skus)])
    for cell in info[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    info.column_dimensions["A"].width = 40
    info.column_dimensions["B"].width = 90
    info.sheet_view.showGridLines = False

    add_unique_prices_sheet(
        wb, calculations, odoo_prices, duplicate_odoo_skus,
        odoo_checked=odoo_client is not None,
        parameters=parameters,
        existing_component_skus=existing_component_skus,
        new_component_skus=new_component_skus,
    )
    add_fpack_breakdown_sheet(wb, lines, calculations, parameters)
    add_parameters_sheet(wb, parameters)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return (
        len(lines),
        len(audit_skus),
        len({canon(line.parent) for line in lines}),
        len(invalid_quantities) + len(calculation_errors) + len(odoo_product_errors) + len(odoo_bom_errors) + len(duplicate_odoo_skus),
    )


def main() -> None:
    from config import load_settings

    print(f"Versija: {SCRIPT_VERSION}")
    settings = load_settings()
    output_dir = environment_output_dir(BASE_DIR)
    source_path = output_dir / SOURCE_NAME
    if not source_path.exists():
        raise FileNotFoundError(
            f"Nerastas 6 veiksmo rezultatas:\n{source_path}\n"
            "Pirmiausia paleiskite 6 veiksmą „Palyginti MAP“."
        )

    parameters_path = BASE_DIR / PARAMETERS_NAME
    parameters = load_parameters(parameters_path)
    print(f"Parametrai: {parameters_path}")
    print(f"Apvalinimas: {parameters.output_decimals} skaitmenys po kablelio")
    output_path = output_dir / OUTPUT_NAME
    client = OdooClient(settings)
    uid = client.authenticate()
    print(f"Prisijungta prie Odoo. UID={uid}")
    rows, unique_parts, fpack_count, diagnostics = build_workbook(
        source_path,
        output_path,
        parameters=parameters,
        odoo_client=client,
    )

    print("CABINET PARTS SAVIKAINOS APSKAIČIUOTOS IR PALYGINTOS SU ODOO")
    print(f"Aplinka: {os.environ.get('FURNIBOX_ENVIRONMENT', 'NEŽINOMA')}")
    print(f"Eilučių: {rows}")
    print(f"Unikalių Cabinet Parts: {unique_parts}")
    print(f"FPACK: {fpack_count}")
    print(f"Diagnostikoje: {diagnostics}")
    print(f"Rezultatas: {output_path}")
    print("Odoo duomenys nepakeisti.")


if __name__ == "__main__":
    main()
