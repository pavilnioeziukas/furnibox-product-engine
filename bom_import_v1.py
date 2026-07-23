"""Generuoja visų naujų BOM Odoo importo failą pagal hierarchijos lygius.

Saugos principai:
- leidžiama vykdyti tik Stage aplinkoje;
- Odoo naudojami tik skaitymo metodai;
- importui paruošiami tik visiškai patikrinti BOM;
- trūkumai pateikiami REVIEW ir DIAGNOSTICS lapuose.
"""

from __future__ import annotations

from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bom_import_pilot_v1 import (
    calculate_levels,
    canon,
    load_bom_types,
    load_new_bom_graph,
    load_odoo_product_ids,
)
from bom_import_pilot_v2 import (
    choose_operation_template,
    load_operation_templates,
)
from config import load_settings
from odoo_client import OdooClient
from output_paths import environment_output_dir, environment_slug


IMPORT_HEADERS = [
    "Product_tmpl_id/id", "qty",
    "BoM Lines/Component/External ID", "product_qty",
    "BoM Type", "Reference", "mo_autodone_by_wo", "auto_plan",
    "operation_ids/name", "operation_ids/workcenter_id/id",
    "operation_ids/time_mode", "operation_ids/time_cycle_manual",
    "operation_ids/sequence",
]


def load_external_ids(
    client: OdooClient, model: str, res_ids: set[int]
) -> dict[int, str]:
    """Grąžina po vieną tikrą External ID kiekvienam Odoo įrašui."""
    if not res_ids:
        return {}
    rows = client.search_read_all(
        "ir.model.data",
        [["model", "=", model], ["res_id", "in", sorted(res_ids)]],
        ["module", "name", "res_id"],
    )
    grouped: dict[int, list[str]] = {}
    for row in rows:
        res_id = int(row["res_id"])
        external_id = f"{row['module']}.{row['name']}"
        grouped.setdefault(res_id, []).append(external_id)

    # Jei įrašas turi kelis External ID, prioritetas teikiamas __export__ ID.
    result = {}
    for res_id, values in grouped.items():
        result[res_id] = sorted(
            set(values),
            key=lambda value: (not value.startswith("__export__."), value),
        )[0]
    return result


def attach_product_external_ids(client: OdooClient, products: dict) -> None:
    """Vidinius produkto ID pakeičia importui skirtais External ID laukais."""
    template_ids = {
        int(product["template_id"]) for product in products.values()
        if product.get("template_id")
    }
    product_ids = {
        int(product["product_id"]) for product in products.values()
        if product.get("product_id")
    }
    template_external_ids = load_external_ids(
        client, "product.template", template_ids
    )
    product_external_ids = load_external_ids(
        client, "product.product", product_ids
    )
    for product in products.values():
        template_id = int(product["template_id"])
        product_id = int(product["product_id"])
        product["template_external_id"] = template_external_ids.get(template_id, "")
        product["product_external_id"] = product_external_ids.get(product_id, "")


def load_stage_subcategories(client: OdooClient, parent_skus: set[str]) -> dict[str, str]:
    """Vienu API nuskaitymu susieja BOM produktus su Stage pakategorėmis."""
    products = client.search_read_all(
        "product.product", [["default_code", "!=", False]],
        ["id", "default_code", "categ_id"], context={"active_test": False},
    )
    category_ids = {
        int(row["categ_id"][0]) for row in products
        if canon(row.get("default_code")) in parent_skus and row.get("categ_id")
    }
    categories = client.search_read_all(
        "product.category", [["id", "in", sorted(category_ids)]],
        ["id", "name", "complete_name"],
    ) if category_ids else []
    category_by_id = {int(row["id"]): row for row in categories}
    result = {}
    for row in products:
        sku = canon(row.get("default_code"))
        if sku not in parent_skus or not row.get("categ_id"):
            continue
        category = category_by_id.get(int(row["categ_id"][0]), {})
        path = str(category.get("complete_name") or category.get("name") or "")
        result[sku] = path.rsplit("/", 1)[-1].strip()
    return result


def load_stage_workcenter_ids(client: OdooClient) -> dict[str, str]:
    """Grąžina unikalių aktyvių Stage darbo centrų External ID pagal vardą."""
    rows = client.search_read_all(
        "mrp.workcenter", [], ["id", "name", "active"],
        context={"active_test": False},
    )
    grouped: dict[str, int] = {}
    duplicates = set()
    for row in rows:
        if not row.get("active", True):
            continue
        name = str(row.get("name") or "").strip()
        if not name:
            continue
        if name in grouped:
            duplicates.add(name)
        grouped[name] = int(row["id"])
    for name in duplicates:
        grouped.pop(name, None)
    external_ids = load_external_ids(client, "mrp.workcenter", set(grouped.values()))
    return {
        name: external_ids[res_id]
        for name, res_id in grouped.items()
        if res_id in external_ids
    }


def prepare_boms(parents, lines, levels, bom_types, products, subcategories,
                 operation_templates, workcenters):
    """Patikrina kiekvieną BOM ir paruošia importo arba diagnostikos įrašą."""
    ready = []
    review = []
    diagnostics = []

    for sku in sorted(parents, key=lambda value: (levels[value], value)):
        bom_lines = lines.get(sku, [])
        bom_type = bom_types.get(sku, "")
        status = "READY"
        messages = []
        template = None
        operations = []

        parent = products.get(sku)
        if not parent:
            messages.append("Stage nerastas BOM produktas")
        elif not parent.get("template_external_id"):
            messages.append("BOM produktas neturi product.template External ID")
        if not bom_lines:
            messages.append("BOM neturi komponentų")
        if not bom_type:
            messages.append("Nėra patvirtinto BOM tipo")

        for line in bom_lines:
            component_sku = line["component"]
            component = products.get(component_sku)
            if not component:
                messages.append(f"Stage nerastas komponentas: {component_sku}")
            elif not component.get("product_external_id"):
                messages.append(
                    f"Komponentas neturi product.product External ID: {component_sku}"
                )

        subcategory = subcategories.get(sku, "")
        if bom_type == "Manufacture this product":
            if not subcategory:
                messages.append("Nenustatyta Stage produkto pakategorė")
            else:
                try:
                    template = choose_operation_template(sku, subcategory, operation_templates)
                    operations = template["operations"]
                except ValueError as exc:
                    messages.append(str(exc))
            if not operations:
                messages.append("Manufacture BOM nerastas Production operacijų etalonas")
            for operation in operations:
                name = operation["workcenter"]
                if name not in workcenters:
                    messages.append(f"Stage nerastas darbo centras: {name}")

        if messages:
            status = "NOT READY"
            for message in dict.fromkeys(messages):
                diagnostics.append([sku, f"lv{levels[sku]}", bom_type, message])

        record = {
            "sku": sku,
            "level": levels[sku],
            "type": bom_type,
            "lines": bom_lines,
            "operations": operations,
            "subcategory": subcategory,
            "template": template,
            "status": status,
            "message": "; ".join(dict.fromkeys(messages)),
        }
        review.append(record)
        if status == "READY":
            ready.append(record)
    return ready, review, diagnostics


def import_rows(record: dict, products: dict, workcenters: dict[str, str]):
    """Vieną BOM paverčia Odoo vienas-prie-daugelio importo eilučių bloku."""
    sku = record["sku"]
    parent = products[sku]
    lines = record["lines"]
    operations = record["operations"]
    row_count = max(len(lines), len(operations), 1)
    reference = f"{date.today():%Y%m%d}_Furnibox_{sku}"
    rows = []
    for index in range(row_count):
        first = index == 0
        line = lines[index] if index < len(lines) else None
        operation = operations[index] if index < len(operations) else None
        component = products[line["component"]] if line else None
        manufacture = record["type"] == "Manufacture this product"
        rows.append([
            parent["template_external_id"] if first else None,
            1 if first else None,
            component["product_external_id"] if component else None,
            line["quantity"] if line else None,
            record["type"] if first else None,
            reference if first else None,
            True if first and manufacture else None,
            True if first and operations else None,
            operation["name"] if operation else None,
            workcenters[operation["workcenter"]] if operation else None,
            operation["time_mode"] if operation else None,
            operation["time"] if operation else None,
            operation["sequence"] if operation else None,
        ])
    return rows


def style_sheet(ws, widths=None):
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index in range(1, ws.max_column + 1):
        if widths and index <= len(widths):
            width = widths[index - 1]
        else:
            values = [
                len(str(ws.cell(row, index).value or ""))
                for row in range(1, min(ws.max_row, 300) + 1)
            ]
            width = min(max(values, default=12) + 2, 55)
        ws.column_dimensions[get_column_letter(index)].width = width


def write_workbook(path: Path, ready, review, diagnostics, products, workcenters):
    wb = Workbook()
    wb.remove(wb.active)

    # 1. Kiekvienam hierarchijos lygiui sukuriamas atskiras importo lapas.
    levels = sorted({record["level"] for record in ready})
    for level in levels:
        ws = wb.create_sheet(f"BOM_import(lv{level})")
        ws.append(IMPORT_HEADERS)
        for record in (row for row in ready if row["level"] == level):
            for row in import_rows(record, products, workcenters):
                ws.append(row)
        style_sheet(ws, [42, 8, 42, 12, 28, 40, 22, 14, 28, 38, 18, 24, 20])

    # 2. REVIEW leidžia prieš importą matyti, kas pateko ir kas buvo sustabdyta.
    ws = wb.create_sheet("REVIEW", 0)
    review_headers = [
        "Status", "Level", "Parent SKU", "BOM Type", "Stage Subcategory",
        "Components", "Operations", "Production Analog", "Analog BOM Reference", "Message",
    ]
    ws.append(review_headers)
    for record in review:
        template = record.get("template") or {}
        ws.append([
            record["status"], f"lv{record['level']}", record["sku"], record["type"],
            record["subcategory"], len(record["lines"]), len(record["operations"]),
            template.get("sku", ""), template.get("reference", ""), record["message"],
        ])
    style_sheet(ws)

    # 3. SUMMARY pateikia kontrolinius skaičius pagal lygį, tipą ir būseną.
    ws = wb.create_sheet("SUMMARY", 1)
    ws.append(["Metric", "Value"])
    ws.append(["All new BOM", len(review)])
    ws.append(["Ready for import", len(ready)])
    ws.append(["Not ready", len(review) - len(ready)])
    ws.append(["KIT ready", sum(record["type"] == "KIT" for record in ready)])
    ws.append([
        "Manufacture ready",
        sum(record["type"] == "Manufacture this product" for record in ready),
    ])
    for (level, status), count in sorted(
        Counter((r["level"], r["status"]) for r in review).items()
    ):
        ws.append([f"lv{level} {status}", count])
    style_sheet(ws)

    ws = wb.create_sheet("DIAGNOSTICS")
    ws.append(["Parent SKU", "Level", "BOM Type", "Message"])
    for row in diagnostics:
        ws.append(row)
    style_sheet(ws)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_import_file(path: Path, records, products, workcenters):
    """Sukuria vieno lygio švarų failą, skirtą tiesioginiam Odoo importui."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM import"
    ws.append(IMPORT_HEADERS)
    for record in records:
        for row in import_rows(record, products, workcenters):
            ws.append(row)
    style_sheet(ws, [42, 8, 42, 12, 28, 40, 22, 14, 28, 38, 18, 24, 20])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    base = Path(__file__).resolve().parent
    if environment_slug() != "stage":
        raise PermissionError("Visų BOM importo generatorius leidžiamas tik Stage.")
    output_dir = environment_output_dir(base)
    comparison_path = output_dir / "MAP_Comparison.xlsx"
    types_path = output_dir / "BOM_Type_Review.xlsx"
    references_path = base / "output" / "production" / "BOM_Operations_Reference.xlsx"
    output_path = output_dir / "BOM_Import_All_Levels.xlsx"
    lv2_output_path = output_dir / "BOM_Import_All_lv2.xlsx"
    lv1_output_path = output_dir / "BOM_Import_All_lv1.xlsx"
    kit_lv2_output_path = output_dir / "BOM_Import_KIT_lv2.xlsx"
    kit_lv1_output_path = output_dir / "BOM_Import_KIT_lv1.xlsx"

    # 1 ŽINGSNIS: nuskaitome naujus BOM, jų tipus ir apskaičiuojame lygius.
    parents, lines = load_new_bom_graph(comparison_path)
    bom_types = load_bom_types(types_path)
    levels = calculate_levels(set(parents), lines)

    # 2 ŽINGSNIS: iš Stage pasiimame aktualius produkto ir komponento ID.
    settings = load_settings()
    client = OdooClient(settings)
    uid = client.authenticate()
    wanted = set(parents) | {
        row["component"] for values in lines.values() for row in values
    }
    products = load_odoo_product_ids(client, wanted)
    attach_product_external_ids(client, products)
    subcategories = load_stage_subcategories(client, set(parents))

    # 3 ŽINGSNIS: Production etalonai nustato Manufacture operacijas ir laikus.
    operation_templates = load_operation_templates(references_path)
    workcenters = load_stage_workcenter_ids(client)

    # 4 ŽINGSNIS: tik visiškai patikrinti BOM patenka į importo lapus.
    ready, review, diagnostics = prepare_boms(
        parents, lines, levels, bom_types, products, subcategories,
        operation_templates, workcenters,
    )
    write_workbook(output_path, ready, review, diagnostics, products, workcenters)

    # 5 ŽINGSNIS: sukuriami atskiri tiesioginio importo failai.
    lv2_ready = [record for record in ready if record["level"] == 2]
    lv1_ready = [record for record in ready if record["level"] == 1]
    write_import_file(lv2_output_path, lv2_ready, products, workcenters)
    write_import_file(lv1_output_path, lv1_ready, products, workcenters)

    # 6 ŽINGSNIS: atskiri KIT-only failai naudojami, kai Manufacture BOM jau
    # buvo importuoti ankstesniu veiksmu. Taip išvengiama jų dubliavimo.
    kit_ready = [record for record in ready if record["type"] == "KIT"]
    kit_lv2_ready = [record for record in kit_ready if record["level"] == 2]
    kit_lv1_ready = [record for record in kit_ready if record["level"] == 1]
    write_import_file(kit_lv2_output_path, kit_lv2_ready, products, workcenters)
    write_import_file(kit_lv1_output_path, kit_lv1_ready, products, workcenters)

    print("Prisijungta prie Stage Odoo. UID=", uid)
    print("\nVISŲ BOM IMPORTO FAILAI SUKURTI")
    print("Kontrolinis failas:", output_path)
    print("Odoo importas lv2:", lv2_output_path, f"({len(lv2_ready)} BOM)")
    print("Odoo importas lv1:", lv1_output_path, f"({len(lv1_ready)} BOM)")
    print("\nKIT-ONLY IMPORTO FAILAI (kai Manufacture jau importuoti)")
    print("KIT importas lv2:", kit_lv2_output_path, f"({len(kit_lv2_ready)} BOM)")
    print("KIT importas lv1:", kit_lv1_output_path, f"({len(kit_lv1_ready)} BOM)")
    print("Iš viso KIT:", len(kit_ready))
    print("Visi nauji BOM:", len(review))
    print("Paruošti importui:", len(ready))
    print("Dar neparuošti:", len(review) - len(ready))
    print(
        "Importo lygiai:",
        ", ".join(f"lv{x}" for x in sorted({r["level"] for r in ready})) or "nėra",
    )
    print("Diagnostikos įrašai:", len(diagnostics))
    print("Odoo pakeitimų neatlikta.")


if __name__ == "__main__":
    main()