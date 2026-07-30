from __future__ import annotations

import tempfile
import unittest
import sys
from types import ModuleType
from pathlib import Path

from openpyxl import load_workbook

if "dotenv" not in sys.modules:
    dotenv_stub = ModuleType("dotenv")
    dotenv_stub.load_dotenv = lambda *_args, **_kwargs: None
    sys.modules["dotenv"] = dotenv_stub

from bom_release.generator import (
    APACK_OPERATION_TYPE_EXTERNAL_ID,
    HRD_OPERATION_TYPE_EXTERNAL_ID,
    generate_release_files,
    required_operation_type_external_id,
)
from bom_release.models import (
    BomReleasePlan,
    BomReleasePlanItem,
    ReleaseAction,
    ReleaseStatus,
)


class FakeClient:
    def search_read_all(self, model, domain, fields, context=None):
        if model == "product.product":
            return [
                {
                    "id": 1,
                    "default_code": "PARENT",
                    "product_tmpl_id": [11, "PARENT"],
                    "active": True,
                },
                {
                    "id": 2,
                    "default_code": "PART",
                    "product_tmpl_id": [12, "PART"],
                    "active": True,
                },
            ]
        if model == "ir.model.data":
            source = {
                ("product.template", 11): ("x", "parent_tmpl"),
                ("product.product", 2): ("x", "part_variant"),
            }
            rows = []
            for (source_model, res_id), (module, name) in source.items():
                if source_model == domain[0][2] and res_id in domain[1][2]:
                    rows.append(
                        {
                            "module": module,
                            "name": name,
                            "res_id": res_id,
                        }
                    )
            return rows
        if model == "mrp.workcenter":
            return [
                {"id": 21, "name": "Pakavimas", "active": True},
            ]
        raise AssertionError(model)


def dataset():
    return {
        "schema_version": "1",
        "dataset_id": "dataset-1",
        "batch_reference": "batch-1",
        "environment": "production",
        "created_at_utc": "2026-07-30T00:00:00Z",
        "source": {},
        "statistics": {},
        "products": [
            {
                "sku": "PARENT",
                "product_type": "OTHER",
                "bom_type": "Manufacture this product",
                "level": 2,
                "source_sku": "PARENT",
                "generated_from": "",
                "reform_category": "OTHER",
                "content_hash": "hash",
                "content_signature": "sig",
                "components": [
                    {
                        "sku": "PART",
                        "quantity": 2.5,
                        "parent_sku": "PARENT",
                        "level": 2,
                    }
                ],
                "operations": [
                    {
                        "name": "Pakavimas",
                        "workcenter": "Pakavimas",
                        "time_mode": "manual",
                        "time_minutes": 3.0,
                        "sequence": 100,
                    }
                ],
            }
        ],
    }


def plan():
    item = BomReleasePlanItem(
        parent_sku="PARENT",
        bom_type="Manufacture this product",
        component_count=1,
        operation_count=1,
        product_exists=True,
        product_id=1,
        product_template_id=11,
        parent_external_id_ready=True,
        active_bom_count=0,
        active_bom_id=None,
        active_reference="",
        active_sequence=None,
        active_bom_type="",
        release_exists=False,
        release_bom_id=None,
        release_reference="RELEASE-1",
        action=ReleaseAction.CREATE,
        status=ReleaseStatus.READY,
    )
    return BomReleasePlan(
        release_id="RELEASE-1",
        release_reference="RELEASE-1",
        environment="production",
        dataset_id="dataset-1",
        dataset_batch_reference="batch-1",
        dataset_path="dataset.json",
        created_at_utc="2026-07-30T00:00:00Z",
        items=[item],
    )


class BomReleaseGeneratorTests(unittest.TestCase):
    def test_generates_import_only_from_dataset_contract(self):
        with tempfile.TemporaryDirectory() as directory:
            files = generate_release_files(
                dataset=dataset(),
                plan=plan(),
                client=FakeClient(),
                output_dir=Path(directory),
            )
            self.assertEqual(len(files), 1)
            release_file = next(item for item in files if item.level > 0)
            self.assertEqual(release_file.bom_count, 1)
            wb = load_workbook(release_file.path, data_only=True)
            try:
                ws = wb["BOM import"]
                self.assertEqual(ws["A2"].value, "PARENT")
                self.assertEqual(ws["B2"].value, "x.parent_tmpl")
                self.assertEqual(ws["D2"].value, "PART")
                self.assertEqual(ws["E2"].value, "x.part_variant")
                self.assertEqual(ws["F2"].value, 2.5)
                self.assertEqual(ws["H2"].value, 10)
                self.assertIsNone(ws["I2"].value)
                self.assertEqual(ws["J2"].value, "RELEASE-1")
                self.assertEqual(ws["P2"].value, 3)
            finally:
                wb.close()

    def test_assigns_required_operation_type_external_id(self):
        source = dataset()
        source["products"][0]["sku"] = "APACK-PARENT-A"
        source["products"][0]["product_type"] = "PREPACK CABINETS"
        self.assertEqual(
            required_operation_type_external_id(source["products"][0]),
            APACK_OPERATION_TYPE_EXTERNAL_ID,
        )
        source["products"][0]["sku"] = "UNI-P-ACC01-HRD201D"
        source["products"][0]["product_type"] = "CABINET HARDWARE"
        self.assertEqual(
            required_operation_type_external_id(source["products"][0]),
            HRD_OPERATION_TYPE_EXTERNAL_ID,
        )

    def test_removes_unsafe_legacy_old_bom_file(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory)
            legacy = output / (
                "BOM_Release_RELEASE-1_00_Old_BOM_Sequence_10.xlsx"
            )
            legacy.write_bytes(b"obsolete")
            generate_release_files(
                dataset=dataset(),
                plan=plan(),
                client=FakeClient(),
                output_dir=output,
            )
            self.assertFalse(legacy.exists())


if __name__ == "__main__":
    unittest.main()
