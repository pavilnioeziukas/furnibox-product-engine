from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from bom_release.generator import (
    APACK_OPERATION_TYPE_EXTERNAL_ID,
    IMPORT_HEADERS,
    MANUFACTURE,
    NEW_BOM_SEQUENCE,
    _style,
    canon,
)


class PilotPreparationError(RuntimeError):
    """Release faile nepavyko saugiai paruošti vieno APACK piloto."""


def _read_audit(path: Path) -> dict[str, Any]:
    audit = json.loads(path.read_text(encoding="utf-8"))
    if audit.get("status") != "PASS":
        raise PilotPreparationError("APACK / HRD-A transformacijos auditas nėra PASS.")
    if not isinstance(audit.get("rows"), list):
        raise PilotPreparationError("Transformacijos audite nėra eilučių.")
    return audit


def _audit_links(
    audit: dict[str, Any],
) -> tuple[dict[str, str], dict[str, set[str]], set[str]]:
    apack_to_hrd: dict[str, str] = {}
    hrd_to_apacks: dict[str, set[str]] = {}
    transferred: set[str] = set()
    review: set[str] = set()
    for row in audit["rows"]:
        apack = canon(row.get("apack_sku"))
        hrd = canon(row.get("hrd_a_sku"))
        decision = canon(row.get("decision"))
        if not apack:
            continue
        if decision == "DEFAULT_HRD_REVIEW":
            review.add(apack)
            continue
        if hrd:
            previous = apack_to_hrd.get(apack)
            if previous and previous != hrd:
                raise PilotPreparationError(
                    f"{apack}: audite susietas su keliais HRD-A."
                )
            apack_to_hrd[apack] = hrd
            hrd_to_apacks.setdefault(hrd, set()).add(apack)
        if decision == "TRANSFER_TO_APACK":
            transferred.add(apack)
    return apack_to_hrd, hrd_to_apacks, transferred - review


def _has_value(value: Any) -> bool:
    return value not in (None, "")


def _read_groups(path: Path) -> list[tuple[str, list[list[Any]]]]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    try:
        if workbook.sheetnames != ["BOM import"]:
            raise PilotPreparationError(
                f"{path.name}: tikėtasi vieno lapo 'BOM import'."
            )
        rows = workbook["BOM import"].iter_rows(values_only=True)
        header = list(next(rows, ()))
        if header != IMPORT_HEADERS:
            raise PilotPreparationError(
                f"{path.name}: neteisingi importo stulpeliai."
            )

        groups: list[tuple[str, list[list[Any]]]] = []
        current_sku = ""
        current_rows: list[list[Any]] = []
        for row_number, raw_row in enumerate(rows, start=2):
            row = list(raw_row[: len(IMPORT_HEADERS)])
            if not any(_has_value(value) for value in row):
                continue
            if any(
                isinstance(value, str) and value.startswith("=")
                for value in row
            ):
                raise PilotPreparationError(
                    f"{path.name}:{row_number}: formulės neleidžiamos."
                )
            parent_sku = canon(row[0])
            if parent_sku:
                if current_rows:
                    groups.append((current_sku, current_rows))
                current_sku = parent_sku
                current_rows = [row]
            elif not current_rows:
                raise PilotPreparationError(
                    f"{path.name}:{row_number}: tęstinė eilutė be parent."
                )
            else:
                current_rows.append(row)
        if current_rows:
            groups.append((current_sku, current_rows))
        return groups
    finally:
        workbook.close()


def _validate_pilot(sku: str, rows: list[list[Any]]) -> None:
    first = rows[0]
    if not sku.startswith("APACK-"):
        raise PilotPreparationError(f"{sku}: tai nėra APACK BOM.")
    if first[6] != MANUFACTURE:
        raise PilotPreparationError(
            f"{sku}: BOM tipas {first[6]!r}, tikėtasi {MANUFACTURE!r}."
        )
    if first[7] != NEW_BOM_SEQUENCE:
        raise PilotPreparationError(
            f"{sku}: Sequence {first[7]!r}, tikėtasi {NEW_BOM_SEQUENCE}."
        )
    if first[8] != APACK_OPERATION_TYPE_EXTERNAL_ID:
        raise PilotPreparationError(
            f"{sku}: Operation Type/External ID {first[8]!r}, "
            f"tikėtasi {APACK_OPERATION_TYPE_EXTERNAL_ID!r}."
        )
    if not str(first[9] or "").strip():
        raise PilotPreparationError(f"{sku}: tuščias Release Reference.")
    if not first[1]:
        raise PilotPreparationError(
            f"{sku}: tuščias parent product.template External ID."
        )
    component_count = sum(
        1 for row in rows if any(_has_value(row[index]) for index in (3, 4, 5))
    )
    operation_count = sum(
        1
        for row in rows
        if any(_has_value(row[index]) for index in range(12, 17))
    )
    if component_count == 0:
        raise PilotPreparationError(f"{sku}: nėra komponentų eilučių.")
    if operation_count == 0:
        raise PilotPreparationError(f"{sku}: nėra operacijų eilučių.")


def _validate_hrd(sku: str, rows: list[list[Any]]) -> None:
    first = rows[0]
    if "HRD" not in sku or not sku.endswith("-A"):
        raise PilotPreparationError(f"{sku}: tai nėra HRD-A BOM.")
    if first[6] != MANUFACTURE:
        raise PilotPreparationError(
            f"{sku}: BOM tipas {first[6]!r}, tikėtasi {MANUFACTURE!r}."
        )
    if first[7] != NEW_BOM_SEQUENCE:
        raise PilotPreparationError(
            f"{sku}: Sequence {first[7]!r}, tikėtasi {NEW_BOM_SEQUENCE}."
        )
    if not first[1] or not first[8] or not str(first[9] or "").strip():
        raise PilotPreparationError(
            f"{sku}: trūksta External ID, Operation Type arba Release Reference."
        )
    if not any(
        any(_has_value(row[index]) for index in (3, 4, 5)) for row in rows
    ):
        raise PilotPreparationError(f"{sku}: nėra komponentų eilučių.")
    if not any(
        any(_has_value(row[index]) for index in range(12, 17)) for row in rows
    ):
        raise PilotPreparationError(f"{sku}: nėra operacijų eilučių.")


def prepare_pilot(
    source: Path,
    audit_path: Path,
    output: Path,
    requested_sku: str | None = None,
) -> tuple[str, str, list[str], int, int]:
    groups = _read_groups(source)
    by_sku: dict[str, list[list[list[Any]]]] = {}
    for sku, rows in groups:
        by_sku.setdefault(sku, []).append(rows)

    duplicates = sorted(sku for sku, matches in by_sku.items() if len(matches) > 1)
    if duplicates:
        raise PilotPreparationError(
            "Release faile dubliuojasi parent SKU: " + ", ".join(duplicates[:10])
        )

    audit = _read_audit(audit_path)
    apack_to_hrd, hrd_to_apacks, transferred = _audit_links(audit)

    requested = canon(requested_sku)
    if requested:
        if requested not in by_sku:
            raise PilotPreparationError(
                f"Release faile nerastas pilotinis SKU {requested}."
            )
        if requested not in transferred:
            raise PilotPreparationError(
                f"{requested}: nėra patvirtintas transformuotas APACK."
            )
        pilot_sku = requested
    else:
        candidates = sorted(
            sku
            for sku in transferred
            if sku in by_sku
            and not sku.startswith("APACK-USB-")
            and apack_to_hrd.get(sku) in by_sku
        )
        if not candidates:
            raise PilotPreparationError(
                "Release faile nerastas saugus APACK piloto kandidatas."
            )
        pilot_sku = min(
            candidates,
            key=lambda sku: (len(hrd_to_apacks[apack_to_hrd[sku]]), sku),
        )

    pilot_hrd = apack_to_hrd.get(pilot_sku)
    if not pilot_hrd or pilot_hrd not in by_sku:
        raise PilotPreparationError(
            f"{pilot_sku}: release faile nerastas susietas HRD-A."
        )
    cohort_apacks = sorted(hrd_to_apacks[pilot_hrd])
    missing = [sku for sku in cohort_apacks if sku not in by_sku]
    if missing:
        raise PilotPreparationError(
            f"{pilot_hrd}: release faile trūksta susietų APACK: "
            + ", ".join(missing)
        )
    for sku in cohort_apacks:
        _validate_pilot(sku, by_sku[sku][0])
    _validate_hrd(pilot_hrd, by_sku[pilot_hrd][0])
    pilot_groups = cohort_apacks + [pilot_hrd]

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOM import"
    sheet.append(IMPORT_HEADERS)
    for sku in pilot_groups:
        for row in by_sku[sku][0]:
            sheet.append(row)
    _style(sheet)
    workbook.save(output)
    workbook.close()

    all_rows = [row for sku in pilot_groups for row in by_sku[sku][0]]
    component_count = sum(
        1 for row in all_rows if any(_has_value(row[index]) for index in (3, 4, 5))
    )
    operation_count = sum(
        1
        for row in all_rows
        if any(_has_value(row[index]) for index in range(12, 17))
    )
    return pilot_sku, pilot_hrd, cohort_apacks, component_count, operation_count


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Iš validuoto BOM Release lv2 Manufacture XLSX paruošia saugią "
            "APACK / HRD-A Stage piloto grupę. Odoo neskaito ir nekeičia."
        )
    )
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--audit", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--pilot-sku")
    args = parser.parse_args()

    sku, hrd, apacks, components, operations = prepare_pilot(
        args.source,
        args.audit,
        args.output,
        args.pilot_sku,
    )
    print("=" * 80)
    print("BOM RELEASE APACK PILOT")
    print("=" * 80)
    print("Pilotinis SKU:", sku)
    print("Susietas HRD-A:", hrd)
    print("Piloto APACK skaičius:", len(apacks))
    print("Piloto BOM skaičius:", len(apacks) + 1)
    print("BOM tipas:", MANUFACTURE)
    print("Sequence:", NEW_BOM_SEQUENCE)
    print(
        "Operation Type/External ID:",
        APACK_OPERATION_TYPE_EXTERNAL_ID,
    )
    print("Komponentų eilutės:", components)
    print("Operacijų eilutės:", operations)
    print("Failas:", args.output.resolve())
    print("Odoo pakeitimai: 0")


if __name__ == "__main__":
    main()
