from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from tamara_adjustments import save_adjustments


BASE_DIR = Path(__file__).resolve().parent

SOURCE_PATH = BASE_DIR / "Purchase prices 2026 08 V1.xlsx"

TARGET_PATH = (
    BASE_DIR
    / "web_state"
    / "shared_data"
    / "tamara_adjustments.json"
)

SOURCE_SKU_COLUMN = "Invoice lines/Product/Internal Reference"
SOURCE_ADJUSTED_COLUMN = "Adjustint kaina"


def headers(sheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value not in (None, "")
    }


def load_source_adjustments() -> dict[str, dict]:
    if not SOURCE_PATH.exists():
        raise FileNotFoundError(
            f"Nerastas šaltinio failas: {SOURCE_PATH}"
        )

    workbook = load_workbook(
        SOURCE_PATH,
        data_only=True,
        read_only=True,
    )

    try:
        sheet = workbook.active
        columns = headers(sheet)

        for required in (
            SOURCE_SKU_COLUMN,
            SOURCE_ADJUSTED_COLUMN,
        ):
            if required not in columns:
                raise RuntimeError(
                    f"Šaltinio faile nerastas stulpelis: {required}"
                )

        adjustments: dict[str, dict] = {}
        duplicates: set[str] = set()

        for row_number in range(2, sheet.max_row + 1):
            sku = sheet.cell(
                row_number,
                columns[SOURCE_SKU_COLUMN],
            ).value

            adjusted_price = sheet.cell(
                row_number,
                columns[SOURCE_ADJUSTED_COLUMN],
            ).value

            if sku in (None, ""):
                continue

            if adjusted_price in (None, ""):
                continue

            sku = str(sku).strip()

            if sku in adjustments:
                duplicates.add(sku)
                continue

            try:
                adjusted_price = float(adjusted_price)
            except (TypeError, ValueError) as exc:
                raise RuntimeError(
                    f"SKU {sku}: Adjustint kaina nėra skaičius: "
                    f"{adjusted_price!r}"
                ) from exc

            adjustments[sku] = {
                "adjusted_purchase_price": adjusted_price,
                "comment": "Migrated from Purchase prices 2026 08 V1.xlsx",
            }

    finally:
        workbook.close()

    if duplicates:
        preview = ", ".join(sorted(duplicates)[:20])
        raise RuntimeError(
            "Šaltinio faile yra pasikartojančių SKU: "
            f"{preview}"
        )

    return adjustments


def migrate() -> None:
    if TARGET_PATH.exists():
        raise RuntimeError(
            "Tamaros korekcijų saugykla jau egzistuoja. "
            "Migracija automatiškai jos neperrašys:\n"
            f"{TARGET_PATH}"
        )

    adjustments = load_source_adjustments()

    save_adjustments(
        TARGET_PATH,
        adjustments,
    )

    print("TAMAROS KOREKCIJŲ MIGRACIJA BAIGTA")
    print("Šaltinis:", SOURCE_PATH)
    print("Tikslas:", TARGET_PATH)
    print("Perkelta korekcijų:", len(adjustments))
    print()
    print(
        "Toliau šis Excel failas nebėra Tamaros korekcijų "
        "duomenų šaltinis."
    )


if __name__ == "__main__":
    migrate()