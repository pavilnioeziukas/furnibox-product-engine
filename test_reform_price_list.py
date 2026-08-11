from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from reform_price_list import build_reform_price_list


class ReformPriceListTests(unittest.TestCase):
    def test_combines_sources_and_prioritizes_cabinet_part_calculation(self):
        with tempfile.TemporaryDirectory() as directory:
            base = Path(directory)
            components = base / "Last_Purchase_Prices.xlsx"
            cabinet_parts = base / "Existing_and_New_Cabinet_Parts_Prices.xlsx"
            output = base / "Reform_Final_Prices.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "COMPONENT PRICES"
            ws.append([
                "Internal Reference", "Name", "Vendor", "Real Purchase Price",
                "Adjusted Purchase Price", "Markup Factor", "Reform Price",
            ])
            ws.append(["ACCS-1", "Accessory", "Reform Supply", 2, 2.5, 1.05, 2.625])
            ws.append(["PART-1", "Old cabinet row", "Furnix", 9, 9, 1, 9])
            adjustments = wb.create_sheet("TAMARA ADJUSTMENTS")
            adjustments.append(["Internal Reference", "Adjusted Purchase Price"])
            adjustments.append(["ACCS-1", 2.75])
            adjustments.append(["PART-1", 9])
            wb.save(components)

            wb = Workbook()
            ws = wb.active
            ws.title = "CABINET PART PRICES"
            ws.append([
                "Internal Reference", "Furnix Unit Cost",
                "Furnix Sales Price to Furnibox", "Product Status", "BOM Source",
            ])
            ws.append(["PART-1", 10, 11.5, "EXISTING", "EXISTING + NEW"])
            wb.save(cabinet_parts)

            products, overlaps = build_reform_price_list(
                components, cabinet_parts, output
            )
            self.assertEqual((products, overlaps), (2, 1))

            result = load_workbook(output, data_only=False)
            self.assertEqual(result.sheetnames, ["REFORM PRICE LIST", "DIAGNOSTICS", "INFO"])
            prices = result["REFORM PRICE LIST"]
            headers = {cell.value: cell.column for cell in prices[1]}
            rows = {
                prices.cell(row, headers["Internal Reference"]).value: row
                for row in range(2, prices.max_row + 1)
            }
            part_row = rows["PART-1"]
            self.assertEqual(prices.cell(part_row, headers["Price Source"]).value, "CABINET PART CALCULATION")
            self.assertEqual(prices.cell(part_row, headers["Adjusted Furnibox Purchase Price"]).value, 11.5)
            self.assertIsNone(
            prices.cell(part_row, headers["Reform Markup Factor"]).value
            )
            self.assertEqual(
            prices.cell(part_row, headers["Reform Purchase Price"]).value,
             11.5,
            )
            component_row = rows["ACCS-1"]
            self.assertEqual(prices.cell(component_row, headers["Adjusted Furnibox Purchase Price"]).value, 2.75)
            self.assertEqual(prices.cell(component_row, headers["Reform Markup Factor"]).value, 1.05)
            self.assertEqual(result["DIAGNOSTICS"]["B2"].value, "PART-1")


if __name__ == "__main__":
    unittest.main()
