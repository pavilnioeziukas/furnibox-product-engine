from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from bom_release.generator import IMPORT_HEADERS, canon

SKU_QUERY_BATCH_SIZE = 40


class StageProductIdMapError(RuntimeError):
    """Stage produkto identifikatorių žodyno sudaryti arba naudoti negalima."""


def read_bom_skus(paths: Iterable[Path]) -> tuple[list[str], list[str]]:
    parents: set[str] = set()
    components: set[str] = set()
    for path in paths:
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            if "BOM import" not in workbook.sheetnames:
                raise StageProductIdMapError(
                    f"{path.name}: nerastas lapas 'BOM import'."
                )
            rows = workbook["BOM import"].iter_rows(values_only=True)
            header = list(next(rows, ()))
            if header != IMPORT_HEADERS:
                raise StageProductIdMapError(
                    f"{path.name}: neteisingi importo stulpeliai."
                )
            for row_number, raw in enumerate(rows, start=2):
                row = list(raw[: len(IMPORT_HEADERS)])
                if not any(value not in (None, "") for value in row):
                    continue
                if any(
                    isinstance(value, str) and value.startswith("=")
                    for value in row
                ):
                    raise StageProductIdMapError(
                        f"{path.name}:{row_number}: formulės neleidžiamos."
                    )
                parent = canon(row[0])
                component = canon(row[3])
                if parent:
                    parents.add(parent)
                if component:
                    components.add(component)
        finally:
            workbook.close()
    if not parents:
        raise StageProductIdMapError("BOM failuose nėra parent produktų.")
    if not components:
        raise StageProductIdMapError("BOM failuose nėra komponentų.")
    return sorted(parents), sorted(components)


def _external_id(row: dict[str, Any]) -> str:
    module = str(row.get("module") or "").strip()
    name = str(row.get("name") or "").strip()
    return f"{module}.{name}" if module and name else ""


def _external_ids(
    client: Any, model: str, record_ids: set[int]
) -> dict[int, str]:
    if not record_ids:
        return {}
    rows = client.search_read_all(
        "ir.model.data",
        [["model", "=", model], ["res_id", "in", sorted(record_ids)]],
        ["module", "name", "res_id"],
    )
    grouped: dict[int, list[str]] = defaultdict(list)
    for row in rows:
        value = _external_id(row)
        if value:
            grouped[int(row["res_id"])].append(value)
    return {
        record_id: sorted(
            set(values),
            key=lambda value: (not value.startswith("__export__."), value),
        )[0]
        for record_id, values in grouped.items()
    }


def _sku_search_domain(skus: list[str]) -> list[Any]:
    """Build a bounded OR query that tolerates case and surrounding whitespace."""
    conditions: list[list[Any]] = [
        ["default_code", "ilike", sku] for sku in skus
    ]
    if not conditions:
        return [["id", "=", 0]]
    return ["|"] * (len(conditions) - 1) + conditions


def _read_sku_rows(
    client: Any, model: str, wanted: list[str], fields: list[str]
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for offset in range(0, len(wanted), SKU_QUERY_BATCH_SIZE):
        batch = wanted[offset : offset + SKU_QUERY_BATCH_SIZE]
        rows.extend(
            client.search_read_all(
                model,
                _sku_search_domain(batch),
                fields,
                context={"active_test": False},
                batch_size=SKU_QUERY_BATCH_SIZE * 4,
            )
        )
    return rows


def export_stage_product_id_map(
    client: Any,
    parent_skus: list[str],
    component_skus: list[str],
    stage_url: str,
) -> dict[str, Any]:
    wanted = sorted(set(parent_skus) | set(component_skus))
    products = _read_sku_rows(
        client,
        "product.product",
        wanted,
        ["id", "default_code", "product_tmpl_id", "active"],
    )
    templates = _read_sku_rows(
        client,
        "product.template",
        wanted,
        ["id", "default_code", "active"],
    )
    products_by_sku: dict[str, list[dict[str, Any]]] = defaultdict(list)
    templates_by_sku: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in products:
        sku = canon(row.get("default_code"))
        if sku in wanted:
            products_by_sku[sku].append(row)
    for row in templates:
        sku = canon(row.get("default_code"))
        if sku in wanted:
            templates_by_sku[sku].append(row)

    template_ids = {
        int(row["id"])
        for sku in component_skus
        for row in templates_by_sku.get(sku, [])
    }
    variants_by_template: dict[int, list[dict[str, Any]]] = defaultdict(list)
    if template_ids:
        variants = client.search_read_all(
            "product.product",
            [["product_tmpl_id", "in", sorted(template_ids)]],
            ["id", "default_code", "product_tmpl_id", "active"],
            context={"active_test": False},
        )
        for row in variants:
            if row.get("product_tmpl_id"):
                variants_by_template[int(row["product_tmpl_id"][0])].append(row)

    resolved_templates: dict[str, int] = {}
    resolved_products: dict[str, int] = {}
    product_active: dict[str, bool] = {}
    missing: list[str] = []
    ambiguous: list[str] = []
    for sku in wanted:
        direct_products = products_by_sku.get(sku, [])
        direct_templates = templates_by_sku.get(sku, [])
        template_candidates = {
            int(row["id"]) for row in direct_templates
        } | {
            int(row["product_tmpl_id"][0])
            for row in direct_products
            if row.get("product_tmpl_id")
        }
        if sku in parent_skus:
            if len(template_candidates) == 1:
                resolved_templates[sku] = next(iter(template_candidates))
            elif not template_candidates:
                missing.append(f"{sku} (product.template)")
            else:
                ambiguous.append(f"{sku} (product.template)")

        if sku in component_skus:
            product_candidates = {int(row["id"]): row for row in direct_products}
            if not product_candidates and len(direct_templates) == 1:
                template_id = int(direct_templates[0]["id"])
                product_candidates = {
                    int(row["id"]): row
                    for row in variants_by_template.get(template_id, [])
                }
            if len(product_candidates) == 1:
                product_id, product_row = next(iter(product_candidates.items()))
                resolved_products[sku] = product_id
                product_active[sku] = bool(product_row.get("active"))
            elif not product_candidates:
                missing.append(f"{sku} (product.product)")
            else:
                ambiguous.append(f"{sku} (product.product)")

    if missing or ambiguous:
        details: list[str] = []
        if missing:
            details.append("Stage nerasti Internal Reference: " + ", ".join(missing))
        if ambiguous:
            details.append(
                "Stage neunikalūs Internal Reference: " + ", ".join(ambiguous)
            )
        raise StageProductIdMapError("; ".join(details))

    product_ids = set(resolved_products.values())
    template_ids = set(resolved_templates.values())
    product_external_ids = _external_ids(client, "product.product", product_ids)
    template_external_ids = _external_ids(
        client, "product.template", template_ids
    )

    records: dict[str, dict[str, Any]] = {}
    missing_ids: list[str] = []
    for sku in wanted:
        product_id = resolved_products.get(sku)
        template_id = resolved_templates.get(sku)
        record = {
            "product_template_id": template_id,
            "product_template_external_id": (
                template_external_ids.get(template_id, "")
                if sku in parent_skus and template_id
                else ""
            ),
            "product_product_id": product_id,
            "product_product_external_id": (
                product_external_ids.get(product_id, "")
                if sku in component_skus and product_id
                else ""
            ),
            "active": product_active.get(sku),
        }
        if sku in parent_skus and not record["product_template_external_id"]:
            missing_ids.append(f"{sku} (product.template)")
        if sku in component_skus and not record["product_product_external_id"]:
            missing_ids.append(f"{sku} (product.product)")
        records[sku] = record
    if missing_ids:
        raise StageProductIdMapError(
            "Stage įrašai neturi External ID: " + ", ".join(missing_ids)
        )

    return {
        "schema_version": 1,
        "environment": "stage",
        "stage_url": stage_url,
        "exported_at_utc": datetime.now(timezone.utc).isoformat(),
        "parent_skus": parent_skus,
        "component_skus": component_skus,
        "records": records,
        "odoo_changes": 0,
    }


def save_map(data: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def load_map(path: Path) -> dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if data.get("schema_version") != 1 or data.get("environment") != "stage":
        raise StageProductIdMapError("Neteisingas Stage ID žodyno formatas.")
    if not isinstance(data.get("records"), dict):
        raise StageProductIdMapError("Stage ID žodyne nėra records objekto.")
    return data
