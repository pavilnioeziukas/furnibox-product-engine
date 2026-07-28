"""Generuoja Manufacture ir KIT BOM importo failus Stage arba Production.

Generatorius remiasi sėkmingai patikrinto Manufacture piloto formatu. Jis tik
nuskaito Odoo ir sukuria Excel failą – jokių BOM Odoo sistemoje nekuria.
"""

from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date
from difflib import SequenceMatcher
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bom_import_pilot_v1 import (
    calculate_levels,
    canon,
    load_bom_types,
    load_new_bom_graph,
)
from bom_import_pilot_v2 import choose_operation_template, load_operation_templates
from bom_import_pilot_v2 import family_token
from bom_import_v1 import load_stage_subcategories
from config import load_settings
from odoo_client import OdooClient
from output_paths import environment_output_dir, environment_slug
from product_import_v3 import apack_sku
from product_import_v7 import (
    cabinet_assembled_sku,
    hrd_assembled_sku,
    is_cabinet_bom_parent,
    is_hrd_bom_parent,
)
from product_detection_v2 import load_reform_universe
from product_detection_v2 import find_bom_input


IMPORT_HEADERS = [
    "Product/Internal Reference", "Product/External ID", "qty",
    "BoM Lines/Component/Internal Reference",
    "BoM Lines/Component/External ID", "product_qty",
    "BoM Type", "Reference", "mo_autodone_by_wo", "auto_plan",
    "operation_ids/name", "operation_ids/workcenter_id",
    "operation_ids/time_mode", "operation_ids/time_cycle_manual",
    "operation_ids/sequence",
]
MANUFACTURE = "Manufacture this product"
KIT = "KIT"
EXCLUDED_APACK_SOURCES = {
    # HRD pakuotė, o ne spintelės FPACK. Iš jos APACK kurti negalima.
    "FPACK-WTP92-HRD001",
}
LEGACY_CABINET_APACKS = {
    # Production legacy SKU naudoja APACK-USB-, nors dabartinė taisyklė iš
    # FPACK-US- generuotų APACK-US-C-. Naudojame realiai esantį produktą.
    "USB-C-CAB01-WAL045": "APACK-USB-C-CAB01-WAL045-A",
}
APACK_FAMILY_FALLBACKS = {
    # Naujos šeimos neturi tiesioginio Production analogo. BOH yra bazinė,
    # HCO – aukšta spintelė, todėl perimami atitinkamų šeimų etalonai.
    "BOH": "BAS",
    "HCO": "HIG",
}
KIT_PARENT_CATEGORIES = {
    # Reform baziniai produktai, kurių BOM Odoo turi būti KIT.
    # CABINET SHELF yra teisėtas KIT, tačiau jam negeneruojama atskira -A versija.
    "CABINETS",
    "CABINET SHELF",
}


def is_kit_bom_parent(sku: str, product: dict) -> bool:
    """Tikras Reform tėvinis produktas, kuriam leidžiama kurti bazinį KIT."""
    code = canon(sku)
    return bool(
        product.get("is_parent")
        and canon(product.get("category")) in KIT_PARENT_CATEGORIES
        and not code.endswith("-A")
    )


def add_generated_apack_boms(parents, lines, levels, bom_types):
    """Kiekvienam naujam FPACK prideda tokios pačios sudėties APACK BOM."""
    generated_from = {}
    for fpack in sorted(parents):
        if not fpack.startswith("FPACK-"):
            continue
        if fpack in EXCLUDED_APACK_SOURCES:
            continue
        if bom_types.get(fpack) != MANUFACTURE:
            continue
        generated = canon(apack_sku(fpack))
        if generated in parents:
            raise ValueError(
                f"APACK jau yra MAP grafe, todėl jo negalima generuoti antrą kartą: "
                f"{generated}"
            )
        parents.add(generated)
        lines[generated] = [dict(line) for line in lines.get(fpack, [])]
        levels[generated] = levels[fpack]
        bom_types[generated] = MANUFACTURE
        generated_from[generated] = fpack
    return generated_from


def add_generated_hrd_assembled_boms(
    parents, lines, levels, bom_types, reform_products, reform_lines
):
    """Visiems Reform HRD BOM tėvams prideda tokios pačios sudėties HRD-A BOM."""
    generated_from = {}
    for source_sku, product in sorted(reform_products.items()):
        source = canon(source_sku)
        if not is_hrd_bom_parent(source, product):
            continue
        source_lines = reform_lines.get(source, [])
        if not source_lines:
            continue
        generated = canon(hrd_assembled_sku(source))
        if generated in parents:
            raise ValueError(
                "HRD-A jau yra MAP grafe, todėl jo negalima generuoti antrą "
                f"kartą: {generated}"
            )
        parents.add(generated)
        lines[generated] = [dict(line) for line in source_lines]
        # HRD yra spintelės BOM komponentas, todėl jo surinkta versija
        # importuojama antrame lygyje, kai visi jos komponentai jau egzistuoja.
        levels[generated] = 2
        bom_types[generated] = MANUFACTURE
        generated_from[generated] = source
    return generated_from


def add_generated_cabinet_assembled_kits(
    parents, lines, levels, bom_types, reform_products, reform_lines
):
    """Iš bazinio CABINET KIT sukuria surinkto CABINET-A KIT.

    Bazinis BOM naudoja FPACK ir HRD, o surinktas BOM privalo naudoti
    atitinkamus APACK ir HRD-A. Generuojami tik KIT, kuriuose yra FPACK.
    """
    generated_from = {}
    for source_sku, product in sorted(reform_products.items()):
        source = canon(source_sku)
        if not is_cabinet_bom_parent(source, product):
            continue
        source_lines = reform_lines.get(source, [])
        if not source_lines:
            continue
        generated = canon(cabinet_assembled_sku(source))
        if generated in parents:
            raise ValueError(
                "CABINET-A jau yra MAP grafe, todėl jo negalima generuoti "
                f"antrą kartą: {generated}"
            )
        transformed = []
        for line in source_lines:
            component = canon(line.get("component"))
            if component.startswith("FPACK-"):
                assembled_component = canon(
                    LEGACY_CABINET_APACKS.get(source, apack_sku(component))
                )
            elif "HRD" in component:
                assembled_component = canon(hrd_assembled_sku(component))
            else:
                raise ValueError(
                    f"{source}: CABINET KIT turi ne FPACK/HRD komponentą "
                    f"{component}; surinkto BOM automatiškai kurti nesaugu"
                )
            transformed.append({
                **line,
                "component": assembled_component,
            })
        apack_count = sum(
            canon(line["component"]).startswith("APACK-")
            for line in transformed
        )
        hrd_a_count = sum(
            "HRD" in canon(line["component"])
            and canon(line["component"]).endswith("-A")
            for line in transformed
        )
        if apack_count != 1 or hrd_a_count < 1:
            raise ValueError(
                f"{source}: surinktam CABINET-A tikėtasi 1 APACK ir bent "
                f"1 HRD-A, gauta APACK={apack_count}, HRD-A={hrd_a_count}"
            )
        parents.add(generated)
        lines[generated] = transformed
        levels[generated] = 1
        bom_types[generated] = KIT
        generated_from[generated] = source
    return generated_from


def load_existing_bom_skus(
    client: OdooClient, products: dict[str, dict]
) -> set[str]:
    """Grąžina Odoo jau egzistuojančių BOM tėvų SKU."""
    by_product_id = {
        int(value["product_id"]): sku for sku, value in products.items()
    }
    by_template_id = {
        int(value["template_id"]): sku for sku, value in products.items()
    }
    rows = client.search_read_all(
        "mrp.bom", [], ["id", "product_tmpl_id", "product_id", "active"],
        context={"active_test": False},
    )
    result = set()
    for row in rows:
        product = row.get("product_id")
        template = row.get("product_tmpl_id")
        sku = ""
        if product:
            sku = by_product_id.get(int(product[0]), "")
        if not sku and template:
            sku = by_template_id.get(int(template[0]), "")
        if sku:
            result.add(sku)
    return result


def remove_existing_generated_boms(
    parents, lines, levels, bom_types, generated_from, existing_bom_skus
):
    """Pašalina sugeneruotas poras, kurių BOM Odoo jau egzistuoja."""
    for sku in list(generated_from):
        if sku not in existing_bom_skus:
            continue
        parents.discard(sku)
        lines.pop(sku, None)
        levels.pop(sku, None)
        bom_types.pop(sku, None)
        generated_from.pop(sku, None)


def load_reform_bom_lines(path: Path) -> dict[str, list[dict]]:
    """Nuskaito visų Reform BOM tėvų komponentus tiesiai iš BOM - Input."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            header_row = None
            headers = None
            for row_no, row in enumerate(
                ws.iter_rows(
                    min_row=1, max_row=min(ws.max_row, 30), values_only=True
                ),
                start=1,
            ):
                values = [
                    str(value).strip() if value is not None else ""
                    for value in row
                ]
                if "BOM SKU Code" in values:
                    header_row, headers = row_no, values
                    break
            if header_row is None:
                continue

            index = {value: i for i, value in enumerate(headers) if value}
            part_columns = []
            for header, code_index in index.items():
                match = re.fullmatch(
                    r"Part\s+(\d+)\s+Code", header, flags=re.IGNORECASE
                )
                if not match:
                    continue
                number = int(match.group(1))
                qty_index = index.get(f"Part {number} Qty")
                part_columns.append((number, code_index, qty_index))
            part_columns.sort()

            result: dict[str, list[dict]] = defaultdict(list)
            for row in ws.iter_rows(
                min_row=header_row + 1, values_only=True
            ):
                parent = canon(row[index["BOM SKU Code"]])
                if not parent:
                    continue
                for _, code_index, qty_index in part_columns:
                    component = canon(row[code_index])
                    if not component:
                        continue
                    quantity = row[qty_index] if qty_index is not None else None
                    result[parent].append({
                        "component": component,
                        "quantity": quantity,
                    })
            return dict(result)
    finally:
        wb.close()
    raise ValueError("Reform faile nerastas 'BOM SKU Code' stulpelis.")


def choose_apack_operation_template(
    sku: str, subcategory: str, templates: dict[int, dict]
) -> dict:
    """Parenka tos pačios SKU šeimos APACK analogą ir sutvarko operacijų eilę."""
    family = family_token(sku)
    reference_family = APACK_FAMILY_FALLBACKS.get(family, family)
    candidates = [
        value
        for value in templates.values()
        if canon(value["subcategory"]) == canon(subcategory)
        and canon(value["sku"]).startswith("APACK-")
        and family_token(value["sku"]) == reference_family
    ]
    if not candidates:
        raise ValueError(
            "Production APACK etalonuose nėra tinkamos SKU šeimos: "
            f"{family} (naudojamas etalonas {reference_family})"
        )
    template = max(
        candidates,
        key=lambda value: SequenceMatcher(
            None, canon(sku), canon(value["sku"])
        ).ratio(),
    )
    assembly = [
        dict(operation)
        for operation in template["operations"]
        if "SURINK" in canon(operation.get("name"))
    ]
    packing = [
        dict(operation)
        for operation in template["operations"]
        if "PAKAV" in canon(operation.get("name"))
    ]
    if len(assembly) != 1 or len(packing) != 1:
        raise ValueError(
            "Production APACK analogas neturi tiksliai vienos surinkimo ir "
            f"vienos pakavimo operacijos: {template['sku']}"
        )
    assembly[0]["sequence"] = 100
    packing[0]["sequence"] = 101
    return {**template, "operations": [assembly[0], packing[0]]}


def choose_hrd_assembled_operation_template(
    templates: dict[int, dict]
) -> dict:
    """Parenka dažniausią Production HRD-A „Komplektavimas“ etaloną."""
    candidates = []
    for template in templates.values():
        template_sku = canon(template.get("sku"))
        if not template_sku.endswith("-A"):
            continue
        for operation in template.get("operations", []):
            if canon(operation.get("name")) != "KOMPLEKTAVIMAS":
                continue
            candidates.append((template, dict(operation)))

    if not candidates:
        raise ValueError(
            "Production etalonuose nerasta HRD-A operacija „Komplektavimas“"
        )

    signatures = Counter(
        (
            operation.get("name"),
            operation.get("workcenter"),
            operation.get("time_mode"),
            operation.get("time"),
            operation.get("sequence"),
        )
        for _, operation in candidates
    )
    selected_signature, _ = signatures.most_common(1)[0]
    template, operation = next(
        (template, operation)
        for template, operation in candidates
        if (
            operation.get("name"),
            operation.get("workcenter"),
            operation.get("time_mode"),
            operation.get("time"),
            operation.get("sequence"),
        ) == selected_signature
    )
    return {**template, "operations": [operation]}


def load_stage_products_by_sku(
    client: OdooClient, wanted_skus: set[str]
) -> tuple[dict[str, dict], set[str]]:
    """Randa aplinkos produktus, jų template/variant External ID ir dublikatus."""
    rows = client.search_read_all(
        "product.product",
        [["default_code", "!=", False]],
        ["id", "default_code", "product_tmpl_id", "active"],
        context={"active_test": False},
    )
    grouped: dict[str, list[dict]] = {}
    for row in rows:
        sku = canon(row.get("default_code"))
        if sku in wanted_skus:
            grouped.setdefault(sku, []).append(row)

    products = {}
    duplicates = set()
    for sku, matches in grouped.items():
        product_ids = {int(row["id"]) for row in matches}
        template_ids = {
            int(row["product_tmpl_id"][0])
            for row in matches
            if row.get("product_tmpl_id")
        }
        if len(product_ids) != 1 or len(template_ids) != 1:
            duplicates.add(sku)
            continue
        row = matches[0]
        products[sku] = {
            "display_sku": str(row["default_code"]).strip(),
            "product_id": next(iter(product_ids)),
            "template_id": next(iter(template_ids)),
        }

    def xmlids(model: str, ids: set[int]) -> dict[int, str]:
        if not ids:
            return {}
        records = client.search_read_all(
            "ir.model.data",
            [["model", "=", model], ["res_id", "in", sorted(ids)]],
            ["module", "name", "res_id"],
        )
        grouped_xmlids: dict[int, list[str]] = defaultdict(list)
        for record in records:
            grouped_xmlids[int(record["res_id"])].append(
                f"{record['module']}.{record['name']}"
            )
        result = {}
        for record_id, values in grouped_xmlids.items():
            values.sort(key=lambda value: (value.startswith("__export__."), value))
            result[record_id] = values[0]
        return result

    product_xmlids = xmlids(
        "product.product", {value["product_id"] for value in products.values()}
    )
    template_xmlids = xmlids(
        "product.template", {value["template_id"] for value in products.values()}
    )
    for value in products.values():
        value["product_xmlid"] = product_xmlids.get(value["product_id"], "")
        value["template_xmlid"] = template_xmlids.get(value["template_id"], "")
    return products, duplicates


def load_stage_workcenters(client: OdooClient) -> tuple[set[str], set[str]]:
    """Grąžina aplinkos aktyvius darbo centrus ir pasikartojančius vardus."""
    rows = client.search_read_all(
        "mrp.workcenter", [], ["id", "name", "active"],
        context={"active_test": False},
    )
    counts = Counter(
        str(row.get("name") or "").strip()
        for row in rows
        if row.get("active", True) and str(row.get("name") or "").strip()
    )
    unique = {name for name, count in counts.items() if count == 1}
    duplicates = {name for name, count in counts.items() if count > 1}
    return unique, duplicates


def prepare_manufacture_boms(
    parents, lines, levels, bom_types, products, subcategories,
    operation_templates, workcenters, duplicate_workcenters, duplicate_skus,
    generated_from, generated_hrd_from,
):
    """Atrenka Manufacture BOM ir atskiria paruoštus nuo blokuojamų."""
    manufacture_skus = {
        sku for sku in parents if bom_types.get(sku) == MANUFACTURE
    }
    ready = []
    review = []
    diagnostics = []

    for sku in sorted(manufacture_skus, key=lambda value: (levels[value], value)):
        bom_lines = lines.get(sku, [])
        messages = []
        operations = []
        template = None

        if sku in duplicate_skus:
            messages.append("Stage Internal Reference neunikalus BOM produktui")
        elif sku not in products:
            messages.append("Stage nerastas BOM produktas")
        elif not products[sku]["template_xmlid"]:
            messages.append("BOM produktas neturi product.template External ID")
        if not bom_lines:
            messages.append("BOM neturi komponentų")
        for line in bom_lines:
            if line["component"] in duplicate_skus:
                messages.append(
                    f"Stage Internal Reference neunikalus komponentui: {line['component']}"
                )
            elif line["component"] not in products:
                messages.append(f"Stage nerastas komponentas: {line['component']}")
            elif not products[line["component"]]["product_xmlid"]:
                messages.append(
                    "Komponentas neturi product.product External ID: "
                    f"{line['component']}"
                )

        subcategory = subcategories.get(sku, "")
        if not subcategory:
            messages.append("Nenustatyta Stage produkto pakategorė")
        elif canon(subcategory) == "LED HARDWARE":
            messages.append("LED HARDWARE operacijų taisyklė nepatvirtinta")
        else:
            try:
                if sku in generated_hrd_from:
                    template = choose_hrd_assembled_operation_template(
                        operation_templates
                    )
                elif sku in generated_from:
                    template = choose_apack_operation_template(
                        sku, subcategory, operation_templates
                    )
                else:
                    template = choose_operation_template(
                        sku, subcategory, operation_templates
                    )
                operations = template["operations"]
            except ValueError as exc:
                messages.append(str(exc))

        if not operations and canon(subcategory) != "LED HARDWARE":
            messages.append("Nerastas Production operacijų etalonas")
        for operation in operations:
            name = str(operation.get("workcenter") or "").strip()
            if name in duplicate_workcenters:
                messages.append(f"Stage darbo centro pavadinimas neunikalus: {name}")
            elif name not in workcenters:
                messages.append(f"Stage nerastas aktyvus darbo centras: {name}")

        messages = list(dict.fromkeys(messages))
        status = "READY" if not messages else "NOT READY"
        record = {
            "sku": sku,
            "level": levels[sku],
            "lines": bom_lines,
            "operations": operations,
            "generated_from": generated_from.get(sku, ""),
            "subcategory": subcategory,
            "template": template,
            "status": status,
            "message": "; ".join(messages),
        }
        review.append(record)
        if status == "READY":
            ready.append(record)
        else:
            diagnostics.extend(
                [sku, f"lv{levels[sku]}", message] for message in messages
            )
    return ready, review, diagnostics


def prepare_kit_boms(
    parents, lines, levels, bom_types, products, duplicate_skus,
    generated_cabinet_from, reform_products,
):
    """Patikrina bazinius ir sugeneruotus surinktų spintelių KIT BOM."""
    kit_skus = {sku for sku in parents if bom_types.get(sku) == KIT}
    ready = []
    review = []
    diagnostics = []
    excluded = []
    for sku in sorted(kit_skus, key=lambda value: (levels[value], value)):
        is_generated_cabinet = sku in generated_cabinet_from
        reform_product = reform_products.get(sku, {})
        if not is_generated_cabinet and not is_kit_bom_parent(sku, reform_product):
            excluded.append({
                "sku": sku,
                "category": str(reform_product.get("category") or "").strip(),
                "reason": (
                    "KIT atmestas: Reform tėvinio produkto kategorija nėra "
                    "CABINETS arba CABINET SHELF"
                ),
            })
            continue
        bom_lines = lines.get(sku, [])
        messages = []
        if sku in duplicate_skus:
            messages.append("Stage Internal Reference neunikalus KIT produktui")
        elif sku not in products:
            messages.append("Stage nerastas KIT produktas")
        elif not products[sku]["template_xmlid"]:
            messages.append("KIT produktas neturi product.template External ID")
        if not bom_lines:
            messages.append("KIT BOM neturi komponentų")
        for line in bom_lines:
            component_sku = line["component"]
            if component_sku in duplicate_skus:
                messages.append(
                    "Stage Internal Reference neunikalus KIT komponentui: "
                    f"{component_sku}"
                )
            elif component_sku not in products:
                messages.append(
                    f"Stage nerastas KIT komponentas: {component_sku}"
                )
            elif not products[component_sku]["product_xmlid"]:
                messages.append(
                    "KIT komponentas neturi product.product External ID: "
                    f"{component_sku}"
                )
        messages = list(dict.fromkeys(messages))
        status = "READY" if not messages else "NOT READY"
        record = {
            "sku": sku,
            "level": levels[sku],
            "lines": bom_lines,
            "operations": [],
            "generated_from": generated_cabinet_from.get(sku, ""),
            "status": status,
            "message": "; ".join(messages),
        }
        review.append(record)
        if status == "READY":
            ready.append(record)
        else:
            diagnostics.extend(
                [sku, f"lv{levels[sku]}", message] for message in messages
            )
    return ready, review, diagnostics, excluded


def import_rows(record: dict, products: dict, batch_reference: str):
    """Vieną Manufacture BOM paverčia importo bloku su tikrais External ID."""
    parent = products[record["sku"]]
    lines = record["lines"]
    operations = record["operations"]
    row_count = max(len(lines), len(operations), 1)

    for index in range(row_count):
        first = index == 0
        line = lines[index] if index < len(lines) else None
        operation = operations[index] if index < len(operations) else None
        component = products[line["component"]] if line else None
        yield [
            parent["display_sku"] if first else None,
            parent["template_xmlid"] if first else None,
            1 if first else None,
            component["display_sku"] if component else None,
            component["product_xmlid"] if component else None,
            line["quantity"] if line else None,
            MANUFACTURE if first else None,
            batch_reference if first else None,
            True if first else None,
            True if first else None,
            operation["name"] if operation else None,
            operation["workcenter"] if operation else None,
            operation["time_mode"] if operation else None,
            operation["time"] if operation else None,
            operation["sequence"] if operation else None,
        ]


def kit_import_rows(record: dict, products: dict, batch_reference: str):
    """Vieną KIT BOM paverčia Odoo importo eilučių bloku."""
    parent = products[record["sku"]]
    for index, line in enumerate(record["lines"]):
        first = index == 0
        component = products[line["component"]]
        yield [
            parent["display_sku"] if first else None,
            parent["template_xmlid"] if first else None,
            1 if first else None,
            component["display_sku"],
            component["product_xmlid"],
            line["quantity"],
            KIT if first else None,
            batch_reference if first else None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
        ]


def style_sheet(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index in range(1, ws.max_column + 1):
        values = [
            len(str(ws.cell(row, index).value or ""))
            for row in range(1, min(ws.max_row, 300) + 1)
        ]
        ws.column_dimensions[get_column_letter(index)].width = min(
            max(values, default=12) + 2, 55
        )


def write_workbook(
    path: Path, ready, review, diagnostics, products, batch_reference: str
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    review_ws = wb.create_sheet("REVIEW")
    review_ws.append([
        "Status", "Level", "Parent SKU", "Generated From",
        "Stage Subcategory", "Components",
        "Operations", "Production Analog", "Analog BOM Reference", "Message",
    ])
    for record in review:
        template = record.get("template") or {}
        review_ws.append([
            record["status"], f"lv{record['level']}", record["sku"],
            record["generated_from"], record["subcategory"], len(record["lines"]),
            len(record["operations"]), template.get("sku", ""),
            template.get("reference", ""), record["message"],
        ])
    style_sheet(review_ws)

    summary_ws = wb.create_sheet("SUMMARY")
    summary_ws.append(["Metric", "Value"])
    summary_ws.append(["All new Manufacture BOM", len(review)])
    summary_ws.append([
        "Generated assembled BOM",
        sum(bool(record["generated_from"]) for record in review),
    ])
    summary_ws.append(["Ready for import", len(ready)])
    summary_ws.append(["Not ready", len(review) - len(ready)])
    for (level, status), count in sorted(
        Counter((row["level"], row["status"]) for row in review).items()
    ):
        summary_ws.append([f"lv{level} {status}", count])
    style_sheet(summary_ws)

    for level in sorted({record["level"] for record in ready}):
        ws = wb.create_sheet(f"BOM_import(lv{level})")
        ws.append(IMPORT_HEADERS)
        for record in (row for row in ready if row["level"] == level):
            for row in import_rows(record, products, batch_reference):
                ws.append(row)
        style_sheet(ws)

    diagnostics_ws = wb.create_sheet("DIAGNOSTICS")
    diagnostics_ws.append(["Parent SKU", "Level", "Message"])
    for row in diagnostics:
        diagnostics_ws.append(row)
    style_sheet(diagnostics_ws)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_import_workbook(
    path: Path, records, products, batch_reference: str
) -> None:
    """Sukuria švarų vieno lapo failą tiesioginiam Odoo importui."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM import"
    ws.append(IMPORT_HEADERS)
    for record in records:
        for row in import_rows(record, products, batch_reference):
            ws.append(row)
    style_sheet(ws)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_kit_import_workbook(
    path: Path, records, products, batch_reference: str
) -> None:
    """Sukuria švarų KIT BOM failą tiesioginiam Odoo importui."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM import"
    ws.append(IMPORT_HEADERS)
    for record in records:
        for row in kit_import_rows(record, products, batch_reference):
            ws.append(row)
    style_sheet(ws)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    base = Path(__file__).resolve().parent
    environment = environment_slug()
    if environment not in {"stage", "production"}:
        raise PermissionError(
            "BOM failų generatorius leidžiamas tik Stage arba Production."
        )

    output_dir = environment_output_dir(base)
    comparison_path = output_dir / "MAP_Comparison.xlsx"
    types_path = output_dir / "BOM_Type_Review.xlsx"
    references_path = base / "output" / "production" / "BOM_Operations_Reference.xlsx"
    output_path = output_dir / "BOM_Import_Manufacture_All.xlsx"
    lv2_output_path = output_dir / "BOM_Import_Manufacture_lv2.xlsx"
    lv1_output_path = output_dir / "BOM_Import_Manufacture_lv1.xlsx"
    apack_output_path = output_dir / "BOM_Import_APACK.xlsx"
    kit_output_path = output_dir / "BOM_Import_KIT_lv1.xlsx"
    reform_input_path = find_bom_input(base)
    batch_reference = f"{date.today():%Y%m%d}_{reform_input_path.stem}"
    reform_products, _, _ = load_reform_universe(reform_input_path)
    reform_lines = load_reform_bom_lines(reform_input_path)

    parents, lines = load_new_bom_graph(comparison_path)
    bom_types = load_bom_types(types_path)
    levels = calculate_levels(set(parents), lines)
    parents = set(parents)
    generated_from = add_generated_apack_boms(
        parents, lines, levels, bom_types
    )
    generated_hrd_from = add_generated_hrd_assembled_boms(
        parents, lines, levels, bom_types, reform_products, reform_lines
    )
    generated_cabinet_from = add_generated_cabinet_assembled_kits(
        parents, lines, levels, bom_types, reform_products, reform_lines
    )
    overlap = set(generated_from) & set(generated_hrd_from)
    if overlap:
        raise ValueError(
            "Tas pats SKU sugeneruotas ir kaip APACK, ir kaip HRD-A: "
            + ", ".join(sorted(overlap))
        )
    generated_from.update(generated_hrd_from)

    settings = load_settings()
    client = OdooClient(settings)
    uid = client.authenticate()
    wanted = set(parents) | {
        row["component"] for values in lines.values() for row in values
    }
    products, duplicate_skus = load_stage_products_by_sku(client, wanted)
    existing_bom_skus = load_existing_bom_skus(client, products)
    remove_existing_generated_boms(
        parents, lines, levels, bom_types, generated_cabinet_from,
        existing_bom_skus,
    )
    subcategories = load_stage_subcategories(client, set(parents))
    operation_templates = load_operation_templates(references_path)
    workcenters, duplicate_workcenters = load_stage_workcenters(client)

    ready, review, diagnostics = prepare_manufacture_boms(
        parents, lines, levels, bom_types, products, subcategories,
        operation_templates, workcenters, duplicate_workcenters, duplicate_skus,
        generated_from, generated_hrd_from,
    )
    kit_ready, kit_review, kit_diagnostics, kit_excluded = prepare_kit_boms(
        parents, lines, levels, bom_types, products, duplicate_skus,
        generated_cabinet_from, reform_products,
    )
    write_workbook(
        output_path, ready, review, diagnostics, products, batch_reference
    )
    lv2_ready = [record for record in ready if record["level"] == 2]
    lv1_ready = [record for record in ready if record["level"] == 1]
    apack_ready = [
        record for record in ready
        if record["sku"].startswith("APACK-")
    ]
    write_import_workbook(
        lv2_output_path, lv2_ready, products, batch_reference
    )
    write_import_workbook(
        lv1_output_path, lv1_ready, products, batch_reference
    )
    write_import_workbook(
        apack_output_path, apack_ready, products, batch_reference
    )
    write_kit_import_workbook(
        kit_output_path, kit_ready, products, batch_reference
    )

    print(f"Prisijungta prie {environment.title()} Odoo. UID=", uid)
    print("\nVISŲ MANUFACTURE BOM IMPORTO FAILAS SUKURTAS")
    print("Failas:", output_path)
    print("Odoo importas lv2:", lv2_output_path, f"({len(lv2_ready)} BOM)")
    print("Odoo importas lv1:", lv1_output_path, f"({len(lv1_ready)} BOM)")
    print("Tik nauji APACK:", apack_output_path, f"({len(apack_ready)} BOM)")
    print(
        "Visi KIT (baziniai + CABINET-A):",
        kit_output_path,
        f"({len(kit_ready)} BOM)",
    )
    print("Bendra BOM nuoroda:", batch_reference)
    print("Visi nauji Manufacture BOM:", len(review))
    print("Sugeneruoti APACK BOM:", len(generated_from) - len(generated_hrd_from))
    print("Sugeneruoti HRD-A BOM:", len(generated_hrd_from))
    print("Sugeneruoti CABINET-A KIT BOM:", len(generated_cabinet_from))
    print("Paruošti importui:", len(ready))
    print("Dar neparuošti:", len(review) - len(ready))
    print("Diagnostikos įrašai:", len(diagnostics))
    print("KIT paruošti importui:", len(kit_ready))
    print("KIT dar neparuošti:", len(kit_review) - len(kit_ready))
    print("KIT diagnostikos įrašai:", len(kit_diagnostics))
    print("Neleistinų kategorijų produktai pašalinti iš KIT:", len(kit_excluded))
    for item in kit_excluded:
        print(
            "  -",
            item["sku"],
            f"(Reform kategorija: {item['category'] or 'nenustatyta'})",
        )
    print("Odoo pakeitimų neatlikta.")


if __name__ == "__main__":
    main()