from pathlib import Path
import sqlite3
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from pricing_control import enrich_pricing_workbook


def _sample_source(path: Path) -> None:
    workbook = Workbook()
    prices = workbook.active
    prices.title = "SO LINE PRICES"
    prices.append([
        "SKU", "Name", "Position Type", "Product Category",
        "Component / Purchase Cost", "Pricing Add-ons Total",
        "Adjustment Amount", "Final Reform SO Unit Price", "Status", "Issues",
    ])
    prices.append([
        "CAB-1", "Cabinet", "BOM", "ALL / CABINETS",
        10.0, 2.0, -0.14, 11.86, "COMPLETE", "",
    ])
    prices.append([
        "CAB-2", "Blocked cabinet", "BOM", "ALL / CABINETS",
        0.0, 2.0, -0.14, None, "BLOCKED", "Missing component price: X",
    ])

    components = workbook.create_sheet("BOM COMPONENT COSTS")
    components.append([
        "Top BOM SKU", "Level II SKU", "Level II Qty",
        "Purchased Component SKU", "Component Qty in Level II",
        "Total Qty in Top BOM", "Purchase Unit Price", "Component Cost",
        "Status", "Cost Source",
    ])
    components.append([
        "CAB-1", "PART-A", 2.0, "PART-A", 1.0, 2.0,
        5.0, 10.0, "OK", "LAST PURCHASE PRICE",
    ])
    components.append([
        "CAB-2", "X", 1.0, "X", 1.0, 1.0,
        None, None, "MISSING PRICE", "MISSING",
    ])

    categories = workbook.create_sheet("BOM CATEGORY BREAKDOWN")
    categories.append([
        "Top SKU", "Application Level", "Pricing Rule SKU", "Category ID",
        "Category Name", "Odoo Product Category", "Multiplier",
        "Assembly", "Storage", "Packaging", "Put on pallet", "Other", "Markup",
        "Add-ons Total", "Adjustment Rate", "Adjusted Add-ons",
    ])
    categories.append([
        "CAB-1", "LEVEL I BOM", "CAB-1", "C01", "Cabinet", "ALL / CABINETS",
        1.0, 1.0, 0.0, 1.0, 0.0, 0.0, 0.0, 2.0, -0.07, 1.86,
    ])
    workbook.save(path)


def test_enrich_pricing_workbook_adds_control_views_without_changing_price(tmp_path):
    source = tmp_path / "source.xlsx"
    output = tmp_path / "controlled.xlsx"
    _sample_source(source)

    enrich_pricing_workbook(
        source,
        output,
        git_commit="abc123",
        run_id="TEST-RUN",
        generated_at="2026-09-02T11:00:00",
    )

    workbook = load_workbook(output, data_only=True)
    assert workbook.sheetnames[:5] == [
        "CONTROL", "PRICE RESULTS", "PRICING RULES", "PRICE TRACE", "EXCEPTIONS"
    ]

    results = workbook["PRICE RESULTS"]
    assert results["H2"].value == 11.86
    assert results["I2"].value == "CALCULATED"
    assert results["I3"].value == "BLOCKED"

    raw = workbook["SO LINE PRICES"]
    assert raw["H2"].value == 11.86
    assert raw["H3"].value is None

    exceptions = workbook["EXCEPTIONS"]
    assert exceptions.max_row == 2
    assert exceptions["A2"].value == "CAB-2"

    trace = workbook["PRICE TRACE"]
    cab1_rows = [
        row for row in trace.iter_rows(min_row=2, values_only=True)
        if row[0] == "CAB-1"
    ]
    assert any(row[2] == "MATERIAL" and row[4] == "PART-A" for row in cab1_rows)
    assert any(row[2] == "PRICING ADD-ON" for row in cab1_rows)
    assert any(row[2] == "FINAL RESULT" and row[7] == 11.86 for row in cab1_rows)
    workbook.close()


def test_enrich_pricing_workbook_adds_changes_when_previous_is_supplied(tmp_path):
    previous = tmp_path / "previous.xlsx"
    current = tmp_path / "current.xlsx"
    output = tmp_path / "controlled.xlsx"
    _sample_source(previous)
    _sample_source(current)

    workbook = load_workbook(previous)
    workbook["SO LINE PRICES"]["H2"] = 10.0
    workbook.save(previous)

    enrich_pricing_workbook(current, output, previous=previous)
    workbook = load_workbook(output, data_only=True)
    assert "CHANGES" in workbook.sheetnames
    changes = workbook["CHANGES"]
    assert changes["A2"].value == "CAB-1"
    assert changes["B2"].value == 10.0
    assert changes["C2"].value == 11.86
    assert round(changes["D2"].value, 2) == 1.86
    workbook.close()


def test_enrich_pricing_workbook_writes_search_index(tmp_path):
    source = tmp_path / "source.xlsx"
    output = tmp_path / "controlled.xlsx"
    index = tmp_path / "Pricing_Explain_Index.sqlite"
    _sample_source(source)

    enrich_pricing_workbook(source, output, search_index=index)

    connection = sqlite3.connect(index)
    try:
        sku = connection.execute(
            "SELECT sku FROM price_result WHERE sku_key = ?", ("cab-1",)
        ).fetchone()
        trace_count = connection.execute(
            "SELECT COUNT(*) FROM price_trace WHERE sku_key = ?", ("cab-1",)
        ).fetchone()[0]
    finally:
        connection.close()
    assert sku == ("CAB-1",)
    assert trace_count == 3


def test_search_index_failure_does_not_block_controlled_workbook(tmp_path):
    source = tmp_path / "source.xlsx"
    output = tmp_path / "controlled.xlsx"
    index = tmp_path / "Pricing_Explain_Index.sqlite"
    _sample_source(source)

    with patch(
        "pricing_control._write_search_index",
        side_effect=sqlite3.OperationalError("database or disk is full"),
    ):
        enrich_pricing_workbook(source, output, search_index=index)

    assert output.is_file()
    workbook = load_workbook(output, data_only=True)
    assert "CONTROL" in workbook.sheetnames
    assert "PRICE RESULTS" in workbook.sheetnames
    workbook.close()
