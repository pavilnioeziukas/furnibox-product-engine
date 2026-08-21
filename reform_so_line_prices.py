"""Generate auditable final Reform SO unit prices without changing Odoo."""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from dataclasses import dataclass, field, replace
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from reform_map import find_sheet, read_edges
from so_pricing_rules import (
    compose_bom_category_rule,
    NonBomRule,
    PricingRule,
    load_config,
    non_bom_rules_from_config,
    pricing_rules_from_config,
)

from bom_import_manufacture_v5 import (
    KIT,
    MANUFACTURE,
    add_generated_apack_boms,
    add_generated_cabinet_assembled_kits,
    add_generated_hrd_assembled_boms,
)
from bom_import_pilot_v1 import calculate_levels
from manifest.manifest_writer import calculate_file_hash


PRICE_FILE = "Reform_Final_Prices.xlsx"
OUTPUT_FILE = "Reform_SO_Line_Prices.xlsx"
ADJUSTMENT = -0.07
TAMARA_PRICING_REFERENCE_PATH = (
    Path(__file__).resolve().parent
    / "manifest"
    / "tamara_pricing_reference.json"
)

ADDONS = (
    "Assembly",
    "Storage",
    "Packaging",
    "Put on pallet",
    "Other",
    "Markup",
)


def text(value):
    return str(value or "").strip()


def key(value):
    return text(value).casefold()


def number(value, default=0.0):
    if value in (None, ""):
        return default

    if isinstance(value, bool) or not isinstance(
        value,
        (int, float),
    ):
        raise ValueError(
            f"Expected number, got {value!r}"
        )

    return float(value)


def headers(sheet):
    return {
        text(cell.value): cell.column
        for cell in sheet[1]
        if cell.value not in (None, "")
    }


@dataclass
class Item:
    sku: str
    qty: float | None
    leaves: list[tuple[str, float]] = field(
        default_factory=list
    )


def load_prices(path: Path):
    """
    Load prepared prices used by the Reform pricing engine.

    Stored value:
        sku -> (name, price, price_source)

    price_source is optional in legacy workbooks.
    """
    wb = load_workbook(
        path,
        data_only=False,
        read_only=True,
    )

    result = {}

    if "REFORM PRICE LIST" in wb.sheetnames:
        ws = wb["REFORM PRICE LIST"]
        h = headers(ws)

        needed = [
            "Internal Reference",
            "Name",
            "Adjusted Furnibox Purchase Price",
            "Reform Markup Factor",
            "Reform Purchase Price",
        ]

        missing = [
            name
            for name in needed
            if name not in h
        ]

        if missing:
            wb.close()
            raise ValueError(
                "REFORM PRICE LIST missing: "
                + ", ".join(missing)
            )

        source_column = h.get("Price Source")

        for row in ws.iter_rows(
            min_row=2,
            values_only=True,
        ):
            sku = text(
                row[h["Internal Reference"] - 1]
            )

            if not sku:
                continue

            price = row[
                h["Reform Purchase Price"] - 1
            ]

            if not isinstance(
                price,
                (int, float),
            ):
                adjusted = row[
                    h["Adjusted Furnibox Purchase Price"] - 1
                ]
                factor = row[
                    h["Reform Markup Factor"] - 1
                ]

                if (
                    not isinstance(adjusted, (int, float))
                    or not isinstance(factor, (int, float))
                ):
                    continue

                price = adjusted * factor

            price_source = ""
            if source_column is not None:
                price_source = text(
                    row[source_column - 1]
                )

            sku_key = key(sku)

            if sku_key in result:
                wb.close()
                raise ValueError(
                    f"Duplicate Reform price SKU: {sku}"
                )

            result[sku_key] = (
                text(row[h["Name"] - 1]),
                float(price),
                price_source,
            )

    elif (
        "Purchase prices" in wb.sheetnames
        or "kainos" in wb.sheetnames
    ):
        source_sheet = (
            "Purchase prices"
            if "Purchase prices" in wb.sheetnames
            else "kainos"
        )

        for row in wb[source_sheet].iter_rows(
            min_row=2,
            values_only=True,
        ):
            if (
                text(row[0])
                and isinstance(row[7], (int, float))
            ):
                result.setdefault(
                    key(row[0]),
                    (
                        text(row[1]),
                        float(row[7]),
                        "LEGACY PURCHASE PRICE",
                    ),
                )

    else:
        wb.close()
        raise ValueError(
            f"No Reform price sheet in {path.name}"
        )

    wb.close()
    return result


def get_price_source(prices, sku):
    row = prices.get(key(sku))

    if not row or len(row) < 3:
        return ""

    return text(row[2])


def is_cabinet_part_price(prices, sku):
    return (
        get_price_source(
            prices,
            sku,
        ).casefold()
        == "cabinet part calculation".casefold()
    )

def load_rules(path: Path):
    wb = load_workbook(
        path,
        data_only=True,
        read_only=True,
    )

    result = {}

    for row in wb[
        "Kainodaros kategorijos"
    ].iter_rows(
        min_row=2,
        values_only=True,
    ):
        sku = text(row[0])

        if not sku:
            continue

        values = tuple(
            number(value)
            for value in row[4:10]
        )

        rule = PricingRule(
            sku,
            text(row[1]),
            text(row[2]),
            text(row[3]),
            *values,
        )

        # Legacy VLOOKUP uses the first match.
        result.setdefault(
            key(sku),
            rule,
        )

    wb.close()

    return result


def rules_from_config(document):
    return {
        key(rule.sku): rule
        for rule
        in pricing_rules_from_config(
            document
        )
    }


def non_bom_from_config(document):
    return [
        (
            rule.sku,
            rule.name,
            rule.product_category,
            rule.pricing_category,
            rule.preparation,
            rule.storage,
            rule.bag,
            rule.sticker,
        )
        for rule
        in non_bom_rules_from_config(
            document
        )
    ]


def normalize_graph(graph):
    """
    Normalize graph parent lookup by SKU.

    Original SKU spelling is preserved in child values,
    while lookups are case-insensitive.
    """
    result = {}

    for parent, children in graph.items():
        parent_key = key(parent)

        result[parent_key] = [
            (
                text(child),
                float(quantity),
            )
            for child, quantity in children
        ]

    return result


def add_generated_boms_to_graph(
    graph,
    products,
):
    """
    Extend the pricing graph with the BOM release generators.

    Keeping this adapter deliberately small means APACK, HRD-A and CABINET-A
    composition has one implementation: bom_import_manufacture_v5.py.
    """
    parents = {
        text(parent).upper()
        for parent in graph
    }
    lines = {
        text(parent).upper(): [
            {"component": child, "quantity": quantity}
            for child, quantity in children
        ]
        for parent, children in graph.items()
    }
    reform_products = {
        parent: {"category": "", "is_parent": True}
        for parent in parents
    }
    for product in products:
        sku = text(product.get("sku")).upper()
        if sku not in parents:
            continue
        category = text(product.get("product_category")).split("/")[-1].strip()
        reform_products[sku] = {
            "category": category,
            "is_parent": True,
        }

    bom_types = {
        parent: (
            KIT
            if reform_products[parent]["category"].upper()
            in {"CABINETS", "CABINET SHELF"}
            else MANUFACTURE
        )
        for parent in parents
    }
    levels = calculate_levels(parents, lines)

    add_generated_apack_boms(parents, lines, levels, bom_types)
    add_generated_hrd_assembled_boms(
        parents, lines, levels, bom_types, reform_products, lines
    )
    add_generated_cabinet_assembled_kits(
        parents, lines, levels, bom_types, reform_products, lines
    )

    return {
        key(parent): [
            (text(line["component"]), float(line["quantity"]))
            for line in lines[parent]
        ]
        for parent in parents
    }


def load_target_dataset_graph(
    dataset_path: Path,
    bom_path: Path,
):
    """Grąžina autoritetingą Furnibox BOM grafiką iš Validated Dataset."""
    dataset = json.loads(dataset_path.read_text(encoding="utf-8"))
    source = dataset.get("source") or {}
    expected_hash = text(source.get("file_hash"))
    actual_hash = calculate_file_hash(bom_path)
    if not expected_hash or expected_hash != actual_hash:
        raise ValueError(
            "Target Dataset sukurtas ne iš pateikto Reform BOM failo. "
            "Pirmiausia sugeneruokite naują pilną Dataset."
        )
    if text(dataset.get("environment")).lower() != "production":
        raise ValueError("Pricing leidžiamas tik iš Production Target Dataset.")

    graph = {}
    for product in dataset.get("products") or []:
        sku = text(product.get("sku"))
        if not sku:
            raise ValueError("Target Dataset turi BOM produktą be SKU.")
        children = []
        for component in product.get("components") or []:
            child = text(component.get("sku"))
            quantity = float(component.get("quantity") or 0)
            if not child or quantity <= 0:
                raise ValueError(f"Target Dataset BOM {sku} turi blogą komponentą.")
            children.append((child, quantity))
        if not children:
            raise ValueError(f"Target Dataset BOM {sku} neturi komponentų.")
        graph[key(sku)] = children

    if not graph:
        raise ValueError("Target Dataset neturi BOM struktūrų.")
    has_apack = any(
        text(product.get("sku")).upper().startswith("APACK-")
        for product in dataset.get("products") or []
    )
    if has_apack and not dataset.get("apack_hrd_transformation"):
        raise ValueError(
            "Target Dataset turi APACK, bet neturi APACK/HRD-A "
            "transformacijos žymos. Pricing iš seno Dataset blokuojamas."
        )
    return dataset, graph


def component_cost_only_manufacture_products(dataset):
    """Internal MANUFACTURE BOM children contribute cost, not extra add-ons."""
    return {
        key(product.get("sku"))
        for product in dataset.get("products") or []
        if text(product.get("bom_type")).upper() == "MANUFACTURE"
        and text(product.get("sku"))
        and (
            text(product.get("generated_from"))
            or text(product.get("sku")).upper().startswith("FPACK-")
            or text(product.get("sku")).upper().endswith("-PP")
        )
    }


def inherit_generated_apack_rules(rules, dataset):
    """Copy a missing APACK add-on profile only from its exact source FPACK."""
    result = dict(rules)

    for product in dataset.get("products") or []:
        sku = text(product.get("sku"))
        source_sku = text(product.get("generated_from"))
        sku_key = key(sku)
        source_key = key(source_sku)

        if (
            not sku.upper().startswith("APACK-")
            or not source_sku.upper().startswith("FPACK-")
            or sku_key in result
            or source_key not in result
        ):
            continue

        result[sku_key] = replace(
            result[source_key],
            sku=sku,
        )

    return result


def _pricing_rule_signature(rule):
    """Return the numeric values that actually determine the price."""
    return rule.addons


def _normalized_product_name(value, ignore_width=False):
    """Normalize Dataset names for conservative analog matching."""
    result = " ".join(text(value).upper().split())
    if ignore_width:
        result = re.sub(
            r"\bW\d+(?:[.,]\d+)?\b",
            "W*",
            result,
        )
    return result


def inherit_unambiguous_analog_rules(rules, dataset):
    """Copy missing BOM rules only from unanimous catalog analogs.

    Exact product type and Product Name 2 are authoritative. Shelf prepack
    products additionally allow the width token to differ, while depth and
    all descriptive text must remain identical. Inferred rules never become
    source candidates for another inference.
    """
    result = dict(rules)
    source_rules = dict(rules)
    catalog = [
        product
        for product in dataset.get("product_catalog") or []
        if product.get("has_bom")
        and text(product.get("sku"))
        and text(product.get("product_type"))
        and text(product.get("name_2"))
    ]

    def matching_rules(target, ignore_width=False):
        target_type = key(target.get("product_type"))
        target_name = _normalized_product_name(
            target.get("name_2"),
            ignore_width=ignore_width,
        )
        matches = []
        for candidate in catalog:
            if key(candidate.get("product_type")) != target_type:
                continue
            if _normalized_product_name(
                candidate.get("name_2"),
                ignore_width=ignore_width,
            ) != target_name:
                continue
            rule = source_rules.get(key(candidate.get("sku")))
            if rule is not None:
                matches.append(rule)
        return matches

    for product in catalog:
        sku = text(product.get("sku"))
        if key(sku) in result:
            continue

        candidates = matching_rules(product)
        if not candidates and sku.upper().endswith("-PP"):
            candidates = matching_rules(
                product,
                ignore_width=True,
            )
        signatures = {
            _pricing_rule_signature(candidate)
            for candidate in candidates
        }
        if len(signatures) != 1:
            continue

        result[key(sku)] = replace(
            candidates[0],
            sku=sku,
        )

    return result


def _target_component_skus(product):
    return [
        text(component.get("sku"))
        for component in product.get("components") or []
        if text(component.get("sku"))
    ]


def _target_market_pack_code(product, eu_code, us_code):
    """Choose the pack category from the transformed physical BOM first."""
    component_skus = {
        key(sku)
        for sku in _target_component_skus(product)
    }
    if key("L0377") in component_skus:
        return us_code
    identity = " ".join(
        text(product.get(field)).upper()
        for field in ("sku", "source_sku", "generated_from")
    )
    if any(
        token.startswith(("US-", "USB-")) or "-US-" in token
        for token in identity.split()
    ):
        return us_code
    return eu_code


def _target_top_market_pack_code(sku, eu_code, us_code):
    normalized = text(sku).upper()
    if normalized.startswith(("US-", "USB-")) or "-US-" in normalized:
        return us_code
    return eu_code


def _shelf_pp_base_category(sku):
    normalized = text(sku).upper()
    if "LED" in normalized:
        return "8.2"
    if "-ROD-" in normalized:
        return "8.1"
    return "8"


def load_tamara_pricing_reference(path=TAMARA_PRICING_REFERENCE_PATH):
    if not Path(path).is_file():
        raise ValueError(f"Nerastas Tamaros kainodaros etalonas: {path}")
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    if payload.get("schema_version") != 1:
        raise ValueError("Nepalaikoma Tamaros kainodaros etalono versija.")
    if payload.get("conflicts"):
        raise ValueError("Tamaros kainodaros etalone yra prieštaringų SKU taisyklių.")
    return {
        key(row.get("sku")): text(row.get("expression"))
        for row in payload.get("sku_expressions") or []
        if text(row.get("sku")) and text(row.get("expression"))
    }


def apply_target_business_category_rules(
    rules,
    dataset,
    document,
    reference=None,
):
    """Overlay transformed products with Tamara CATEGORY/BOM PAP logic.

    Returned ``authoritative`` products carry the complete add-on expression
    for that BOM.  Their child material costs are still resolved recursively,
    but child add-on rules must not be added again.
    """
    result = dict(rules)
    authoritative = set()
    products = dataset.get("products") or []
    if reference is None:
        reference = load_tamara_pricing_reference()

    def assign(sku, expression):
        result[key(sku)] = compose_bom_category_rule(
            sku,
            expression,
            document,
        )
        authoritative.add(key(sku))

    for product in products:
        sku = text(product.get("sku"))
        normalized = sku.upper()
        product_type = text(product.get("product_type")).upper()
        children = _target_component_skus(product)

        exact_expression = reference.get(key(sku))
        if exact_expression:
            assign(sku, exact_expression)
            continue

        if product_type == "SHELF PREPACK" or (
            normalized.endswith("-PP") and "SHELF" in normalized
        ):
            pack = _target_market_pack_code(product, "25.1", "26.1")
            assign(sku, f"{_shelf_pp_base_category(sku)}+{pack}")
            continue

        if normalized.startswith("APACK-"):
            pack = _target_market_pack_code(product, "22.1", "23.1")
            assign(sku, f"12+{pack}+24.1")
            continue

        shelf_pp = next(
            (
                child
                for child in children
                if child.upper().endswith("-PP") and "SHELF" in child.upper()
            ),
            "",
        )
        if product_type == "CABINET SHELF" and shelf_pp:
            pack = _target_top_market_pack_code(sku, "25.1", "26.1")
            base = _shelf_pp_base_category(shelf_pp)
            if base == "8":
                expression = f"8+{pack}+7+30+35"
            elif base == "8.1":
                expression = f"8.1+{pack}+7+35"
            else:
                expression = f"8.2+{pack}+7+35+11"
            assign(sku, expression)
            continue

        has_apack = any(child.upper().startswith("APACK-") for child in children)
        if normalized.endswith("-A") and has_apack:
            pack = _target_top_market_pack_code(sku, "22.1", "23.1")
            assign(sku, f"12+9+{pack}+24.1")

    return result, authoritative


def exclude_bom_products_from_non_bom(items, graph):
    """A current Target BOM parent cannot also be priced as a non-BOM SKU."""
    bom_skus = {key(sku) for sku in graph}
    return [
        item
        for item in items
        if item and key(item[0]) not in bom_skus
    ]


def load_reform_boms(
    path: Path,
    products,
    rules=None,
    dataset_path: Path | None = None,
):
    """
    Build pricing input and preserve the full Reform BOM graph.

    Pricing coverage:
    - configured Reform top-level BOM products;
    - FPACK BOMs found in the Reform BOM graph and having a pricing rule;
    - APACK BOMs found in the Reform BOM graph and having a pricing rule.

    The graph is not limited to Level I -> Level II -> Level III.
    It can be recursively resolved to any depth.
    """
    workbook = load_workbook(
        path,
        data_only=True,
        read_only=True,
    )

    sheet = find_sheet(workbook)

    _, raw_graph, diagnostics = (
        read_edges(sheet)
    )

    workbook.close()

    invalid = [
        row
        for row in diagnostics
        if row[0]
        in {
            "INVALID QUANTITY",
            "NON-POSITIVE QUANTITY",
        }
    ]

    if invalid:
        raise ValueError(
            f"Reform BOM turi "
            f"{len(invalid)} "
            "neteisingų kiekių. "
            "Pirmiausia pataisykite įvestį."
        )

    graph = normalize_graph(raw_graph)
    if dataset_path is not None:
        _, graph = load_target_dataset_graph(dataset_path, path)
    else:
        graph = add_generated_boms_to_graph(graph, products)

    pricing_products = {}

    for product in products:
        sku = text(
            product.get("sku")
        )

        if not sku:
            continue

        pricing_products[
            key(sku)
        ] = {
            "sku": sku,
            "product_category": text(
                product.get(
                    "product_category"
                )
            ),
        }

    # FPACK/APACK are valid standalone pricing objects as well as BOM nodes.
    # Synthetic APACK graph nodes above follow the same FPACK -> APACK rule
    # already used by the BOM import pipeline.
    # Dataset gali turėti FPACK/APACK/Shelf-PP tarpinius kainos objektus.
    if rules:
        for parent_key in graph:
            if parent_key in pricing_products:
                continue

            rule = rules.get(
                parent_key
            )

            if rule is None:
                continue

            original_sku = text(
                rule.sku
            )

            upper_sku = (
                original_sku.upper()
            )

            if not (
                upper_sku.startswith(
                    "FPACK-"
                )
                or upper_sku.startswith(
                    "APACK-"
                )
                or upper_sku.endswith("-PP")
            ):
                continue

            pricing_products[
                parent_key
            ] = {
                "sku": original_sku,
                "product_category": (
                    text(
                        rule.odoo_category
                    )
                    or text(
                        rule.category_name
                    )
                ),
            }

    result = {}

    for product in pricing_products.values():
        top = text(
            product.get("sku")
        )

        items = []

        for child, quantity in graph.get(
            key(top),
            [],
        ):
            items.append(
                Item(
                    sku=child,
                    qty=float(quantity),
                    leaves=[
                        (
                            sku,
                            float(qty),
                        )
                        for sku, qty
                        in graph.get(
                            key(child),
                            [],
                        )
                    ],
                )
            )

        result[top] = (
            text(
                product.get(
                    "product_category"
                )
            ),
            items,
        )

    return result, graph
def load_boms(path: Path):
    """
    Load legacy pricing workbook BOM layout.

    This path is kept for legacy tests and backward compatibility.
    """
    wb = load_workbook(
        path,
        data_only=True,
        read_only=True,
    )

    result = {}
    current_top = ""
    current_item = None

    for row in wb[
        "bomai"
    ].iter_rows(
        min_row=3,
        values_only=True,
    ):
        top = text(row[1])

        if not top:
            continue

        if top != current_top:
            current_top = top
            current_item = None

            result.setdefault(
                top,
                (
                    text(row[2]),
                    [],
                ),
            )

        if text(row[3]):
            current_item = Item(
                text(row[3]),
                (
                    number(row[5])
                    if row[5]
                    not in (None, "")
                    else None
                ),
            )

            result[top][1].append(
                current_item
            )

        if text(row[6]):
            if current_item is None:
                current_item = Item(
                    "",
                    None,
                )

                result[top][1].append(
                    current_item
                )

            current_item.leaves.append(
                (
                    text(row[6]),
                    number(
                        row[7],
                        1.0,
                    ),
                )
            )

    wb.close()

    return result


def build_graph_from_legacy_boms(
    boms,
):
    """
    Build a minimal graph from the old workbook structure.

    This keeps existing tests and the legacy calculation path working
    while the Production application uses the full Reform BOM graph.
    """
    graph = {}

    for _, (_, items) in boms.items():
        for item in items:
            if (
                not item.sku
                or not item.leaves
            ):
                continue

            graph[
                key(item.sku)
            ] = [
                (
                    text(child),
                    float(quantity),
                )
                for child, quantity
                in item.leaves
            ]

    return graph


def load_non_bom(path: Path):
    wb = load_workbook(
        path,
        data_only=True,
        read_only=True,
    )

    rows = []

    for row in wb[
        "Ne BOM pozicijos"
    ].iter_rows(
        min_row=2,
        values_only=True,
    ):
        if text(row[0]):
            rows.append(
                (
                    text(row[0]),
                    text(row[1]),
                    text(row[2]),
                    text(row[3]),
                    *[
                        number(value)
                        for value
                        in row[6:10]
                    ],
                )
            )

    wb.close()

    return rows


def load_rule_conflicts(
    path: Path,
):
    wb = load_workbook(
        path,
        data_only=True,
        read_only=True,
    )

    seen = {}
    conflicts = []

    for excel_row, row in enumerate(
        wb[
            "Kainodaros kategorijos"
        ].iter_rows(
            min_row=2,
            values_only=True,
        ),
        2,
    ):
        sku = text(row[0])

        if not sku:
            continue

        values = tuple(
            number(value)
            for value in row[4:10]
        )

        sku_key = key(sku)

        if (
            sku_key in seen
            and seen[sku_key][1]
            != values
        ):
            conflicts.append(
                (
                    sku,
                    seen[sku_key][0],
                    excel_row,
                    seen[sku_key][1],
                    values,
                )
            )

        else:
            seen.setdefault(
                sku_key,
                (
                    excel_row,
                    values,
                ),
            )

    wb.close()

    return conflicts


def breakdown(
    rule,
    multiplier,
    level,
):
    return {
        "level": level,
        "rule": rule,
        "multiplier": multiplier,
        "addons": tuple(
            value * multiplier
            for value
            in rule.addons
        ),
    }


def resolve_component_cost(
    sku,
    prices,
    graph,
    trail=None,
    cache=None,
    bom_cost_skus=None,
):
    """
    Resolve one SKU unit cost.

    Business priority:

    1. If SKU has a prepared direct price in Reform_Final_Prices,
       use that price.

       This price already comes from the purchasing-price layer:
       last purchase price plus the configured purchase-price
       adjustment / Reform transfer logic.

    2. If there is no prepared direct price but SKU has a BOM,
       recursively calculate:
           SUM(child resolved cost * child quantity)

    3. If SKU has neither a prepared direct price nor a BOM,
       return MISSING PRICE.

    Odoo Standard Price is intentionally NOT used as a fallback.
    """
    sku = text(sku)
    sku_key = key(sku)

    if not sku:
        return {
            "cost": None,
            "source": "MISSING",
            "issues": [
                "Missing SKU"
            ],
            "leaves": [],
        }

    if cache is None:
        cache = {}

    if bom_cost_skus is None:
        bom_cost_skus = set()
    elif not isinstance(bom_cost_skus, set):
        bom_cost_skus = {
            key(value)
            for value in bom_cost_skus
        }

    if sku_key in cache:
        return cache[sku_key]

    # Priority 1:
    # prepared direct purchase / transfer price.
    if sku_key in prices and sku_key not in bom_cost_skus:
        unit_price = float(
            prices[sku_key][1]
        )

        if unit_price <= 0:
            result = {
                "cost": None,
                "source": "NON-POSITIVE DIRECT PRICE",
                "issues": [
                    f"Non-positive component price: {sku} ({unit_price:g})"
                ],
                "leaves": [
                    {
                        "sku": sku,
                        "qty": 1.0,
                        "unit_price": None,
                        "source": "NON-POSITIVE DIRECT PRICE",
                    }
                ],
            }
            cache[sku_key] = result
            return result

        price_source = (
            get_price_source(
                prices,
                sku,
            )
            or "DIRECT PRICE"
        )

        result = {
            "cost": unit_price,
            "source": price_source,
            "issues": [],
            "leaves": [
                {
                    "sku": sku,
                    "qty": 1.0,
                    "unit_price": unit_price,
                    "source": price_source,
                }
            ],
        }

        cache[sku_key] = result

        return result

    children = graph.get(
        sku_key,
        [],
    )

    # Priority 3:
    # no direct price and no BOM.
    if not children:
        result = {
            "cost": None,
            "source": "MISSING",
            "issues": [
                f"Missing component price: {sku}"
            ],
            "leaves": [
                {
                    "sku": sku,
                    "qty": 1.0,
                    "unit_price": None,
                    "source": "MISSING",
                }
            ],
        }

        cache[sku_key] = result

        return result

    trail = tuple(
        trail or ()
    )

    trail_keys = {
        key(value)
        for value in trail
    }

    if sku_key in trail_keys:
        return {
            "cost": None,
            "source": "CYCLE",
            "issues": [
                "BOM pricing cycle: "
                + " -> ".join(
                    trail
                    + (sku,)
                )
            ],
            "leaves": [],
        }

    next_trail = (
        trail
        + (sku,)
    )

    # Priority 2:
    # recursively calculate BOM.
    total = 0.0
    issues = []
    leaves = []

    for (
        child_sku,
        child_qty,
    ) in children:
        child_qty = float(
            child_qty
        )

        child_result = (
            resolve_component_cost(
                child_sku,
                prices,
                graph,
                trail=next_trail,
                cache=cache,
                bom_cost_skus=bom_cost_skus,
            )
        )

        for leaf in child_result[
            "leaves"
        ]:
            leaves.append(
                {
                    "sku": leaf["sku"],
                    "qty": (
                        float(
                            leaf["qty"]
                        )
                        * child_qty
                    ),
                    "unit_price": (
                        leaf[
                            "unit_price"
                        ]
                    ),
                    "source": (
                        leaf[
                            "source"
                        ]
                    ),
                }
            )

        if child_result[
            "cost"
        ] is None:
            issues.extend(
                child_result[
                    "issues"
                ]
            )

            continue

        total += (
            float(
                child_result[
                    "cost"
                ]
            )
            * child_qty
        )

    issues = list(
        dict.fromkeys(
            issues
        )
    )

    if issues:
        result = {
            "cost": None,
            "source": "BOM BLOCKED",
            "issues": issues,
            "leaves": leaves,
        }

        cache[sku_key] = result

        return result

    result = {
        "cost": total,
        "source": "BOM CALCULATION",
        "issues": [],
        "leaves": leaves,
    }

    cache[sku_key] = result

    return result


def calculate_boms(
    boms,
    prices,
    rules,
    adjustment=ADJUSTMENT,
    graph=None,
    component_cost_only_tops=None,
    authoritative_rule_tops=None,
):
    """
    Calculate BOM sale prices.

    Cost resolution is recursive.

    Pricing add-on logic remains the existing business logic:
    - Level II BOM add-ons multiplied by Level II quantity.
    - Direct Level II add-ons applied once.
    - Level I add-ons applied once.
    """
    if graph is None:
        graph = (
            build_graph_from_legacy_boms(
                boms
            )
        )

    else:
        graph = normalize_graph(
            graph
        )

    component_cost_only_tops = {
        key(value)
        for value in (
            component_cost_only_tops
            or set()
        )
    }
    bom_cost_skus = set(component_cost_only_tops)
    authoritative_rule_tops = {
        key(value)
        for value in (
            authoritative_rule_tops
            or set()
        )
    }

    results = []
    details = []

    # Shared cache prevents recalculating the same HRD / sub-BOM
    # hundreds of times.
    cost_cache = {}

    for top, (
        category,
        items,
    ) in boms.items():
        cost = 0.0
        issues = []
        applied = []
        component_details = []

        if not items:
            issues.append(
                f"Target BOM has no components: {top}"
            )

        for item in items:
            if not item.sku:
                issues.append(
                    "Leaf components have no "
                    "Level II BOM item"
                )

            if item.qty is None:
                issues.append(
                    "Missing Level II quantity: "
                    f"{item.sku or '[unknown]'}"
                )

            item_qty = (
                item.qty
                or 0.0
            )

            if not item.sku:
                continue

            resolved = (
                resolve_component_cost(
                    item.sku,
                    prices,
                    graph,
                    cache=cost_cache,
                    bom_cost_skus=bom_cost_skus,
                )
            )

            if resolved[
                "cost"
            ] is None:
                issues.extend(
                    resolved[
                        "issues"
                    ]
                )

            else:
                cost += (
                    float(
                        resolved[
                            "cost"
                        ]
                    )
                    * item_qty
                )

            # Flatten recursive BOM into priced leaf components
            # for an auditable cost breakdown.
            if resolved[
                "leaves"
            ]:
                for leaf in resolved[
                    "leaves"
                ]:
                    leaf_qty = float(
                        leaf["qty"]
                    )

                    unit_price = leaf[
                        "unit_price"
                    ]

                    total_qty = (
                        item_qty
                        * leaf_qty
                    )

                    component_details.append(
                        {
                            "top": top,
                            "level_ii": (
                                item.sku
                            ),
                            "level_ii_qty": (
                                item_qty
                            ),
                            "component": (
                                leaf["sku"]
                            ),
                            "component_qty": (
                                leaf_qty
                            ),
                            "total_qty": (
                                total_qty
                            ),
                            "unit_price": (
                                unit_price
                            ),
                            "line_cost": (
                                None
                                if unit_price
                                is None
                                else (
                                    total_qty
                                    * unit_price
                                )
                            ),
                            "status": (
                                "MISSING PRICE"
                                if unit_price
                                is None
                                else "OK"
                            ),
                            "cost_source": (
                                leaf[
                                    "source"
                                ]
                            ),
                        }
                    )

            else:
                component_details.append(
                    {
                        "top": top,
                        "level_ii": (
                            item.sku
                        ),
                        "level_ii_qty": (
                            item_qty
                        ),
                        "component": (
                            item.sku
                        ),
                        "component_qty": (
                            1.0
                        ),
                        "total_qty": (
                            item_qty
                        ),
                        "unit_price": None,
                        "line_cost": None,
                        "status": (
                            "MISSING PRICE"
                        ),
                        "cost_source": (
                            resolved[
                                "source"
                            ]
                        ),
                    }
                )

            if key(top) in authoritative_rule_tops:
                # Tamara's product category expression already represents
                # the complete add-on combination for this BOM.  Child
                # material cost remains recursive, but child add-ons would
                # duplicate that authoritative product-category total.
                continue

            if key(top) in component_cost_only_tops:
                # APACK, HRD-A and Shelf-PP are generated internal
                # MANUFACTURE products. Their child materials already enter
                # recursive component cost. Requiring another sales add-on
                # rule for every screw, package or sticker would apply the
                # legacy Level II rule at the wrong structural level.
                continue

            has_bom = bool(
                graph.get(
                    key(item.sku),
                    [],
                )
            )

            if has_bom:
                multiplier = (
                    item_qty
                )

                level = (
                    "LEVEL II BOM"
                )

            else:
                # Existing legacy rule:
                # direct item pricing add-on is
                # applied once, not by quantity.
                multiplier = 1.0

                level = (
                    "DIRECT LEVEL II"
                )

            item_key = key(
                item.sku
            )

            cabinet_part_direct = (
                level == "DIRECT LEVEL II"
                and is_cabinet_part_price(
                    prices,
                    item.sku,
                )
            )

            if cabinet_part_direct:
                # Furnix Cabinet Part pricing is already represented
                # by the calculated Furnix -> Furnibox transfer price.
                # Do not require or apply a separate DIRECT LEVEL II
                # pricing rule, otherwise the same detail is priced twice.
                pass

            elif item_key not in rules:
                issues.append(
                    f"Missing {level} "
                    "pricing rule: "
                    f"{item.sku}"
                )

            else:
                applied.append(
                    breakdown(
                        rules[
                            item_key
                        ],
                        multiplier,
                        level,
                    )
                )

        top_key = key(
            top
        )

        if top_key not in rules:
            issues.append(
                "Missing LEVEL I BOM "
                f"pricing rule: {top}"
            )

        else:
            applied.append(
                breakdown(
                    rules[
                        top_key
                    ],
                    1.0,
                    "LEVEL I BOM",
                )
            )

        addon_values = tuple(
            sum(
                row["addons"][index]
                for row in applied
            )
            for index
            in range(6)
        )

        addon_total = sum(
            addon_values
        )

        issues = list(
            dict.fromkeys(
                issues
            )
        )

        issue_text = "; ".join(
            issues
        )

        results.append(
            {
                "sku": top,
                "name": prices.get(
                    top_key,
                    ("", 0),
                )[0],
                "type": "BOM",
                "category": category,
                "cost": cost,
                "addons": (
                    addon_values
                ),
                "adjustment": (
                    addon_total
                    * adjustment
                ),
                "final": (
                    cost
                    + addon_total
                    * (
                        1
                        + adjustment
                    )
                    if not issues
                    else None
                ),
                "status": (
                    "COMPLETE"
                    if not issues
                    else "BLOCKED"
                ),
                "issues": (
                    issue_text
                ),
                "component_details": (
                    component_details
                ),
            }
        )

        for row in applied:
            row["top"] = top

            details.append(
                row
            )

    return (
        results,
        details,
    )


def write_component_cost_breakdown(
    workbook,
    bom_rows,
):
    """
    Write the auditable cost calculation behind every BOM.

    Recursive BOMs are flattened to the priced leaf components.
    """
    ws = workbook.create_sheet(
        "BOM COMPONENT COSTS",
        1,
    )

    ws.append(
        [
            "Top BOM SKU",
            "Level II SKU",
            "Level II Qty",
            "Purchased Component SKU",
            "Component Qty in Level II",
            "Total Qty in Top BOM",
            "Purchase Unit Price",
            "Component Cost",
            "Status",
            "Cost Source",
        ]
    )

    for row in bom_rows:
        for detail in row.get(
            "component_details",
            [],
        ):
            ws.append(
                [
                    detail[
                        "top"
                    ],
                    detail[
                        "level_ii"
                    ],
                    detail[
                        "level_ii_qty"
                    ],
                    detail[
                        "component"
                    ],
                    detail[
                        "component_qty"
                    ],
                    detail[
                        "total_qty"
                    ],
                    detail[
                        "unit_price"
                    ],
                    detail[
                        "line_cost"
                    ],
                    detail[
                        "status"
                    ],
                    detail.get(
                        "cost_source",
                        "",
                    ),
                ]
            )

    style(
        ws,
        "4472C4",
    )

    widths(
        ws,
        [
            31,
            31,
            14,
            31,
            22,
            20,
            20,
            18,
            16,
            18,
        ],
    )

    for row_number in range(
        2,
        ws.max_row + 1,
    ):
        ws.cell(
            row_number,
            3,
        ).number_format = "0.####"

        ws.cell(
            row_number,
            5,
        ).number_format = "0.####"

        ws.cell(
            row_number,
            6,
        ).number_format = "0.####"

        ws.cell(
            row_number,
            7,
        ).number_format = (
            '0.0000 [$€-x-euro2]'
        )

        ws.cell(
            row_number,
            8,
        ).number_format = (
            '0.0000 [$€-x-euro2]'
        )


def calculate_non_bom(
    items,
    prices,
):
    results = []

    for (
        sku,
        name,
        category,
        pricing_category,
        preparation,
        storage,
        bag,
        sticker,
    ) in items:
        missing = (
            key(sku)
            not in prices
        )

        cost = prices.get(
            key(sku),
            ("", 0),
        )[1]

        addon_values = (
            0.0,
            storage,
            (
                preparation
                + bag
                + sticker
            ),
            0.0,
            0.0,
            0.0,
        )

        results.append(
            {
                "sku": sku,
                "name": name,
                "type": "NON-BOM",
                "category": (
                    category
                ),
                "pricing_category": (
                    pricing_category
                ),
                "cost": cost,
                "addons": (
                    addon_values
                ),
                "preparation": (
                    preparation
                ),
                "bag": bag,
                "sticker": sticker,
                "adjustment": 0.0,
                "final": (
                    cost
                    + sum(
                        addon_values
                    )
                    if not missing
                    else None
                ),
                "status": (
                    "BLOCKED"
                    if missing
                    else "COMPLETE"
                ),
                "issues": (
                    f"Missing purchase price: {sku}"
                    if missing
                    else ""
                ),
            }
        )

    return results


def style(
    sheet,
    color="1F4E78",
):
    for cell in sheet[1]:
        cell.fill = PatternFill(
            "solid",
            fgColor=color,
        )

        cell.font = Font(
            color="FFFFFF",
            bold=True,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    sheet.row_dimensions[
        1
    ].height = 42

    sheet.freeze_panes = "A2"

    sheet.auto_filter.ref = (
        sheet.dimensions
    )

    sheet.sheet_view.showGridLines = (
        False
    )


def widths(
    sheet,
    values,
):
    for index, value in enumerate(
        values,
        1,
    ):
        sheet.column_dimensions[
            sheet.cell(
                1,
                index,
            ).column_letter
        ].width = value


def build_reform_so_line_prices(
    model_path: Path,
    price_path: Path,
    output_path: Path,
    adjustment=ADJUSTMENT,
):
    """
    Legacy workbook-based entry point.

    Kept for compatibility and tests.
    """
    if not -1 < adjustment <= 0:
        raise ValueError(
            "Adjustment must be "
            "between -100% and 0%"
        )

    prices = load_prices(
        price_path
    )

    rules = load_rules(
        model_path
    )

    legacy_boms = load_boms(
        model_path
    )

    bom_rows, details = (
        calculate_boms(
            legacy_boms,
            prices,
            rules,
            adjustment=adjustment,
        )
    )

    non_rows = (
        calculate_non_bom(
            load_non_bom(
                model_path
            ),
            prices,
        )
    )

    all_rows = sorted(
        bom_rows + non_rows,
        key=lambda row: (
            row["type"],
            row[
                "sku"
            ].casefold(),
        ),
    )

    wb = Workbook()

    ws = wb.active
    ws.title = "SO LINE PRICES"

    ws.append(
        [
            "SKU",
            "Name",
            "Position Type",
            "Product Category",
            "Component / Purchase Cost",
            *ADDONS,
            "Pricing Add-ons Total",
            "Adjustment Rate",
            "Adjustment Amount",
            "Final Reform SO Unit Price",
            "Status",
            "Issues",
        ]
    )

    for row in all_rows:
        ws.append(
            [
                row["sku"],
                row["name"],
                row["type"],
                row["category"],
                row["cost"],
                *row["addons"],
                sum(
                    row["addons"]
                ),
                (
                    adjustment
                    if row[
                        "type"
                    ]
                    == "BOM"
                    else 0
                ),
                row[
                    "adjustment"
                ],
                row["final"],
                row["status"],
                row["issues"],
            ]
        )

    style(ws)

    widths(
        ws,
        [
            31,
            42,
            14,
            30,
            22,
            14,
            14,
            14,
            16,
            14,
            14,
            20,
            16,
            19,
            25,
            14,
            75,
        ],
    )

    for row_number in range(
        2,
        ws.max_row + 1,
    ):
        for column in (
            list(
                range(
                    5,
                    13,
                )
            )
            + [
                14,
                15,
            ]
        ):
            ws.cell(
                row_number,
                column,
            ).number_format = (
                '0.0000 [$€-x-euro2]'
            )

        ws.cell(
            row_number,
            13,
        ).number_format = (
            "0.0%"
        )

    write_component_cost_breakdown(
        wb,
        bom_rows,
    )

    ws = wb.create_sheet(
        "BOM CATEGORY BREAKDOWN"
    )

    ws.append(
        [
            "Top SKU",
            "Application Level",
            "Pricing Rule SKU",
            "Category ID",
            "Category Name",
            "Odoo Product Category",
            "Multiplier",
            *ADDONS,
            "Add-ons Total",
            "Adjustment Rate",
            "Adjusted Add-ons",
        ]
    )

    for row in details:
        rule = row["rule"]

        total = sum(
            row["addons"]
        )

        ws.append(
            [
                row["top"],
                row["level"],
                rule.sku,
                rule.category_id,
                rule.category_name,
                rule.odoo_category,
                row["multiplier"],
                *row["addons"],
                total,
                adjustment,
                total
                * (
                    1
                    + adjustment
                ),
            ]
        )

    style(
        ws,
        "5B9BD5",
    )

    widths(
        ws,
        [
            31,
            19,
            31,
            13,
            24,
            31,
            12,
        ]
        + [14] * 6
        + [
            18,
            16,
            18,
        ],
    )

    ws = wb.create_sheet(
        "CATEGORY RULES"
    )

    ws.append(
        [
            "Category ID",
            "Category Name",
            "Odoo Product Category",
            "Products",
            *ADDONS,
            "Total",
            *[
                f"{name} Applied"
                for name
                in ADDONS
            ],
        ]
    )

    variants = Counter(
        (
            rule.category_id,
            rule.category_name,
            rule.odoo_category,
            *rule.addons,
        )
        for rule
        in rules.values()
    )

    for values, count in sorted(
        variants.items(),
        key=lambda item: tuple(
            str(value)
            for value
            in item[0][:3]
        ),
    ):
        (
            category_id,
            name,
            odoo,
            *addon_values,
        ) = values

        ws.append(
            [
                category_id,
                name,
                odoo,
                count,
                *addon_values,
                sum(
                    addon_values
                ),
                *[
                    (
                        "YES"
                        if value
                        else "NO"
                    )
                    for value
                    in addon_values
                ],
            ]
        )

    style(
        ws,
        "70AD47",
    )

    widths(
        ws,
        [
            13,
            25,
            32,
            12,
        ]
        + [14] * 7
        + [18] * 6,
    )

    ws = wb.create_sheet(
        "NON-BOM RULES"
    )

    ws.append(
        [
            "SKU",
            "Name",
            "Product Category",
            "Pricing Category",
            "Purchase Price",
            "Pack Preparation",
            "Storage",
            "Bag",
            "Sticker",
            "Final Unit Price",
            "Status",
            "Issues",
        ]
    )

    for row in non_rows:
        ws.append(
            [
                row["sku"],
                row["name"],
                row["category"],
                row[
                    "pricing_category"
                ],
                row["cost"],
                row[
                    "preparation"
                ],
                row["addons"][1],
                row["bag"],
                row["sticker"],
                row["final"],
                row["status"],
                row["issues"],
            ]
        )

    style(
        ws,
        "8064A2",
    )

    widths(
        ws,
        [
            31,
            42,
            27,
            17,
            18,
            18,
            14,
            12,
            12,
            18,
            14,
            65,
        ],
    )

    ws = wb.create_sheet(
        "DIAGNOSTICS"
    )

    ws.append(
        [
            "Position Type",
            "SKU",
            "Status",
            "Issues",
        ]
    )

    for row in all_rows:
        if (
            row["status"]
            == "BLOCKED"
        ):
            ws.append(
                [
                    row["type"],
                    row["sku"],
                    row["status"],
                    row["issues"],
                ]
            )

    for (
        sku,
        first_row,
        duplicate_row,
        first_values,
        duplicate_values,
    ) in load_rule_conflicts(
        model_path
    ):
        ws.append(
            [
                "PRICING RULE",
                sku,
                "CONFLICT",
                (
                    f"Rows {first_row} "
                    f"and {duplicate_row} "
                    "differ; first row used. "
                    f"{first_values} vs "
                    f"{duplicate_values}"
                ),
            ]
        )

    style(
        ws,
        "C00000",
    )

    widths(
        ws,
        [
            18,
            34,
            14,
            100,
        ],
    )

    ws = wb.create_sheet(
        "INFO"
    )

    for row in [
        (
            "Purpose",
            "Final Reform SO line unit price",
        ),
        (
            "BOM rule",
            (
                "Components + Assembly + Storage + "
                "Packaging + Put on pallet + Other + Markup"
            ),
        ),
        (
            "Component price priority",
            (
                "Prepared direct price; otherwise recursive BOM; "
                "otherwise BLOCKED"
            ),
        ),
        (
            "Odoo Standard Price fallback",
            "NO",
        ),
        (
            "Adjustment",
            adjustment,
        ),
        (
            "Markup meaning",
            "Additive monetary amount, not percentage",
        ),
        (
            "Non-BOM rule",
            (
                "Purchase price + pack preparation + "
                "storage + bag + sticker"
            ),
        ),
        (
            "BOM products",
            len(
                bom_rows
            ),
        ),
        (
            "Non-BOM products",
            len(
                non_rows
            ),
        ),
        (
            "Blocked",
            sum(
                row["status"]
                == "BLOCKED"
                for row
                in all_rows
            ),
        ),
        (
            "Odoo changed",
            "NO",
        ),
    ]:
        ws.append(row)

    ws.column_dimensions[
        "A"
    ].width = 30

    ws.column_dimensions[
        "B"
    ].width = 110

    ws.sheet_view.showGridLines = (
        False
    )

    wb.calculation.fullCalcOnLoad = (
        True
    )

    wb.calculation.forceFullCalc = (
        True
    )

    wb.calculation.calcMode = (
        "auto"
    )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    wb.save(
        output_path
    )

    return (
        len(bom_rows),
        len(non_rows),
        sum(
            row["status"]
            == "BLOCKED"
            for row
            in all_rows
        ),
    )


def write_price_workbook(
    bom_rows,
    non_rows,
    details,
    rules,
    adjustment,
    output_path,
):
    """
    Write auditable workbook for application-owned pricing inputs.
    """
    all_rows = sorted(
        bom_rows + non_rows,
        key=lambda row: (
            row["type"],
            row[
                "sku"
            ].casefold(),
        ),
    )

    wb = Workbook()

    ws = wb.active
    ws.title = "SO LINE PRICES"

    ws.append(
        [
            "SKU",
            "Name",
            "Position Type",
            "Product Category",
            "Component / Purchase Cost",
            *ADDONS,
            "Pricing Add-ons Total",
            "Adjustment Rate",
            "Adjustment Amount",
            "Final Reform SO Unit Price",
            "Status",
            "Issues",
        ]
    )

    for row in all_rows:
        ws.append(
            [
                row["sku"],
                row["name"],
                row["type"],
                row["category"],
                row["cost"],
                *row["addons"],
                sum(
                    row["addons"]
                ),
                (
                    adjustment
                    if row[
                        "type"
                    ]
                    == "BOM"
                    else 0
                ),
                row[
                    "adjustment"
                ],
                row["final"],
                row["status"],
                row["issues"],
            ]
        )

    style(ws)

    widths(
        ws,
        [
            31,
            42,
            14,
            30,
            22,
            14,
            14,
            14,
            16,
            14,
            14,
            20,
            16,
            19,
            25,
            14,
            75,
        ],
    )

    for row_number in range(
        2,
        ws.max_row + 1,
    ):
        for column in (
            list(
                range(
                    5,
                    13,
                )
            )
            + [
                14,
                15,
            ]
        ):
            ws.cell(
                row_number,
                column,
            ).number_format = (
                '0.0000 [$€-x-euro2]'
            )

        ws.cell(
            row_number,
            13,
        ).number_format = (
            "0.0%"
        )

    write_component_cost_breakdown(
        wb,
        bom_rows,
    )

    ws = wb.create_sheet(
        "BOM CATEGORY BREAKDOWN"
    )

    ws.append(
        [
            "Top SKU",
            "Application Level",
            "Pricing Rule SKU",
            "Category ID",
            "Category Name",
            "Odoo Product Category",
            "Multiplier",
            *ADDONS,
            "Add-ons Total",
            "Adjustment Rate",
            "Adjusted Add-ons",
        ]
    )

    for row in details:
        rule = row["rule"]

        total = sum(
            row["addons"]
        )

        ws.append(
            [
                row["top"],
                row["level"],
                rule.sku,
                rule.category_id,
                rule.category_name,
                rule.odoo_category,
                row[
                    "multiplier"
                ],
                *row["addons"],
                total,
                adjustment,
                total
                * (
                    1
                    + adjustment
                ),
            ]
        )

    style(
        ws,
        "5B9BD5",
    )

    widths(
        ws,
        [
            31,
            19,
            31,
            13,
            24,
            31,
            12,
        ]
        + [14] * 6
        + [
            18,
            16,
            18,
        ],
    )

    ws = wb.create_sheet(
        "CATEGORY RULES"
    )

    ws.append(
        [
            "Category ID",
            "Category Name",
            "Odoo Product Category",
            "Products",
            *ADDONS,
            "Total",
            *[
                f"{name} Applied"
                for name
                in ADDONS
            ],
        ]
    )

    variants = Counter(
        (
            rule.category_id,
            rule.category_name,
            rule.odoo_category,
            *rule.addons,
        )
        for rule
        in rules.values()
    )

    for values, count in sorted(
        variants.items(),
        key=lambda item: tuple(
            str(value)
            for value
            in item[0][:3]
        ),
    ):
        (
            category_id,
            name,
            odoo,
            *addon_values,
        ) = values

        ws.append(
            [
                category_id,
                name,
                odoo,
                count,
                *addon_values,
                sum(
                    addon_values
                ),
                *[
                    (
                        "YES"
                        if value
                        else "NO"
                    )
                    for value
                    in addon_values
                ],
            ]
        )

    style(
        ws,
        "70AD47",
    )

    widths(
        ws,
        [
            13,
            25,
            32,
            12,
        ]
        + [14] * 7
        + [18] * 6,
    )

    ws = wb.create_sheet(
        "NON-BOM RULES"
    )

    ws.append(
        [
            "SKU",
            "Name",
            "Product Category",
            "Pricing Category",
            "Purchase Price",
            "Pack Preparation",
            "Storage",
            "Bag",
            "Sticker",
            "Final Unit Price",
            "Status",
            "Issues",
        ]
    )

    for row in non_rows:
        ws.append(
            [
                row["sku"],
                row["name"],
                row[
                    "category"
                ],
                row[
                    "pricing_category"
                ],
                row["cost"],
                row[
                    "preparation"
                ],
                row["addons"][1],
                row["bag"],
                row["sticker"],
                row["final"],
                row["status"],
                row["issues"],
            ]
        )

    style(
        ws,
        "8064A2",
    )

    widths(
        ws,
        [
            31,
            42,
            27,
            17,
            18,
            18,
            14,
            12,
            12,
            18,
            14,
            65,
        ],
    )

    ws = wb.create_sheet(
        "DIAGNOSTICS"
    )

    ws.append(
        [
            "Position Type",
            "SKU",
            "Status",
            "Issues",
        ]
    )

    for row in all_rows:
        if (
            row["status"]
            == "BLOCKED"
        ):
            ws.append(
                [
                    row["type"],
                    row["sku"],
                    row["status"],
                    row["issues"],
                ]
            )

    style(
        ws,
        "C00000",
    )

    widths(
        ws,
        [
            18,
            34,
            14,
            100,
        ],
    )

    ws = wb.create_sheet(
        "INFO"
    )

    for row in [
        (
            "Purpose",
            "Final Reform SO line unit price",
        ),
        (
            "Rules source",
            (
                "Furnibox Product Engine "
                "application configuration"
            ),
        ),
        (
            "BOM rule",
            (
                "Components + Assembly + Storage + Packaging + "
                "Put on pallet + Other + Markup"
            ),
        ),
        (
            "Component price priority",
            (
                "Prepared direct price; otherwise recursive BOM; "
                "otherwise BLOCKED"
            ),
        ),
        (
            "Odoo Standard Price fallback",
            "NO",
        ),
        (
            "Adjustment",
            adjustment,
        ),
        (
            "Markup meaning",
            (
                "Additive monetary amount, "
                "not percentage"
            ),
        ),
        (
            "Non-BOM rule",
            (
                "Purchase price + pack preparation + "
                "storage + bag + sticker"
            ),
        ),
        (
            "BOM products",
            len(
                bom_rows
            ),
        ),
        (
            "FPACK products",
            sum(
                row["sku"].upper().startswith("FPACK-")
                for row in bom_rows
            ),
        ),
        (
            "APACK products",
            sum(
                row["sku"].upper().startswith("APACK-")
                for row in bom_rows
            ),
        ),
        (
            "APACK BOM source",
            (
                "Synthetic copy of corresponding FPACK BOM using existing "
                "BOM import convention; APACK/HRD-A transfer enrichment deferred"
            ),
        ),
        (
            "Shelf Prepack standalone coverage",
            "DEFERRED (-PP not promoted to final pricing rows)",
        ),
        (
            "Non-BOM products",
            len(
                non_rows
            ),
        ),
        (
            "Blocked",
            sum(
                row["status"]
                == "BLOCKED"
                for row
                in all_rows
            ),
        ),
        (
            "Odoo changed",
            "NO",
        ),
    ]:
        ws.append(row)

    ws.column_dimensions[
        "A"
    ].width = 30

    ws.column_dimensions[
        "B"
    ].width = 110

    ws.sheet_view.showGridLines = (
        False
    )

    wb.calculation.fullCalcOnLoad = (
        True
    )

    wb.calculation.forceFullCalc = (
        True
    )

    wb.calculation.calcMode = (
        "auto"
    )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    wb.save(
        output_path
    )

    return (
        len(bom_rows),
        len(non_rows),
        sum(
            row["status"]
            == "BLOCKED"
            for row
            in all_rows
        ),
    )


def build_from_application_config(
    bom_path: Path,
    price_path: Path,
    config_path: Path,
    output_path: Path,
    dataset_path: Path | None = None,
):
    document = load_config(
        config_path
    )

    if (
        not document["bom_skus"]
        or not document[
            "bom_products"
        ]
    ):
        raise ValueError(
            "SO kainodaros taisyklės dar "
            "nesukonfigūruotos aplikacijoje."
        )

    adjustment = float(
        document[
            "adjustment_rate"
        ]
    )

    prices = load_prices(
        price_path
    )

    rules = rules_from_config(
        document
    )

    component_cost_only_tops = set()
    authoritative_rule_tops = set()
    if dataset_path is not None:
        target_dataset, _ = load_target_dataset_graph(
            dataset_path,
            bom_path,
        )
        rules = inherit_generated_apack_rules(
            rules,
            target_dataset,
        )
        rules = inherit_unambiguous_analog_rules(
            rules,
            target_dataset,
        )
        rules, authoritative_rule_tops = apply_target_business_category_rules(
            rules,
            target_dataset,
            document,
        )
        component_cost_only_tops = (
            component_cost_only_manufacture_products(
                target_dataset
            )
        )

    boms, graph = load_reform_boms(
        bom_path,
        document[
            "bom_products"
        ],
        rules=rules,
        dataset_path=dataset_path,
    )

    bom_rows, details = (
        calculate_boms(
            boms,
            prices,
            rules,
            adjustment=adjustment,
            graph=graph,
            component_cost_only_tops=(
                component_cost_only_tops
            ),
            authoritative_rule_tops=(
                authoritative_rule_tops
            ),
        )
    )

    non_rows = (
        calculate_non_bom(
            exclude_bom_products_from_non_bom(
                non_bom_from_config(
                    document
                ),
                graph,
            ),
            prices,
        )
    )

    return write_price_workbook(
        bom_rows,
        non_rows,
        details,
        rules,
        adjustment,
        output_path,
    )


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Generuoti galutines Reform SO kainas."
        )
    )

    parser.add_argument(
        "--bom-input",
        required=True,
        type=Path,
        help=(
            "Aktualus Reform BOM .xlsx"
        ),
    )

    parser.add_argument(
        "--dataset",
        type=Path,
        help="Pilnas iš to paties Reform failo sugeneruotas Target Dataset.",
    )

    parser.add_argument(
        "--price-input",
        type=Path,
        default=(
            Path(
                "output/production"
            )
            / PRICE_FILE
        ),
    )

    parser.add_argument(
        "--rules",
        type=Path,
        default=Path(
            "web_state/shared_data/"
            "so_pricing_rules.json"
        ),
    )

    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path(
            "output/production"
        ),
    )

    args = parser.parse_args()

    for path in (
        args.bom_input,
        args.price_input,
        args.rules,
    ):
        if not path.exists():
            raise FileNotFoundError(
                "Nerastas šaltinio failas: "
                f"{path.resolve()}"
            )

    if args.dataset is not None and not args.dataset.exists():
        raise FileNotFoundError(
            "Nerastas Target Dataset: "
            f"{args.dataset.resolve()}"
        )

    output = (
        args.output_dir
        / OUTPUT_FILE
    )

    (
        bom,
        non_bom,
        blocked,
    ) = build_from_application_config(
        args.bom_input,
        args.price_input,
        args.rules,
        output,
        dataset_path=args.dataset,
    )

    print(
        "GALUTINĖS REFORM SO EILUČIŲ "
        "KAINOS APSKAIČIUOTOS"
    )

    print(
        "Failas:",
        output,
    )

    print(
        "BOM pozicijos:",
        bom,
    )

    print(
        "Ne BOM pozicijos:",
        non_bom,
    )

    print(
        "BLOCKED:",
        blocked,
    )

    print(
        "Odoo duomenys nepakeisti."
    )


if __name__ == "__main__":
    main()
