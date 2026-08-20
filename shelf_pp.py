"""Furnibox Shelf PP transformacijos taisyklės.

Reform galutinė lentyna turi Shelf Part ir likusius komponentus. Furnibox
struktūroje Shelf Part pakeičiamas atskiru Manufacture ``-PP`` produktu, kurio
BOM sudaro ta pati detalė bei iš Production analogo perimta pakuotė ir
lipdukas. Neaiškūs analogai niekada nėra spėjami.
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook


def canon(value: Any) -> str:
    return str(value or "").strip().upper()


class ShelfPpError(RuntimeError):
    """Shelf PP negalima saugiai sugeneruoti iš pateiktų etalonų."""


@dataclass(frozen=True)
class ShelfPpTemplate:
    source_pp_sku: str
    source_part_sku: str
    extra_components: tuple[tuple[str, float], ...]


# Naujos Reform geometrijos, kurioms Production dar neturi tiesioginio -PP.
# Kiekvienas etalonas parinktas iš tos pačios rinkos ir patikrintos pakuotės
# klasės. Sąrašas sąmoningai baigtinis: naujas neatpažintas dydis turi sustabdyti
# generavimą, o ne tyliai paveldėti galimai netinkamą pakuotę.
SHELF_PP_PROFILE_FALLBACKS = {
    "EU-SREW-SHELF-CORNER-R_LEFT-963X564-{COLOR}":
        "EU-SREW-SHELF-CORNER-R_LEFT-1238X564-{COLOR}",
    "EU-SREW-SHELF-CORNER-R_RIGHT-963X564-{COLOR}":
        "EU-SREW-SHELF-CORNER-R_RIGHT-1238X564-{COLOR}",
    "EU-SREW-SHELF-CORNER-RW_LEFT-963X340-{COLOR}":
        "EU-SREW-SHELF-FIX-1163X340-{COLOR}",
    "EU-SREW-SHELF-CORNER-RW_RIGHT-963X340-{COLOR}":
        "EU-SREW-SHELF-FIX-1163X340-{COLOR}",
    "US-SREW-SHELF-163X339-{COLOR}":
        "US-SREW-SHELF-268X339-{COLOR}",
    "US-SREW-SHELF-FIX-878X339-{COLOR}":
        "US-SREW-SHELF-FIX-726X339-{COLOR}",
    "US-SREW-SHELF-FIX-878X574-{COLOR}":
        "US-SREW-SHELF-FIX-726X574-{COLOR}",
    "US-SREW-SHELF-CORNER-R_LEFT-963X574-{COLOR}":
        "US-SREW-SHELF-CORNER-R_LEFT-1157X574-{COLOR}",
    "US-SREW-SHELF-CORNER-R_RIGHT-963X574-{COLOR}":
        "US-SREW-SHELF-CORNER-R_RIGHT-1157X574-{COLOR}",
}


def shelf_pp_sku(part_sku: str) -> str:
    code = canon(part_sku)
    if not code:
        raise ShelfPpError("Shelf Part SKU yra tuščias.")
    return code if code.endswith("-PP") else f"{code}-PP"


def shelf_profile(sku: str) -> str:
    """Profilis, leidžiantis perimti pakuotę tik tarp tos pačios geometrijos."""
    code = canon(sku)
    if code.endswith("-PP"):
        code = code[:-3]
    # Reform v10 turi vieną „163 x339“ variantą su tarpu; SKU geometrijoje
    # tarpai nėra semantiniai.
    code = re.sub(r"\s+", "", code)
    # Spalva nekeičia pakuotės profilio, bet regionas, šeima ir matmenys lieka.
    return re.sub(r"-(WW|BB|NO)$", "-{COLOR}", code)


def shelf_reference_profile(sku: str) -> str:
    """Grąžina tiesioginį arba aiškiai patvirtintą Production profilį."""
    profile = shelf_profile(sku)
    return SHELF_PP_PROFILE_FALLBACKS.get(profile, profile)


def load_odoo_edges(path: Path) -> dict[str, list[dict[str, Any]]]:
    if not path.is_file():
        raise ShelfPpError(f"Nerastas Production Odoo MAP: {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if "ODOO EDGES" not in wb.sheetnames:
            raise ShelfPpError("Odoo MAP neturi lapo 'ODOO EDGES'.")
        rows = wb["ODOO EDGES"].iter_rows(values_only=True)
        headers = next(rows)
        index = {str(value or "").strip(): i for i, value in enumerate(headers)}
        required = {"Parent SKU", "Component SKU", "Quantity"}
        missing = required - set(index)
        if missing:
            raise ShelfPpError(
                "Odoo MAP trūksta stulpelių: " + ", ".join(sorted(missing))
            )
        result: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            parent = canon(row[index["Parent SKU"]])
            component = canon(row[index["Component SKU"]])
            if not parent or not component:
                continue
            try:
                quantity = float(row[index["Quantity"]] or 0)
            except (TypeError, ValueError) as exc:
                raise ShelfPpError(
                    f"{parent} / {component}: netinkamas kiekis Production MAP."
                ) from exc
            result[parent].append({"component": component, "quantity": quantity})
        return dict(result)
    finally:
        wb.close()


def build_shelf_pp_templates(
    odoo_lines: dict[str, list[dict[str, Any]]],
) -> dict[str, ShelfPpTemplate]:
    """Sukuria tik struktūriškai įrodytus esamų Shelf PP etalonus."""
    templates: dict[str, ShelfPpTemplate] = {}
    for raw_parent, raw_lines in sorted(odoo_lines.items()):
        parent = canon(raw_parent)
        if not parent.endswith("-PP"):
            continue
        expected_part = parent[:-3]
        normalized = [
            (canon(line.get("component")), float(line.get("quantity") or 0))
            for line in raw_lines
            if canon(line.get("component"))
        ]
        matching_parts = [row for row in normalized if row[0] == expected_part]
        if len(matching_parts) != 1:
            # Kiti -PP produktai (pvz., ne lentynos) nėra Shelf PP etalonai.
            continue
        extras = tuple(row for row in normalized if row[0] != expected_part)
        if len(extras) < 2 or any(quantity <= 0 for _, quantity in normalized):
            continue
        templates[parent] = ShelfPpTemplate(
            source_pp_sku=parent,
            source_part_sku=expected_part,
            extra_components=extras,
        )
    return templates


def choose_shelf_pp_template(
    target_pp_sku: str,
    templates: dict[str, ShelfPpTemplate],
) -> ShelfPpTemplate:
    target = canon(target_pp_sku)
    if target in templates:
        return templates[target]
    profile = shelf_reference_profile(target)
    matches = [
        template
        for sku, template in templates.items()
        if shelf_profile(sku) == profile
    ]
    signatures = {
        template.extra_components
        for template in matches
    }
    if not matches:
        raise ShelfPpError(
            f"{target}: Production nerastas tos pačios Shelf geometrijos PP analogas."
        )
    if len(signatures) != 1:
        sources = ", ".join(sorted(item.source_pp_sku for item in matches))
        raise ShelfPpError(
            f"{target}: Production PP analogai turi skirtingas pakuotes: {sources}."
        )
    return sorted(matches, key=lambda item: item.source_pp_sku)[0]


def add_generated_shelf_pp_boms(
    parents: set[str],
    lines: dict[str, list[dict]],
    levels: dict[str, int],
    bom_types: dict[str, str],
    reform_products: dict[str, dict],
    templates: dict[str, ShelfPpTemplate],
    *,
    kit_type: str,
    manufacture_type: str,
) -> dict[str, str]:
    """Transformuoja Reform Shelf struktūras ir prideda Shelf PP BOM."""
    generated_from: dict[str, str] = {}
    products_by_sku = {
        canon(sku): product
        for sku, product in reform_products.items()
    }
    for raw_parent in sorted(list(parents)):
        parent = canon(raw_parent)
        product = products_by_sku.get(parent, {})
        if canon(product.get("category")) != "CABINET SHELF":
            continue
        # Reform jau gali turėti atskirą specialios lentynos prepack BOM.
        # Jis yra Manufacture produkto šaltinis, ne galutinė Shelf struktūra.
        if parent.endswith("-PP"):
            continue
        source_lines = lines.get(raw_parent, lines.get(parent, []))
        shelf_part_lines = []
        for line in source_lines:
            component = canon(line.get("component"))
            component_product = products_by_sku.get(component, {})
            if canon(component_product.get("part_group")) == "SHELF PART":
                shelf_part_lines.append(line)
        if not shelf_part_lines:
            # Kai Reform jau pateikia atskirą prepack komponentą, galutinė
            # lentyna yra tinkama Furnibox struktūra ir jos antrą kartą
            # neperpakuojame.
            prepacked = [
                line for line in source_lines
                if canon(line.get("component")).endswith("-PP")
            ]
            if len(prepacked) == 1:
                bom_types[parent] = kit_type
                continue
        if len(shelf_part_lines) != 1:
            raise ShelfPpError(
                f"{parent}: tikėtasi vieno SHELF PART komponento, "
                f"rasta {len(shelf_part_lines)}."
            )
        part_line = shelf_part_lines[0]
        part_sku = canon(part_line.get("component"))
        pp_sku = shelf_pp_sku(part_sku)
        template = choose_shelf_pp_template(pp_sku, templates)
        if pp_sku in parents and pp_sku not in generated_from:
            raise ShelfPpError(f"{pp_sku}: PP BOM jau yra Reform grafe.")

        transformed = []
        for line in source_lines:
            if line is part_line:
                transformed.append({**line, "component": pp_sku})
            else:
                transformed.append(dict(line))
        lines[parent] = transformed
        bom_types[parent] = kit_type

        pp_lines = [{**part_line, "component": part_sku}]
        pp_lines.extend(
            {"component": component, "quantity": quantity}
            for component, quantity in template.extra_components
        )
        parents.add(pp_sku)
        lines[pp_sku] = pp_lines
        levels[pp_sku] = 2
        bom_types[pp_sku] = manufacture_type
        generated_from[pp_sku] = part_sku
    return generated_from


def choose_shelf_pp_operation_template(
    target_pp_sku: str,
    operation_templates: dict[int, dict],
) -> dict:
    """Parenka vieną to paties profilio PP pakavimo operaciją."""
    target = canon(target_pp_sku)
    candidates = [
        template
        for template in operation_templates.values()
        if canon(template.get("sku")).endswith("-PP")
        and shelf_profile(template.get("sku")) == shelf_reference_profile(target)
    ]
    valid = []
    for template in candidates:
        operations = list(template.get("operations") or [])
        packing = [op for op in operations if "PAKAV" in canon(op.get("name"))]
        if len(operations) == 1 and len(packing) == 1:
            valid.append(template)

    # Esamas Production SKU yra stipresnis etalonas už tos pačios geometrijos
    # kitų spalvų analogus. Spalvų parametrų vienodumo reikalaujame tik tada,
    # kai tikslinio PP Production dar nėra ir operaciją tenka paveldėti.
    exact = [
        template for template in valid
        if canon(template.get("sku")) == target
    ]
    if exact:
        exact_signatures = {
            tuple(sorted((key, str(value)) for key, value in template["operations"][0].items()))
            for template in exact
        }
        if len(exact_signatures) != 1:
            raise ShelfPpError(
                f"{target}: tikslūs Production PP operacijų etalonai nesutampa."
            )
        return exact[0]

    signatures = {
        tuple(sorted((key, str(value)) for key, value in template["operations"][0].items()))
        for template in valid
    }
    if not valid:
        raise ShelfPpError(
            f"{target}: nerastas PP analogas su tiksliai viena pakavimo operacija."
        )
    if len(signatures) != 1:
        raise ShelfPpError(
            f"{target}: analogiškų PP pakavimo operacijų parametrai nesutampa."
        )
    return sorted(valid, key=lambda item: canon(item.get("sku")))[0]
