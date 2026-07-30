from __future__ import annotations

from dataclasses import dataclass, field
import json
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from bom_release.generator import (
    IMPORT_HEADERS,
    KIT,
    MANUFACTURE,
    NEW_BOM_SEQUENCE,
    canon,
    normalized_bom_type,
    required_operation_type_external_id,
)
from operation_contract import (
    manufacture_operations_required,
    product_category,
)


class BomReleaseImportValidationError(RuntimeError):
    """Sugeneruoti importo failai neatitinka Dataset sutarties."""


@dataclass
class ImportValidationResult:
    expected_boms: int
    actual_boms: int = 0
    files_checked: int = 0
    component_rows: int = 0
    operation_rows: int = 0
    errors: list[str] = field(default_factory=list)

    @property
    def passed(self) -> bool:
        return not self.errors and self.actual_boms == self.expected_boms


def _text(value: Any) -> str:
    return str(value or "").strip()


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _same_number(left: Any, right: Any) -> bool:
    left_number = _number(left)
    right_number = _number(right)
    if left_number is None or right_number is None:
        return left_number is right_number
    return abs(left_number - right_number) < 1e-9


def _expected_group(record: dict[str, Any]) -> tuple[int, str]:
    return (
        int(record.get("level") or 0),
        normalized_bom_type(record.get("bom_type")),
    )


def _expected_filename(
    release_id: str,
    level: int,
    bom_type: str,
) -> str:
    type_slug = "KIT" if bom_type == KIT else "Manufacture"
    return f"BOM_Release_{release_id}_lv{level}_{type_slug}.xlsx"


def _row_has_formula(values: Iterable[Any]) -> bool:
    return any(
        isinstance(value, str) and value.startswith("=")
        for value in values
    )


def _workbook_records(
    path: Path,
    result: ImportValidationResult,
) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    try:
        if workbook.sheetnames != ["BOM import"]:
            result.errors.append(
                f"{path.name}: lapai turi būti tik ['BOM import'], "
                f"gauta {workbook.sheetnames}"
            )
            if "BOM import" not in workbook.sheetnames:
                return []
        sheet = workbook["BOM import"]
        rows = sheet.iter_rows(values_only=True)
        header = list(next(rows, ()))
        if header != IMPORT_HEADERS:
            result.errors.append(
                f"{path.name}: neteisingi importo stulpeliai"
            )
            return []

        records: list[dict[str, Any]] = []
        current: dict[str, Any] | None = None
        for row_number, row in enumerate(rows, start=2):
            values = list(row[: len(IMPORT_HEADERS)])
            if not any(value not in (None, "") for value in values):
                continue
            if _row_has_formula(values):
                result.errors.append(
                    f"{path.name}:{row_number}: formulės neleidžiamos"
                )
            parent_sku = canon(values[0])
            if parent_sku:
                if current:
                    records.append(current)
                current = {
                    "row": row_number,
                    "sku": parent_sku,
                    "parent_external_id": _text(values[1]),
                    "qty": values[2],
                    "bom_type": _text(values[6]),
                    "sequence": values[7],
                    "operation_type_external_id": _text(values[8]),
                    "reference": _text(values[9]),
                    "mo_autodone_by_wo": values[10],
                    "auto_plan": values[11],
                    "components": [],
                    "operations": [],
                }
            elif current is None:
                result.errors.append(
                    f"{path.name}:{row_number}: tęstinė eilutė be parent"
                )
                continue
            elif any(
                values[index] not in (None, "")
                for index in (1, 2, 6, 7, 8, 9, 10, 11)
            ):
                result.errors.append(
                    f"{path.name}:{row_number}: parent laukai kartojami "
                    "tęstinėje eilutėje"
                )

            if current is None:
                continue
            component_sku = canon(values[3])
            component_external_id = _text(values[4])
            component_qty = values[5]
            if component_sku or component_external_id or component_qty not in (None, ""):
                current["components"].append(
                    (component_sku, component_external_id, component_qty)
                )
                result.component_rows += 1

            operation_values = values[12:17]
            if any(value not in (None, "") for value in operation_values):
                current["operations"].append(tuple(operation_values))
                result.operation_rows += 1
        if current:
            records.append(current)
        return records
    finally:
        workbook.close()


def _compare_record(
    *,
    actual: dict[str, Any],
    expected: dict[str, Any],
    release_reference: str,
    source: str,
    result: ImportValidationResult,
) -> None:
    sku = canon(expected.get("sku"))

    def error(message: str) -> None:
        result.errors.append(f"{source}: {sku}: {message}")

    if not actual["parent_external_id"]:
        error("tuščias parent product.template External ID")
    if not _same_number(actual["qty"], 1):
        error(f"BOM qty turi būti 1, gauta {actual['qty']!r}")
    expected_type = normalized_bom_type(expected.get("bom_type"))
    if actual["bom_type"] != expected_type:
        error(
            f"BOM tipas {actual['bom_type']!r}, tikėtasi {expected_type!r}"
        )
    if not _same_number(actual["sequence"], NEW_BOM_SEQUENCE):
        error(
            f"naujo BOM sequence {actual['sequence']!r}, "
            f"tikėtasi {NEW_BOM_SEQUENCE}"
        )
    expected_operation_type = required_operation_type_external_id(expected)
    if actual["operation_type_external_id"] != expected_operation_type:
        error(
            "Operation Type/External ID "
            f"{actual['operation_type_external_id']!r}, "
            f"tikėtasi {expected_operation_type!r}"
        )
    if actual["reference"] != release_reference:
        error(
            f"Reference {actual['reference']!r}, "
            f"tikėtasi {release_reference!r}"
        )
    manufacture = expected_type == MANUFACTURE
    if bool(actual["mo_autodone_by_wo"]) != manufacture:
        error("neteisingas mo_autodone_by_wo")
    if bool(actual["auto_plan"]) != manufacture:
        error("neteisingas auto_plan")

    expected_components = list(expected.get("components") or [])
    actual_components = actual["components"]
    if len(actual_components) != len(expected_components):
        error(
            f"komponentų eilučių {len(actual_components)}, "
            f"tikėtasi {len(expected_components)}"
        )
    for index, (expected_component, actual_component) in enumerate(
        zip(expected_components, actual_components),
        start=1,
    ):
        actual_sku, external_id, quantity = actual_component
        expected_sku = canon(expected_component.get("sku"))
        if actual_sku != expected_sku:
            error(
                f"komponentas #{index} {actual_sku!r}, "
                f"tikėtasi {expected_sku!r}"
            )
        if not external_id:
            error(f"komponentas #{index} neturi product.product External ID")
        if not _same_number(quantity, expected_component.get("quantity")):
            error(
                f"komponento #{index} kiekis {quantity!r}, tikėtasi "
                f"{expected_component.get('quantity')!r}"
            )

    expected_operations = list(expected.get("operations") or [])
    actual_operations = actual["operations"]
    if len(actual_operations) != len(expected_operations):
        error(
            f"operacijų eilučių {len(actual_operations)}, "
            f"tikėtasi {len(expected_operations)}"
        )
    for index, (expected_operation, actual_operation) in enumerate(
        zip(expected_operations, actual_operations),
        start=1,
    ):
        expected_values = (
            _text(expected_operation.get("name")),
            _text(expected_operation.get("workcenter")),
            _text(expected_operation.get("time_mode")),
            expected_operation.get("time_minutes"),
            expected_operation.get("sequence"),
        )
        for label, actual_value, expected_value in zip(
            ("name", "workcenter", "time_mode", "time", "sequence"),
            actual_operation,
            expected_values,
        ):
            equal = (
                _same_number(actual_value, expected_value)
                if label in {"time", "sequence"}
                else _text(actual_value) == _text(expected_value)
            )
            if not equal:
                error(
                    f"operacija #{index} {label} {actual_value!r}, "
                    f"tikėtasi {expected_value!r}"
                )


def validate_release_imports(
    *,
    dataset: dict[str, Any],
    release_id: str,
    release_reference: str,
    import_dir: Path,
) -> ImportValidationResult:
    expected_by_sku = {
        canon(record.get("sku")): record
        for record in dataset.get("products", [])
        if canon(record.get("sku"))
    }
    result = ImportValidationResult(expected_boms=len(expected_by_sku))
    manifest_path = import_dir / f"BOM_Release_{release_id}_manifest.json"
    if not manifest_path.is_file():
        result.errors.append(f"trūksta manifesto: {manifest_path.name}")
    else:
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            if manifest.get("release_id") != release_id:
                result.errors.append("manifesto release_id nesutampa")
            if manifest.get("release_reference") != release_reference:
                result.errors.append("manifesto release_reference nesutampa")
            if manifest.get("new_bom_sequence") != NEW_BOM_SEQUENCE:
                result.errors.append(
                    f"manifesto new_bom_sequence turi būti "
                    f"{NEW_BOM_SEQUENCE}"
                )
            if manifest.get("activation_required") is not True:
                result.errors.append(
                    "manifestas turi reikalauti atskiro aktyvavimo"
                )
        except (OSError, ValueError, TypeError, json.JSONDecodeError) as exc:
            result.errors.append(f"nepavyko perskaityti manifesto: {exc}")

    for sku, record in sorted(expected_by_sku.items()):
        bom_type = normalized_bom_type(record.get("bom_type"))
        operations = list(record.get("operations") or [])
        if (
            bom_type == MANUFACTURE
            and not operations
            and manufacture_operations_required(
                sku=sku,
                category=product_category(record),
            )
        ):
            result.errors.append(
                f"Dataset: {sku}: Manufacture BOM neturi operacijų"
            )
        if bom_type == KIT and operations:
            result.errors.append(
                f"Dataset: {sku}: KIT BOM turi operacijų"
            )
    expected_groups: dict[tuple[int, str], set[str]] = {}
    for sku, record in expected_by_sku.items():
        expected_groups.setdefault(_expected_group(record), set()).add(sku)

    expected_paths = {
        group: import_dir / _expected_filename(release_id, *group)
        for group in expected_groups
    }
    actual_xlsx = set(import_dir.glob("*.xlsx"))
    missing_files = [
        path.name for path in expected_paths.values() if not path.is_file()
    ]
    extra_files = sorted(
        path.name
        for path in (
            actual_xlsx
            - set(expected_paths.values())
        )
    )
    if missing_files:
        result.errors.append(
            "trūksta failų: " + ", ".join(sorted(missing_files))
        )
    if extra_files:
        result.errors.append(
            "netikėti XLSX failai: " + ", ".join(extra_files)
        )

    seen: dict[str, str] = {}
    for group, path in expected_paths.items():
        if not path.is_file():
            continue
        result.files_checked += 1
        records = _workbook_records(path, result)
        actual_group_skus: set[str] = set()
        for actual in records:
            sku = actual["sku"]
            if sku in seen:
                result.errors.append(
                    f"{path.name}: {sku}: BOM kartojasi; "
                    f"pirmas failas {seen[sku]}"
                )
                continue
            seen[sku] = path.name
            actual_group_skus.add(sku)
            expected = expected_by_sku.get(sku)
            if expected is None:
                result.errors.append(
                    f"{path.name}: {sku}: SKU nėra Dataset"
                )
                continue
            if _expected_group(expected) != group:
                result.errors.append(
                    f"{path.name}: {sku}: įrašas neteisingame faile"
                )
            _compare_record(
                actual=actual,
                expected=expected,
                release_reference=release_reference,
                source=path.name,
                result=result,
            )
        missing_group = expected_groups[group] - actual_group_skus
        if missing_group:
            preview = ", ".join(sorted(missing_group)[:10])
            suffix = (
                f" ir dar {len(missing_group) - 10}"
                if len(missing_group) > 10
                else ""
            )
            result.errors.append(
                f"{path.name}: trūksta BOM: {preview}{suffix}"
            )

    result.actual_boms = len(seen)
    missing_all = set(expected_by_sku) - set(seen)
    if missing_all and not missing_files:
        result.errors.append(
            f"bendrai trūksta {len(missing_all)} Dataset BOM"
        )
    if len(seen) != len(expected_by_sku):
        result.errors.append(
            f"BOM aprėptis {len(seen)}/{len(expected_by_sku)}"
        )
    return result


def require_valid_release_imports(**kwargs: Any) -> ImportValidationResult:
    result = validate_release_imports(**kwargs)
    if not result.passed:
        preview = "; ".join(result.errors[:10])
        suffix = (
            f"; dar {len(result.errors) - 10}"
            if len(result.errors) > 10
            else ""
        )
        raise BomReleaseImportValidationError(preview + suffix)
    return result
