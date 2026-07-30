from __future__ import annotations

from dataclasses import dataclass
import json
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bom_release.models import BomReleasePlan, ReleaseAction
from odoo_client import OdooClient


IMPORT_HEADERS = [
    "Product/Internal Reference",
    "Product/External ID",
    "qty",
    "BoM Lines/Component/Internal Reference",
    "BoM Lines/Component/External ID",
    "product_qty",
    "BoM Type",
    "sequence",
    "Operation Type/External ID",
    "Reference",
    "mo_autodone_by_wo",
    "auto_plan",
    "operation_ids/name",
    "operation_ids/workcenter_id",
    "operation_ids/time_mode",
    "operation_ids/time_cycle_manual",
    "operation_ids/sequence",
]
NEW_BOM_SEQUENCE = 10
MANUFACTURE = "Manufacture this product"
KIT = "KIT"
HRD_OPERATION_TYPE_EXTERNAL_ID = (
    "__export__.stock_picking_type_19_2883a6b3"
)
APACK_OPERATION_TYPE_EXTERNAL_ID = (
    "__export__.stock_picking_type_21_0e009783"
)


class BomReleaseGenerationError(RuntimeError):
    """Release paketo negalima saugiai sugeneruoti."""


@dataclass(frozen=True)
class GeneratedReleaseFile:
    path: Path
    level: int
    bom_type: str
    bom_count: int


def canon(value: Any) -> str:
    return str(value or "").strip().upper()


def normalized_bom_type(value: Any) -> str:
    normalized = canon(value)
    if normalized in {"KIT", "PHANTOM"}:
        return KIT
    if normalized in {
        "MANUFACTURE",
        "MANUFACTURE THIS PRODUCT",
        "NORMAL",
    }:
        return MANUFACTURE
    raise BomReleaseGenerationError(
        f"Nepalaikomas Dataset BOM tipas: {value!r}"
    )


def required_operation_type_external_id(record: dict[str, Any]) -> str:
    if normalized_bom_type(record.get("bom_type")) != MANUFACTURE:
        return ""
    sku = canon(record.get("sku"))
    category = canon(
        record.get("product_type") or record.get("reform_category")
    )
    if category == "CABINET HARDWARE":
        return HRD_OPERATION_TYPE_EXTERNAL_ID
    if sku.startswith("APACK-"):
        return APACK_OPERATION_TYPE_EXTERNAL_ID
    return ""


def validate_dataset_acceptance(dataset: dict[str, Any]) -> None:
    # Importuojama čia, kad generatoriaus branduolys nepriklausytų nuo CLI.
    from pre_activation_acceptance import Acceptance

    acceptance = Acceptance(dataset)
    acceptance.run(acceptance.dataset_structures())
    errors = [
        issue
        for issue in acceptance.issues
        if issue.severity == "ERROR"
    ]
    if errors:
        first = errors[0]
        raise BomReleaseGenerationError(
            "Dataset acceptance yra FAIL: "
            f"{len(errors)} klaidų; pirma: "
            f"{first.test_code} {first.parent_sku} {first.message}"
        )


def _validate_contract(
    dataset: dict[str, Any],
    plan: BomReleasePlan,
) -> list[dict[str, Any]]:
    if not plan.can_generate:
        raise BomReleaseGenerationError(
            f"Release planas turi {plan.blocked_count} BLOCKED BOM."
        )
    if str(dataset.get("dataset_id") or "") != plan.dataset_id:
        raise BomReleaseGenerationError(
            "Release planas sukurtas ne iš pateikto Dataset."
        )
    if not plan.release_reference.strip():
        raise BomReleaseGenerationError(
            "Release reference negali būti tuščias."
        )

    by_sku = {
        canon(row.get("sku")): row
        for row in dataset.get("products", [])
        if canon(row.get("sku"))
    }
    create_skus = {
        item.parent_sku
        for item in plan.items
        if item.action == ReleaseAction.CREATE
    }
    missing = sorted(create_skus - set(by_sku))
    if missing:
        raise BomReleaseGenerationError(
            "Release plane yra Dataset nerastų SKU: "
            + ", ".join(missing)
        )
    return [by_sku[sku] for sku in sorted(create_skus)]


def _workcenters(client: OdooClient) -> tuple[set[str], set[str]]:
    rows = client.search_read_all(
        "mrp.workcenter",
        [],
        ["id", "name", "active"],
        context={"active_test": False},
    )
    counts: dict[str, int] = {}
    for row in rows:
        if not row.get("active", True):
            continue
        name = str(row.get("name") or "").strip()
        if name:
            counts[name] = counts.get(name, 0) + 1
    return (
        {name for name, count in counts.items() if count == 1},
        {name for name, count in counts.items() if count > 1},
    )


def _products_by_sku(
    client: OdooClient,
    wanted_skus: set[str],
) -> tuple[dict[str, dict[str, Any]], set[str]]:
    rows = client.search_read_all(
        "product.product",
        [["default_code", "!=", False]],
        ["id", "default_code", "product_tmpl_id", "active"],
        context={"active_test": False},
    )
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        sku = canon(row.get("default_code"))
        if sku in wanted_skus:
            grouped.setdefault(sku, []).append(row)

    products: dict[str, dict[str, Any]] = {}
    duplicates: set[str] = set()
    for sku, matches in grouped.items():
        product_ids = {int(row["id"]) for row in matches}
        template_ids = {
            int(row["product_tmpl_id"][0])
            for row in matches
            if row.get("product_tmpl_id")
        }
        if len(product_ids) != 1 or len(template_ids) != 1:
            duplicates.add(sku)
            continue
        row = matches[0]
        products[sku] = {
            "display_sku": str(row["default_code"]).strip(),
            "product_id": next(iter(product_ids)),
            "template_id": next(iter(template_ids)),
        }

    def xmlids(model: str, ids: set[int]) -> dict[int, str]:
        if not ids:
            return {}
        records = client.search_read_all(
            "ir.model.data",
            [["model", "=", model], ["res_id", "in", sorted(ids)]],
            ["module", "name", "res_id"],
        )
        grouped_ids: dict[int, list[str]] = {}
        for record in records:
            grouped_ids.setdefault(int(record["res_id"]), []).append(
                f"{record['module']}.{record['name']}"
            )
        return {
            record_id: sorted(
                values,
                key=lambda value: (
                    value.startswith("__export__."),
                    value,
                ),
            )[0]
            for record_id, values in grouped_ids.items()
        }

    product_xmlids = xmlids(
        "product.product",
        {value["product_id"] for value in products.values()},
    )
    template_xmlids = xmlids(
        "product.template",
        {value["template_id"] for value in products.values()},
    )
    for value in products.values():
        value["product_xmlid"] = product_xmlids.get(
            value["product_id"],
            "",
        )
        value["template_xmlid"] = template_xmlids.get(
            value["template_id"],
            "",
        )
    return products, duplicates


def _validate_odoo_mappings(
    records: Iterable[dict[str, Any]],
    products: dict[str, dict[str, Any]],
    duplicates: set[str],
    workcenters: set[str],
    duplicate_workcenters: set[str],
) -> None:
    errors: list[str] = []
    for record in records:
        parent = canon(record.get("sku"))
        required = {parent} | {
            canon(component.get("sku"))
            for component in record.get("components", [])
            if canon(component.get("sku"))
        }
        for sku in sorted(required):
            if sku in duplicates:
                errors.append(f"{parent}: neunikalus produktas {sku}")
            elif sku not in products:
                errors.append(f"{parent}: Odoo nerastas produktas {sku}")
            elif sku == parent and not products[sku].get("template_xmlid"):
                errors.append(
                    f"{parent}: parent neturi product.template External ID"
                )
            elif sku != parent and not products[sku].get("product_xmlid"):
                errors.append(
                    f"{parent}: komponentas {sku} neturi "
                    "product.product External ID"
                )
        for operation in record.get("operations", []):
            name = str(operation.get("workcenter") or "").strip()
            if name in duplicate_workcenters:
                errors.append(f"{parent}: neunikalus darbo centras {name}")
            elif name not in workcenters:
                errors.append(f"{parent}: nerastas darbo centras {name}")
    if errors:
        preview = "; ".join(errors[:10])
        suffix = f"; dar {len(errors) - 10}" if len(errors) > 10 else ""
        raise BomReleaseGenerationError(preview + suffix)


def _rows(
    record: dict[str, Any],
    products: dict[str, dict[str, Any]],
    release_reference: str,
) -> Iterable[list[Any]]:
    parent_sku = canon(record.get("sku"))
    parent = products[parent_sku]
    components = list(record.get("components") or [])
    operations = list(record.get("operations") or [])
    bom_type = normalized_bom_type(record.get("bom_type"))
    row_count = max(len(components), len(operations), 1)

    for index in range(row_count):
        first = index == 0
        component_row = (
            components[index] if index < len(components) else None
        )
        operation = operations[index] if index < len(operations) else None
        component = (
            products[canon(component_row.get("sku"))]
            if component_row
            else None
        )
        yield [
            parent["display_sku"] if first else None,
            parent["template_xmlid"] if first else None,
            1 if first else None,
            component["display_sku"] if component else None,
            component["product_xmlid"] if component else None,
            component_row.get("quantity") if component_row else None,
            bom_type if first else None,
            NEW_BOM_SEQUENCE if first else None,
            required_operation_type_external_id(record) if first else None,
            release_reference if first else None,
            True if first and bom_type == MANUFACTURE else None,
            True if first and bom_type == MANUFACTURE else None,
            operation.get("name") if operation else None,
            operation.get("workcenter") if operation else None,
            operation.get("time_mode") if operation else None,
            (
                operation.get("time_minutes")
                if operation
                else None
            ),
            operation.get("sequence") if operation else None,
        ]


def _style(ws: Any) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index in range(1, ws.max_column + 1):
        values = [
            len(str(ws.cell(row, index).value or ""))
            for row in range(1, min(ws.max_row, 300) + 1)
        ]
        ws.column_dimensions[get_column_letter(index)].width = min(
            max(values, default=12) + 2,
            55,
        )


def generate_release_files(
    *,
    dataset: dict[str, Any],
    plan: BomReleasePlan,
    client: OdooClient,
    output_dir: Path,
) -> list[GeneratedReleaseFile]:
    """Tik skaito Odoo ir sukuria neperdengiančius importo failus."""
    validate_dataset_acceptance(dataset)
    records = _validate_contract(dataset, plan)
    wanted = {
        canon(record.get("sku"))
        for record in records
    } | {
        canon(component.get("sku"))
        for record in records
        for component in record.get("components", [])
        if canon(component.get("sku"))
    }
    products, duplicates = _products_by_sku(client, wanted)
    workcenters, duplicate_workcenters = _workcenters(client)
    _validate_odoo_mappings(
        records,
        products,
        duplicates,
        workcenters,
        duplicate_workcenters,
    )

    groups: dict[tuple[int, str], list[dict[str, Any]]] = {}
    for record in records:
        key = (
            int(record.get("level") or 0),
            normalized_bom_type(record.get("bom_type")),
        )
        groups.setdefault(key, []).append(record)

    output_dir.mkdir(parents=True, exist_ok=True)
    generated: list[GeneratedReleaseFile] = []
    legacy_old_bom_path = output_dir / (
        f"BOM_Release_{plan.release_id}_00_Old_BOM_Sequence_10.xlsx"
    )
    if legacy_old_bom_path.is_file():
        legacy_old_bom_path.unlink()

    manifest_path = output_dir / (
        f"BOM_Release_{plan.release_id}_manifest.json"
    )
    manifest_path.write_text(
        json.dumps(
            {
                "release_id": plan.release_id,
                "release_reference": plan.release_reference,
                "dataset_id": plan.dataset_id,
                "new_bom_sequence": NEW_BOM_SEQUENCE,
                "activation_required": True,
                "operation_type_external_id_rules": {
                    "CABINET HARDWARE": HRD_OPERATION_TYPE_EXTERNAL_ID,
                    "APACK": APACK_OPERATION_TYPE_EXTERNAL_ID,
                },
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    for (level, bom_type), group in sorted(
        groups.items(),
        key=lambda item: (-item[0][0], item[0][1]),
    ):
        type_slug = "KIT" if bom_type == KIT else "Manufacture"
        path = output_dir / (
            f"BOM_Release_{plan.release_id}_lv{level}_{type_slug}.xlsx"
        )
        wb = Workbook()
        ws = wb.active
        ws.title = "BOM import"
        ws.append(IMPORT_HEADERS)
        for record in sorted(group, key=lambda row: canon(row.get("sku"))):
            for row in _rows(
                record,
                products,
                plan.release_reference,
            ):
                ws.append(row)
        _style(ws)
        wb.save(path)
        generated.append(
            GeneratedReleaseFile(
                path=path,
                level=level,
                bom_type=bom_type,
                bom_count=len(group),
            )
        )
    return generated
