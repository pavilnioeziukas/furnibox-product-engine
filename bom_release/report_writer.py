from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bom_release.models import BomReleasePlan


HEADER_FILL = PatternFill(
    "solid",
    fgColor="1F4E78",
)
BLOCKED_FILL = PatternFill(
    "solid",
    fgColor="F4CCCC",
)
READY_FILL = PatternFill(
    "solid",
    fgColor="D9EAD3",
)
EXISTS_FILL = PatternFill(
    "solid",
    fgColor="FFF2CC",
)


def _style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(
            color="FFFFFF",
            bold=True,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for column_index in range(
        1,
        ws.max_column + 1,
    ):
        values = [
            len(
                str(
                    ws.cell(
                        row=row_index,
                        column=column_index,
                    ).value
                    or ""
                )
            )
            for row_index in range(
                1,
                min(
                    ws.max_row,
                    500,
                )
                + 1,
            )
        ]

        ws.column_dimensions[
            get_column_letter(
                column_index
            )
        ].width = min(
            max(
                values,
                default=12,
            )
            + 2,
            60,
        )


def write_release_plan(
    plan: BomReleasePlan,
    output_path: Path,
) -> Path:
    output_path = output_path.resolve()
    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "SUMMARY"
    summary_ws.append(
        ["Metric", "Value"]
    )
    summary_ws.append(
        ["Release ID", plan.release_id]
    )
    summary_ws.append(
        [
            "Release Reference",
            plan.release_reference,
        ]
    )
    summary_ws.append(
        ["Environment", plan.environment]
    )
    summary_ws.append(
        ["Dataset ID", plan.dataset_id]
    )
    summary_ws.append(
        [
            "Dataset Batch",
            plan.dataset_batch_reference,
        ]
    )
    summary_ws.append(
        [
            "Dataset Path",
            plan.dataset_path,
        ]
    )
    summary_ws.append(
        [
            "Created At UTC",
            plan.created_at_utc,
        ]
    )
    summary_ws.append(
        ["Total BOM", plan.total_count]
    )
    summary_ws.append(
        ["READY", plan.ready_count]
    )
    summary_ws.append(
        [
            "ALREADY EXISTS",
            plan.already_exists_count,
        ]
    )
    summary_ws.append(
        ["BLOCKED", plan.blocked_count]
    )
    summary_ws.append(
        [
            "Missing Parent Product",
            plan.missing_parent_count,
        ]
    )
    summary_ws.append(
        [
            "Parents with Missing Components",
            plan.missing_component_parent_count,
        ]
    )
    summary_ws.append(
        [
            "Multiple Sequence 0",
            plan.multiple_sequence_zero_count,
        ]
    )
    summary_ws.append(
        [
            "Can Generate",
            "YES"
            if plan.can_generate
            else "NO",
        ]
    )

    plan_ws = wb.create_sheet(
        "RELEASE_PLAN"
    )
    plan_ws.append(
        [
            "Status",
            "Action",
            "Parent SKU",
            "BOM Type",
            "Components",
            "Operations",
            "Product Exists",
            "Product ID",
            "Product Template ID",
            "Sequence 0 BOM Count",
            "Active BOM ID",
            "Active Reference",
            "Active Sequence",
            "Active BOM Type",
            "Release Exists",
            "Release BOM ID",
            "Release Reference",
            "Missing Component Count",
            "Missing Components",
            "Duplicate Product IDs",
            "Blocking Reasons",
            "Warnings",
        ]
    )

    for row_index, item in enumerate(
        plan.items,
        start=2,
    ):
        plan_ws.append(
            [
                item.status.value,
                item.action.value,
                item.parent_sku,
                item.bom_type,
                item.component_count,
                item.operation_count,
                "YES"
                if item.product_exists
                else "NO",
                item.product_id,
                item.product_template_id,
                item.active_bom_count,
                item.active_bom_id,
                item.active_reference,
                item.active_sequence,
                item.active_bom_type,
                "YES"
                if item.release_exists
                else "NO",
                item.release_bom_id,
                item.release_reference,
                item.missing_component_count,
                "; ".join(
                    item.missing_components
                ),
                "; ".join(
                    str(value)
                    for value
                    in item.duplicate_product_ids
                ),
                "; ".join(
                    item.blocking_reasons
                ),
                "; ".join(
                    item.warnings
                ),
            ]
        )

        fill = {
            "BLOCKED": BLOCKED_FILL,
            "READY": READY_FILL,
            "ALREADY EXISTS": (
                EXISTS_FILL
            ),
        }.get(item.status.value)

        if fill is not None:
            for cell in plan_ws[
                row_index
            ]:
                cell.fill = fill

    blocked_ws = wb.create_sheet(
        "BLOCKED"
    )
    blocked_ws.append(
        [
            "Parent SKU",
            "Blocking Reasons",
            "Missing Components",
            "Sequence 0 BOM Count",
            "Duplicate Product IDs",
        ]
    )

    for item in plan.items:
        if item.status.value != "BLOCKED":
            continue

        blocked_ws.append(
            [
                item.parent_sku,
                "; ".join(
                    item.blocking_reasons
                ),
                "; ".join(
                    item.missing_components
                ),
                item.active_bom_count,
                "; ".join(
                    str(value)
                    for value
                    in item.duplicate_product_ids
                ),
            ]
        )

    for ws in (
        summary_ws,
        plan_ws,
        blocked_ws,
    ):
        _style_sheet(ws)

    wb.save(output_path)
    return output_path
