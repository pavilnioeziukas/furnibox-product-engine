"""Generuoja visų naujų Manufacture BOM importo failą Stage aplinkai.

Generatorius remiasi sėkmingai patikrinto Manufacture piloto formatu. Jis tik
nuskaito Odoo ir sukuria Excel failą – jokių BOM Odoo sistemoje nekuria.
"""

from __future__ import annotations

from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bom_import_pilot_v1 import (
    calculate_levels,
    canon,
    load_bom_types,
    load_new_bom_graph,
)
from bom_import_pilot_v2 import choose_operation_template, load_operation_templates
from bom_import_v1 import load_stage_subcategories
from config import load_settings
from odoo_client import OdooClient
from output_paths import environment_output_dir, environment_slug


IMPORT_HEADERS = [
    "Product/Internal Reference", "qty",
    "BoM Lines/Component/Internal Reference", "product_qty",
    "BoM Type", "Reference", "mo_autodone_by_wo", "auto_plan",
    "operation_ids/name", "operation_ids/workcenter_id",
    "operation_ids/time_mode", "operation_ids/time_cycle_manual",
    "operation_ids/sequence",
]
MANUFACTURE = "Manufacture this product"


def load_stage_products_by_sku(
    client: OdooClient, wanted_skus: set[str]
) -> tuple[dict[str, dict], set[str]]:
    """Randa Stage produktus pagal Internal Reference ir aptinka dublikatus.

    Importo failas sąmoningai nenaudoja aplinkai specifinių External ID. Vienas
    Internal Reference turi rodyti į lygiai vieną produkto variantą ir vieną
    produkto šabloną; kitu atveju Odoo importo susiejimas būtų dviprasmis.
    """
    rows = client.search_read_all(
        "product.product",
        [["default_code", "!=", False]],
        ["id", "default_code", "product_tmpl_id", "active"],
        context={"active_test": False},
    )
    grouped: dict[str, list[dict]] = {}
    for row in rows:
        sku = canon(row.get("default_code"))
        if sku in wanted_skus:
            grouped.setdefault(sku, []).append(row)

    products = {}
    duplicates = set()
    for sku, matches in grouped.items():
        product_ids = {int(row["id"]) for row in matches}
        template_ids = {
            int(row["product_tmpl_id"][0])
            for row in matches
            if row.get("product_tmpl_id")
        }
        if len(product_ids) != 1 or len(template_ids) != 1:
            duplicates.add(sku)
            continue
        row = matches[0]
        products[sku] = {
            "display_sku": str(row["default_code"]).strip(),
            "product_id": next(iter(product_ids)),
            "template_id": next(iter(template_ids)),
        }
    return products, duplicates


def load_stage_workcenters(client: OdooClient) -> tuple[set[str], set[str]]:
    """Grąžina unikalius aktyvius darbo centrus ir pasikartojančius vardus."""
    rows = client.search_read_all(
        "mrp.workcenter", [], ["id", "name", "active"],
        context={"active_test": False},
    )
    counts = Counter(
        str(row.get("name") or "").strip()
        for row in rows
        if row.get("active", True) and str(row.get("name") or "").strip()
    )
    unique = {name for name, count in counts.items() if count == 1}
    duplicates = {name for name, count in counts.items() if count > 1}
    return unique, duplicates


def prepare_manufacture_boms(
    parents, lines, levels, bom_types, products, subcategories,
    operation_templates, workcenters, duplicate_workcenters, duplicate_skus,
):
    """Atrenka Manufacture BOM ir atskiria paruoštus nuo blokuojamų."""
    manufacture_skus = {
        sku for sku in parents if bom_types.get(sku) == MANUFACTURE
    }
    ready = []
    review = []
    diagnostics = []

    for sku in sorted(manufacture_skus, key=lambda value: (levels[value], value)):
        bom_lines = lines.get(sku, [])
        messages = []
        operations = []
        template = None

        if sku in duplicate_skus:
            messages.append("Stage Internal Reference neunikalus BOM produktui")
        elif sku not in products:
            messages.append("Stage nerastas BOM produktas")
        if not bom_lines:
            messages.append("BOM neturi komponentų")
        for line in bom_lines:
            if line["component"] in duplicate_skus:
                messages.append(
                    f"Stage Internal Reference neunikalus komponentui: {line['component']}"
                )
            elif line["component"] not in products:
                messages.append(f"Stage nerastas komponentas: {line['component']}")

        subcategory = subcategories.get(sku, "")
        if not subcategory:
            messages.append("Nenustatyta Stage produkto pakategorė")
        elif canon(subcategory) == "LED HARDWARE":
            messages.append(
                "LED HARDWARE operacijų taisyklė nepatvirtinta"
            )
        else:
            try:
                template = choose_operation_template(
                    sku, subcategory, operation_templates
                )
                operations = template["operations"]
            except ValueError as exc:
                messages.append(str(exc))

        if not operations and canon(subcategory) != "LED HARDWARE":
            messages.append("Nerastas Production operacijų etalonas")
        for operation in operations:
            name = str(operation.get("workcenter") or "").strip()
            if name in duplicate_workcenters:
                messages.append(f"Stage darbo centro pavadinimas neunikalus: {name}")
            elif name not in workcenters:
                messages.append(f"Stage nerastas aktyvus darbo centras: {name}")

        messages = list(dict.fromkeys(messages))
        status = "READY" if not messages else "NOT READY"
        record = {
            "sku": sku,
            "level": levels[sku],
            "lines": bom_lines,
            "operations": operations,
            "subcategory": subcategory,
            "template": template,
            "status": status,
            "message": "; ".join(messages),
        }
        review.append(record)
        if status == "READY":
            ready.append(record)
        else:
            diagnostics.extend(
                [sku, f"lv{levels[sku]}", message] for message in messages
            )
    return ready, review, diagnostics


def import_rows(record: dict, products: dict):
    """Vieną Manufacture BOM paverčia piloto formato importo bloku."""
    parent = products[record["sku"]]
    lines = record["lines"]
    operations = record["operations"]
    row_count = max(len(lines), len(operations), 1)
    reference = f"{date.today():%Y%m%d}_MFG_{record['sku']}"

    for index in range(row_count):
        first = index == 0
        line = lines[index] if index < len(lines) else None
        operation = operations[index] if index < len(operations) else None
        component = products[line["component"]] if line else None
        yield [
            parent["display_sku"] if first else None,
            1 if first else None,
            component["display_sku"] if component else None,
            line["quantity"] if line else None,
            MANUFACTURE if first else None,
            reference if first else None,
            True if first else None,
            True if first else None,
            operation["name"] if operation else None,
            operation["workcenter"] if operation else None,
            operation["time_mode"] if operation else None,
            operation["time"] if operation else None,
            operation["sequence"] if operation else None,
        ]


def style_sheet(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index in range(1, ws.max_column + 1):
        values = [
            len(str(ws.cell(row, index).value or ""))
            for row in range(1, min(ws.max_row, 300) + 1)
        ]
        ws.column_dimensions[get_column_letter(index)].width = min(
            max(values, default=12) + 2, 55
        )


def write_workbook(path: Path, ready, review, diagnostics, products) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    review_ws = wb.create_sheet("REVIEW")
    review_ws.append([
        "Status", "Level", "Parent SKU", "Stage Subcategory", "Components",
        "Operations", "Production Analog", "Analog BOM Reference", "Message",
    ])
    for record in review:
        template = record.get("template") or {}
        review_ws.append([
            record["status"], f"lv{record['level']}", record["sku"],
            record["subcategory"], len(record["lines"]),
            len(record["operations"]), template.get("sku", ""),
            template.get("reference", ""), record["message"],
        ])
    style_sheet(review_ws)

    summary_ws = wb.create_sheet("SUMMARY")
    summary_ws.append(["Metric", "Value"])
    summary_ws.append(["All new Manufacture BOM", len(review)])
    summary_ws.append(["Ready for import", len(ready)])
    summary_ws.append(["Not ready", len(review) - len(ready)])
    for (level, status), count in sorted(
        Counter((row["level"], row["status"]) for row in review).items()
    ):
        summary_ws.append([f"lv{level} {status}", count])
    style_sheet(summary_ws)

    for level in sorted({record["level"] for record in ready}):
        ws = wb.create_sheet(f"BOM_import(lv{level})")
        ws.append(IMPORT_HEADERS)
        for record in (row for row in ready if row["level"] == level):
            for row in import_rows(record, products):
                ws.append(row)
        style_sheet(ws)

    diagnostics_ws = wb.create_sheet("DIAGNOSTICS")
    diagnostics_ws.append(["Parent SKU", "Level", "Message"])
    for row in diagnostics:
        diagnostics_ws.append(row)
    style_sheet(diagnostics_ws)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_import_workbook(path: Path, records, products) -> None:
    """Sukuria vieno lapo failą, paruoštą tiesioginiam Odoo importui."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM_import"
    ws.append(IMPORT_HEADERS)
    for record in records:
        for row in import_rows(record, products):
            ws.append(row)
    style_sheet(ws)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    base = Path(__file__).resolve().parent
    if environment_slug() != "stage":
        raise PermissionError(
            "Manufacture BOM importo generatorius leidžiamas tik Stage."
        )

    output_dir = environment_output_dir(base)
    comparison_path = output_dir / "MAP_Comparison.xlsx"
    types_path = output_dir / "BOM_Type_Review.xlsx"
    references_path = base / "output" / "production" / "BOM_Operations_Reference.xlsx"
    output_path = output_dir / "BOM_Import_Manufacture_All.xlsx"
    lv2_output_path = output_dir / "BOM_Import_Manufacture_lv2.xlsx"
    lv1_output_path = output_dir / "BOM_Import_Manufacture_lv1.xlsx"

    parents, lines = load_new_bom_graph(comparison_path)
    bom_types = load_bom_types(types_path)
    levels = calculate_levels(set(parents), lines)

    settings = load_settings()
    client = OdooClient(settings)
    uid = client.authenticate()
    wanted = set(parents) | {
        row["component"] for values in lines.values() for row in values
    }
    products, duplicate_skus = load_stage_products_by_sku(client, wanted)
    subcategories = load_stage_subcategories(client, set(parents))
    operation_templates = load_operation_templates(references_path)
    workcenters, duplicate_workcenters = load_stage_workcenters(client)

    ready, review, diagnostics = prepare_manufacture_boms(
        parents, lines, levels, bom_types, products, subcategories,
        operation_templates, workcenters, duplicate_workcenters, duplicate_skus,
    )
    write_workbook(output_path, ready, review, diagnostics, products)
    ready_by_level = {
        level: [record for record in ready if record["level"] == level]
        for level in (2, 1)
    }
    write_import_workbook(lv2_output_path, ready_by_level[2], products)
    write_import_workbook(lv1_output_path, ready_by_level[1], products)

    print("Prisijungta prie Stage Odoo. UID=", uid)
    print("\nVISŲ MANUFACTURE BOM IMPORTO FAILAS SUKURTAS")
    print("Failas:", output_path)
    print("Odoo importas lv2:", lv2_output_path)
    print("Odoo importas lv1:", lv1_output_path)
    print("Visi nauji Manufacture BOM:", len(review))
    print("Paruošti importui:", len(ready))
    print("Dar neparuošti:", len(review) - len(ready))
    print("Diagnostikos įrašai:", len(diagnostics))
    print("Importo eiliškumas: pirmiausia lv2, po to lv1.")
    print("Odoo pakeitimų neatlikta.")


if __name__ == "__main__":
    main()
