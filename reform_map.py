"""Generate a canonical Reform BOM map from the `BOM - Input` sheet.

The generator dynamically detects all `Part N Code` / `Part N Qty` pairs,
so it is not tied to v08 or to a fixed number of Part columns.

Outputs:
- REFORM EDGES: canonical Parent SKU -> Component SKU relationships.
- REFORM MAP: human-readable lv1/lv2/lv3 layout.
- DIAGNOSTICS: invalid quantities, duplicate rows and hierarchy cycles.
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


@dataclass(frozen=True)
class PartColumns:
    number: int
    code_index: int
    qty_index: int


def normalize_sku(value: Any) -> str:
    return str(value or "").strip()


def find_input_file(base_dir: Path) -> Path:
    patterns = (
        "BOM for Furnibox*.xlsx",
        "BOM_for Furnibox*.xlsx",
        "Reform_BOM_Input*.xlsx",
    )
    candidates: list[Path] = []
    for pattern in patterns:
        candidates.extend(base_dir.glob(pattern))
    candidates = [p for p in candidates if not p.name.startswith("~$")]
    if not candidates:
        raise FileNotFoundError(
            f"Nerastas Reform BOM failas aplanke {base_dir}. "
            "Naudokite --input su tiksliu failo keliu."
        )
    return max(candidates, key=lambda p: p.stat().st_mtime)


def find_sheet(workbook) -> Any:
    for name in workbook.sheetnames:
        if "bom" in name.lower() and "input" in name.lower():
            return workbook[name]
    if len(workbook.sheetnames) == 1:
        return workbook[workbook.sheetnames[0]]
    raise ValueError("Nerastas lapas 'BOM - Input'.")


def find_header(ws, max_scan_rows: int = 50) -> tuple[int, tuple[Any, ...]]:
    for row_number, values in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True),
        start=1,
    ):
        normalized = [normalize_sku(v) for v in values]
        if "BOM SKU Code" in normalized:
            return row_number, values
    raise ValueError("Nerasta antraštė 'BOM SKU Code'.")


def detect_columns(headers: tuple[Any, ...]) -> tuple[int, list[PartColumns]]:
    positions = {
        normalize_sku(value): index
        for index, value in enumerate(headers)
        if normalize_sku(value)
    }
    sku_index = positions["BOM SKU Code"]
    parts: list[PartColumns] = []
    for label, code_index in positions.items():
        match = re.fullmatch(r"Part (\d+) Code", label)
        if not match:
            continue
        number = int(match.group(1))
        qty_label = f"Part {number} Qty"
        if qty_label not in positions:
            raise ValueError(f"Trūksta stulpelio '{qty_label}'.")
        parts.append(PartColumns(number, code_index, positions[qty_label]))
    parts.sort(key=lambda item: item.number)
    if not parts:
        raise ValueError("Nerasti 'Part N Code' / 'Part N Qty' stulpeliai.")
    return sku_index, parts


def read_edges(ws) -> tuple[list[list[Any]], dict[str, list[tuple[str, float]]], list[list[Any]]]:
    header_row, headers = find_header(ws)
    sku_index, part_columns = detect_columns(headers)

    raw_edges: dict[tuple[str, str], dict[str, Any]] = {}
    parent_rows: dict[str, int] = {}
    diagnostics: list[list[Any]] = []

    for row_number, values in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        parent = normalize_sku(values[sku_index] if sku_index < len(values) else None)
        if not parent:
            continue
        if parent in parent_rows:
            diagnostics.append([
                "DUPLICATE PARENT", row_number, parent, "",
                f"Pirmoji eilutė: {parent_rows[parent]}",
            ])
        else:
            parent_rows[parent] = row_number

        for part in part_columns:
            component = normalize_sku(
                values[part.code_index] if part.code_index < len(values) else None
            )
            if not component:
                continue
            raw_qty = values[part.qty_index] if part.qty_index < len(values) else None
            try:
                quantity = float(raw_qty)
            except (TypeError, ValueError):
                diagnostics.append([
                    "INVALID QUANTITY", row_number, parent, component,
                    f"Part {part.number}: {raw_qty!r}",
                ])
                continue
            if quantity <= 0:
                diagnostics.append([
                    "NON-POSITIVE QUANTITY", row_number, parent, component,
                    f"Part {part.number}: {quantity}",
                ])
                continue

            key = (parent, component)
            if key not in raw_edges:
                raw_edges[key] = {
                    "quantity": 0.0,
                    "rows": [],
                    "parts": [],
                }
            raw_edges[key]["quantity"] += quantity
            raw_edges[key]["rows"].append(row_number)
            raw_edges[key]["parts"].append(part.number)

    edges: list[list[Any]] = []
    graph: dict[str, list[tuple[str, float]]] = defaultdict(list)
    for (parent, component), info in sorted(raw_edges.items()):
        quantity = info["quantity"]
        if quantity.is_integer():
            quantity = int(quantity)
        rows_text = ", ".join(map(str, info["rows"]))
        parts_text = ", ".join(map(str, info["parts"]))
        edges.append([parent, component, quantity, rows_text, parts_text])
        graph[parent].append((component, quantity))
        if len(info["rows"]) > 1:
            diagnostics.append([
                "AGGREGATED EDGE", rows_text, parent, component,
                f"Suminis kiekis: {quantity}",
            ])
    return edges, graph, diagnostics


def build_display_map(
    edges: list[list[Any]],
    graph: dict[str, list[tuple[str, float]]],
) -> tuple[list[list[Any]], list[list[Any]]]:
    direct_by_parent: dict[str, list[tuple[str, float]]] = defaultdict(list)
    for row in edges:
        direct_by_parent[row[0]].append((row[1], row[2]))

    output: list[list[Any]] = []
    diagnostics: list[list[Any]] = []
    for lv1 in sorted(direct_by_parent):
        for lv2, qty_lv2 in direct_by_parent[lv1]:
            children = graph.get(lv2, [])
            if not children:
                output.append([lv1, lv2, qty_lv2, "", ""])
                continue
            for index, (lv3, qty_lv3) in enumerate(children):
                if lv3 in {lv1, lv2}:
                    diagnostics.append([
                        "CYCLE", "", lv1, lv3,
                        f"{lv1} -> {lv2} -> {lv3}",
                    ])
                output.append([
                    lv1 if index == 0 else "",
                    lv2 if index == 0 else "",
                    qty_lv2 if index == 0 else "",
                    lv3,
                    qty_lv3,
                ])
    return output, diagnostics


def add_table(ws, headers: list[str], rows: list[list[Any]], table_name: str) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    if rows:
        table = Table(displayName=table_name, ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    for column in ws.columns:
        lengths = [len(str(cell.value or "")) for cell in column[:200]]
        ws.column_dimensions[column[0].column_letter].width = min(
            max(lengths, default=10) + 2,
            45,
        )


def write_workbook(
    output_path: Path,
    edges: list[list[Any]],
    display_map: list[list[Any]],
    diagnostics: list[list[Any]],
    source_file: Path,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    add_table(
        wb.create_sheet("REFORM EDGES"),
        ["Parent SKU", "Component SKU", "Quantity", "Source Row", "Part Number"],
        edges,
        "ReformEdges",
    )
    add_table(
        wb.create_sheet("REFORM MAP"),
        ["lv1", "lv2", "qty_lv2", "lv3", "qty_lv3"],
        display_map,
        "ReformMap",
    )
    add_table(
        wb.create_sheet("DIAGNOSTICS"),
        ["Type", "Source Row", "Parent SKU", "Component SKU", "Message"],
        diagnostics,
        "ReformDiagnostics",
    )
    info = wb.create_sheet("INFO")
    info.append(["Parameter", "Value"])
    info.append(["Source file", str(source_file)])
    info.append(["Direct edges", len(edges)])
    info.append(["MAP rows", len(display_map)])
    info.append(["Diagnostics", len(diagnostics)])
    for cell in info[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 80

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate Reform BOM MAP")
    parser.add_argument("--input", type=Path, help="Reform BOM .xlsx file")
    parser.add_argument("--output", type=Path, help="Output .xlsx file")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    base_dir = Path(__file__).resolve().parent
    input_path = args.input or find_input_file(base_dir / "data")
    output_path = args.output or (base_dir / "output" / "Reform_MAP.xlsx")

    print("Nuskaitomas Reform failas:", input_path)
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    worksheet = find_sheet(workbook)
    print("Naudojamas lapas:", worksheet.title)

    edges, graph, diagnostics = read_edges(worksheet)
    display_map, map_diagnostics = build_display_map(edges, graph)
    diagnostics.extend(map_diagnostics)
    write_workbook(output_path, edges, display_map, diagnostics, input_path)

    print("\nREFORM MAP SUKURTAS")
    print("Failas:", output_path)
    print("Tiesioginiai ryšiai:", len(edges))
    print("MAP eilutės:", len(display_map))
    print("Diagnostikos įrašai:", len(diagnostics))


if __name__ == "__main__":
    main()
