from __future__ import annotations

import hashlib
import json
import os
import secrets
import shutil
import subprocess
import sys
import threading
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from flask import (
    Flask,
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from werkzeug.utils import secure_filename
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from so_pricing_rules import (
    load_config,
    migrate_legacy_workbook,
    save_config,
    validate_config,
)
from cabinet_parts_price_parameters import (
    load_parameters as load_cabinet_parts_parameters,
    save_parameters as save_cabinet_parts_parameters,
    validate_parameters as validate_cabinet_parts_parameters,
)
from purchase_price_adjustments import (
    load_adjustments as load_purchase_price_adjustments,
    save_adjustments as save_purchase_price_adjustments,
    validate_adjustment as validate_purchase_price_adjustment,
)
from purchase_price_adjustments_import import (
    build_adjustment_preview,
    load_purchase_price_excel_adjustments,
    summarize_preview,
)
from webapp.product_engine import ProductEngineSettings, load_actions


SETTINGS = ProductEngineSettings.from_env(BASE_DIR)
STATE_DIR = SETTINGS.state_dir

UPLOAD_DIR = STATE_DIR / "uploads"
CHUNK_UPLOAD_DIR = UPLOAD_DIR / ".chunks"
RUN_DIR = STATE_DIR / "runs"

PRODUCTION_DATASET_DIR = (
    STATE_DIR
    / "shared_data"
    / "validated_datasets"
    / "production"
)
PRODUCTION_DATASET_PATH = (
    PRODUCTION_DATASET_DIR
    / "latest.json"
)

SO_PRICING_CONFIG_PATH = (
    STATE_DIR
    / "shared_data"
    / "so_pricing_rules.json"
)

CABINET_PARTS_PARAMETERS_PATH = (
    STATE_DIR
    / "shared_data"
    / "cabinet_parts_price_parameters.json"
)

PURCHASE_PRICE_ADJUSTMENTS_PATH = (
    STATE_DIR
    / "shared_data"
    / "purchase_price_adjustments.json"
)
LEGACY_PURCHASE_PRICE_ADJUSTMENTS_PATH = (
    STATE_DIR
    / "shared_data"
    / "tamara_adjustments.json"
)
PURCHASE_PRICE_IMPORT_DIR = (
    STATE_DIR
    / "purchase_price_imports"
)

DEFAULT_SO_PRICING_CONFIG_PATH = (
    BASE_DIR
    / "manifest"
    / "so_pricing_rules.json"
)

MAX_UPLOAD_BYTES = SETTINGS.max_upload_mb * 1024 * 1024


app = Flask(__name__)

app.secret_key = SETTINGS.web_secret or secrets.token_hex(32)

app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_BYTES


for directory in (
    UPLOAD_DIR,
    CHUNK_UPLOAD_DIR,
    RUN_DIR,
    PRODUCTION_DATASET_DIR,
    PURCHASE_PRICE_IMPORT_DIR,
):
    directory.mkdir(
        parents=True,
        exist_ok=True,
    )


def _chunk_upload_paths(upload_id: str) -> tuple[Path, Path]:
    try:
        normalized = uuid.UUID(upload_id).hex
    except (ValueError, AttributeError):
        abort(404)
    return (
        CHUNK_UPLOAD_DIR / f"{normalized}.json",
        CHUNK_UPLOAD_DIR / f"{normalized}.part",
    )


def _read_chunk_upload(upload_id: str) -> tuple[dict[str, Any], Path, Path]:
    metadata_path, part_path = _chunk_upload_paths(upload_id)
    try:
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    except FileNotFoundError:
        abort(404, "Įkėlimo sesija nerasta.")
    return metadata, metadata_path, part_path
if (
    not PURCHASE_PRICE_ADJUSTMENTS_PATH.exists()
    and LEGACY_PURCHASE_PRICE_ADJUSTMENTS_PATH.exists()
):
    PURCHASE_PRICE_ADJUSTMENTS_PATH.parent.mkdir(
        parents=True,
        exist_ok=True,
    )
    shutil.copy2(
        LEGACY_PURCHASE_PRICE_ADJUSTMENTS_PATH,
        PURCHASE_PRICE_ADJUSTMENTS_PATH,
    )


if (
    not SO_PRICING_CONFIG_PATH.exists()
    and DEFAULT_SO_PRICING_CONFIG_PATH.exists()
):
    SO_PRICING_CONFIG_PATH.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    shutil.copy2(
        DEFAULT_SO_PRICING_CONFIG_PATH,
        SO_PRICING_CONFIG_PATH,
    )


BUILTIN_ACTIONS: dict[str, dict[str, Any]] = {
    "mo_component_consumption_audit": {
        "title": "Audituoti užbaigtų MO komponentų sunaudojimą",
        "description": (
            "Tik skaitymo būdu randa užbaigtus MO, kuriuose faktiškai "
            "sunaudota mažiau komponentų negu buvo suplanuota pagal MO BOM. "
            "Parodo planuotą, sunaudotą ir trūkstamą kiekį."
        ),
        "script": "run_mo_component_consumption_audit.py",
        "requires_upload": False,
        "collect_changed_outputs": False,
        "args": ["--output-dir", "{output_dir}", "--days", "550"],
    },
    "odoo_supply_chain_audit": {
        "title": "Audituoti Odoo tiekimo grandines",
        "description": (
            "Tik skaitymo būdu atskiria galiojančius WH/INT pagal WH/INPC "
            "būseną, randa aktyvius MO su Invoiced SO ir patikrina "
            "WH/Input-Custom likučius. Odoo duomenų nekeičia."
        ),
        "script": "run_odoo_supply_chain_audit.py",
        "requires_upload": False,
        "collect_changed_outputs": False,
        "args": [
            "--output-dir",
            "{output_dir}",
        ],
    },
    "stock_by_location": {
        "title": "Generuoti SKU likučius pagal lokaciją",
        "description": (
            "Nuskaito Production likučius WH/Stock ir C/Stock "
            "lokacijose bei paskutinius faktinius pirkimų gavimus."
        ),
        "module": "run_stock_by_location",
        "requires_upload": False,
    },
    "so_supply_status": {
        "title": "Patikrinti SO tiekimo būklę",
        "description": (
            "Pagal SO parodo, ar faktinis MO poreikis rezervuotas "
            "ir kur dabar yra Furnix detalės: PO, gavime, "
            "rūšiavime ar WH/Stock."
        ),
        "script": "run_so_supply_status.py",
        "requires_upload": False,
        "requires_so_number": True,
        "args": [
            "--so-number",
            "{so_number}",
            "--output-dir",
            "{output_dir}",
        ],
    },
    "so_reservation_audit": {
        "title": "Audituoti SO komponentų rezervacijas",
        "description": (
            "Tiesiogiai ir tik skaitymo būdu patikrina konkretaus SO "
            "gamybos užsakymų komponentų poreikį, rezervuotą kiekį ir trūkumą. "
            "Validated Dataset nereikalingas."
        ),
        "script": "run_so_reservation_audit.py",
        "requires_upload": False,
        "requires_so_number": True,
        "collect_changed_outputs": False,
        "args": [
            "--so-number",
            "{so_number}",
            "--output-dir",
            "{output_dir}",
        ],
    },
    "odoo_snapshot": {
        "title": "Nuskaityti Odoo duomenis",
        "description": (
            "Atnaujina tik skaitomą Odoo produktų, BOM ir "
            "kainų momentinę kopiją."
        ),
        "script": "main.py",
        "requires_upload": False,
    },
    "odoo_map": {
        "title": "Generuoti Odoo MAP",
        "description": (
            "Nuskaito aktyvius Odoo BOM ir sukuria Dataset "
            "generavimui reikalingą Odoo_MAP.xlsx."
        ),
        "script": "odoo_map.py",
        "requires_upload": False,
    },
    "bom_operations_reference": {
        "title": "Generuoti BOM operacijų etaloną",
        "description": (
            "Nuskaito Production BOM operacijas ir sukuria "
            "Validated Dataset reikalingą operacijų etaloną."
        ),
        "script": "bom_operations_reference_v1.py",
        "requires_upload": False,
    },
    "product_lifecycle_audit": {
        "title": "Audituoti nebeaktualius Odoo produktus ir BOM",
        "description": (
            "Tik skaitymo būdu palygina Production Odoo su Target Dataset, "
            "patikrina likučius, aktyvius dokumentus ir BOM priklausomybes. "
            "Pateikia archyvavimo kandidatus, bet Odoo nieko nearchyvuoja."
        ),
        "script": "product_lifecycle_audit.py",
        "requires_upload": False,
        "collect_changed_outputs": False,
        "needs_full_target_dataset_arg": True,
        "args": [
            "--pricing-config",
            str(SO_PRICING_CONFIG_PATH),
        ],
    },
    "purchase_prices": {
        "title": "Nuskaityti paskutines pirkimo kainas",
        "description": (
            "Atnaujina komponentų paskutinių pirkimų kainų failą."
        ),
        "script": "last_purchase_prices.py",
        "requires_upload": False,
    },
    "refresh_reform_pricing": {
        "title": "Atnaujinti Reform kainodarą",
        "description": (
            "Vienu paleidimu atnaujina faktines ir Furnibox (Tamaros) pirkimo "
            "kainas, Cabinet Parts kainas, "
            "perskaičiuoja visą Reform BOM kainodarą. Jei yra BLOCKED pozicijų, "
            "pateikia saugų COMPLETE_ONLY failą be jų; pilną galutinį failą "
            "pateikia tik tada, kai BLOCKED nėra. Odoo nekeičia."
        ),
        "script": "refresh_reform_pricing.py",
        "requires_upload": True,
        "collect_changed_outputs": False,
        "blocked_return_codes": [2],
        "args": [
            "--bom-input",
            "{upload}",
            "--rules",
            str(SO_PRICING_CONFIG_PATH),
            "--output-dir",
            "{output_dir}",
        ],
    },
    "so_line_prices": {
        "title": "Generuoti Reform SO kainas",
        "description": (
            "Pagal aktualų Reform BOM, komponentų kainas ir "
            "aplikacijoje valdomas taisykles sukuria "
            "audituojamą SO kainoraštį."
        ),
        "script": "reform_so_line_prices.py",
        "requires_upload": True,
        "args": [
            "--bom-input",
            "{upload}",
            "--rules",
            str(SO_PRICING_CONFIG_PATH),
            "--price-input",
            str(
                BASE_DIR
                / "output"
                / "production"
                / "Reform_Final_Prices.xlsx"
            ),
            "--output-dir",
            "{output_dir}",
        ],
    },
    "validated_dataset": {
        "title": "Generuoti Validated Dataset",
        "description": (
            "Iš įkelto Reform BOM ir Production etalonų "
            "sukuria patikrintą duomenų rinkinį."
        ),
        "script": "generate_full_validated_dataset.py",
        "requires_upload": True,
        "args": [
            "--bom-input",
            "{upload}",
        ],
    },
    "diagnostic_validated_dataset": {
        "title": "Generuoti diagnostinį Validated Dataset",
        "description": (
            "Tik SO tiekimo diagnostikai sukuria Dataset, aiškiai "
            "praleisdamas neišspręsto BOM tipo pozicijas. "
            "Nenaudoti BOM release ar Odoo importui."
        ),
        "script": "generate_full_validated_dataset.py",
        "requires_upload": True,
        "args": [
            "--bom-input",
            "{upload}",
            "--skip-unresolved-bom-types",
        ],
    },
    "product_import": {
        "title": "Paruošti produktų importo failą",
        "description": (
            "Generuoja Odoo importui skirtą Excel; "
            "Odoo duomenų nekeičia."
        ),
        "script": "product_import_v10.py",
        "requires_upload": True,
        "args": [
            "--bom-input",
            "{upload}",
        ],
    },
    "release_analysis": {
        "title": "Analizuoti BOM release",
        "description": (
            "Palygina naujausią Validated Dataset su Production "
            "ir parengia release planą."
        ),
        "script": "analyze_bom_release.py",
        "requires_upload": False,
    },
    "acceptance": {
        "title": "Paleisti acceptance patikrą",
        "description": (
            "Patikrina naujausią Validated Dataset ir pateikia "
            "PASS arba FAIL ataskaitą."
        ),
        "script": "pre_activation_acceptance.py",
        "requires_upload": False,
        "args": [
            "--source",
            "dataset",
        ],
    },
    "release_files": {
        "title": "Generuoti BOM importo failus",
        "description": (
            "Generuoja Excel importo failus iš patvirtinto Dataset; "
            "Odoo duomenų nekeičia."
        ),
        "script": "generate_bom_release.py",
        "requires_upload": False,
        "needs_dataset_arg": True,
    },
}

ACTIONS = load_actions(
    BUILTIN_ACTIONS,
    enabled_actions=SETTINGS.enabled_actions,
    action_modules=SETTINGS.action_modules,
)


_jobs_lock = threading.Lock()

_active_processes: dict[
    str,
    subprocess.Popen[str],
] = {}

_reserved_jobs: set[str] = set()


def utc_now() -> str:
    return datetime.now(
        timezone.utc
    ).isoformat()


def auth_enabled() -> bool:
    return bool(SETTINGS.web_password)


@app.context_processor
def product_engine_context() -> dict[str, Any]:
    return {"product_engine": SETTINGS}


@app.before_request
def require_login():
    if (
        request.endpoint
        in {
            "login",
            "health",
            "static",
        }
        or not auth_enabled()
    ):
        return None

    if not session.get("authenticated"):
        return redirect(
            url_for(
                "login",
                next=request.path,
            )
        )

    return None


@app.route(
    "/login",
    methods=["GET", "POST"],
)
def login():
    error = None

    if request.method == "POST":
        expected = SETTINGS.web_password

        if secrets.compare_digest(
            request.form.get(
                "password",
                "",
            ),
            expected,
        ):
            session["authenticated"] = True

            return redirect(
                request.args.get("next")
                or url_for("index")
            )

        error = "Neteisingas slaptažodis."

    return render_template(
        "login.html",
        error=error,
    )


@app.post("/logout")
def logout():
    session.clear()

    return redirect(
        url_for("login")
    )


def read_job(
    job_dir: Path,
) -> dict[str, Any]:
    return json.loads(
        (
            job_dir
            / "job.json"
        ).read_text(
            encoding="utf-8"
        )
    )


def write_job(
    job_dir: Path,
    payload: dict[str, Any],
) -> None:
    temporary = (
        job_dir
        / "job.json.tmp"
    )

    temporary.write_text(
        json.dumps(
            payload,
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    temporary.replace(
        job_dir
        / "job.json"
    )


def list_jobs() -> list[dict[str, Any]]:
    jobs = []

    for job_dir in RUN_DIR.iterdir():
        if (
            job_dir.is_dir()
            and (
                job_dir
                / "job.json"
            ).exists()
        ):
            try:
                jobs.append(
                    read_job(job_dir)
                )
            except (
                OSError,
                json.JSONDecodeError,
            ):
                continue

    return sorted(
        jobs,
        key=lambda item: item["created_at"],
        reverse=True,
    )[:50]


def latest_upload() -> Path | None:
    files = [
        path
        for path in UPLOAD_DIR.glob("*.xlsx")
        if path.is_file()
    ]

    if not files:
        return None

    return max(
        files,
        key=lambda path: path.stat().st_mtime,
    )


def prune_completed_jobs(keep_per_action: int = 10) -> list[str]:
    """Delete only old completed run directories inside RUN_DIR."""
    grouped: dict[str, list[tuple[str, Path]]] = {}
    protected = set(_reserved_jobs) | set(_active_processes)
    latest_target = latest_full_target_dataset()
    if latest_target is not None:
        protected.add(latest_target.parent.parent.name)

    for job_dir in RUN_DIR.iterdir():
        if not job_dir.is_dir() or job_dir.name in protected:
            continue
        metadata = job_dir / "job.json"
        if not metadata.exists():
            continue
        try:
            job = read_job(job_dir)
        except (OSError, json.JSONDecodeError):
            continue
        if job.get("status") in {"QUEUED", "RUNNING"}:
            continue
        grouped.setdefault(text := str(job.get("action") or "unknown"), []).append(
            (str(job.get("created_at") or ""), job_dir)
        )

    removed = []
    run_root = RUN_DIR.resolve()
    for jobs in grouped.values():
        jobs.sort(key=lambda item: item[0], reverse=True)
        for _, job_dir in jobs[keep_per_action:]:
            resolved = job_dir.resolve()
            if resolved.parent != run_root:
                raise RuntimeError(f"Nesaugi run valymo vieta: {resolved}")
            shutil.rmtree(resolved)
            removed.append(job_dir.name)
    return removed


def latest_dataset() -> Path | None:
    shared_roots = {SETTINGS.shared_data_dir}
    legacy_shared_data = os.getenv("FURNIBOX_SHARED_DATA", "").strip()
    if legacy_shared_data:
        shared_roots.add(Path(legacy_shared_data))

    candidates: list[Path] = []

    for root in shared_roots:
        if root.exists():
            candidates.extend(
                root.glob(
                    "validated_datasets/**/*.json"
                )
            )

    candidates = [
        path
        for path in candidates
        if (
            "Validated_Product_Dataset"
            in path.name
            or path.name == "latest.json"
        )
    ]

    if not candidates:
        return None

    return max(
        candidates,
        key=lambda path: path.stat().st_mtime,
    )


def latest_full_target_dataset() -> Path | None:
    candidates = []
    for path in RUN_DIR.glob("*/files/Furnibox_Target_Dataset.json"):
        metadata = path.parent.parent / "job.json"
        try:
            job = read_job(path.parent.parent)
        except (OSError, json.JSONDecodeError):
            continue
        if metadata.exists() and job.get("status") in {"PASS", "BLOCKED"}:
            candidates.append(path)
    if not candidates:
        return None
    return max(candidates, key=lambda path: path.stat().st_mtime)


def snapshot_outputs() -> dict[str, int]:
    result: dict[str, int] = {}

    roots = [
        BASE_DIR / "output",
        BASE_DIR / "validated_datasets",
    ]

    roots.append(SETTINGS.shared_data_dir)

    for root in roots:
        if not root.exists():
            continue

        for path in root.rglob("*"):
            if path.is_file():
                result[
                    str(path.resolve())
                ] = path.stat().st_mtime_ns

    return result


def collect_outputs(
    before: dict[str, int],
    destination: Path,
) -> list[dict[str, str]]:
    files = []

    for absolute, mtime in snapshot_outputs().items():
        source = Path(absolute)

        if before.get(absolute) == mtime:
            continue

        target = (
            destination
            / source.name
        )

        counter = 2

        while target.exists():
            target = (
                destination
                / (
                    f"{source.stem}_"
                    f"{counter}"
                    f"{source.suffix}"
                )
            )

            counter += 1

        shutil.copy2(
            source,
            target,
        )

        files.append(
            {
                "name": target.name,
                "path": str(target),
            }
        )

    return sorted(
        files,
        key=lambda item: item["name"],
    )


def run_job(
    job_id: str,
    action_key: str,
    upload: Path | None,
    so_number: str | None = None,
) -> None:
    job_dir = (
        RUN_DIR
        / job_id
    )

    output_dir = (
        job_dir
        / "files"
    )

    output_dir.mkdir(
        exist_ok=True
    )

    job = read_job(job_dir)

    action = ACTIONS[
        action_key
    ]

    before = snapshot_outputs()

    if action.get("module"):
        command = [
            sys.executable,
            "-u",
            "-m",
            action["module"],
        ]
    else:
        command = [
            sys.executable,
            "-u",
            str(
                BASE_DIR
                / action["script"]
            ),
        ]

    for argument in action.get(
        "args",
        [],
    ):
        command.append(
            argument.format(
                upload=(
                    str(upload)
                    if upload
                    else ""
                ),
                so_number=(
                    so_number
                    or ""
                ),
                output_dir=str(
                    output_dir
                ),
            )
        )

    if action.get("needs_dataset_arg") or action.get("needs_full_target_dataset_arg"):
        dataset = (
            latest_full_target_dataset()
            if action.get("needs_full_target_dataset_arg")
            else latest_dataset()
        )

        if dataset is None:
            job.update(
                status="ERROR",
                finished_at=utc_now(),
                error=(
                    "Pilnas Target Dataset dar nesukurtas. "
                    "Pirmiausia paleiskite Reform kainodaros atnaujinimą."
                ),
            )

            write_job(
                job_dir,
                job,
            )

            with _jobs_lock:
                _reserved_jobs.discard(
                    job_id
                )

            return

        command.extend(
            [
                "--dataset",
                str(dataset),
                "--output-dir",
                str(output_dir),
            ]
        )

    env = os.environ.copy()

    env["PRODUCT_ENGINE_ENVIRONMENT"] = SETTINGS.environment
    env["FURNIBOX_ENVIRONMENT"] = SETTINGS.environment

    env[
        "PYTHONUTF8"
    ] = "1"

    shared_data = str(
        STATE_DIR
        / "shared_data"
    )

    env.setdefault("PRODUCT_ENGINE_SHARED_DATA_DIR", shared_data)
    env.setdefault("FURNIBOX_SHARED_DATA", shared_data)
    env.setdefault("FURNIBOX_SHARED_DATA_DIR", shared_data)

    log_path = (
        job_dir
        / "run.log"
    )

    job.update(
        status="RUNNING",
        started_at=utc_now(),
    )

    write_job(
        job_dir,
        job,
    )

    try:
        with log_path.open(
            "w",
            encoding="utf-8",
        ) as log:
            process = subprocess.Popen(
                command,
                cwd=BASE_DIR,
                env=env,
                stdout=log,
                stderr=subprocess.STDOUT,
                text=True,
            )

            with _jobs_lock:
                _active_processes[
                    job_id
                ] = process

            return_code = (
                process.wait()
            )

        files = (
            collect_outputs(before, output_dir)
            if action.get("collect_changed_outputs", True)
            else []
        )

        known_paths = {
            item["path"]
            for item in files
        }

        for path in sorted(
            output_dir.iterdir()
        ):
            if (
                path.is_file()
                and str(path)
                not in known_paths
            ):
                files.append(
                    {
                        "name": path.name,
                        "path": str(path),
                    }
                )

        blocked_return_codes = set(action.get("blocked_return_codes", []))
        job.update(
            status=(
                "PASS"
                if return_code == 0
                else (
                    "BLOCKED"
                    if return_code in blocked_return_codes
                    else "FAIL"
                )
            ),
            return_code=return_code,
            finished_at=utc_now(),
            files=files,
        )

    except Exception as exc:
        job.update(
            status="ERROR",
            finished_at=utc_now(),
            error=str(exc),
        )

    finally:
        with _jobs_lock:
            _active_processes.pop(
                job_id,
                None,
            )

            _reserved_jobs.discard(
                job_id
            )

        write_job(
            job_dir,
            job,
        )


@app.get("/")
def index():
    return render_template(
        "index.html",
        actions=ACTIONS,
        jobs=list_jobs(),
        upload=latest_upload(),
        dataset=latest_dataset(),
        auth_enabled=auth_enabled(),
        show_bom_workspace=SETTINGS.show_bom_workspace,
    )


@app.get("/purchase-pricing")
def purchase_pricing():
    parameters = (
        load_cabinet_parts_parameters(
            CABINET_PARTS_PARAMETERS_PATH
        )
    )

    purchase_price_adjustments = (
        load_purchase_price_adjustments(
            PURCHASE_PRICE_ADJUSTMENTS_PATH
        )
    )

    adjustment_query = request.args.get(
        "adjustment_q",
        "",
    ).strip()

    adjustment_rows = [
        {
            "sku": sku,
            **document,
        }
        for sku, document
        in purchase_price_adjustments.items()
    ]

    if adjustment_query:
        query = adjustment_query.casefold()

        adjustment_rows = [
            row
            for row in adjustment_rows
            if (
                query
                in row["sku"].casefold()
                or query
                in row["comment"].casefold()
            )
        ]

    return render_template(
        "purchase_pricing.html",
        parameters=parameters,
        adjustment_rows=adjustment_rows[:200],
        adjustment_total=len(
            purchase_price_adjustments
        ),
        adjustment_query=adjustment_query,
    )


def _latest_job_for(action: str) -> dict[str, Any] | None:
    return next(
        (
            job
            for job in list_jobs()
            if job.get("action") == action
        ),
        None,
    )


def _job_file(job: dict[str, Any] | None, *names: str) -> Path | None:
    if not job:
        return None
    by_name = {
        str(item.get("name")): Path(str(item.get("path")))
        for item in job.get("files", [])
        if item.get("name") and item.get("path")
    }
    for name in names:
        path = by_name.get(name)
        if path and path.exists() and path.is_file():
            return path
    return None


def _sheet_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return []
        rows = workbook[sheet_name].iter_rows(values_only=True)
        header = next(rows, ())
        names = [str(value or "").strip() for value in header]
        return [
            dict(zip(names, values))
            for values in rows
            if any(value not in (None, "") for value in values)
        ]
    finally:
        workbook.close()


def _read_pricing_workbook_match(
    path: Path,
    normalized: str,
) -> tuple[dict[str, Any] | None, list[dict[str, Any]], list[dict[str, Any]]]:
    """Read only matching rows; never materialize the large trace sheet."""
    workbook = load_workbook(path, data_only=True, read_only=True)
    match = None
    trace: list[dict[str, Any]] = []
    candidates: list[dict[str, Any]] = []
    try:
        if "PRICE RESULTS" not in workbook.sheetnames:
            return match, trace, candidates
        rows = workbook["PRICE RESULTS"].iter_rows(values_only=True)
        names = [str(value or "").strip() for value in next(rows, ())]
        for values in rows:
            row = dict(zip(names, values))
            sku = str(row.get("SKU") or "").strip()
            if normalized in sku.casefold() and len(candidates) < 20:
                candidates.append({"sku": sku, "name": row.get("Name") or ""})
            if sku.casefold() == normalized:
                match = {
                    "sku": sku,
                    "name": row.get("Name") or "",
                    "position_type": row.get("Position Type") or "",
                    "category": row.get("Product Category") or "",
                    "cost": row.get("Component / Purchase Cost"),
                    "addons": row.get("Pricing Add-ons Total"),
                    "adjustment": row.get("Adjustment Amount"),
                    "final": row.get("Final Reform SO Unit Price"),
                    "status": row.get("Control Status") or "",
                    "rules": row.get("Applied Rule IDs") or "",
                    "issues": row.get("Issues / Review Reason") or "",
                }

        if match and "PRICE TRACE" in workbook.sheetnames:
            rows = workbook["PRICE TRACE"].iter_rows(values_only=True)
            names = [str(value or "").strip() for value in next(rows, ())]
            found = False
            for values in rows:
                row = dict(zip(names, values))
                sku = str(row.get("SKU") or "").strip().casefold()
                if sku == normalized:
                    trace.append(row)
                    found = True
                elif found:
                    # pricing_control writes every SKU trace as one contiguous block.
                    break
        return match, trace, candidates
    finally:
        workbook.close()


def _search_latest_pricing(job: dict[str, Any] | None, query: str) -> dict[str, Any]:
    result: dict[str, Any] = {
        "query": query,
        "match": None,
        "trace": [],
        "suggestions": [],
        "source_job": job,
    }
    normalized = query.strip().casefold()
    if not normalized or not job:
        return result

    pricing_path = _job_file(
        job,
        "Reform_SO_Line_Prices.xlsx",
        "Reform_SO_Line_Prices_COMPLETE_ONLY.xlsx",
    )
    blocker_path = _job_file(job, "Reform_Pricing_BLOCKED.xlsx")

    pricing_match = None
    pricing_trace: list[dict[str, Any]] = []
    pricing_candidates: list[dict[str, Any]] = []
    if pricing_path:
        pricing_match, pricing_trace, pricing_candidates = (
            _read_pricing_workbook_match(pricing_path, normalized)
        )
    blockers = _sheet_rows(blocker_path, "BLOCKERS") if blocker_path else []
    candidates: list[dict[str, Any]] = pricing_candidates
    result["match"] = pricing_match

    for row in blockers:
        sku = str(row.get("SKU") or "").strip()
        if normalized in sku.casefold():
            candidates.append({"sku": sku, "name": ""})
        if sku.casefold() == normalized:
            result["match"] = {
                "sku": sku,
                "name": "",
                "position_type": row.get("Position Type") or "",
                "category": "",
                "cost": None,
                "addons": None,
                "adjustment": None,
                "final": None,
                "status": row.get("Status") or "BLOCKED",
                "rules": "R006",
                "issues": row.get("Issues") or "",
            }

    if result["match"] and pricing_path:
        result["trace"] = pricing_trace
        result["materials"] = [
            row for row in result["trace"]
            if str(row.get("Step Type") or "").upper() == "MATERIAL"
        ]
        result["tariffs"] = [
            row for row in result["trace"]
            if str(row.get("Step Type") or "").upper() == "PRICING ADD-ON"
        ]
    else:
        result["materials"] = []
        result["tariffs"] = []

    seen = set()
    for candidate in candidates:
        key = candidate["sku"].casefold()
        if key not in seen:
            result["suggestions"].append(candidate)
            seen.add(key)
        if len(result["suggestions"]) == 20:
            break
    return result


PRICING_RULE_LABELS = {
    "R001": (
        "Tiesioginė komponento kaina",
        "Panaudota paruošta faktinė arba patvirtinta pakoreguota pirkimo kaina.",
    ),
    "R002": (
        "Komponentų savikaina pagal BOM",
        "Savikaina gauta sudėjus BOM komponentų kainas, padaugintas iš jų kiekių.",
    ),
    "R003": (
        "Kategorijos darbų ir aptarnavimo priedai",
        "Pridėti kategorijai nustatyti surinkimo, sandėliavimo, pakavimo ir kiti tarifai.",
    ),
    "R004": (
        "Bendra priedų korekcija",
        "Bendras korekcijos procentas pritaikytas tik kainodaros priedams, ne komponentų savikainai.",
    ),
    "R005": (
        "Vidinio gamybos produkto savikaina",
        "APACK, HRD-A arba Shelf-PP komponentai įtraukti per jų vidinę BOM struktūrą.",
    ),
    "R006": (
        "Trūkstamų duomenų saugiklis",
        "Kaina blokuojama, kai nėra patikimos tiesioginės kainos arba išsprendžiamos BOM struktūros.",
    ),
    "R007": (
        "Ne BOM produkto kaina",
        "Prie pirkimo kainos pridėti paruošimo, sandėliavimo, maišelio ir lipduko tarifai.",
    ),
}


def _rule_explanations(match: dict[str, Any] | None) -> list[dict[str, str]]:
    if not match:
        return []
    result = []
    for rule_id in str(match.get("rules") or "").split(","):
        rule_id = rule_id.strip()
        if not rule_id:
            continue
        title, explanation = PRICING_RULE_LABELS.get(
            rule_id,
            ("Techninė kainodaros taisyklė", "Taisyklės paaiškinimas dar neaprašytas."),
        )
        result.append({"id": rule_id, "title": title, "explanation": explanation})
    return result


@app.get("/pricing-control")
def pricing_control():
    config = load_config(SO_PRICING_CONFIG_PATH)
    parameters = load_cabinet_parts_parameters(
        CABINET_PARTS_PARAMETERS_PATH
    )
    adjustments = load_purchase_price_adjustments(
        PURCHASE_PRICE_ADJUSTMENTS_PATH
    )

    parameter_updated_at = None
    if CABINET_PARTS_PARAMETERS_PATH.exists():
        parameter_updated_at = datetime.fromtimestamp(
            CABINET_PARTS_PARAMETERS_PATH.stat().st_mtime,
            timezone.utc,
        ).isoformat()

    pricing_job = _latest_job_for("refresh_reform_pricing")
    sku_query = request.args.get("sku", "").strip()
    sku_search = _search_latest_pricing(pricing_job, sku_query)

    return render_template(
        "pricing_control.html",
        config=config,
        parameters=parameters,
        adjustment_total=len(adjustments),
        pricing_job=pricing_job,
        sku_search=sku_search,
        rule_explanations=_rule_explanations(sku_search["match"]),
        lifecycle_job=_latest_job_for("product_lifecycle_audit"),
        parameter_updated_at=parameter_updated_at,
    )


@app.post("/pricing-control/manual-values")
def update_pricing_control_manual_values():
    document = {
        "back_rate_per_m2": form_number("back_rate_per_m2"),
        "processing_rate_per_m2": form_number("processing_rate_per_m2"),
        "ww_material_rate_per_m2": form_number("ww_material_rate_per_m2"),
        "bb_material_rate_per_m2": form_number("bb_material_rate_per_m2"),
        "no_material_rate_per_m2": form_number("no_material_rate_per_m2"),
        "small_part_threshold_m2": form_number("small_part_threshold_m2"),
        "small_part_surcharge": form_number("small_part_surcharge"),
        "furnix_markup_percent": form_number("furnix_markup_percent"),
        "output_decimals": request.form.get("output_decimals", "").strip(),
    }

    try:
        parameters = validate_cabinet_parts_parameters(document)
        adjustment_percent = form_number("adjustment_percent")
        if not -100 < adjustment_percent <= 0:
            raise ValueError(
                "Bendra BOM korekcija turi būti didesnė nei -100 % "
                "ir ne didesnė nei 0 %."
            )

        config = load_config(SO_PRICING_CONFIG_PATH)
        config["adjustment_rate"] = adjustment_percent / 100

        # Validate both groups before either application-owned file is changed.
        config = validate_config(config)

        save_cabinet_parts_parameters(
            CABINET_PARTS_PARAMETERS_PATH,
            parameters,
        )
        save_config(SO_PRICING_CONFIG_PATH, config)
    except ValueError as exc:
        abort(400, str(exc))

    flash(
        "Rankiniu būdu valdomos kainodaros reikšmės išsaugotos. "
        "Naujos reikšmės bus naudojamos kitame kainodaros paleidime."
    )
    return redirect(url_for("pricing_control") + "#manual-values")


@app.post("/purchase-pricing")
def update_purchase_pricing():
    document = {
        "back_rate_per_m2": form_number(
            "back_rate_per_m2"
        ),
        "processing_rate_per_m2": form_number(
            "processing_rate_per_m2"
        ),
        "ww_material_rate_per_m2": form_number(
            "ww_material_rate_per_m2"
        ),
        "bb_material_rate_per_m2": form_number(
            "bb_material_rate_per_m2"
        ),
        "no_material_rate_per_m2": form_number(
            "no_material_rate_per_m2"
        ),
        "small_part_threshold_m2": form_number(
            "small_part_threshold_m2"
        ),
        "small_part_surcharge": form_number(
            "small_part_surcharge"
        ),
        "furnix_markup_percent": form_number(
            "furnix_markup_percent"
        ),
        "output_decimals": request.form.get(
            "output_decimals",
            "",
        ).strip(),
    }

    try:
        parameters = (
            validate_cabinet_parts_parameters(
                document
            )
        )

        save_cabinet_parts_parameters(
            CABINET_PARTS_PARAMETERS_PATH,
            parameters,
        )

    except ValueError as exc:
        abort(
            400,
            str(exc),
        )

    flash(
        "Pirkimo kainodaros parametrai išsaugoti."
    )

    return redirect(
        url_for(
            "purchase_pricing"
        )
    )


@app.post(
    "/purchase-pricing/adjustments/excel-preview"
)
def preview_purchase_price_excel():
    file = request.files.get("file")

    if file is None or not file.filename:
        abort(
            400,
            "Pasirinkite pirkimo kainų Excel failą.",
        )

    filename = secure_filename(file.filename)

    if Path(filename).suffix.lower() != ".xlsx":
        abort(
            400,
            "Leidžiamas tik .xlsx failas.",
        )

    import_id = uuid.uuid4().hex

    excel_path = (
        PURCHASE_PRICE_IMPORT_DIR
        / f"{import_id}.xlsx"
    )

    metadata_path = (
        PURCHASE_PRICE_IMPORT_DIR
        / f"{import_id}.json"
    )

    file.save(excel_path)

    try:
        excel_adjustments = (
            load_purchase_price_excel_adjustments(
                excel_path
            )
        )

        current_adjustments = (
            load_purchase_price_adjustments(
                PURCHASE_PRICE_ADJUSTMENTS_PATH
            )
        )

        preview_rows = build_adjustment_preview(
            excel_adjustments,
            current_adjustments,
        )

        summary = summarize_preview(
            preview_rows
        )

        changed_rows = [
            row
            for row in preview_rows
            if row.status != "SAME"
        ]

        metadata_path.write_text(
            json.dumps(
                {
                    "source_filename": filename,
                    "created_at": utc_now(),
                },
                ensure_ascii=False,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )

        return render_template(
            "purchase_price_adjustments_preview.html",
            import_id=import_id,
            source_filename=filename,
            summary=summary,
            rows=changed_rows[:500],
            changed_total=len(changed_rows),
        )

    except (ValueError, OSError) as exc:
        excel_path.unlink(missing_ok=True)
        metadata_path.unlink(missing_ok=True)
        abort(400, str(exc))


@app.post(
    "/purchase-pricing/adjustments/excel-apply/<import_id>"
)
def apply_purchase_price_excel(import_id: str):
    safe_import_id = secure_filename(import_id)

    if (
        not safe_import_id
        or safe_import_id != import_id
    ):
        abort(404)

    excel_path = (
        PURCHASE_PRICE_IMPORT_DIR
        / f"{import_id}.xlsx"
    )

    metadata_path = (
        PURCHASE_PRICE_IMPORT_DIR
        / f"{import_id}.json"
    )

    if (
        not excel_path.exists()
        or not metadata_path.exists()
    ):
        abort(
            404,
            "Excel peržiūra neberasta. Įkelkite failą iš naujo.",
        )

    try:
        metadata = json.loads(
            metadata_path.read_text(
                encoding="utf-8"
            )
        )

        excel_adjustments = (
            load_purchase_price_excel_adjustments(
                excel_path
            )
        )

        current_adjustments = (
            load_purchase_price_adjustments(
                PURCHASE_PRICE_ADJUSTMENTS_PATH
            )
        )

        preview_rows = build_adjustment_preview(
            excel_adjustments,
            current_adjustments,
        )

        rows_to_apply = [
            row
            for row in preview_rows
            if row.status in {
                "NEW",
                "CHANGED",
            }
        ]

        if not rows_to_apply:
            flash(
                "Pirkimo kainų Excel neturi naujų ar pakeistų kainų."
            )
            return redirect(
                url_for("purchase_pricing")
                + "#purchase-price-adjustments"
            )

        if PURCHASE_PRICE_ADJUSTMENTS_PATH.exists():
            backup_dir = (
                STATE_DIR
                / "shared_data"
                / "backups"
            )
            backup_dir.mkdir(
                parents=True,
                exist_ok=True,
            )

            timestamp = datetime.now().strftime(
                "%Y%m%d_%H%M%S_%f"
            )

            backup_path = (
                backup_dir
                / f"purchase_price_adjustments_{timestamp}.json"
            )

            shutil.copy2(
                PURCHASE_PRICE_ADJUSTMENTS_PATH,
                backup_path,
            )

        source_filename = str(
            metadata.get(
                "source_filename",
                "Pirkimo kainų Excel",
            )
        )

        for row in rows_to_apply:
            current_adjustments[row.sku] = {
                "adjusted_purchase_price":
                    row.new_adjustment,
                "comment": (
                    "Imported from "
                    f"{source_filename}"
                ),
            }

        save_purchase_price_adjustments(
            PURCHASE_PRICE_ADJUSTMENTS_PATH,
            current_adjustments,
        )

        flash(
            f"Pritaikyta {len(rows_to_apply)} "
            "Pirkimo kainų pakeitimų."
        )

    except (
        ValueError,
        OSError,
        json.JSONDecodeError,
    ) as exc:
        abort(400, str(exc))

    finally:
        excel_path.unlink(missing_ok=True)
        metadata_path.unlink(missing_ok=True)

    return redirect(
        url_for("purchase_pricing")
        + "#purchase-price-adjustments"
    )


@app.post(
    "/purchase-pricing/adjustments/<path:sku>"
)
def update_purchase_price_adjustment(
    sku: str,
):
    adjustments = (
        load_purchase_price_adjustments(
            PURCHASE_PRICE_ADJUSTMENTS_PATH
        )
    )

    if sku not in adjustments:
        abort(404)

    try:
        adjustments[sku] = (
            validate_purchase_price_adjustment(
                sku,
                {
                    "adjusted_purchase_price":
                        request.form.get(
                            "adjusted_purchase_price",
                            "",
                        ),
                    "comment":
                        request.form.get(
                            "comment",
                            "",
                        ),
                },
            )
        )

        save_purchase_price_adjustments(
            PURCHASE_PRICE_ADJUSTMENTS_PATH,
            adjustments,
        )

    except ValueError as exc:
        abort(
            400,
            str(exc),
        )

    flash(
        f"Pirkimo kainos korekcija "
        f"{sku} išsaugota."
    )

    query = request.form.get(
        "return_query",
        "",
    ).strip()

    return redirect(
        url_for(
            "purchase_pricing",
            adjustment_q=query,
        )
        + "#purchase-price-adjustments"
    )


@app.get(
    "/purchase-pricing/adjustments/export"
)
def export_purchase_price_adjustments():
    if not PURCHASE_PRICE_ADJUSTMENTS_PATH.exists():
        abort(
            404,
            "Tamaros korekcijų saugykla dar nesukurta.",
        )

    return send_file(
        PURCHASE_PRICE_ADJUSTMENTS_PATH,
        as_attachment=True,
        download_name=(
            "purchase_price_adjustments.json"
        ),
    )


@app.post(
    "/purchase-pricing/adjustments/import"
)
def import_purchase_price_adjustments():
    file = request.files.get(
        "file"
    )

    if (
        file is None
        or not file.filename
    ):
        abort(
            400,
            "Pasirinkite Tamaros korekcijų JSON failą.",
        )

    filename = secure_filename(
        file.filename
    )

    if (
        Path(filename).suffix.lower()
        != ".json"
    ):
        abort(
            400,
            "Leidžiamas tik .json failas.",
        )

    temporary_path = (
        UPLOAD_DIR
        / (
            "purchase_price_adjustments_import_"
            f"{uuid.uuid4().hex}.json"
        )
    )

    file.save(
        temporary_path
    )

    try:
        adjustments = (
            load_purchase_price_adjustments(
                temporary_path
            )
        )

        if not adjustments:
            abort(
                400,
                (
                    "Importuojamas Tamaros korekcijų "
                    "failas neturi nė vienos korekcijos."
                ),
            )

        if PURCHASE_PRICE_ADJUSTMENTS_PATH.exists():
            backup_dir = (
                STATE_DIR
                / "shared_data"
                / "backups"
            )

            backup_dir.mkdir(
                parents=True,
                exist_ok=True,
            )

            timestamp = (
                datetime.now().strftime(
                    "%Y%m%d_%H%M%S"
                )
            )

            backup_path = (
                backup_dir
                / (
                    "purchase_price_adjustments_"
                    f"{timestamp}.json"
                )
            )

            shutil.copy2(
                PURCHASE_PRICE_ADJUSTMENTS_PATH,
                backup_path,
            )

        save_purchase_price_adjustments(
            PURCHASE_PRICE_ADJUSTMENTS_PATH,
            adjustments,
        )

    except ValueError as exc:
        abort(
            400,
            str(exc),
        )

    finally:
        temporary_path.unlink(
            missing_ok=True
        )

    flash(
        f"Importuotos {len(adjustments)} "
        "Tamaros pirkimo kainų korekcijos."
    )

    return redirect(
        url_for(
            "purchase_pricing"
        )
        + "#purchase-price-adjustments"
    )


@app.get("/pricing-rules")
def pricing_rules():
    document = load_config(
        SO_PRICING_CONFIG_PATH
    )

    query = request.args.get(
        "q",
        "",
    ).strip().casefold()

    assignments = document[
        "bom_skus"
    ]

    if query:
        assignments = [
            row
            for row in assignments
            if query
            in " ".join(
                str(value)
                for value
                in row.values()
            ).casefold()
        ]

    bom_counts = {}

    for row in document[
        "bom_skus"
    ]:
        category_id = row[
            "category_id"
        ]

        bom_counts[
            category_id
        ] = (
            bom_counts.get(
                category_id,
                0,
            )
            + 1
        )

    non_bom_counts = {}

    for row in document[
        "non_bom_skus"
    ]:
        category_id = row[
            "category_id"
        ]

        non_bom_counts[
            category_id
        ] = (
            non_bom_counts.get(
                category_id,
                0,
            )
            + 1
        )

    return render_template(
        "pricing_rules.html",
        config=document,
        assignments=assignments,
        bom_counts=bom_counts,
        non_bom_counts=non_bom_counts,
        configured=(
            SO_PRICING_CONFIG_PATH.exists()
        ),
        query=request.args.get(
            "q",
            "",
        ).strip(),
    )


@app.post(
    "/pricing-rules/migrate"
)
def migrate_pricing_rules():
    file = request.files.get(
        "file"
    )

    if (
        file is None
        or not file.filename
        or Path(
            secure_filename(
                file.filename
            )
        ).suffix.lower()
        != ".xlsx"
    ):
        abort(
            400,
            "Pasirinkite seną kainodaros .xlsx failą.",
        )

    temporary = (
        UPLOAD_DIR
        / (
            "pricing_migration_"
            f"{uuid.uuid4().hex}.xlsx"
        )
    )

    file.save(
        temporary
    )

    try:
        document = (
            migrate_legacy_workbook(
                temporary
            )
        )

        save_config(
            SO_PRICING_CONFIG_PATH,
            document,
        )

    except ValueError as exc:
        abort(
            400,
            str(exc),
        )

    finally:
        temporary.unlink(
            missing_ok=True
        )

    flash(
        "Taisyklės perkeltos į aplikaciją. "
        "Seno Excel kasdieniam darbui nebereikia."
    )

    return redirect(
        url_for(
            "pricing_rules"
        )
    )


def form_number(
    name: str,
) -> float:
    raw = (
        request.form.get(
            name,
            "",
        )
        .strip()
        .replace(
            ",",
            ".",
        )
    )

    try:
        return float(raw)

    except ValueError:
        abort(
            400,
            (
                f"Laukas „{name}“ "
                "turi būti skaičius."
            ),
        )


@app.post(
    "/pricing-rules/adjustment"
)
def update_pricing_adjustment():
    document = load_config(
        SO_PRICING_CONFIG_PATH
    )

    document[
        "adjustment_rate"
    ] = (
        form_number(
            "adjustment_percent"
        )
        / 100
    )

    try:
        save_config(
            SO_PRICING_CONFIG_PATH,
            document,
        )

    except ValueError as exc:
        abort(
            400,
            str(exc),
        )

    flash(
        "Bendra BOM kainodaros korekcija išsaugota."
    )

    return redirect(
        url_for(
            "pricing_rules"
        )
    )


@app.post(
    "/pricing-rules/business-categories/<code>"
)
def update_business_pricing_category(
    code: str,
):
    document = load_config(
        SO_PRICING_CONFIG_PATH
    )
    match = next(
        (
            row
            for row in document.get("bom_category_rates", [])
            if str(row.get("code") or "").casefold() == code.casefold()
        ),
        None,
    )
    if match is None:
        abort(404)
    match["name"] = request.form.get("name", "").strip()
    for name in (
        "assembly",
        "storage",
        "packaging",
        "put_on_pallet",
        "other",
        "markup",
    ):
        match[name] = form_number(name)
    try:
        save_config(SO_PRICING_CONFIG_PATH, document)
    except ValueError as exc:
        abort(400, str(exc))
    flash(f"Verslo BOM kategorija {match['code']} išsaugota.")
    return redirect(url_for("pricing_rules") + "#business-categories")


@app.post(
    "/pricing-rules/categories/<category_id>"
)
def update_pricing_category(
    category_id: str,
):
    document = load_config(
        SO_PRICING_CONFIG_PATH
    )

    match = next(
        (
            row
            for row
            in document[
                "bom_categories"
            ]
            if row["id"]
            == category_id
        ),
        None,
    )

    if match is None:
        abort(404)

    for name in (
        "name",
        "source_category_id",
        "odoo_category",
    ):
        match[name] = (
            request.form.get(
                name,
                "",
            ).strip()
        )

    for name in (
        "assembly",
        "storage",
        "packaging",
        "put_on_pallet",
        "other",
        "markup",
    ):
        match[name] = (
            form_number(name)
        )

    save_config(
        SO_PRICING_CONFIG_PATH,
        document,
    )

    flash(
        f"BOM kategorija "
        f"{match['name']} išsaugota."
    )

    return redirect(
        url_for(
            "pricing_rules"
        )
        + "#bom-categories"
    )


@app.post(
    "/pricing-rules/bom-skus/<path:sku>"
)
def update_bom_sku(
    sku: str,
):
    document = load_config(
        SO_PRICING_CONFIG_PATH
    )

    match = next(
        (
            row
            for row
            in document[
                "bom_skus"
            ]
            if row[
                "sku"
            ].casefold()
            == sku.casefold()
        ),
        None,
    )

    if match is None:
        abort(404)

    match[
        "category_id"
    ] = request.form.get(
        "category_id",
        "",
    ).strip()

    try:
        save_config(
            SO_PRICING_CONFIG_PATH,
            document,
        )

    except ValueError as exc:
        abort(
            400,
            str(exc),
        )

    flash(
        f"BOM SKU "
        f"{match['sku']} "
        "kategorija išsaugota."
    )

    return redirect(
        url_for(
            "pricing_rules",
            q=request.form.get(
                "return_query",
                "",
            ),
        )
        + "#bom-skus"
    )


@app.post(
    "/pricing-rules/non-bom-categories/<category_id>"
)
def update_non_bom_category(
    category_id: str,
):
    document = load_config(
        SO_PRICING_CONFIG_PATH
    )

    match = next(
        (
            row
            for row
            in document[
                "non_bom_categories"
            ]
            if row["id"]
            == category_id
        ),
        None,
    )

    if match is None:
        abort(404)

    match["name"] = (
        request.form.get(
            "name",
            "",
        ).strip()
    )

    for name in (
        "preparation",
        "storage",
        "bag",
        "sticker",
    ):
        match[name] = (
            form_number(name)
        )

    save_config(
        SO_PRICING_CONFIG_PATH,
        document,
    )

    flash(
        f"Ne BOM kategorija "
        f"{match['name']} išsaugota."
    )

    return redirect(
        url_for(
            "pricing_rules"
        )
        + "#non-bom-categories"
    )


@app.post(
    "/pricing-rules/non-bom-skus/<path:sku>"
)
def update_non_bom_sku(
    sku: str,
):
    document = load_config(
        SO_PRICING_CONFIG_PATH
    )

    match = next(
        (
            row
            for row
            in document[
                "non_bom_skus"
            ]
            if row[
                "sku"
            ].casefold()
            == sku.casefold()
        ),
        None,
    )

    if match is None:
        abort(404)

    match[
        "category_id"
    ] = request.form.get(
        "category_id",
        "",
    ).strip()

    try:
        save_config(
            SO_PRICING_CONFIG_PATH,
            document,
        )

    except ValueError as exc:
        abort(
            400,
            str(exc),
        )

    flash(
        f"Ne BOM SKU "
        f"{match['sku']} "
        "kategorija išsaugota."
    )

    return redirect(
        url_for(
            "pricing_rules"
        )
        + "#non-bom-skus"
    )


@app.post("/upload")
def upload():
    file = request.files.get(
        "file"
    )

    if (
        file is None
        or not file.filename
    ):
        abort(
            400,
            "Nepasirinktas failas.",
        )

    filename = secure_filename(
        file.filename
    )

    if (
        Path(filename).suffix.lower()
        != ".xlsx"
    ):
        abort(
            400,
            "Leidžiamas tik .xlsx failas.",
        )

    timestamp = (
        datetime.now().strftime(
            "%Y%m%d_%H%M%S"
        )
    )

    target = (
        UPLOAD_DIR
        / (
            f"{timestamp}_"
            f"{filename}"
        )
    )

    file.save(
        target
    )

    return redirect(
        url_for("index")
    )


@app.post("/upload/chunked")
def start_chunked_upload():
    payload = request.get_json(silent=True) or {}
    filename = secure_filename(str(payload.get("filename", "")))
    try:
        size = int(payload.get("size", -1))
    except (TypeError, ValueError):
        size = -1

    if not filename or Path(filename).suffix.lower() != ".xlsx":
        abort(400, "Leidžiamas tik .xlsx failas.")
    if size <= 0 or size > MAX_UPLOAD_BYTES:
        abort(413, f"Failas turi būti ne didesnis kaip {SETTINGS.max_upload_mb} MB.")

    upload_id = uuid.uuid4().hex
    metadata_path, part_path = _chunk_upload_paths(upload_id)
    metadata = {"filename": filename, "size": size, "created_at": utc_now()}
    metadata_path.write_text(json.dumps(metadata), encoding="utf-8")
    part_path.touch(exist_ok=False)
    return jsonify({"upload_id": upload_id, "offset": 0})


@app.put("/upload/chunked/<upload_id>")
def append_chunked_upload(upload_id: str):
    metadata, _, part_path = _read_chunk_upload(upload_id)
    try:
        offset = int(request.headers.get("Upload-Offset", "-1"))
    except ValueError:
        offset = -1
    chunk = request.get_data(cache=False)
    current_size = part_path.stat().st_size

    if not chunk:
        abort(400, "Tuščia failo dalis.")
    if offset < 0 or offset + len(chunk) > metadata["size"]:
        abort(400, "Neteisingos failo dalies ribos.")

    if offset == current_size:
        with part_path.open("ab") as target:
            target.write(chunk)
            target.flush()
            os.fsync(target.fileno())
        current_size += len(chunk)
    elif offset + len(chunk) <= current_size:
        with part_path.open("rb") as target:
            target.seek(offset)
            if target.read(len(chunk)) != chunk:
                abort(409, "Failo dalis nesutampa su jau įkeltais duomenimis.")
    else:
        abort(409, f"Tikėtasi failo pozicijos {current_size}.")

    return jsonify({"offset": current_size, "size": metadata["size"]})


@app.post("/upload/chunked/<upload_id>/complete")
def complete_chunked_upload(upload_id: str):
    metadata, metadata_path, part_path = _read_chunk_upload(upload_id)
    actual_size = part_path.stat().st_size
    if actual_size != metadata["size"]:
        abort(409, f"Įkelta {actual_size} iš {metadata['size']} baitų.")

    expected_hash = str((request.get_json(silent=True) or {}).get("sha256", "")).lower()
    if expected_hash:
        digest = hashlib.sha256()
        with part_path.open("rb") as source:
            for block in iter(lambda: source.read(1024 * 1024), b""):
                digest.update(block)
        actual_hash = digest.hexdigest()
        if not secrets.compare_digest(actual_hash, expected_hash):
            abort(409, "Įkelto failo kontrolinė suma nesutampa.")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    target = UPLOAD_DIR / f"{timestamp}_{metadata['filename']}"
    part_path.replace(target)
    metadata_path.unlink(missing_ok=True)
    return jsonify({"filename": target.name, "size": actual_size})


@app.post(
    "/upload/production-dataset"
)
def upload_production_dataset():
    """
    Store the active Production dataset in persistent web state.

    The validator resolves this exact path, so the original
    uploaded filename must not be used as the runtime filename.
    """

    file = request.files.get(
        "file"
    )

    if (
        file is None
        or not file.filename
    ):
        abort(
            400,
            "Nepasirinktas Validated Dataset JSON failas.",
        )

    if (
        Path(
            secure_filename(
                file.filename
            )
        ).suffix.lower()
        != ".json"
    ):
        abort(
            400,
            "Leidžiamas tik .json failas.",
        )

    payload = file.read()

    if not payload:
        abort(
            400,
            "Įkeltas failas tuščias.",
        )

    try:
        document = json.loads(
            payload
        )

    except (
        UnicodeDecodeError,
        json.JSONDecodeError,
    ):
        abort(
            400,
            "Failas nėra korektiškas JSON.",
        )

    if not isinstance(
        document,
        (
            dict,
            list,
        ),
    ):
        abort(
            400,
            (
                "Validated Dataset JSON "
                "šaknis turi būti objektas "
                "arba sąrašas."
            ),
        )

    PRODUCTION_DATASET_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    temporary = (
        PRODUCTION_DATASET_PATH
        .with_suffix(
            ".json.tmp"
        )
    )

    temporary.write_bytes(
        payload
    )

    temporary.replace(
        PRODUCTION_DATASET_PATH
    )

    return redirect(
        url_for("index")
    )


@app.post(
    "/run/<action_key>"
)
def start_job(
    action_key: str,
):
    action = ACTIONS.get(
        action_key
    )

    if action is None:
        abort(404)

    prune_completed_jobs()

    upload_path = (
        latest_upload()
    )

    if (
        action[
            "requires_upload"
        ]
        and upload_path is None
    ):
        abort(
            409,
            (
                "Pirmiausia įkelkite "
                "Reform BOM .xlsx failą."
            ),
        )

    so_number = (
        request.form.get(
            "so_number",
            "",
        )
        .strip()
        .upper()
    )

    if (
        action.get(
            "requires_so_number"
        )
        and not so_number
    ):
        abort(
            400,
            "Įveskite SO numerį.",
        )

    job_id = (
        uuid.uuid4()
        .hex[:12]
    )

    with _jobs_lock:
        if _reserved_jobs:
            abort(
                409,
                (
                    "Kitas veiksmas jau vykdomas. "
                    "Palaukite, kol jis baigsis."
                ),
            )

        _reserved_jobs.add(
            job_id
        )

    job_dir = (
        RUN_DIR
        / job_id
    )

    job_dir.mkdir()

    write_job(
        job_dir,
        {
            "id": job_id,
            "action": action_key,
            "title": action["title"],
            "status": "QUEUED",
            "created_at": utc_now(),
            "upload": (
                upload_path.name
                if upload_path
                else None
            ),
            "so_number": (
                so_number
                or None
            ),
            "files": [],
        },
    )

    threading.Thread(
        target=run_job,
        args=(
            job_id,
            action_key,
            upload_path,
            so_number or None,
        ),
        daemon=True,
    ).start()

    return redirect(
        url_for(
            "job_detail",
            job_id=job_id,
        )
    )


@app.get(
    "/jobs/<job_id>"
)
def job_detail(
    job_id: str,
):
    job_dir = (
        RUN_DIR
        / secure_filename(
            job_id
        )
    )

    if not (
        job_dir
        / "job.json"
    ).exists():
        abort(404)

    job = read_job(
        job_dir
    )

    log_path = (
        job_dir
        / "run.log"
    )

    log = (
        log_path.read_text(
            encoding="utf-8",
            errors="replace",
        )
        if log_path.exists()
        else ""
    )

    return render_template(
        "job.html",
        job=job,
        log=log,
    )


@app.get(
    "/api/jobs/<job_id>"
)
def job_api(
    job_id: str,
):
    job_dir = (
        RUN_DIR
        / secure_filename(
            job_id
        )
    )

    if not (
        job_dir
        / "job.json"
    ).exists():
        abort(404)

    job = read_job(
        job_dir
    )

    log_path = (
        job_dir
        / "run.log"
    )

    job["log"] = (
        log_path.read_text(
            encoding="utf-8",
            errors="replace",
        )[-30000:]
        if log_path.exists()
        else ""
    )

    return jsonify(job)


@app.post(
    "/jobs/<job_id>/stop"
)
def stop_job(
    job_id: str,
):
    with _jobs_lock:
        process = (
            _active_processes.get(
                job_id
            )
        )

    if (
        process
        and process.poll()
        is None
    ):
        process.terminate()

    return redirect(
        url_for(
            "job_detail",
            job_id=job_id,
        )
    )


@app.get(
    "/jobs/<job_id>/files/<filename>"
)
def download(
    job_id: str,
    filename: str,
):
    path = (
        RUN_DIR
        / secure_filename(
            job_id
        )
        / "files"
        / secure_filename(
            filename
        )
    )

    if not path.is_file():
        abort(404)

    return send_file(
        path,
        as_attachment=True,
        download_name=path.name,
    )


@app.get("/health")
def health():
    return jsonify(
        status="ok"
    )


if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=int(
            os.getenv(
                "PORT",
                "8080",
            )
        ),
        debug=False,
    )
