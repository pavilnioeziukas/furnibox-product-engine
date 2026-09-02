"""Read-only audit of Production Odoo products absent from Target Dataset."""

from __future__ import annotations

import argparse
import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from bom_release.analyzer import load_latest_dataset_record, many2one_id
from config import load_settings
from odoo_client import OdooClient
from so_pricing_rules import load_config
from target_reconciliation.odoo_reader import read_production_snapshot


OUTPUT_FILE = "Odoo_Product_Lifecycle_Audit.xlsx"
JSON_FILE = "Odoo_Product_Lifecycle_Audit.json"

STATUS_KEEP = "KEEP"
STATUS_ARCHIVED = "ARCHIVED ALREADY"
STATUS_IN_USE = "DO NOT ARCHIVE - IN USE"
STATUS_PRICING = "REVIEW - PRICING CONFIG"
STATUS_CANDIDATE = "ARCHIVE CANDIDATE"


def text(value: Any) -> str:
    return str(value or "").strip()


def key(value: Any) -> str:
    return text(value).casefold()


def _product_ids(rows: list[dict[str, Any]]) -> set[int]:
    return {
        product_id
        for row in rows
        if (product_id := many2one_id(row.get("product_id")))
    }


def read_usage(client: OdooClient, product_ids: list[int]) -> dict[str, Any]:
    """Read only evidence that prevents safe archival recommendations."""
    if not product_ids:
        return {
            "stock": {}, "sales": set(), "purchases": set(),
            "manufacturing": set(), "raw_material": set(),
        }

    locations = client.search_read_all(
        "stock.location", [["usage", "=", "internal"]], ["id"]
    )
    location_ids = [int(row["id"]) for row in locations]
    quants = client.search_read_all(
        "stock.quant",
        [["product_id", "in", product_ids], ["location_id", "in", location_ids]],
        ["product_id", "quantity", "reserved_quantity"],
    ) if location_ids else []
    stock: defaultdict[int, float] = defaultdict(float)
    for row in quants:
        product_id = many2one_id(row.get("product_id"))
        if product_id:
            stock[product_id] += float(row.get("quantity") or 0)

    sales = client.search_read_all(
        "sale.order.line",
        [["product_id", "in", product_ids], ["state", "not in", ["cancel", "done"]]],
        ["product_id"],
    )
    purchases = client.search_read_all(
        "purchase.order.line",
        [["product_id", "in", product_ids], ["state", "not in", ["cancel", "done"]]],
        ["product_id"],
    )
    manufacturing = client.search_read_all(
        "mrp.production",
        [["product_id", "in", product_ids], ["state", "not in", ["cancel", "done"]]],
        ["product_id"],
    )
    raw_material = client.search_read_all(
        "stock.move",
        [
            ["product_id", "in", product_ids],
            ["raw_material_production_id", "!=", False],
            ["state", "not in", ["cancel", "done"]],
        ],
        ["product_id"],
    )
    return {
        "stock": dict(stock),
        "sales": _product_ids(sales),
        "purchases": _product_ids(purchases),
        "manufacturing": _product_ids(manufacturing),
        "raw_material": _product_ids(raw_material),
    }


def classify_product(
    product: dict[str, Any],
    *,
    in_target: bool,
    in_pricing: bool,
    active_bom_ids: list[int],
    used_by_boms: list[str],
    usage: dict[str, Any],
) -> dict[str, Any]:
    product_id = int(product["id"])
    stock_quantity = float(usage.get("stock", {}).get(product_id, 0))
    evidence = []
    if stock_quantity:
        evidence.append(f"Internal stock: {stock_quantity:g}")
    if product_id in usage.get("sales", set()):
        evidence.append("Open sales order line")
    if product_id in usage.get("purchases", set()):
        evidence.append("Open purchase order line")
    if product_id in usage.get("manufacturing", set()):
        evidence.append("Open manufacturing order")
    if product_id in usage.get("raw_material", set()):
        evidence.append("Required by open manufacturing order")
    if used_by_boms:
        evidence.append(f"Used by {len(used_by_boms)} active BOM(s)")

    if in_target:
        status = STATUS_KEEP
        decision = "Current Target Dataset product."
    elif not product.get("active", True):
        status = STATUS_ARCHIVED
        decision = "Product is already inactive in Production Odoo."
    elif evidence:
        status = STATUS_IN_USE
        decision = "Archival is unsafe while active usage exists."
    elif in_pricing:
        status = STATUS_PRICING
        decision = "Remove or confirm the pricing assignment before archival."
    else:
        status = STATUS_CANDIDATE
        decision = "No current Target Dataset membership or active usage was found."

    return {
        "sku": text(product.get("sku")),
        "name": text(product.get("name")),
        "status": status,
        "decision": decision,
        "active": bool(product.get("active", True)),
        "in_target_dataset": in_target,
        "in_pricing_config": in_pricing,
        "active_bom_ids": active_bom_ids,
        "used_by_active_boms": used_by_boms,
        "internal_stock": stock_quantity,
        "open_sales": product_id in usage.get("sales", set()),
        "open_purchases": product_id in usage.get("purchases", set()),
        "open_manufacturing": product_id in usage.get("manufacturing", set()),
        "required_by_open_manufacturing": product_id in usage.get("raw_material", set()),
        "evidence": "; ".join(evidence),
    }


def build_audit(
    dataset: dict[str, Any],
    production: dict[str, Any],
    pricing_config: dict[str, Any],
    usage: dict[str, Any],
) -> list[dict[str, Any]]:
    target_skus = {
        key(row.get("sku"))
        for row in dataset.get("product_catalog") or []
        if text(row.get("sku"))
    }
    pricing_skus = {
        key(row.get("sku"))
        for field in ("bom_products", "bom_skus", "non_bom_skus")
        for row in pricing_config.get(field) or []
        if text(row.get("sku"))
    }
    active_boms: defaultdict[str, list[int]] = defaultdict(list)
    used_by: defaultdict[str, set[str]] = defaultdict(set)
    for bom in production.get("boms") or []:
        if not bom.get("active", True) or not text(bom.get("sku")):
            continue
        parent = text(bom.get("sku"))
        active_boms[key(parent)].append(int(bom["id"]))
        for component in bom.get("components") or []:
            component_sku = text(component.get("component_sku"))
            if component_sku:
                used_by[key(component_sku)].add(parent)

    result = []
    for product in production.get("products") or []:
        sku = text(product.get("sku"))
        if not sku:
            continue
        sku_key = key(sku)
        result.append(classify_product(
            product,
            in_target=sku_key in target_skus,
            in_pricing=sku_key in pricing_skus,
            active_bom_ids=sorted(active_boms.get(sku_key, [])),
            used_by_boms=sorted(used_by.get(sku_key, set())),
            usage=usage,
        ))
    return sorted(result, key=lambda row: (row["status"], row["sku"].casefold()))


def write_workbook(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "SUMMARY"
    summary.append(["Status", "Count"])
    counts = Counter(row["status"] for row in rows)
    for status in (STATUS_KEEP, STATUS_ARCHIVED, STATUS_IN_USE, STATUS_PRICING, STATUS_CANDIDATE):
        summary.append([status, counts.get(status, 0)])
    summary.append([])
    summary.append(["Control", "READ ONLY - Odoo data was not changed"])

    columns = [
        ("SKU", "sku"), ("Name", "name"), ("Status", "status"),
        ("Decision", "decision"), ("Active", "active"),
        ("In Target Dataset", "in_target_dataset"),
        ("In Pricing Config", "in_pricing_config"),
        ("Active BOM IDs", "active_bom_ids"),
        ("Used By Active BOMs", "used_by_active_boms"),
        ("Internal Stock", "internal_stock"), ("Evidence", "evidence"),
    ]
    sheets = {
        "ARCHIVE CANDIDATES": {STATUS_CANDIDATE},
        "IN USE": {STATUS_IN_USE},
        "PRICING CONFIG REVIEW": {STATUS_PRICING},
        "EVIDENCE": {STATUS_KEEP, STATUS_ARCHIVED, STATUS_IN_USE, STATUS_PRICING, STATUS_CANDIDATE},
    }
    thin = Side(style="thin", color="D6DED8")
    fills = {
        STATUS_KEEP: "E2F0D9", STATUS_ARCHIVED: "D9EAF7",
        STATUS_IN_USE: "F4CCCC", STATUS_PRICING: "FFF2CC",
        STATUS_CANDIDATE: "FCE5CD",
    }
    for name, statuses in sheets.items():
        sheet = workbook.create_sheet(name)
        sheet.append([label for label, _ in columns])
        for row in rows:
            if row["status"] not in statuses:
                continue
            sheet.append([
                ", ".join(str(value) for value in row[field])
                if isinstance(row[field], list) else row[field]
                for _, field in columns
            ])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="174C35")
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row[0].row > 1:
                row[2].fill = PatternFill("solid", fgColor=fills.get(text(row[2].value), "FFFFFF"))
        widths = [28, 42, 30, 52, 10, 18, 18, 20, 42, 16, 58]
        for index, width in enumerate(widths, 1):
            sheet.column_dimensions[chr(64 + index)].width = width

    summary.sheet_view.showGridLines = False
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 42
    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="174C35")
    for row in summary.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Read-only Production Odoo product lifecycle audit.")
    parser.add_argument("--dataset", required=True, type=Path)
    parser.add_argument("--pricing-config", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    args = parser.parse_args()

    settings = load_settings()
    if "stage" in settings.url.lower():
        raise PermissionError("Auditui būtina Production Odoo aplinka.")
    dataset, _ = load_latest_dataset_record(args.dataset)
    pricing_config = load_config(args.pricing_config)
    client = OdooClient(settings)
    client.authenticate()
    production = read_production_snapshot(client)
    target_skus = {
        key(row.get("sku"))
        for row in dataset.get("product_catalog") or []
        if text(row.get("sku"))
    }
    ids = [
        int(row["id"])
        for row in production.get("products") or []
        if text(row.get("sku")) and key(row.get("sku")) not in target_skus
    ]
    usage = read_usage(client, ids)
    rows = build_audit(dataset, production, pricing_config, usage)
    output = args.output_dir / OUTPUT_FILE
    write_workbook(output, rows)
    payload = {
        "environment": "production",
        "odoo_changed": False,
        "statuses": dict(Counter(row["status"] for row in rows)),
        "rows": rows,
    }
    (args.output_dir / JSON_FILE).write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(payload["statuses"], ensure_ascii=False, indent=2))
    print("READ-ONLY rezultatas:", output.resolve())
    print("Odoo duomenys nepakeisti.")


if __name__ == "__main__":
    main()
