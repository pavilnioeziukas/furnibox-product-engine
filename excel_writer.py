from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)

def write_rows(ws, rows):
    if not rows:
        ws.append(["No data"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in ws.iter_rows(min_col=idx, max_col=idx, min_row=2, max_row=min(ws.max_row, 200)):
            val = row[0].value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 45)

def write_snapshot(path, products, boms, bom_lines, purchase_prices, metadata):
    wb = Workbook()
    wb.remove(wb.active)

    info = wb.create_sheet("INFO")
    for row in [
        ("Generated", datetime.now().isoformat(sep=" ", timespec="seconds")),
        ("Odoo URL", metadata["url"]),
        ("Database", metadata["db"]),
        ("User", metadata["login"]),
        ("Odoo UID", metadata["uid"]),
        ("Products", len(products)),
        ("BOM", len(boms)),
        ("BOM Lines", len(bom_lines)),
        ("Last Purchase Prices", len(purchase_prices)),
    ]:
        info.append(row)
    info.column_dimensions["A"].width = 24
    info.column_dimensions["B"].width = 40

    for title, rows in [
        ("ODOO PRODUCTS", products),
        ("ODOO BOM", boms),
        ("ODOO BOM LINES", bom_lines),
        ("LAST PURCHASE PRICES", purchase_prices),
    ]:
        ws = wb.create_sheet(title)
        write_rows(ws, rows)

    wb.save(path)