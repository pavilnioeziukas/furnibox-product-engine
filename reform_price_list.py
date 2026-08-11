"""Sujungia Cabinet Parts ir kitų komponentų kainas į Reform kainoraštį.

Scenarijus skaito dviejų jau patikrintų generatorių rezultatus ir sukuria vieną
kontrolinį Excel failą. Odoo duomenų nekeičia.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


COMPONENT_FILE = "Last_Purchase_Prices.xlsx"
CABINET_PART_FILE = "Existing_and_New_Cabinet_Parts_Prices.xlsx"
OUTPUT_FILE = "Reform_Final_Prices.xlsx"
REFORM_MARKUP_FACTOR = 1.05


def _headers(sheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value not in (None, "")
    }


def _required(headers: dict[str, int], name: str, sheet_name: str) -> int:
    if name not in headers:
        raise ValueError(f"Lape '{sheet_name}' nerastas stulpelis: {name}")
    return headers[name]


def load_component_prices(path: Path) -> list[dict]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    if "COMPONENT PRICES" not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Faile '{path.name}' nerastas lapas COMPONENT PRICES.")
    sheet = workbook["COMPONENT PRICES"]
    headers = _headers(sheet)
    columns = {
        name: _required(headers, name, sheet.title)
        for name in (
            "Internal Reference", "Name", "Vendor", "Real Purchase Price",
            "Adjusted Purchase Price", "Markup Factor", "Reform Price",
        )
    }
    adjustments = {}
    if "TAMARA ADJUSTMENTS" not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Faile '{path.name}' nerastas lapas TAMARA ADJUSTMENTS.")
    adjustment_sheet = workbook["TAMARA ADJUSTMENTS"]
    adjustment_headers = _headers(adjustment_sheet)
    adjustment_sku = _required(
        adjustment_headers, "Internal Reference", adjustment_sheet.title
    )
    adjustment_price = _required(
        adjustment_headers, "Adjusted Purchase Price", adjustment_sheet.title
    )
    for values in adjustment_sheet.iter_rows(min_row=2, values_only=True):
        sku = str(values[adjustment_sku - 1] or "").strip()
        if not sku:
            continue
        key = sku.casefold()
        if key in adjustments:
            workbook.close()
            raise ValueError(f"Tamaros korekcijose kartojasi SKU: {sku}")
        adjustments[key] = values[adjustment_price - 1]

    rows = []
    seen = set()
    for values in sheet.iter_rows(min_row=2, values_only=True):
        sku = str(values[columns["Internal Reference"] - 1] or "").strip()
        if not sku:
            continue
        key = sku.casefold()
        if key in seen:
            workbook.close()
            raise ValueError(f"Komponentų kainose kartojasi SKU: {sku}")
        seen.add(key)
        row = {name: values[column - 1] for name, column in columns.items()}
        row["Adjusted Purchase Price"] = adjustments.get(
            key, row["Real Purchase Price"]
        )
        rows.append(row)
    workbook.close()
    return rows


def load_cabinet_part_prices(path: Path) -> list[dict]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    if "CABINET PART PRICES" not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Faile '{path.name}' nerastas lapas CABINET PART PRICES.")
    sheet = workbook["CABINET PART PRICES"]
    headers = _headers(sheet)
    columns = {
        name: _required(headers, name, sheet.title)
        for name in (
            "Internal Reference", "Furnix Unit Cost",
            "Furnix Sales Price to Furnibox", "Product Status", "BOM Source",
        )
    }
    rows = []
    seen = set()
    for values in sheet.iter_rows(min_row=2, values_only=True):
        sku = str(values[columns["Internal Reference"] - 1] or "").strip()
        if not sku:
            continue
        key = sku.casefold()
        if key in seen:
            workbook.close()
            raise ValueError(f"Cabinet Parts kainose kartojasi SKU: {sku}")
        seen.add(key)
        rows.append({name: values[column - 1] for name, column in columns.items()})
    workbook.close()
    return rows


def build_reform_price_list(
    component_path: Path,
    cabinet_part_path: Path,
    output_path: Path,
    reform_markup_factor: float = REFORM_MARKUP_FACTOR,
) -> tuple[int, int]:
    if reform_markup_factor < 1:
        raise ValueError("Reform antkainio koeficientas negali būti mažesnis už 1.")

    components = load_component_prices(component_path)
    cabinet_parts = load_cabinet_part_prices(cabinet_part_path)
    cabinet_skus = {
        str(row["Internal Reference"]).strip().casefold()
        for row in cabinet_parts
    }
    overlaps = [
        row for row in components
        if str(row["Internal Reference"]).strip().casefold() in cabinet_skus
    ]

    workbook = Workbook()
    prices = workbook.active
    prices.title = "REFORM PRICE LIST"
    prices.append([
        "Internal Reference", "Name", "Price Source", "Vendor / Supply Source",
        "Real Furnibox Purchase Price", "Adjusted Furnibox Purchase Price",
        "Reform Markup Factor", "Reform Purchase Price", "Status / BOM Source",
    ])

    for row in sorted(cabinet_parts, key=lambda item: str(item["Internal Reference"]).casefold()):
        purchase_price = row["Furnix Sales Price to Furnibox"]
        prices.append([
            row["Internal Reference"], "", "CABINET PART CALCULATION", "Furnix",
            purchase_price, purchase_price, None,
            purchase_price, f'{row["Product Status"]} / {row["BOM Source"]}',
        ])

    for row in sorted(components, key=lambda item: str(item["Internal Reference"]).casefold()):
        sku = str(row["Internal Reference"]).strip()
        if sku.casefold() in cabinet_skus:
            continue
        prices.append([
            sku, row["Name"], "LAST PURCHASE PRICE", row["Vendor"],
            row["Real Purchase Price"], row["Adjusted Purchase Price"],
            row["Markup Factor"], None, "",
        ])
        excel_row = prices.max_row
        prices.cell(excel_row, 8).value = f"=F{excel_row}*G{excel_row}"

    dark_fill = PatternFill("solid", fgColor="1F4E78")
    white_font = Font(color="FFFFFF", bold=True)
    green_fill = PatternFill("solid", fgColor="E2F0D9")
    for cell in prices[1]:
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    prices.row_dimensions[1].height = 38
    prices.freeze_panes = "A2"
    prices.auto_filter.ref = prices.dimensions
    prices.sheet_view.showGridLines = False
    widths = [32, 45, 27, 35, 29, 31, 22, 25, 31]
    for index, width in enumerate(widths, start=1):
        prices.column_dimensions[prices.cell(1, index).column_letter].width = width
    for row_number in range(2, prices.max_row + 1):
        for column in (5, 6, 8):
            prices.cell(row_number, column).number_format = '0.0000 [$€-x-euro2]'
        prices.cell(row_number, 7).number_format = "0.00"
        prices.cell(row_number, 8).fill = green_fill

    diagnostics = workbook.create_sheet("DIAGNOSTICS")
    diagnostics.append(["Type", "Internal Reference", "Decision"])
    for row in overlaps:
        diagnostics.append([
            "SKU IN BOTH SOURCES", row["Internal Reference"],
            "Used CABINET PART CALCULATION; component purchase-history row excluded.",
        ])
    for cell in diagnostics[1]:
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.font = white_font
    diagnostics.freeze_panes = "A2"
    diagnostics.auto_filter.ref = diagnostics.dimensions
    diagnostics.column_dimensions["A"].width = 28
    diagnostics.column_dimensions["B"].width = 34
    diagnostics.column_dimensions["C"].width = 78
    diagnostics.sheet_view.showGridLines = False

    info = workbook.create_sheet("INFO")
    for row in [
        ("Generated", datetime.now().isoformat(sep=" ", timespec="seconds")),
        ("Component source", str(component_path)),
        ("Cabinet Part source", str(cabinet_part_path)),
        ("Cabinet Part rule", "Furnix Sales Price to Furnibox × Reform markup factor"),
        ("Other component rule", "Tamara Adjusted Purchase Price × source markup factor"),
        ("Default Reform markup factor for Cabinet Parts", reform_markup_factor),
        ("Final unique products", prices.max_row - 1),
        ("Overlapping SKU", len(overlaps)),
        ("Odoo changed", "NO"),
    ]:
        info.append(row)
    for cell in info[1]:
        cell.fill = dark_fill
        cell.font = white_font
    info.column_dimensions["A"].width = 48
    info.column_dimensions["B"].width = 100
    info.sheet_view.showGridLines = False

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return prices.max_row - 1, len(overlaps)


def main() -> None:
    from config import load_settings

    settings = load_settings()
    component_path = settings.output_dir / COMPONENT_FILE
    cabinet_part_path = settings.output_dir / CABINET_PART_FILE
    for path in (component_path, cabinet_part_path):
        if not path.exists():
            raise FileNotFoundError(f"Nerastas šaltinio failas: {path}")
    output_path = settings.output_dir / OUTPUT_FILE
    products, overlaps = build_reform_price_list(
        component_path, cabinet_part_path, output_path
    )
    print("GALUTINIS REFORM KAINŲ FAILAS SUKURTAS")
    print("Failas:", output_path)
    print("Unikalūs produktai:", products)
    print("SKU abiejuose šaltiniuose:", overlaps)
    print("Odoo duomenys nepakeisti.")


if __name__ == "__main__":
    main()
