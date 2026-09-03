from pathlib import Path
import json
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from refresh_reform_pricing import (
    read_pricing_status,
    read_current_sales_prices,
    refresh,
    write_blocker_report,
    write_complete_only_price_workbook,
    write_furnibox_purchase_prices,
    write_furnix_parts_price_review,
    write_pricing_chain_audit,
    write_tamara_product_classification_review,
)


class RefreshReformPricingTests(unittest.TestCase):
    def test_refresh_passes_generated_target_dataset_to_pricing(self):
        with tempfile.TemporaryDirectory() as directory:
            base = Path(directory)
            bom = base / "BOM_for Furnibox_v10.xlsx"
            bom.write_bytes(b"source")
            output = base / "result"
            calls = []

            def fake_run_step(title, *args):
                calls.append((title, args))
                if "reconciliation" in title.lower():
                    reconciliation_path = Path(args[args.index("--output") + 1])
                    reconciliation_path.write_text(
                        json.dumps({
                            "mode": "READ_ONLY",
                            "environment": "production",
                            "summary": {
                                "product_statuses": {"BLOCKED": 4},
                                "bom_statuses": {"BLOCKED": 46},
                            },
                        }),
                        encoding="utf-8",
                    )

            with (
                patch(
                    "refresh_reform_pricing.run_step",
                    side_effect=fake_run_step,
                ),
                patch(
                    "refresh_reform_pricing.read_pricing_status",
                    return_value=({"COMPLETE": 1}, []),
                ),
                patch("refresh_reform_pricing.write_furnibox_purchase_prices"),
                patch("refresh_reform_pricing.write_furnix_parts_price_review"),
                patch("refresh_reform_pricing.write_pricing_chain_audit"),
                patch("refresh_reform_pricing.shutil.copy2"),
            ):
                self.assertEqual(refresh(bom, output), 0)

            self.assertIn("Pilnas Furnibox Target Dataset", calls[0][0])
            self.assertIn("--local-only", calls[0][1])
            self.assertIn("reconciliation", calls[1][0].lower())
            pricing_args = calls[-1][1]
            self.assertIn("--dataset", pricing_args)
            dataset_index = pricing_args.index("--dataset") + 1
            self.assertEqual(
                Path(pricing_args[dataset_index]),
                output / "Furnibox_Target_Dataset.json",
            )
            result = json.loads(
                (output / "Reform_Pricing_Result.json").read_text(encoding="utf-8")
            )
            self.assertEqual(
                result["target_reconciliation"]["bom_statuses"]["BLOCKED"],
                46,
            )

    def test_blocked_refresh_publishes_safe_complete_only_outputs(self):
        with tempfile.TemporaryDirectory() as directory:
            base = Path(directory)
            bom = base / "BOM_for Furnibox_v10.xlsx"
            bom.write_bytes(b"source")
            output = base / "result"

            def fake_run_step(title, *args):
                if "reconciliation" in title.lower():
                    reconciliation_path = Path(args[args.index("--output") + 1])
                    reconciliation_path.write_text(
                        json.dumps({
                            "mode": "READ_ONLY",
                            "environment": "production",
                            "summary": {},
                        }),
                        encoding="utf-8",
                    )

            blocked = [{
                "sku": "BAD-BOM",
                "position_type": "BOM",
                "status": "BLOCKED",
                "issues": "Missing component price: PART",
            }]
            with (
                patch("refresh_reform_pricing.run_step", side_effect=fake_run_step),
                patch(
                    "refresh_reform_pricing.read_pricing_status",
                    return_value=({"COMPLETE": 10, "BLOCKED": 1}, blocked),
                ),
                patch("refresh_reform_pricing.write_blocker_report"),
                patch(
                    "refresh_reform_pricing.write_complete_only_price_workbook"
                ) as partial_writer,
                patch(
                    "refresh_reform_pricing.write_furnibox_purchase_prices"
                ) as purchase_writer,
                patch("refresh_reform_pricing.write_furnix_parts_price_review"),
                patch("refresh_reform_pricing.write_pricing_chain_audit"),
                patch("refresh_reform_pricing.shutil.copy2") as copy_file,
            ):
                self.assertEqual(refresh(bom, output), 2)

            partial_writer.assert_called_once()
            purchase_writer.assert_called_once()
            self.assertEqual(copy_file.call_count, 2)
            copied_outputs = {
                call.args[1].name
                for call in copy_file.call_args_list
            }
            self.assertEqual(copied_outputs, {
                "Reform_Pricing_Source.xlsx",
                "Cabinet_Parts_Pricing.xlsx",
            })
            result = json.loads(
                (output / "Reform_Pricing_Result.json").read_text(encoding="utf-8")
            )
            self.assertFalse(result["released"])
            self.assertTrue(result["partial_released"])
            self.assertEqual(result["excluded_blocked_count"], 1)
            self.assertEqual(
                result["partial_file"],
                "Reform_SO_Line_Prices_COMPLETE_ONLY.xlsx",
            )

    def test_pricing_chain_audit_checks_primary_rollup(self):
        with tempfile.TemporaryDirectory() as directory:
            base = Path(directory)
            source = base / "pricing.xlsx"
            cabinet_parts = base / "cabinet_parts.xlsx"
            output = base / "audit.xlsx"

            workbook = Workbook()
            prices = workbook.active
            prices.title = "SO LINE PRICES"
            prices.append([
                "SKU", "Product Category", "Component / Purchase Cost",
                "Pricing Add-ons Total", "Adjustment Amount",
                "Final Reform SO Unit Price", "Status", "Issues",
            ])
            prices.append([
                "FPACK-EU-CAB01-BAS001", "PREPACK", 10, 1, -0.07,
                10.93, "COMPLETE", "",
            ])
            components = workbook.create_sheet("BOM COMPONENT COSTS")
            components.append([
                "Top BOM SKU", "Purchased Component SKU", "Purchase Unit Price",
                "Component Cost", "Cost Source",
            ])
            components.append([
                "FPACK-EU-CAB01-BAS001", "PART-1", 5, 10,
                "CABINET PART CALCULATION",
            ])
            workbook.save(source)

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "CABINET PART PRICES"
            sheet.append(["Internal Reference", "Furnix Sales Price to Furnibox"])
            sheet.append(["PART-1", 5])
            workbook.save(cabinet_parts)

            write_pricing_chain_audit(
                source,
                cabinet_parts,
                output,
                current_sales_prices={"FPACK-EU-CAB01-BAS001": 0.01},
            )

            result = load_workbook(output, data_only=True)
            summary = result["SUMMARY"]
            self.assertEqual(summary["A2"].value, "CABINET PARTS")
            self.assertEqual(summary["B2"].value, 1)
            detail = result["PRIMARY CHAIN"]
            headers = {cell.value: cell.column for cell in detail[1]}
            self.assertEqual(detail.cell(2, headers["Audit Status"]).value, "PASS")
            self.assertEqual(detail.cell(2, headers["Cabinet Part Lines"]).value, 1)
            self.assertEqual(
                detail.cell(2, headers["Price Review"]).value,
                "SO PRICE CORRECTION REQUIRED",
            )

    def test_furnix_parts_review_uses_lpp_not_odoo_list_price(self):
        with tempfile.TemporaryDirectory() as directory:
            base = Path(directory)
            cabinet_parts = base / "cabinet_parts.xlsx"
            last_purchases = base / "last_purchases.xlsx"
            output = base / "review.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "CABINET PART PRICES"
            sheet.append([
                "Internal Reference", "Odoo Product ID", "Odoo Active",
                "Furnix Unit Cost", "Furnix Markup, %",
                "Furnix Sales Price to Furnibox", "Product Status", "BOM Source",
            ])
            sheet.append([
                "PART-1", 101, "YES", 8.0, 0.0, 8.0, "EXISTING", "EXISTING",
            ])
            workbook.save(cabinet_parts)

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "COMPONENT PRICES"
            sheet.append([
                "Internal Reference", "Real Purchase Price", "Vendor",
                "Purchase Order", "Order Date",
            ])
            sheet.append(["PART-1", 7.5, "Furnix", "PO001", "2026-08-20"])
            workbook.save(last_purchases)

            write_furnix_parts_price_review(
                cabinet_parts,
                last_purchases,
                output,
                current_sales_prices={"PART-1": 0.01},
            )

            result = load_workbook(output, data_only=True, read_only=True)
            detail = result["FURNIX PARTS REVIEW"]
            headers = {cell.value: cell.column for cell in detail[1]}
            self.assertEqual(
                detail.cell(2, headers["Production Last Purchase Price"]).value,
                7.5,
            )
            self.assertAlmostEqual(
                detail.cell(2, headers["Purchase Price Change"]).value,
                0.5,
            )
            self.assertEqual(
                detail.cell(2, headers["Current Odoo SO/List Price"]).value,
                0.01,
            )
            self.assertEqual(
                detail.cell(2, headers["SO Price Status"]).value,
                "SO PRICE CORRECTION REQUIRED",
            )
            result.close()

    def test_current_sales_prices_are_loaded_from_read_only_reconciliation(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "reconciliation.json"
            path.write_text(json.dumps({
                "current_sales_prices": [
                    {"sku": "CABINET-1", "price": 100.0},
                    {"sku": "APACK-1", "price": 0.01},
                    {"sku": "NO-PRICE", "price": None},
                ],
            }), encoding="utf-8")

            self.assertEqual(read_current_sales_prices(path), {
                "CABINET-1": 100.0,
                "APACK-1": 0.01,
            })

    def make_result(self, path: Path, rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SO LINE PRICES"
        sheet.append(["SKU", "Position Type", "Status", "Issues"])
        for row in rows:
            sheet.append(row)
        workbook.save(path)

    def test_complete_result_passes_release_gate(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "result.xlsx"
            self.make_result(path, [["SKU-1", "BOM", "COMPLETE", ""]])
            statuses, blocked = read_pricing_status(path)
            self.assertEqual(statuses["COMPLETE"], 1)
            self.assertEqual(blocked, [])

    def test_blocked_result_creates_report_and_fails_gate(self):
        with tempfile.TemporaryDirectory() as directory:
            base = Path(directory)
            result = base / "result.xlsx"
            report = base / "blocked.xlsx"
            self.make_result(
                result,
                [["SKU-1", "BOM", "BLOCKED", "Missing component price: PART-1"]],
            )
            statuses, blocked = read_pricing_status(result)
            write_blocker_report(report, statuses, blocked)
            self.assertEqual(len(blocked), 1)
            self.assertTrue(report.exists())

    def test_complete_only_output_excludes_blocked_from_every_price_sheet(self):
        with tempfile.TemporaryDirectory() as directory:
            base = Path(directory)
            source = base / "source.xlsx"
            destination = base / "complete_only.xlsx"
            workbook = Workbook()
            prices = workbook.active
            prices.title = "SO LINE PRICES"
            prices.append(["SKU", "Status"])
            prices.append(["GOOD-BOM", "COMPLETE"])
            prices.append(["BAD-BOM", "BLOCKED"])
            costs = workbook.create_sheet("BOM COMPONENT COSTS")
            costs.append(["Top BOM SKU", "Status"])
            costs.append(["GOOD-BOM", "COMPLETE"])
            costs.append(["BAD-BOM", "BLOCKED"])
            breakdown = workbook.create_sheet("BOM CATEGORY BREAKDOWN")
            breakdown.append(["Top SKU", "Category"])
            breakdown.append(["GOOD-BOM", "GOOD"])
            breakdown.append(["BAD-BOM", "BAD"])
            rules = workbook.create_sheet("CATEGORY RULES")
            rules.append(["Category ID"])
            non_bom = workbook.create_sheet("NON-BOM RULES")
            non_bom.append(["SKU", "Status"])
            non_bom.append(["GOOD-NON", "COMPLETE"])
            non_bom.append(["BAD-NON", "BLOCKED"])
            diagnostics = workbook.create_sheet("DIAGNOSTICS")
            diagnostics.append(["Position Type", "SKU", "Status", "Issues"])
            diagnostics.append(["BOM", "BAD-BOM", "BLOCKED", "Missing"])
            info = workbook.create_sheet("INFO")
            info.append(["Purpose", "Final prices"])
            workbook.save(source)

            write_complete_only_price_workbook(
                source,
                destination,
                [
                    {"sku": "BAD-BOM"},
                    {"sku": "BAD-NON"},
                ],
            )

            result = load_workbook(destination, data_only=True, read_only=True)
            self.assertEqual(
                [row[0] for row in result["SO LINE PRICES"].iter_rows(
                    min_row=2, values_only=True
                )],
                ["GOOD-BOM"],
            )
            self.assertEqual(
                [row[0] for row in result["BOM COMPONENT COSTS"].iter_rows(
                    min_row=2, values_only=True
                )],
                ["GOOD-BOM"],
            )
            self.assertEqual(
                [row[0] for row in result["NON-BOM RULES"].iter_rows(
                    min_row=2, values_only=True
                )],
                ["GOOD-NON"],
            )
            self.assertIn("EXCLUDED BLOCKED", result.sheetnames)
            self.assertEqual(result["INFO"]["B1"].value, "COMPLETE POSITIONS ONLY")
            self.assertEqual(result["INFO"]["B3"].value, 2)
            result.close()

    def test_publishes_tamara_purchase_price_as_explicit_column(self):
        with tempfile.TemporaryDirectory() as directory:
            base = Path(directory)
            source = base / "source.xlsx"
            destination = base / "purchase.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "REFORM PRICE LIST"
            sheet.append([
                "Internal Reference", "Name", "Price Source",
                "Vendor / Supply Source", "Real Furnibox Purchase Price",
                "Adjusted Furnibox Purchase Price", "Reform Markup Factor",
                "Reform Purchase Price", "Status / BOM Source",
            ])
            sheet.append([
                "PART-1", "Part", "PO", "Vendor", 10, 12, 1.05,
                "=F2*G2", "OK",
            ])
            workbook.save(source)

            write_furnibox_purchase_prices(source, destination)
            published = load_workbook(destination, data_only=True, read_only=True)
            self.assertEqual(
                published.sheetnames,
                ["CABINET PARTS", "COMPONENTS", "INFO"],
            )
            result = published["COMPONENTS"]
            header = [cell.value for cell in result[1]]
            self.assertEqual(header, [
                "Internal Reference", "Name", "Price Source",
                "Vendor / Supply Source", "Real Furnibox Purchase Price",
                "Furnibox (Tamara) Purchase Price", "Reform Markup Factor",
                "Reform Purchase Price", "Status / BOM Source",
            ])
            self.assertEqual(result.cell(2, 6).value, 12)
            self.assertEqual(result.cell(2, 7).value, 1.05)
            self.assertEqual(result.cell(2, 8).value, 12.6)
            published.close()

    def test_splits_cabinet_parts_from_bought_components(self):
        with tempfile.TemporaryDirectory() as directory:
            base = Path(directory)
            source = base / "source.xlsx"
            destination = base / "purchase.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "REFORM PRICE LIST"
            sheet.append([
                "Internal Reference", "Name", "Price Source",
                "Vendor / Supply Source", "Real Furnibox Purchase Price",
                "Adjusted Furnibox Purchase Price", "Reform Markup Factor",
                "Reform Purchase Price", "Status / BOM Source",
            ])
            sheet.append([
                "PART-FURNIX", "Panel", "CABINET PART CALCULATION", "Furnix",
                5, 5, None, 5, "EXISTING",
            ])
            sheet.append([
                "COMP-1", "Hinge", "LAST PURCHASE PRICE", "Vendor",
                2, 2, 1, 2, "",
            ])
            workbook.save(source)

            write_furnibox_purchase_prices(source, destination)
            published = load_workbook(destination, data_only=True, read_only=True)
            self.assertEqual(published["CABINET PARTS"]["A2"].value, "PART-FURNIX")
            self.assertEqual(published["COMPONENTS"]["A2"].value, "COMP-1")
            published.close()

    def test_exports_full_tamara_product_classification_review(self):
        with tempfile.TemporaryDirectory() as directory:
            base = Path(directory)
            dataset = base / "dataset.json"
            reconciliation = base / "reconciliation.json"
            destination = base / "review.xlsx"
            dataset.write_text(json.dumps({
                "products": [{
                    "sku": "MADE-1",
                    "bom_type": "MANUFACTURE",
                    "product_type": "ACCESSORIES",
                }],
                "product_catalog": [
                    {
                        "sku": "MADE-1", "origin": "REFORM",
                        "role": "BOM PARENT", "has_bom": True,
                        "name_2": "Made",
                    },
                    {
                        "sku": "BUY-1", "origin": "REFORM",
                        "role": "NON-BOM COMPONENT", "has_bom": False,
                        "name_2": "Bought",
                    },
                ],
            }), encoding="utf-8")
            reconciliation.write_text(json.dumps({
                "products": [
                    {"sku": "MADE-1", "status": "PRODUCT UNCHANGED"},
                    {"sku": "BUY-1", "status": "UPDATE PRODUCT"},
                ],
            }), encoding="utf-8")

            write_tamara_product_classification_review(
                dataset,
                reconciliation,
                destination,
            )
            workbook = load_workbook(destination, data_only=True)
            review = workbook["PRODUCT REVIEW"]
            rows = {
                review.cell(row, 1).value: row
                for row in range(2, review.max_row + 1)
            }
            self.assertEqual(review.cell(rows["MADE-1"], 5).value, "MANUFACTURE")
            self.assertEqual(review.cell(rows["MADE-1"], 9).value, "GAMINAMAS")
            self.assertEqual(review.cell(rows["BUY-1"], 9).value, "PERKAMAS")
            self.assertTrue(review.data_validations.count)
            workbook.close()


if __name__ == "__main__":
    unittest.main()
