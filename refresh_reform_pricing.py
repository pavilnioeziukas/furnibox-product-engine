"""Run the complete read-only Reform pricing refresh pipeline."""

from __future__ import annotations

import argparse
import json
import os
import shutil
import subprocess
import sys
import tempfile
from collections import Counter
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from pricing_control import enrich_pricing_workbook


BASE_DIR = Path(__file__).resolve().parent
PRODUCTION_DIR = BASE_DIR / "output" / "production"
RULES_PATH = BASE_DIR / "web_state" / "shared_data" / "so_pricing_rules.json"


def run_step(title: str, *arguments: str) -> None:
    print(f"\n=== {title} ===", flush=True)
    result = subprocess.run(
        [sys.executable, "-u", *arguments],
        cwd=BASE_DIR,
        env={
            **os.environ,
            "PYTHONUTF8": "1",
            "PYTHONIOENCODING": "utf-8",
            "FURNIBOX_ENVIRONMENT": "PRODUCTION",
        },
        check=False,
    )
    if result.returncode:
        raise RuntimeError(f"Žingsnis nepavyko: {title} (kodas {result.returncode})")


def read_pricing_status(path: Path) -> tuple[Counter, list[dict]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["SO LINE PRICES"]
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    columns = {value: index for index, value in enumerate(header)}
    statuses: Counter = Counter()
    blocked = []
    for row in rows:
        status = str(row[columns["Status"]] or "")
        statuses[status] += 1
        if status != "COMPLETE":
            blocked.append({
                "sku": row[columns["SKU"]],
                "position_type": row[columns["Position Type"]],
                "status": status,
                "issues": row[columns["Issues"]],
            })
    workbook.close()
    return statuses, blocked


def write_blocker_report(path: Path, statuses: Counter, blocked: list[dict]) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "SUMMARY"
    summary.append(["Status", "Count"])
    for status, count in sorted(statuses.items()):
        summary.append([status, count])
    summary.append([])
    summary.append(["Decision", "FULL FINAL FILE NOT RELEASED"])
    summary.append(["Reason", "Pricing contains BLOCKED positions"])
    summary.append([
        "Safe output",
        "Reform_SO_Line_Prices_COMPLETE_ONLY.xlsx excludes every BLOCKED SKU",
    ])

    details = workbook.create_sheet("BLOCKERS")
    details.append(["SKU", "Position Type", "Status", "Issues"])
    for row in blocked:
        details.append([
            row["sku"], row["position_type"], row["status"], row["issues"]
        ])
    details.freeze_panes = "A2"
    details.auto_filter.ref = details.dimensions
    details.column_dimensions["A"].width = 38
    details.column_dimensions["B"].width = 18
    details.column_dimensions["C"].width = 15
    details.column_dimensions["D"].width = 100
    workbook.save(path)


def write_complete_only_price_workbook(
    source: Path,
    destination: Path,
    blocked: list[dict],
) -> None:
    """Publish a safe partial price list without any blocked SKU."""
    workbook = load_workbook(source)
    blocked_skus = {
        str(row.get("sku") or "").strip().casefold()
        for row in blocked
        if str(row.get("sku") or "").strip()
    }

    def remove_rows(sheet_name: str, column_name: str, remove_value) -> None:
        sheet = workbook[sheet_name]
        columns = {
            str(cell.value or ""): cell.column
            for cell in sheet[1]
        }
        column = columns[column_name]
        for row_number in range(sheet.max_row, 1, -1):
            if remove_value(sheet.cell(row_number, column).value):
                sheet.delete_rows(row_number)
        sheet.auto_filter.ref = sheet.dimensions

    remove_rows(
        "SO LINE PRICES",
        "Status",
        lambda value: str(value or "").strip().upper() != "COMPLETE",
    )
    remove_rows(
        "BOM COMPONENT COSTS",
        "Top BOM SKU",
        lambda value: str(value or "").strip().casefold() in blocked_skus,
    )
    remove_rows(
        "BOM CATEGORY BREAKDOWN",
        "Top SKU",
        lambda value: str(value or "").strip().casefold() in blocked_skus,
    )
    remove_rows(
        "NON-BOM RULES",
        "Status",
        lambda value: str(value or "").strip().upper() != "COMPLETE",
    )

    diagnostics = workbook["DIAGNOSTICS"]
    diagnostics.title = "EXCLUDED BLOCKED"

    info = workbook["INFO"]
    info.insert_rows(1, 4)
    info.cell(1, 1, "Release type")
    info.cell(1, 2, "COMPLETE POSITIONS ONLY")
    info.cell(2, 1, "Safe usage")
    info.cell(
        2,
        2,
        "Use only listed SKUs; excluded SKUs must keep their previous price",
    )
    info.cell(3, 1, "Excluded blocked positions")
    info.cell(3, 2, len(blocked_skus))
    info.cell(4, 1, "Full final release")
    info.cell(4, 2, "NO")

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)


def _primary_pricing_layer(sku: str, category: str) -> str:
    normalized = str(sku or "").strip().upper()
    category = str(category or "").strip().upper()
    if normalized.startswith("FPACK-"):
        return "FPACK"
    if normalized.startswith("APACK-"):
        return "APACK"
    if normalized.endswith("-A") and "-C-CAB" in normalized:
        return "CABINETS-A"
    if category == "ALL / CABINET SHELF":
        return "SHELF"
    if category == "ALL / CABINETS":
        return "CABINETS"
    return ""


def write_pricing_chain_audit(
    source: Path,
    cabinet_part_source: Path,
    destination: Path,
    current_sales_prices: dict[str, float] | None = None,
) -> None:
    """Audit the five pricing layers that determine cabinet and shelf prices."""
    current_sales_prices = {
        str(sku or "").strip().casefold(): price
        for sku, price in (current_sales_prices or {}).items()
        if str(sku or "").strip() and isinstance(price, (int, float))
    }
    source_workbook = load_workbook(source, data_only=True, read_only=True)
    prices = source_workbook["SO LINE PRICES"]
    price_rows = prices.iter_rows(values_only=True)
    price_header = next(price_rows)
    price_columns = {value: index for index, value in enumerate(price_header)}

    components = source_workbook["BOM COMPONENT COSTS"]
    component_rows = components.iter_rows(values_only=True)
    component_header = next(component_rows)
    component_columns = {value: index for index, value in enumerate(component_header)}
    by_top: dict[str, list[tuple]] = {}
    for row in component_rows:
        top = str(row[component_columns["Top BOM SKU"]] or "").strip()
        if top:
            by_top.setdefault(top.casefold(), []).append(row)

    audited = []
    for row in price_rows:
        sku = str(row[price_columns["SKU"]] or "").strip()
        category = str(row[price_columns["Product Category"]] or "").strip()
        layer = _primary_pricing_layer(sku, category)
        if not layer:
            continue
        details = by_top.get(sku.casefold(), [])
        component_cost = row[price_columns["Component / Purchase Cost"]]
        rollup = sum(
            float(detail[component_columns["Component Cost"]] or 0)
            for detail in details
        )
        non_positive = sum(
            1
            for detail in details
            if isinstance(
                detail[component_columns["Purchase Unit Price"]],
                (int, float),
            )
            and detail[component_columns["Purchase Unit Price"]] <= 0
        )
        sources = Counter(
            str(detail[component_columns["Cost Source"]] or "").strip()
            for detail in details
        )
        status = str(row[price_columns["Status"]] or "").strip().upper()
        final_price = row[price_columns["Final Reform SO Unit Price"]]
        current_price = current_sales_prices.get(sku.casefold())
        current_is_placeholder = (
            isinstance(current_price, (int, float))
            and current_price <= 0.01
        )
        price_change = (
            final_price - current_price
            if isinstance(final_price, (int, float))
            and isinstance(current_price, (int, float))
            and current_price > 0
            and not current_is_placeholder
            else None
        )
        price_change_percent = (
            price_change / current_price
            if price_change is not None
            else None
        )
        if current_is_placeholder:
            price_review = "SO PRICE CORRECTION REQUIRED"
        elif price_change_percent is None:
            price_review = "NO ODOO BASELINE"
        elif abs(price_change_percent) > 0.10:
            price_review = "REVIEW >10%"
        else:
            price_review = "WITHIN 10%"
        checks = []
        if status != "COMPLETE":
            checks.append(str(row[price_columns["Issues"]] or "Pricing BLOCKED"))
        if not details:
            checks.append("No component cost breakdown")
        if not isinstance(component_cost, (int, float)) or component_cost <= 0:
            checks.append("Non-positive component cost")
        if non_positive:
            checks.append(f"Non-positive component prices: {non_positive}")
        if isinstance(component_cost, (int, float)) and abs(rollup - component_cost) > 1e-6:
            checks.append(
                f"Component roll-up mismatch: {rollup:.6f} != {component_cost:.6f}"
            )
        audited.append({
            "layer": layer,
            "sku": sku,
            "category": category,
            "component_cost": component_cost,
            "addons": row[price_columns["Pricing Add-ons Total"]],
            "adjustment": row[price_columns["Adjustment Amount"]],
            "final": final_price,
            "current_price": current_price,
            "price_change": price_change,
            "price_change_percent": price_change_percent,
            "price_review": price_review,
            "detail_count": len(details),
            "cabinet_part_lines": sources["CABINET PART CALCULATION"],
            "purchase_price_lines": sources["LAST PURCHASE PRICE"],
            "rollup": rollup,
            "status": "PASS" if not checks else "BLOCKED",
            "issues": "; ".join(checks),
        })
    source_workbook.close()

    cabinet_workbook = load_workbook(
        cabinet_part_source, data_only=True, read_only=True
    )
    cabinet_sheet = cabinet_workbook["CABINET PART PRICES"]
    cabinet_rows = max(cabinet_sheet.max_row - 1, 0)
    cabinet_workbook.close()

    workbook = Workbook()
    summary = workbook.active
    summary.title = "SUMMARY"
    summary.append(["Layer", "Rows", "PASS", "BLOCKED", "Price review >10%"])
    summary.append(["CABINET PARTS", cabinet_rows, cabinet_rows, 0, "N/A"])
    for layer in ("FPACK", "CABINETS", "SHELF", "APACK", "CABINETS-A"):
        layer_rows = [row for row in audited if row["layer"] == layer]
        summary.append([
            layer,
            len(layer_rows),
            sum(row["status"] == "PASS" for row in layer_rows),
            sum(row["status"] != "PASS" for row in layer_rows),
            sum(row["price_review"] == "REVIEW >10%" for row in layer_rows),
        ])
    summary.append([])
    summary.append(["Rule", "FPACK, APACK and Shelf-PP use recursive BOM cost"])
    summary.append(["Odoo use", "Read-only purchase prices; Odoo is not changed"])
    summary.append([
        "Important",
        "Odoo sales price 0.01 is an intentional SO correction indicator, not a purchase-price benchmark",
    ])

    detail = workbook.create_sheet("PRIMARY CHAIN")
    headers = [
        "Layer", "SKU", "Product Category", "Component Cost", "Pricing Add-ons",
        "Adjustment", "Final Reform Price", "Current Odoo Sales Price",
        "Price Change", "Price Change %", "Price Review", "Component Lines",
        "Cabinet Part Lines", "Odoo Purchase Price Lines", "Roll-up Cost",
        "Audit Status", "Issues",
    ]
    detail.append(headers)
    for row in audited:
        detail.append([
            row["layer"], row["sku"], row["category"], row["component_cost"],
            row["addons"], row["adjustment"], row["final"], row["current_price"],
            row["price_change"], row["price_change_percent"], row["price_review"],
            row["detail_count"],
            row["cabinet_part_lines"], row["purchase_price_lines"], row["rollup"],
            row["status"], row["issues"],
        ])
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions
    for column in ("D", "E", "F", "G", "H", "I", "O"):
        for cell in detail[column][1:]:
            cell.number_format = '0.0000 [$€-x-euro2]'
    for cell in detail["J"][1:]:
        cell.number_format = "0.00%"
    for sheet in (summary, detail):
        for cell in sheet[1]:
            font = copy(cell.font)
            font.bold = True
            cell.font = font
    detail.column_dimensions["A"].width = 18
    detail.column_dimensions["B"].width = 38
    detail.column_dimensions["C"].width = 38
    detail.column_dimensions["Q"].width = 80
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)


def _read_last_purchase_prices(path: Path) -> dict[str, dict]:
    """Load the latest approved Production purchase price by SKU."""
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["COMPONENT PRICES"]
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    columns = {value: index for index, value in enumerate(header)}
    required = (
        "Internal Reference",
        "Real Purchase Price",
        "Vendor",
        "Purchase Order",
        "Order Date",
    )
    missing = [name for name in required if name not in columns]
    if missing:
        workbook.close()
        raise ValueError(
            "Last Purchase Prices faile trūksta stulpelių: "
            + ", ".join(missing)
        )

    result = {}
    for row in rows:
        sku = str(row[columns["Internal Reference"]] or "").strip()
        if not sku:
            continue
        result[sku.casefold()] = {
            "price": row[columns["Real Purchase Price"]],
            "vendor": row[columns["Vendor"]],
            "purchase_order": row[columns["Purchase Order"]],
            "order_date": row[columns["Order Date"]],
        }
    workbook.close()
    return result


def write_furnix_parts_price_review(
    cabinet_part_source: Path,
    last_purchase_source: Path,
    destination: Path,
    current_sales_prices: dict[str, float] | None = None,
) -> None:
    """Compare calculated Furnix part prices with Production LPP.

    Production ``list_price`` is displayed only as an SO correction indicator.
    It is never used as the purchase-price comparison baseline.
    """
    last_purchases = _read_last_purchase_prices(last_purchase_source)
    current_sales_prices = {
        str(sku or "").strip().casefold(): price
        for sku, price in (current_sales_prices or {}).items()
        if str(sku or "").strip() and isinstance(price, (int, float))
    }

    source_workbook = load_workbook(
        cabinet_part_source, data_only=True, read_only=True
    )
    source_sheet = source_workbook["CABINET PART PRICES"]
    rows = source_sheet.iter_rows(values_only=True)
    header = next(rows)
    columns = {value: index for index, value in enumerate(header)}
    required = (
        "Internal Reference",
        "Odoo Product ID",
        "Odoo Active",
        "Furnix Unit Cost",
        "Furnix Markup, %",
        "Furnix Sales Price to Furnibox",
        "Product Status",
        "BOM Source",
    )
    missing = [name for name in required if name not in columns]
    if missing:
        source_workbook.close()
        raise ValueError(
            "Cabinet Parts faile trūksta stulpelių: " + ", ".join(missing)
        )

    prepared = []
    for row in rows:
        sku = str(row[columns["Internal Reference"]] or "").strip()
        if not sku:
            continue
        key = sku.casefold()
        last_purchase = last_purchases.get(key, {})
        lpp = last_purchase.get("price")
        new_purchase = row[columns["Furnix Sales Price to Furnibox"]]
        current_sales_price = current_sales_prices.get(key)
        delta = (
            new_purchase - lpp
            if isinstance(new_purchase, (int, float))
            and isinstance(lpp, (int, float))
            else None
        )
        delta_percent = (
            delta / lpp
            if delta is not None and lpp != 0
            else None
        )
        if current_sales_price is None:
            so_price_status = "NO ODOO LIST PRICE"
        elif current_sales_price <= 0.01:
            so_price_status = "SO PRICE CORRECTION REQUIRED"
        else:
            so_price_status = "ODOO LIST PRICE PRESENT"
        if not isinstance(lpp, (int, float)):
            review = "NO PURCHASE HISTORY / NEW"
        elif delta > 0:
            review = "NEW PURCHASE PRICE HIGHER"
        elif delta < 0:
            review = "NEW PURCHASE PRICE LOWER"
        else:
            review = "UNCHANGED"
        prepared.append([
            sku,
            row[columns["Odoo Product ID"]],
            row[columns["Odoo Active"]],
            row[columns["Product Status"]],
            row[columns["BOM Source"]],
            lpp,
            last_purchase.get("vendor"),
            last_purchase.get("purchase_order"),
            last_purchase.get("order_date"),
            row[columns["Furnix Unit Cost"]],
            row[columns["Furnix Markup, %"]],
            new_purchase,
            delta,
            delta_percent,
            new_purchase,
            current_sales_price,
            so_price_status,
            review,
        ])
    source_workbook.close()

    workbook = Workbook()
    summary = workbook.active
    summary.title = "SUMMARY"
    summary.append(["Metric", "Count"])
    summary.append(["Furnix dimensional parts", len(prepared)])
    summary.append([
        "With Production Last Purchase Price",
        sum(isinstance(row[5], (int, float)) for row in prepared),
    ])
    summary.append([
        "No purchase history / new",
        sum(row[17] == "NO PURCHASE HISTORY / NEW" for row in prepared),
    ])
    summary.append([
        "New purchase price higher",
        sum(row[17] == "NEW PURCHASE PRICE HIGHER" for row in prepared),
    ])
    summary.append([
        "New purchase price lower",
        sum(row[17] == "NEW PURCHASE PRICE LOWER" for row in prepared),
    ])
    summary.append([
        "SO price correction required (0.01 or lower)",
        sum(row[16] == "SO PRICE CORRECTION REQUIRED" for row in prepared),
    ])
    summary.append([])
    summary.append([
        "Purchase comparison baseline",
        "Production Odoo latest approved purchase order line (Last Purchase Price)",
    ])
    summary.append([
        "Important",
        "Odoo list_price 0.01 is only an SO correction indicator and is excluded from purchase-price deltas",
    ])
    summary.append(["Odoo changed", "NO"])

    detail = workbook.create_sheet("FURNIX PARTS REVIEW")
    detail.append([
        "Internal Reference",
        "Odoo Product ID",
        "Odoo Active",
        "Product Status",
        "BOM Source",
        "Production Last Purchase Price",
        "Last Purchase Vendor",
        "Last Purchase Order",
        "Last Purchase Date",
        "New Furnix Unit Cost",
        "Furnix Markup, %",
        "New Furnix to Furnibox Purchase Price",
        "Purchase Price Change",
        "Purchase Price Change, %",
        "New Reform Price",
        "Current Odoo SO/List Price",
        "SO Price Status",
        "Purchase Price Review",
    ])
    for row in prepared:
        detail.append(row)
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions
    for column in ("F", "J", "L", "M", "O", "P"):
        for cell in detail[column][1:]:
            cell.number_format = '0.0000 [$€-x-euro2]'
    for column in ("K", "N"):
        for cell in detail[column][1:]:
            cell.number_format = "0.00%"
    widths = {
        "A": 42, "B": 18, "C": 13, "D": 18, "E": 24, "F": 28,
        "G": 34, "H": 22, "I": 21, "J": 22, "K": 18, "L": 34,
        "M": 22, "N": 23, "O": 22, "P": 26, "Q": 34, "R": 31,
    }
    for column, width in widths.items():
        detail.column_dimensions[column].width = width
    for sheet in (summary, detail):
        for cell in sheet[1]:
            font = copy(cell.font)
            font.bold = True
            cell.font = font
    summary.column_dimensions["A"].width = 48
    summary.column_dimensions["B"].width = 100
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)


def write_furnibox_purchase_prices(source: Path, destination: Path) -> None:
    """Publish the full Furnibox-to-Reform purchase price chain."""
    source_workbook = load_workbook(source, data_only=True, read_only=True)
    source_sheet = source_workbook["REFORM PRICE LIST"]
    rows = source_sheet.iter_rows(values_only=True)
    source_header = next(rows)
    columns = {value: index for index, value in enumerate(source_header)}
    selected = [
        ("Internal Reference", "Internal Reference"),
        ("Name", "Name"),
        ("Price Source", "Price Source"),
        ("Vendor / Supply Source", "Vendor / Supply Source"),
        ("Real Furnibox Purchase Price", "Real Furnibox Purchase Price"),
        ("Adjusted Furnibox Purchase Price", "Furnibox (Tamara) Purchase Price"),
        ("Reform Markup Factor", "Reform Markup Factor"),
        ("Reform Purchase Price", "Reform Purchase Price"),
        ("Status / BOM Source", "Status / BOM Source"),
    ]

    result = Workbook()
    sheet = result.active
    sheet.title = "FURNIBOX PURCHASE PRICES"
    sheet.append([output_name for _, output_name in selected])
    for row in rows:
        published_row = [
            row[columns[source_name]]
            for source_name, _ in selected
        ]
        reform_price_index = 7
        reform_price = published_row[reform_price_index]
        adjusted_price = published_row[5]
        markup_factor = published_row[6]
        if (
            not isinstance(reform_price, (int, float))
            and isinstance(adjusted_price, (int, float))
            and isinstance(markup_factor, (int, float))
        ):
            published_row[reform_price_index] = adjusted_price * markup_factor
        sheet.append(published_row)
    source_workbook.close()

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 50
    sheet.column_dimensions["C"].width = 28
    sheet.column_dimensions["D"].width = 35
    for column in ("E", "F", "H"):
        sheet.column_dimensions[column].width = 28
        for cell in sheet[column][1:]:
            cell.number_format = '0.0000 [$€-x-euro2]'
    sheet.column_dimensions["G"].width = 22
    sheet.column_dimensions["I"].width = 30
    for cell in sheet["G"][1:]:
        cell.number_format = "0.0000"

    info = result.create_sheet("INFO")
    info.append(["Parameter", "Value"])
    info.append(["Real purchase price", "Latest approved vendor purchase price"])
    info.append([
        "Furnibox (Tamara) Purchase Price",
        "Tamara-adjusted Furnibox purchase price used by Reform pricing",
    ])
    info.append([
        "Reform Markup Factor",
        "Multiplier applied to the Tamara-adjusted purchase price",
    ])
    info.append([
        "Reform Purchase Price",
        "Purchase price shown to Reform after applying the markup factor",
    ])
    info.append(["Odoo changed", "NO"])
    result.save(destination)


def read_reconciliation_summary(path: Path) -> dict:
    document = json.loads(path.read_text(encoding="utf-8"))
    if document.get("mode") != "READ_ONLY":
        raise ValueError("Target reconciliation rezultatas nėra READ_ONLY.")
    if str(document.get("environment") or "").lower() != "production":
        raise ValueError("Target reconciliation atliktas ne su Production Odoo.")
    summary = document.get("summary")
    if not isinstance(summary, dict):
        raise ValueError("Target reconciliation rezultatas neturi summary.")
    return summary


def read_current_sales_prices(path: Path) -> dict[str, float]:
    document = json.loads(path.read_text(encoding="utf-8"))
    result = {}
    for row in document.get("current_sales_prices") or []:
        sku = str(row.get("sku") or "").strip()
        price = row.get("price")
        if sku and isinstance(price, (int, float)):
            result[sku] = float(price)
    return result


def refresh(bom_input: Path, output_dir: Path, rules_path: Path = RULES_PATH) -> int:
    output_dir.mkdir(parents=True, exist_ok=True)
    PRODUCTION_DIR.mkdir(parents=True, exist_ok=True)

    reform_map = PRODUCTION_DIR / "Reform_MAP.xlsx"
    odoo_map = PRODUCTION_DIR / "Odoo_MAP.xlsx"
    detection = PRODUCTION_DIR / "Product_Detection_All.xlsx"
    comparison = PRODUCTION_DIR / "MAP_Comparison.xlsx"
    cabinet_part_prices = (
        PRODUCTION_DIR / "Existing_and_New_Cabinet_Parts_Prices.xlsx"
    )
    target_dataset = output_dir / "Furnibox_Target_Dataset.json"
    target_reconciliation = output_dir / "Target_Odoo_Reconciliation.json"

    run_step(
        "1/9 Pilnas Furnibox Target Dataset",
        "generate_full_validated_dataset.py",
        "--bom-input", str(bom_input),
        "--output-path", str(target_dataset),
        "--local-only",
    )
    run_step(
        "2/9 Target Dataset ir Production Odoo reconciliation",
        "reconcile_target_odoo.py",
        "--dataset", str(target_dataset),
        "--output", str(target_reconciliation),
    )
    reconciliation_summary = read_reconciliation_summary(target_reconciliation)
    run_step("3/9 Reform BOM paruošimas", "reform_map.py", "--input", str(bom_input), "--output", str(reform_map))
    run_step("4/9 Reform SKU patikra Odoo", "product_detection_v3.py", "--bom-input", str(bom_input))
    run_step(
        "5/9 Legacy MAP palyginimas detalių kainodarai",
        "map_comparison_v5.py", "--reform", str(reform_map), "--odoo", str(odoo_map),
        "--products", str(detection), "--bom-input", str(bom_input), "--output", str(comparison),
    )
    run_step("6/9 Cabinet ir Shelf detalių kainos", "cabinet_parts_price_v1.py")
    run_step("7/9 Naujausios pirkimo kainos", "last_purchase_prices.py")
    run_step("8/9 Bendras Reform pirkimo kainų šaltinis", "reform_price_list.py")

    with tempfile.TemporaryDirectory(prefix="reform-pricing-") as temporary:
        candidate_dir = Path(temporary)
        run_step(
            "9/9 Galutinės Reform pardavimo kainos",
            "reform_so_line_prices.py", "--bom-input", str(bom_input),
            "--dataset", str(target_dataset),
            "--price-input", str(PRODUCTION_DIR / "Reform_Final_Prices.xlsx"),
            "--rules", str(rules_path), "--output-dir", str(candidate_dir),
        )
        candidate = candidate_dir / "Reform_SO_Line_Prices.xlsx"
        statuses, blocked = read_pricing_status(candidate)
        write_pricing_chain_audit(
            candidate,
            cabinet_part_prices,
            output_dir / "Pricing_Chain_Audit.xlsx",
            current_sales_prices=read_current_sales_prices(
                target_reconciliation
            ),
        )
        write_furnix_parts_price_review(
            cabinet_part_prices,
            PRODUCTION_DIR / "Last_Purchase_Prices.xlsx",
            output_dir / "Furnix_Parts_Price_Review.xlsx",
            current_sales_prices=read_current_sales_prices(
                target_reconciliation
            ),
        )

        partial_name = (
            "Reform_SO_Line_Prices_COMPLETE_ONLY.xlsx"
            if blocked
            else None
        )
        generated_at = datetime.now().isoformat(timespec="seconds")
        run_id = f"REFORM-{generated_at.replace(':', '').replace('-', '')}"
        git_commit = os.getenv("RAILWAY_GIT_COMMIT_SHA", "")
        result = {
            "generated_at": generated_at,
            "bom_input": str(bom_input),
            "statuses": dict(statuses),
            "blocked": blocked,
            "target_reconciliation": reconciliation_summary,
            "released": not blocked,
            "partial_released": bool(blocked),
            "partial_file": partial_name,
            "excluded_blocked_count": len(blocked),
            "pricing_control_version": "pricing-control-v1",
            "pricing_run_id": run_id,
            "git_commit": git_commit,
            "odoo_changed": False,
        }
        (output_dir / "Reform_Pricing_Result.json").write_text(
            json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8"
        )

        if blocked:
            write_blocker_report(
                output_dir / "Reform_Pricing_BLOCKED.xlsx", statuses, blocked
            )
            partial_path = output_dir / partial_name
            write_complete_only_price_workbook(
                candidate,
                partial_path,
                blocked,
            )
            enrich_pricing_workbook(
                partial_path,
                git_commit=git_commit,
                run_id=run_id,
                generated_at=generated_at,
            )
            write_furnibox_purchase_prices(
                PRODUCTION_DIR / "Reform_Final_Prices.xlsx",
                output_dir / "Furnibox_Tamara_Purchase_Prices.xlsx",
            )
            shutil.copy2(
                PRODUCTION_DIR / "Reform_Final_Prices.xlsx",
                output_dir / "Reform_Pricing_Source.xlsx",
            )
            shutil.copy2(
                cabinet_part_prices,
                output_dir / "Cabinet_Parts_Pricing.xlsx",
            )
            print(f"\nSUSTABDYTA: {len(blocked)} BLOCKED pozicijos.")
            print(
                "Pilnas galutinis failas nepateiktas; "
                "sukurtas COMPLETE_ONLY failas be BLOCKED pozicijų."
            )
            return 2

        enrich_pricing_workbook(
            candidate,
            git_commit=git_commit,
            run_id=run_id,
            generated_at=generated_at,
        )
        write_furnibox_purchase_prices(
            PRODUCTION_DIR / "Reform_Final_Prices.xlsx",
            output_dir / "Furnibox_Tamara_Purchase_Prices.xlsx",
        )
        shutil.copy2(
            PRODUCTION_DIR / "Reform_Final_Prices.xlsx",
            output_dir / "Reform_Pricing_Source.xlsx",
        )
        shutil.copy2(
            cabinet_part_prices,
            output_dir / "Cabinet_Parts_Pricing.xlsx",
        )
        shutil.copy2(candidate, output_dir / "Reform_SO_Line_Prices.xlsx")

    print("\nREFORM KAINODARA ATNAUJINTA: 0 BLOCKED")
    print("Parengtos faktinės, Furnibox (Tamara) pirkimo ir Reform pardavimo kainos.")
    print("Odoo nekeistas.")
    return 0


def main() -> None:
    parser = argparse.ArgumentParser(description="Atnaujinti visą Reform kainodarą")
    parser.add_argument("--bom-input", required=True, type=Path)
    parser.add_argument("--rules", type=Path, default=RULES_PATH)
    parser.add_argument("--output-dir", required=True, type=Path)
    args = parser.parse_args()
    if not args.bom_input.exists():
        raise FileNotFoundError(args.bom_input)
    raise SystemExit(refresh(args.bom_input, args.output_dir, args.rules))


if __name__ == "__main__":
    main()
