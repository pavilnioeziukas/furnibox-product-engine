"""Human-readable control and explain-price layer for Reform pricing workbooks.

This module deliberately does not calculate or change prices.  It enriches an
already generated Reform_SO_Line_Prices workbook with control, rule, trace,
exception and change views so a reviewer can explain where every result came
from without reading Python code.
"""

from __future__ import annotations

import argparse
import json
import sqlite3
from collections import Counter, defaultdict
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


RULES_VERSION = "pricing-control-v1"

# These IDs describe business logic already implemented by the pricing engine.
# This module never reapplies the formulas; IDs are audit labels only.
PRICING_RULES = (
    {
        "id": "R001",
        "name": "Prepared purchase / transfer price",
        "applies_to": "Purchased component",
        "logic": (
            "Use the prepared direct price from the purchase-price layer when "
            "available. The source is carried into the component trace."
        ),
        "implemented_in": "reform_so_line_prices.resolve_component_cost",
    },
    {
        "id": "R002",
        "name": "Recursive BOM material cost",
        "applies_to": "BOM component without direct prepared price",
        "logic": "Material cost = SUM(resolved child cost x child quantity).",
        "implemented_in": "reform_so_line_prices.resolve_component_cost",
    },
    {
        "id": "R003",
        "name": "BOM pricing add-ons",
        "applies_to": "BOM",
        "logic": (
            "Apply configured Level II and Level I pricing add-ons using the "
            "existing category-rule logic."
        ),
        "implemented_in": "reform_so_line_prices.calculate_boms",
    },
    {
        "id": "R004",
        "name": "BOM add-on adjustment",
        "applies_to": "BOM add-ons",
        "logic": (
            "The configured adjustment is applied to pricing add-ons only; "
            "recursive material cost is not reduced by the adjustment."
        ),
        "implemented_in": "reform_so_line_prices.calculate_boms",
    },
    {
        "id": "R005",
        "name": "Generated internal manufacture cost",
        "applies_to": "APACK / HRD-A / Shelf-PP internal MANUFACTURE products",
        "logic": (
            "Child materials enter recursive component cost; do not require or "
            "apply an additional child sales add-on rule at the wrong level."
        ),
        "implemented_in": "reform_so_line_prices.calculate_boms",
    },
    {
        "id": "R006",
        "name": "No Odoo Standard Price fallback",
        "applies_to": "Missing component price",
        "logic": (
            "If neither a prepared direct price nor a resolvable BOM exists, "
            "the position is BLOCKED. Odoo Standard Price is not a fallback."
        ),
        "implemented_in": "reform_so_line_prices.resolve_component_cost",
    },
    {
        "id": "R007",
        "name": "Non-BOM price",
        "applies_to": "Non-BOM position",
        "logic": (
            "Final unit price = purchase price + pack preparation + storage + "
            "bag + sticker."
        ),
        "implemented_in": "reform_so_line_prices.calculate_non_bom",
    },
)

HEADER_FILL = "1F4E78"
HEADER_FONT = "FFFFFF"
STATUS_FILL = {
    "CALCULATED": "C6EFCE",
    "REVIEW": "FFEB9C",
    "BLOCKED": "F4CCCC",
}
THIN = Side(style="thin", color="B7B7B7")
TABLE_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _text(value) -> str:
    return str(value or "").strip()


def _key(value) -> str:
    return _text(value).casefold()


def _columns(sheet) -> dict[str, int]:
    return {
        _text(cell.value): cell.column
        for cell in sheet[1]
        if _text(cell.value)
    }


def _style_table(sheet, widths: dict[int, float] | None = None) -> None:
    if sheet.max_row < 1:
        return
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(color=HEADER_FONT, bold=True)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = TABLE_BORDER
    sheet.row_dimensions[1].height = 36
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = TABLE_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, width in (widths or {}).items():
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width


def _fill_status(cell, status: str) -> None:
    color = STATUS_FILL.get(status)
    if color:
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = copy(cell.font)
        cell.font = Font(
            name=cell.font.name,
            size=cell.font.sz,
            bold=True,
            italic=cell.font.italic,
            color=cell.font.color,
        )


def _control_status(engine_status: str, issues: str = "") -> str:
    engine_status = _text(engine_status).upper()
    if engine_status == "BLOCKED":
        return "BLOCKED"
    if engine_status != "COMPLETE":
        return "REVIEW"
    if _text(issues):
        return "REVIEW"
    return "CALCULATED"


def _generated_internal_sku(sku: str) -> bool:
    normalized = _text(sku).upper()
    return (
        normalized.startswith("APACK-")
        or normalized.endswith("-PP")
        or normalized.endswith("-HRD-A")
        or normalized.startswith("HRD-") and normalized.endswith("-A")
    )


def _row_rule_ids(position_type: str, sku: str, status: str) -> list[str]:
    result: list[str] = []
    if _text(position_type).upper() == "BOM":
        result.extend(["R002", "R003", "R004"])
        if _generated_internal_sku(sku):
            result.append("R005")
    else:
        result.append("R007")
    if _text(status).upper() == "BLOCKED":
        result.append("R006")
    return result


def _remove_if_exists(workbook, name: str) -> None:
    if name in workbook.sheetnames:
        del workbook[name]


def _insert_control_sheets_first(workbook, names: list[str]) -> None:
    for name in reversed(names):
        sheet = workbook[name]
        workbook._sheets.remove(sheet)  # openpyxl has no public reorder API
        workbook._sheets.insert(0, sheet)


def _read_results(workbook) -> list[dict]:
    source = workbook["SO LINE PRICES"]
    columns = _columns(source)
    required = (
        "SKU",
        "Name",
        "Position Type",
        "Product Category",
        "Component / Purchase Cost",
        "Pricing Add-ons Total",
        "Adjustment Amount",
        "Final Reform SO Unit Price",
        "Status",
        "Issues",
    )
    missing = [name for name in required if name not in columns]
    if missing:
        raise ValueError("SO LINE PRICES missing columns: " + ", ".join(missing))

    rows = []
    for values in source.iter_rows(min_row=2, values_only=True):
        sku = _text(values[columns["SKU"] - 1])
        if not sku:
            continue
        engine_status = _text(values[columns["Status"] - 1])
        issues = _text(values[columns["Issues"] - 1])
        rows.append(
            {
                "sku": sku,
                "name": _text(values[columns["Name"] - 1]),
                "position_type": _text(values[columns["Position Type"] - 1]),
                "category": _text(values[columns["Product Category"] - 1]),
                "cost": values[columns["Component / Purchase Cost"] - 1],
                "addons": values[columns["Pricing Add-ons Total"] - 1],
                "adjustment": values[columns["Adjustment Amount"] - 1],
                "final": values[columns["Final Reform SO Unit Price"] - 1],
                "engine_status": engine_status,
                "issues": issues,
                "control_status": _control_status(engine_status, issues),
                "rule_ids": _row_rule_ids(
                    _text(values[columns["Position Type"] - 1]), sku, engine_status
                ),
            }
        )
    return rows


def _component_trace(workbook) -> dict[str, list[dict]]:
    result: dict[str, list[dict]] = defaultdict(list)
    if "BOM COMPONENT COSTS" not in workbook.sheetnames:
        return result
    sheet = workbook["BOM COMPONENT COSTS"]
    columns = _columns(sheet)
    required = (
        "Top BOM SKU",
        "Level II SKU",
        "Total Qty in Top BOM",
        "Purchased Component SKU",
        "Purchase Unit Price",
        "Component Cost",
        "Status",
        "Cost Source",
    )
    if any(name not in columns for name in required):
        return result
    for values in sheet.iter_rows(min_row=2, values_only=True):
        top = _text(values[columns["Top BOM SKU"] - 1])
        if not top:
            continue
        result[_key(top)].append(
            {
                "level_ii": _text(values[columns["Level II SKU"] - 1]),
                "component": _text(
                    values[columns["Purchased Component SKU"] - 1]
                ),
                "qty": values[columns["Total Qty in Top BOM"] - 1],
                "unit_price": values[columns["Purchase Unit Price"] - 1],
                "amount": values[columns["Component Cost"] - 1],
                "status": _text(values[columns["Status"] - 1]),
                "source": _text(values[columns["Cost Source"] - 1]),
            }
        )
    return result


def _category_trace(workbook) -> dict[str, list[dict]]:
    result: dict[str, list[dict]] = defaultdict(list)
    if "BOM CATEGORY BREAKDOWN" not in workbook.sheetnames:
        return result
    sheet = workbook["BOM CATEGORY BREAKDOWN"]
    columns = _columns(sheet)
    required = (
        "Top SKU",
        "Application Level",
        "Pricing Rule SKU",
        "Category ID",
        "Category Name",
        "Multiplier",
        "Add-ons Total",
        "Adjusted Add-ons",
    )
    if any(name not in columns for name in required):
        return result
    for values in sheet.iter_rows(min_row=2, values_only=True):
        top = _text(values[columns["Top SKU"] - 1])
        if not top:
            continue
        result[_key(top)].append(
            {
                "level": _text(values[columns["Application Level"] - 1]),
                "rule_sku": _text(values[columns["Pricing Rule SKU"] - 1]),
                "category_id": _text(values[columns["Category ID"] - 1]),
                "category_name": _text(values[columns["Category Name"] - 1]),
                "multiplier": values[columns["Multiplier"] - 1],
                "addons": values[columns["Add-ons Total"] - 1],
                "adjusted_addons": values[columns["Adjusted Add-ons"] - 1],
                "calculation": _text(
                    values[columns["Calculation Basis"] - 1]
                ) if "Calculation Basis" in columns else "",
            }
        )
    return result


def _write_control(
    workbook,
    results: list[dict],
    source_name: str,
    run_id: str,
    generated_at: str,
    git_commit: str,
) -> None:
    _remove_if_exists(workbook, "CONTROL")
    sheet = workbook.create_sheet("CONTROL")
    counts = Counter(row["control_status"] for row in results)
    rows = [
        ("Pricing run", run_id),
        ("Generated at", generated_at),
        ("Rules version", RULES_VERSION),
        ("Git commit", git_commit or "NOT PROVIDED"),
        ("Source workbook", source_name),
        ("Price calculation changed by this layer", "NO"),
        ("Odoo changed", "NO"),
        ("Total positions", len(results)),
        ("CALCULATED", counts["CALCULATED"]),
        ("REVIEW", counts["REVIEW"]),
        ("BLOCKED", counts["BLOCKED"]),
        (
            "Release principle",
            "Only CALCULATED positions are ready without further review; "
            "BLOCKED positions must not receive a newly calculated price.",
        ),
        (
            "How to explain a price",
            "Find SKU in PRICE RESULTS, then filter PRICE TRACE by the same SKU. "
            "The trace shows material sources, category-rule steps and final result.",
        ),
    ]
    for row in rows:
        sheet.append(row)
    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 105
    sheet.sheet_view.showGridLines = False
    for row in sheet.iter_rows():
        row[0].font = Font(bold=True)
        for cell in row:
            cell.border = TABLE_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_number in range(1, sheet.max_row + 1):
        status = _text(sheet.cell(row_number, 1).value).upper()
        if status in STATUS_FILL:
            _fill_status(sheet.cell(row_number, 1), status)
            _fill_status(sheet.cell(row_number, 2), status)


def _write_rules(workbook) -> None:
    _remove_if_exists(workbook, "PRICING RULES")
    sheet = workbook.create_sheet("PRICING RULES")
    sheet.append(["Rule ID", "Rule", "Applies to", "Logic", "Implementation"])
    for rule in PRICING_RULES:
        sheet.append(
            [
                rule["id"],
                rule["name"],
                rule["applies_to"],
                rule["logic"],
                rule["implemented_in"],
            ]
        )
    _style_table(sheet, {1: 11, 2: 34, 3: 40, 4: 88, 5: 55})


def _write_results(workbook, results: list[dict]) -> None:
    _remove_if_exists(workbook, "PRICE RESULTS")
    sheet = workbook.create_sheet("PRICE RESULTS")
    sheet.append(
        [
            "SKU",
            "Name",
            "Position Type",
            "Product Category",
            "Component / Purchase Cost",
            "Pricing Add-ons Total",
            "Adjustment Amount",
            "Final Reform SO Unit Price",
            "Control Status",
            "Applied Rule IDs",
            "Issues / Review Reason",
        ]
    )
    for row in results:
        sheet.append(
            [
                row["sku"],
                row["name"],
                row["position_type"],
                row["category"],
                row["cost"],
                row["addons"],
                row["adjustment"],
                row["final"],
                row["control_status"],
                ", ".join(row["rule_ids"]),
                row["issues"],
            ]
        )
    _style_table(
        sheet,
        {1: 34, 2: 42, 3: 16, 4: 34, 5: 21, 6: 20, 7: 19, 8: 24,
         9: 16, 10: 22, 11: 75},
    )
    for row_number in range(2, sheet.max_row + 1):
        for column in (5, 6, 7, 8):
            sheet.cell(row_number, column).number_format = '0.0000 [$€-x-euro2]'
        _fill_status(sheet.cell(row_number, 9), _text(sheet.cell(row_number, 9).value))


def _write_trace(workbook, results: list[dict]) -> None:
    _remove_if_exists(workbook, "PRICE TRACE")
    components = _component_trace(workbook)
    categories = _category_trace(workbook)
    sheet = workbook.create_sheet("PRICE TRACE")
    sheet.append(
        [
            "SKU",
            "Step #",
            "Step Type",
            "Rule ID",
            "Input / Component / Rule",
            "Qty / Multiplier",
            "Unit Price",
            "Amount",
            "Source",
            "Step Status",
            "Explanation",
        ]
    )
    for result in results:
        sku = result["sku"]
        step = 1
        for detail in components.get(_key(sku), []):
            status = "BLOCKED" if detail["status"].upper() != "OK" else "CALCULATED"
            sheet.append(
                [
                    sku,
                    step,
                    "MATERIAL",
                    "R001",
                    detail["component"],
                    detail["qty"],
                    detail["unit_price"],
                    detail["amount"],
                    detail["source"],
                    status,
                    (
                        f"{detail['level_ii']} -> {detail['component']}; "
                        "prepared component price or recursively resolved leaf."
                    ),
                ]
            )
            step += 1
        for detail in categories.get(_key(sku), []):
            sheet.append(
                [
                    sku,
                    step,
                    "PRICING ADD-ON",
                    "R003/R004",
                    (
                        f"{detail['rule_sku']} | {detail['category_id']} | "
                        f"{detail['category_name']}"
                    ),
                    detail["multiplier"],
                    None,
                    detail["adjusted_addons"],
                    detail["level"],
                    "CALCULATED",
                    (
                        (detail["calculation"] + "; ")
                        if detail["calculation"] else ""
                    ) + (
                        f"raw add-ons {detail['addons']}; adjusted add-ons "
                        f"{detail['adjusted_addons']}."
                    ),
                ]
            )
            step += 1
        sheet.append(
            [
                sku,
                step,
                "FINAL RESULT",
                ", ".join(result["rule_ids"]),
                "Final Reform SO Unit Price",
                None,
                None,
                result["final"],
                "SO LINE PRICES",
                result["control_status"],
                (
                    f"Material / purchase cost {result['cost']}; add-ons "
                    f"{result['addons']}; adjustment {result['adjustment']}; "
                    f"issues: {result['issues'] or 'none'}."
                ),
            ]
        )
    _style_table(
        sheet,
        {1: 34, 2: 9, 3: 20, 4: 16, 5: 50, 6: 18, 7: 18, 8: 18,
         9: 28, 10: 16, 11: 95},
    )
    for row_number in range(2, sheet.max_row + 1):
        for column in (7, 8):
            sheet.cell(row_number, column).number_format = '0.0000 [$€-x-euro2]'
        _fill_status(
            sheet.cell(row_number, 10), _text(sheet.cell(row_number, 10).value)
        )


def _write_exceptions(workbook, results: list[dict]) -> None:
    _remove_if_exists(workbook, "EXCEPTIONS")
    sheet = workbook.create_sheet("EXCEPTIONS")
    sheet.append(
        ["SKU", "Position Type", "Control Status", "Final Price", "Rule IDs", "Reason"]
    )
    for row in results:
        if row["control_status"] == "CALCULATED":
            continue
        sheet.append(
            [
                row["sku"],
                row["position_type"],
                row["control_status"],
                row["final"],
                ", ".join(row["rule_ids"]),
                row["issues"] or "Engine status requires review",
            ]
        )
    _style_table(sheet, {1: 34, 2: 17, 3: 16, 4: 20, 5: 22, 6: 100})
    for row_number in range(2, sheet.max_row + 1):
        sheet.cell(row_number, 4).number_format = '0.0000 [$€-x-euro2]'
        _fill_status(
            sheet.cell(row_number, 3), _text(sheet.cell(row_number, 3).value)
        )


def _previous_prices(path: Path | None) -> dict[str, float]:
    if path is None or not path.exists():
        return {}
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet_name = "PRICE RESULTS" if "PRICE RESULTS" in workbook.sheetnames else "SO LINE PRICES"
    sheet = workbook[sheet_name]
    columns = _columns(sheet)
    price_name = "Final Reform SO Unit Price"
    if "SKU" not in columns or price_name not in columns:
        workbook.close()
        return {}
    result = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        sku = _text(values[columns["SKU"] - 1])
        price = values[columns[price_name] - 1]
        if sku and isinstance(price, (int, float)):
            result[_key(sku)] = float(price)
    workbook.close()
    return result


def _write_changes(workbook, results: list[dict], previous: Path | None) -> bool:
    old = _previous_prices(previous)
    if not old:
        return False
    _remove_if_exists(workbook, "CHANGES")
    sheet = workbook.create_sheet("CHANGES")
    sheet.append(
        [
            "SKU",
            "Previous Price",
            "Current Price",
            "Change",
            "Change %",
            "Control Status",
            "Interpretation",
        ]
    )
    for row in results:
        current = row["final"]
        previous_price = old.get(_key(row["sku"]))
        if not isinstance(current, (int, float)) or previous_price is None:
            change = None
            change_percent = None
            interpretation = "NO COMPARABLE PRICE"
        else:
            change = float(current) - previous_price
            change_percent = change / previous_price if previous_price else None
            interpretation = "UNCHANGED" if abs(change) < 1e-9 else "CHANGED"
        sheet.append(
            [
                row["sku"],
                previous_price,
                current,
                change,
                change_percent,
                row["control_status"],
                interpretation,
            ]
        )
    _style_table(sheet, {1: 34, 2: 20, 3: 20, 4: 18, 5: 14, 6: 16, 7: 24})
    for row_number in range(2, sheet.max_row + 1):
        for column in (2, 3, 4):
            sheet.cell(row_number, column).number_format = '0.0000 [$€-x-euro2]'
        sheet.cell(row_number, 5).number_format = "0.00%"
        _fill_status(
            sheet.cell(row_number, 6), _text(sheet.cell(row_number, 6).value)
        )
    return True


def enrich_pricing_workbook(
    source: Path,
    destination: Path | None = None,
    *,
    previous: Path | None = None,
    git_commit: str = "",
    run_id: str = "",
    generated_at: str = "",
    search_index: Path | None = None,
) -> Path:
    """Add human-readable control sheets without changing calculated values."""
    source = Path(source)
    destination = Path(destination or source)
    if not source.exists():
        raise FileNotFoundError(source)

    workbook = load_workbook(source)
    results = _read_results(workbook)
    generated_at = generated_at or datetime.now().isoformat(timespec="seconds")
    run_id = run_id or f"PRICING-{generated_at.replace(':', '').replace('-', '')}"

    _write_control(
        workbook,
        results,
        source.name,
        run_id,
        generated_at,
        git_commit,
    )
    _write_results(workbook, results)
    _write_rules(workbook)
    _write_trace(workbook, results)
    _write_exceptions(workbook, results)
    has_changes = _write_changes(workbook, results, previous)

    control_names = [
        "CONTROL",
        "PRICE RESULTS",
        "PRICING RULES",
        "PRICE TRACE",
        "EXCEPTIONS",
    ]
    if has_changes:
        control_names.append("CHANGES")
    _insert_control_sheets_first(workbook, control_names)

    if search_index is not None:
        _write_search_index(workbook, Path(search_index))

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def _write_search_index(workbook, path: Path) -> None:
    """Create a compact random-access index for the web Explain Price view."""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.unlink(missing_ok=True)
    connection = sqlite3.connect(path)
    try:
        connection.executescript(
            """
            CREATE TABLE price_result (sku_key TEXT PRIMARY KEY, sku TEXT NOT NULL, payload TEXT NOT NULL);
            CREATE TABLE price_trace (sku_key TEXT NOT NULL, step_number INTEGER NOT NULL, payload TEXT NOT NULL);
            CREATE INDEX price_trace_sku ON price_trace (sku_key, step_number);
            """
        )
        for sheet_name, table_name in (("PRICE RESULTS", "price_result"), ("PRICE TRACE", "price_trace")):
            sheet = workbook[sheet_name]
            headers = [_text(cell.value) for cell in sheet[1]]
            for values in sheet.iter_rows(min_row=2, values_only=True):
                row = dict(zip(headers, values))
                sku = _text(row.get("SKU"))
                if not sku:
                    continue
                payload = json.dumps(row, ensure_ascii=False, default=str)
                if table_name == "price_result":
                    connection.execute("INSERT INTO price_result VALUES (?, ?, ?)", (_key(sku), sku, payload))
                else:
                    connection.execute("INSERT INTO price_trace VALUES (?, ?, ?)", (_key(sku), int(row.get("Step #") or 0), payload))
        connection.commit()
    finally:
        connection.close()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Add Pricing Control / Explain Price sheets to a Reform workbook"
    )
    parser.add_argument("--source", required=True, type=Path)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--previous", type=Path)
    parser.add_argument("--git-commit", default="")
    parser.add_argument("--run-id", default="")
    args = parser.parse_args()
    output = enrich_pricing_workbook(
        args.source,
        args.output,
        previous=args.previous,
        git_commit=args.git_commit,
        run_id=args.run_id,
    )
    print(output)


if __name__ == "__main__":
    main()
