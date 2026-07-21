"""Paruošia archyvuotų Reform produktų aktyvavimo importą Stage aplinkai.

Programa Odoo duomenis tik nuskaito. Produktai aktyvuojami tik tada, kai
vartotojas sugeneruotą Excel failą patikrina ir pats importuoja į Stage.
"""

from __future__ import annotations

import argparse
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import load_settings
from odoo_client import OdooClient
from output_paths import environment_output_dir, environment_slug


BASE_DIR = Path(__file__).resolve().parent


def truthy(value) -> bool:
    """Supranta ir Excel TRUE/FALSE, ir tekstines reikšmes."""
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "taip"}


def find_detection_file(base: Path) -> Path:
    """Randa naujausią Stage produkto palyginimo rezultatą."""
    candidates = list((base / "output" / "stage").glob("Product_Detection_All*.xlsx"))
    candidates = [path for path in candidates if not path.name.startswith("~$")]
    if not candidates:
        raise FileNotFoundError(
            "Nerastas output\\stage\\Product_Detection_All.xlsx. "
            "Pirmiausia GUI paleiskite 3 veiksmą."
        )
    return max(candidates, key=lambda path: path.stat().st_mtime)


def load_archived_products(path: Path) -> list[dict]:
    """Atrenka tik Odoo rastus, bet šiuo metu archyvuotus Reform produktus."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["PRODUCT DETECTION"]
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows)
    index = {str(value).strip(): position for position, value in enumerate(headers)}
    required = {
        "Reform SKU", "Reform Role", "Reform Product Category",
        "Reform Part Group", "Used By BOM Count", "Odoo Template ID",
        "Odoo Active", "Exists in Odoo",
    }
    missing = required - set(index)
    if missing:
        workbook.close()
        raise ValueError("PRODUCT DETECTION lape trūksta stulpelių: " + ", ".join(sorted(missing)))

    result = []
    seen_templates = set()
    for row in rows:
        exists = str(row[index["Exists in Odoo"]] or "").strip().upper() in {"YES", "TAIP", "TRUE", "1"}
        active = truthy(row[index["Odoo Active"]])
        template_id = row[index["Odoo Template ID"]]
        if not exists or active or not template_id:
            continue
        template_id = int(template_id)
        if template_id in seen_templates:
            continue
        seen_templates.add(template_id)
        result.append({
            "sku": str(row[index["Reform SKU"]] or "").strip(),
            "template_id": template_id,
            "role": str(row[index["Reform Role"]] or "").strip(),
            "category": str(row[index["Reform Product Category"]] or "").strip(),
            "part_group": str(row[index["Reform Part Group"]] or "").strip(),
            "used_by": int(row[index["Used By BOM Count"]] or 0),
        })
    workbook.close()
    return result


def load_external_ids(client: OdooClient, template_ids: set[int]) -> dict[int, str]:
    """Per API susieja product.template Database ID su importuojamu External ID."""
    if not template_ids:
        return {}
    rows = client.search_read_all(
        "ir.model.data",
        [["model", "=", "product.template"], ["res_id", "in", sorted(template_ids)]],
        ["module", "name", "res_id"],
    )
    grouped: dict[int, list[str]] = {}
    for row in rows:
        if row.get("res_id") and row.get("module") and row.get("name"):
            grouped.setdefault(int(row["res_id"]), []).append(f"{row['module']}.{row['name']}")

    result = {}
    for template_id, values in grouped.items():
        # Stabilus modulio External ID yra geresnis už automatinį __export__ ID.
        values.sort(key=lambda value: (value.startswith("__export__."), value))
        result[template_id] = values[0]
    return result


def format_sheet(worksheet) -> None:
    """Padaro rezultatą lengvai perskaitomą prieš importą."""
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column in worksheet.columns:
        letter = get_column_letter(column[0].column)
        width = max(len(str(cell.value or "")) for cell in column) + 2
        worksheet.column_dimensions[letter].width = min(max(width, 12), 55)


def write_output(path: Path, products: list[dict], external_ids: dict[int, str]) -> tuple[int, int]:
    """Sukuria importo, peržiūros, diagnostikos ir informacijos lapus."""
    workbook = Workbook()
    import_sheet = workbook.active
    import_sheet.title = "PRODUCT REACTIVATE"
    import_sheet.append(["id", "active"])

    review = workbook.create_sheet("REVIEW")
    review.append([
        "Internal Reference", "External ID", "Template Database ID",
        "Reform Role", "Used By BOM Count", "Product Category", "Part Group", "Action",
    ])
    diagnostics = workbook.create_sheet("DIAGNOSTICS")
    diagnostics.append(["Internal Reference", "Template Database ID", "Problem"])

    ready = 0
    missing = 0
    for product in sorted(products, key=lambda item: item["sku"]):
        external_id = external_ids.get(product["template_id"], "")
        if external_id:
            import_sheet.append([external_id, True])
            action = "READY - activate in Stage"
            ready += 1
        else:
            diagnostics.append([product["sku"], product["template_id"], "External ID nerastas"])
            action = "STOP - External ID nerastas"
            missing += 1
        review.append([
            product["sku"], external_id, product["template_id"], product["role"],
            product["used_by"], product["category"], product["part_group"], action,
        ])

    info = workbook.create_sheet("INFO")
    info.append(["Rodiklis", "Reikšmė"])
    info.append(["Archyvuoti Reform produktų šablonai", len(products)])
    info.append(["Paruošta aktyvuoti", ready])
    info.append(["Trūksta External ID", missing])
    info.append(["Odoo pakeitimai atlikti", "NE"])

    for worksheet in workbook.worksheets:
        format_sheet(worksheet)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return ready, missing


def main() -> None:
    parser = argparse.ArgumentParser(description="Paruošti archyvuotų Reform produktų aktyvavimo importą")
    parser.add_argument("--env-file", default=".env.stage", help="Stage aplinkos .env failas")
    parser.add_argument("--detection", type=Path, help="Product_Detection_All.xlsx kelias")
    args = parser.parse_args()

    # 1 ŽINGSNIS: užkrauname tik Stage prisijungimą ir užblokuojame Production.
    env_path = Path(args.env_file)
    if not env_path.is_absolute():
        env_path = BASE_DIR / env_path
    load_dotenv(env_path, override=True)
    if environment_slug() != "stage":
        raise PermissionError("Šis generatorius leidžiamas tik Stage aplinkoje.")

    # 2 ŽINGSNIS: iš 3 veiksmo rezultato atrenkame archyvuotus Reform produktus.
    detection_path = args.detection or find_detection_file(BASE_DIR)
    products = load_archived_products(detection_path)

    # 3 ŽINGSNIS: Stage API tik nuskaito produktų External ID. Jokių write/create nėra.
    settings = load_settings()
    client = OdooClient(settings)
    uid = client.authenticate()
    external_ids = load_external_ids(client, {item["template_id"] for item in products})

    # 4 ŽINGSNIS: sukuriame Odoo importo failą su External ID ir active=True.
    output_path = environment_output_dir(BASE_DIR) / "Stage_Reactivate_Reform_Products_External_ID.xlsx"
    ready, missing = write_output(output_path, products, external_ids)

    print("Prisijungta prie Stage Odoo. UID =", uid)
    print("\nPRODUKTŲ AKTYVAVIMO IMPORTAS SUKURTAS")
    print("Failas:", output_path)
    print("Archyvuoti Reform produktai:", len(products))
    print("Paruošta aktyvuoti:", ready)
    print("Trūksta External ID:", missing)
    print("Odoo pakeitimų neatlikta.")


if __name__ == "__main__":
    main()
