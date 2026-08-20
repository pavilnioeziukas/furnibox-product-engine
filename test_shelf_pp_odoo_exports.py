from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from bom_release.generator import generate_release_files
from bom_release.import_validator import validate_release_imports
from bom_release.models import (
    BomReleasePlan,
    BomReleasePlanItem,
    ReleaseAction,
    ReleaseStatus,
)


SHELF = "EUB-C-CAB01-SLF001"
PART = "EU-SREW-SHELF-163X564-WW"
PP = f"{PART}-PP"
PINS = "SLF-PINS-HRD-6"
PACKAGE = "N9569A"
STICKER = "TERMO 90X48"


def component(sku: str, quantity: float, parent: str, level: int) -> dict:
    return {
        "sku": sku,
        "quantity": quantity,
        "parent_sku": parent,
        "level": level,
    }


def shelf_dataset() -> dict:
    return {
        "schema_version": "1",
        "dataset_id": "shelf-dataset",
        "batch_reference": "shelf-batch",
        "environment": "stage",
        "created_at_utc": "2026-08-20T00:00:00Z",
        "source": {"file_name": "Reform BOM vXX.xlsx", "file_hash": "hash"},
        "statistics": {},
        "products": [
            {
                "sku": SHELF,
                "product_type": "CABINET SHELF",
                "reform_category": "CABINET SHELF",
                "bom_type": "KIT",
                "level": 1,
                "components": [
                    component(PP, 1, SHELF, 1),
                    component(PINS, 1, SHELF, 1),
                ],
                "operations": [],
            },
            {
                "sku": PP,
                "product_type": "SHELF PREPACK",
                "reform_category": "SHELF PREPACK",
                "bom_type": "MANUFACTURE",
                "level": 2,
                "components": [
                    component(PART, 1, PP, 2),
                    component(PACKAGE, 0.4, PP, 2),
                    component(STICKER, 2, PP, 2),
                ],
                "operations": [{
                    "name": "Lentynų pakavimas",
                    "workcenter": "Pakuotojai",
                    "time_mode": "manual",
                    "time_minutes": 1,
                    "sequence": 0,
                }],
            },
        ],
    }


def plan_item(sku: str, bom_type: str, components: int, operations: int) -> BomReleasePlanItem:
    return BomReleasePlanItem(
        parent_sku=sku,
        bom_type=bom_type,
        component_count=components,
        operation_count=operations,
        product_exists=True,
        product_id=1,
        product_template_id=11,
        parent_external_id_ready=True,
        active_bom_count=0,
        active_bom_id=None,
        active_reference="",
        active_sequence=None,
        active_bom_type="",
        release_exists=False,
        release_bom_id=None,
        release_reference="SHELF-RELEASE",
        action=ReleaseAction.CREATE,
        status=ReleaseStatus.READY,
    )


def shelf_plan() -> BomReleasePlan:
    return BomReleasePlan(
        release_id="SHELF-RELEASE",
        release_reference="SHELF-RELEASE",
        environment="stage",
        dataset_id="shelf-dataset",
        dataset_batch_reference="shelf-batch",
        dataset_path="dataset.json",
        created_at_utc="2026-08-20T00:00:00Z",
        items=[
            plan_item(SHELF, "KIT", 2, 0),
            plan_item(PP, "MANUFACTURE", 3, 1),
        ],
    )


class ShelfClient:
    def search_read_all(self, model, domain, fields, context=None):
        skus = [SHELF, PP, PART, PINS, PACKAGE, STICKER]
        if model == "product.product":
            return [
                {
                    "id": index,
                    "default_code": sku,
                    "product_tmpl_id": [100 + index, sku],
                    "active": True,
                }
                for index, sku in enumerate(skus, start=1)
            ]
        if model == "ir.model.data":
            source_model = domain[0][2]
            wanted = set(domain[1][2])
            rows = []
            for index, _sku in enumerate(skus, start=1):
                record_id = 100 + index if source_model == "product.template" else index
                if record_id in wanted:
                    rows.append({
                        "module": "x",
                        "name": f"{source_model.replace('.', '_')}_{record_id}",
                        "res_id": record_id,
                    })
            return rows
        if model == "mrp.workcenter":
            return [{"id": 1, "name": "Pakuotojai", "active": True}]
        raise AssertionError(model)


class ShelfPpOdooExportTests(unittest.TestCase):
    def test_product_dataset_round_trips_to_odoo_bom_files(self):
        dataset = shelf_dataset()
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory)
            files = generate_release_files(
                dataset=dataset,
                plan=shelf_plan(),
                client=ShelfClient(),
                output_dir=output,
            )
            manufacture = next(file.path for file in files if file.bom_type != "KIT")
            kit = next(file.path for file in files if file.bom_type == "KIT")

            wb = load_workbook(manufacture, data_only=True)
            try:
                ws = wb["BOM import"]
                self.assertEqual(ws["A2"].value, PP)
                self.assertEqual(
                    [(ws[f"D{row}"].value, ws[f"F{row}"].value) for row in range(2, 5)],
                    [(PART, 1), (PACKAGE, 0.4), (STICKER, 2)],
                )
                self.assertTrue(ws["K2"].value)
                self.assertTrue(ws["L2"].value)
                self.assertEqual(ws["M2"].value, "Lentynų pakavimas")
                self.assertEqual(ws["N2"].value, "Pakuotojai")
                self.assertEqual(ws["P2"].value, 1)
            finally:
                wb.close()

            wb = load_workbook(kit, data_only=True)
            try:
                ws = wb["BOM import"]
                self.assertEqual(ws["A2"].value, SHELF)
                self.assertEqual(
                    [(ws[f"D{row}"].value, ws[f"F{row}"].value) for row in range(2, 4)],
                    [(PP, 1), (PINS, 1)],
                )
                self.assertIsNone(ws["K2"].value)
                self.assertIsNone(ws["L2"].value)
            finally:
                wb.close()

            validation = validate_release_imports(
                dataset=dataset,
                release_id="SHELF-RELEASE",
                release_reference="SHELF-RELEASE",
                import_dir=output,
            )
            self.assertTrue(validation.passed, validation.errors)


if __name__ == "__main__":
    unittest.main()
