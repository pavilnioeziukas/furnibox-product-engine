"""Paruošia vieno Manufacture BOM pilotą su Production operacijų etalonu.

Programa skaito Stage Odoo ir Production etalonų Excel, bet Odoo nieko nekeičia.
"""

from __future__ import annotations

import re
from collections import defaultdict
from datetime import date
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bom_import_pilot_v1 import (
    calculate_levels,
    canon,
    load_bom_types,
    load_new_bom_graph,
    load_odoo_product_ids,
)
from config import load_settings
from odoo_client import OdooClient
from output_paths import environment_output_dir, environment_slug


IMPORT_HEADERS = [
    "default_code", "Product_tmpl_id", "qty",
    "BoM Lines/Component/Internal Reference", "Product_id", "product_qty",
    "BoM Type", "Reference", "mo_autodone_by_wo", "auto_plan",
    "operation_ids/name", "operation_ids/workcenter_id",
    "operation_ids/time_mode", "operation_ids/time_cycle_manual",
    "operation_ids/sequence",
]


def family_token(sku: str) -> str:
    """Grąžina pakategorės šeimą iš SKU, pvz. HIG, BAS, WAL arba HRD."""
    text = canon(sku)
    match = re.search(r"CAB\d{2}-([A-Z]+)", text)
    if match:
        return match.group(1)
    match = re.search(r"-(HRD|MIS|SLF)(\d+)", text)
    if match:
        return match.group(1)
    return ""


def choose_manufacture_pilot(parents, lines, bom_types) -> str:
    """Parenka mažiausią Manufacture BOM, kad pirmas testas būtų lengvas."""
    candidates = [
        sku for sku in parents
        if bom_types.get(sku) == "Manufacture this product" and lines.get(sku)
    ]
    if not candidates:
        raise ValueError("Nerastas nė vienas naujas Manufacture BOM su komponentais.")
    return min(candidates, key=lambda sku: (len(lines[sku]), sku))


def load_operation_templates(path: Path) -> dict[int, dict]:
    """Sugrupuoja Production operacijas į pilnus vieno BOM etalonus."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["BOM OPERATIONS"]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    grouped = defaultdict(list)
    metadata = {}
    for row in rows:
        record = dict(zip(headers, row))
        bom_id = int(record["BOM ID"])
        grouped[bom_id].append({
            "name": record["Operation"],
            "workcenter": record["Work Center"],
            "time_mode": record["Time Mode"],
            "time": record["Manual Time"],
            "sequence": record["Operation Sequence"],
        })
        metadata[bom_id] = {
            "sku": str(record["Parent SKU"] or "").strip(),
            "category_path": str(record["Category Path"] or "").strip(),
            "subcategory": str(record["Subcategory"] or "").strip(),
            "reference": str(record["BOM Reference"] or "").strip(),
        }
    wb.close()
    return {
        bom_id: {**metadata[bom_id], "operations": operations}
        for bom_id, operations in grouped.items()
    }


def stage_subcategory(client: OdooClient, pilot: str) -> str:
    """Nustato tikrą Stage produkto Odoo pakategorę pagal jos pilną kelią."""
    products = client.search_read_all(
        "product.product", [["default_code", "=ilike", pilot]],
        ["id", "default_code", "categ_id"], context={"active_test": False},
    )
    exact = next((row for row in products if canon(row.get("default_code")) == pilot), None)
    if not exact or not exact.get("categ_id"):
        raise ValueError(f"Stage nepavyko nustatyti produkto kategorijos: {pilot}")
    category_id = int(exact["categ_id"][0])
    categories = client.search_read_all(
        "product.category", [["id", "=", category_id]],
        ["id", "name", "complete_name"],
    )
    if not categories:
        raise ValueError(f"Stage nerasta produkto kategorija ID={category_id}")
    path = str(categories[0].get("complete_name") or categories[0].get("name") or "")
    return path.rsplit("/", 1)[-1].strip()


def choose_operation_template(pilot: str, subcategory: str, templates: dict[int, dict]):
    """Parenka artimiausią Production SKU tos pačios pakategorės ir šeimos."""
    candidates = [value for value in templates.values() if canon(value["subcategory"]) == canon(subcategory)]
    if not candidates:
        raise ValueError(f"Production etalonuose nėra operacijų pakategorei: {subcategory}")
    token = family_token(pilot)
    same_family = [value for value in candidates if family_token(value["sku"]) == token]
    pool = same_family or candidates
    pool.sort(
        key=lambda value: SequenceMatcher(None, canon(pilot), canon(value["sku"])).ratio(),
        reverse=True,
    )
    return pool[0]


def workcenter_references(client: OdooClient, names: set[str]) -> dict[str, str]:
    """Patikrina, kad kiekvienas darbo centro pavadinimas Stage yra unikalus.

    Stage darbo centrai neturi External ID, todėl Odoo importo many2one lauką
    susiejame pagal tikslų unikalų pavadinimą.
    """
    rows = client.search_read_all("mrp.workcenter", [], ["id", "name", "active"], context={"active_test": False})
    result = {}
    for name in names:
        matches = [
            row for row in rows
            if str(row.get("name") or "").strip() == name and row.get("active", True)
        ]
        if len(matches) > 1:
            raise ValueError(f"Stage yra keli aktyvūs darbo centrai tuo pačiu pavadinimu: {name}")
        if len(matches) == 1:
            result[name] = name
    return result


def write_output(path: Path, pilot: str, level: int, lines: list[dict], products,
                 subcategory: str, template: dict, workcenters: dict[str, str]) -> None:
    """Viename BOM bloke lygiagrečiai surašo komponentus ir operacijas."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"BOM_import(lv{level})"
    ws.append(IMPORT_HEADERS)
    parent = products[pilot]
    operations = template["operations"]
    row_count = max(len(lines), len(operations))
    reference = f"{date.today():%Y%m%d}_Pilot_MFG_{pilot}"
    for index in range(row_count):
        first = index == 0
        line = lines[index] if index < len(lines) else None
        operation = operations[index] if index < len(operations) else None
        component = products[line["component"]] if line else None
        ws.append([
            parent["display_sku"] if first else None,
            parent["template_xmlid"] if first else None,
            1 if first else None,
            component["display_sku"] if component else None,
            component["product_xmlid"] if component else None,
            line["quantity"] if line else None,
            "Manufacture this product" if first else None,
            reference if first else None,
            True if first else None,
            True if first and operations else None,
            operation["name"] if operation else None,
            workcenters[operation["workcenter"]] if operation else None,
            operation["time_mode"] if operation else None,
            operation["time"] if operation else None,
            operation["sequence"] if operation else None,
        ])

    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index in range(1, len(IMPORT_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(index)].width = 28

    check = wb.create_sheet("CHECK")
    check.append(["Patikra", "Reikšmė"])
    checks = [
        ("Aplinka", "Stage"), ("Pilotinis BOM", pilot), ("Lygis", f"lv{level}"),
        ("Pakategorė", subcategory), ("Production analogas", template["sku"]),
        ("Production BOM ID", template.get("reference") or ""),
        ("BOM tipas", "Manufacture this product"),
        ("Auto Finish MO by Last WO", True), ("Auto Plan", bool(operations)),
        ("Komponentų", len(lines)), ("Operacijų", len(operations)),
        ("Svarbu", "Odoo pakeitimų neatlikta – failas paruoštas importo testui"),
    ]
    for row in checks:
        check.append(row)
    for cell in check[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
    check.column_dimensions["A"].width = 34
    check.column_dimensions["B"].width = 75
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    base = Path(__file__).resolve().parent
    if environment_slug() != "stage":
        raise PermissionError("Manufacture BOM pilotas leidžiamas tik Stage.")
    output_dir = environment_output_dir(base)
    comparison_path = output_dir / "MAP_Comparison.xlsx"
    types_path = output_dir / "BOM_Type_Review.xlsx"
    references_path = base / "output" / "production" / "BOM_Operations_Reference.xlsx"
    output_path = output_dir / "BOM_Import_Manufacture_Pilot.xlsx"

    # 1 ŽINGSNIS: parenkame mažiausią naują Manufacture BOM.
    parents, all_lines = load_new_bom_graph(comparison_path)
    bom_types = load_bom_types(types_path)
    pilot = choose_manufacture_pilot(set(parents), all_lines, bom_types)
    levels = calculate_levels(set(parents), all_lines)
    lines = all_lines[pilot]

    # 2 ŽINGSNIS: Stage API pateikia produktų External ID ir patikrina
    # unikalius darbo centrų pavadinimus.
    settings = load_settings()
    client = OdooClient(settings)
    uid = client.authenticate()
    wanted = {pilot} | {row["component"] for row in lines}
    products = load_odoo_product_ids(client, wanted)
    missing = sorted(wanted - set(products))
    if missing:
        raise ValueError("Stage nerasti produktai: " + ", ".join(missing))
    subcategory = stage_subcategory(client, pilot)

    # 3 ŽINGSNIS: pasirenkame tos pačios pakategorės ir SKU šeimos Production
    # analogą bei kopijuojame visą jo operacijų rinkinį ir laikus.
    templates = load_operation_templates(references_path)
    template = choose_operation_template(pilot, subcategory, templates)
    workcenter_names = {operation["workcenter"] for operation in template["operations"]}
    workcenters = workcenter_references(client, workcenter_names)
    missing_workcenters = sorted(workcenter_names - set(workcenters))
    if missing_workcenters:
        raise ValueError("Stage nerasti aktyvūs darbo centrai: " + ", ".join(missing_workcenters))

    # 4 ŽINGSNIS: sukuriame importo failą. Jokių Odoo create/write nėra.
    write_output(
        output_path, pilot, levels[pilot], lines, products,
        subcategory, template, workcenters,
    )
    print("Prisijungta prie Stage Odoo. UID=", uid)
    print("\nMANUFACTURE BOM IMPORTO PILOTAS SUKURTAS")
    print("Failas:", output_path)
    print("Pilotinis SKU:", pilot)
    print("Pakategorė:", subcategory)
    print("Production analogas:", template["sku"])
    print("Auto Finish MO by Last WO: TAIP")
    print("Auto Plan:", "TAIP" if template["operations"] else "NE")
    print("Operacijų:", len(template["operations"]))
    print("Odoo pakeitimų neatlikta.")


if __name__ == "__main__":
    main()
