#!/usr/bin/env python3
"""Compare unique Reform product SKUs across three BOM Transformer versions.

Scope:
- includes every unique ``BOM SKU Code``;
- includes every unique component ``Part Code`` (therefore CABINET PART and
  SHELF PART are included);
- excludes Product Engine-only derived variants;
- counts a SKU once even when it appears in several BOMs or roles.

The script accepts both source Reform workbooks and transformed workbooks.
Source workbooks are read from ``BOM - Full DB`` (all ``Part N Code`` fields),
while transformed-only workbooks are read from ``BOM VERTICAL``.  This order is
intentional: a transformer file can contain stale data from another version.
It optionally checks SKU presence in an Odoo Snapshot workbook.
"""

from __future__ import annotations

import argparse
from io import BytesIO
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


NAVY = "17365D"
WHITE = "FFFFFF"
LIGHT_BLUE = "D9EAF7"
LIGHT_GREEN = "E2F0D9"
LIGHT_RED = "FCE4D6"


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def category(value: Any) -> str:
    result = clean(value).upper()
    return result if result and result != "0" else "NENURODYTA"


@dataclass
class Occurrences:
    bom_categories: Counter[str] = field(default_factory=Counter)
    part_categories: Counter[str] = field(default_factory=Counter)
    references: set[str] = field(default_factory=set)


@dataclass(frozen=True)
class Product:
    sku: str
    category: str
    source_type: str
    all_categories: str
    category_conflict: bool
    references: str


def header_map(ws, row_number: int) -> dict[str, int]:
    values = next(ws.iter_rows(min_row=row_number, max_row=row_number, values_only=True))
    return {clean(value): index for index, value in enumerate(values) if clean(value)}


def most_common(counter: Counter[str]) -> str:
    return sorted(counter.items(), key=lambda item: (-item[1], item[0]))[0][0]


def choose_sheet(workbook, requested: str | None = None) -> str:
    if requested:
        if requested not in workbook.sheetnames:
            raise ValueError(f"nėra nurodyto lapo {requested!r}")
        return requested
    for candidate in ("BOM - Full DB", "BOM VERTICAL"):
        if candidate in workbook.sheetnames:
            return candidate
    raise ValueError("nėra nei 'BOM - Full DB', nei 'BOM VERTICAL' lapo")


def read_reform(path: Path, sheet_name: str | None = None) -> dict[str, Product]:
    workbook = load_workbook(
        path, read_only=True, data_only=True, keep_vba=path.suffix.lower() == ".xlsm"
    )
    try:
        selected_sheet = choose_sheet(workbook, sheet_name)
    except ValueError as exc:
        raise ValueError(f"{path.name}: {exc}") from exc
    ws = workbook[selected_sheet]

    headers: dict[str, int] | None = None
    header_row = 0
    for candidate in range(1, min(ws.max_row, 30) + 1):
        mapped = header_map(ws, candidate)
        has_bom = "BOM SKU Code" in mapped or "SKU Code" in mapped
        has_parts = (
            {"Part Code", "Part Group"}.issubset(mapped)
            or any(
                key.startswith("Part ") and key.endswith(" Code")
                for key in mapped
            )
        )
        if has_bom and has_parts:
            headers, header_row = mapped, candidate
            break
    if headers is None:
        raise ValueError(
            f"{path.name} / {selected_sheet}: nerasti BOM SKU ir komponentų stulpeliai"
        )

    bom_code = headers.get("BOM SKU Code", headers.get("SKU Code"))
    assert bom_code is not None
    bom_category = headers.get("Product Category", headers.get("Product Catagory"))
    part_pairs: list[tuple[int, int]] = []
    if "Part Code" in headers and "Part Group" in headers:
        part_pairs.append((headers["Part Code"], headers["Part Group"]))
    for heading, code_index in headers.items():
        if heading.startswith("Part ") and heading.endswith(" Code"):
            group_heading = heading[:-4] + "Group"
            if group_heading in headers:
                pair = (code_index, headers[group_heading])
                if pair not in part_pairs:
                    part_pairs.append(pair)
    ref_index = headers.get("REF")
    if bom_category is None:
        raise ValueError(f"{path.name}: nerastas Product Category / Product Catagory")

    found: dict[str, Occurrences] = defaultdict(Occurrences)
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        ref = clean(row[ref_index]) if ref_index is not None and ref_index < len(row) else ""
        bom_sku = clean(row[bom_code]) if bom_code < len(row) else ""
        if bom_sku:
            found[bom_sku].bom_categories[category(row[bom_category])] += 1
            if ref:
                found[bom_sku].references.add(ref)

        for part_code, part_group in part_pairs:
            component_sku = clean(row[part_code]) if part_code < len(row) else ""
            if component_sku:
                found[component_sku].part_categories[category(row[part_group])] += 1
                if ref:
                    found[component_sku].references.add(ref)

    products: dict[str, Product] = {}
    for sku, occurrences in found.items():
        # A SKU's own BOM category is authoritative. Part Group is used only
        # when the SKU never appears as a BOM product.
        preferred = occurrences.bom_categories or occurrences.part_categories
        categories = sorted(
            set(occurrences.bom_categories) | set(occurrences.part_categories)
        )
        products[sku] = Product(
            sku=sku,
            category=most_common(preferred),
            source_type=(
                "BOM + PART"
                if occurrences.bom_categories and occurrences.part_categories
                else "BOM SKU"
                if occurrences.bom_categories
                else "PART CODE"
            ),
            all_categories=" | ".join(categories),
            category_conflict=len(categories) > 1,
            references=", ".join(sorted(occurrences.references)[:10]),
        )
    return products


def read_odoo_snapshot(path: Path | None) -> dict[str, dict[str, Any]]:
    if path is None:
        return {}
    # A Library materialized file can have a transfer suffix after ``.xlsx``.
    # Passing a binary stream avoids rejecting an otherwise valid workbook
    # solely because of that local filename suffix.
    workbook = load_workbook(BytesIO(path.read_bytes()), read_only=True, data_only=True)
    if "ODOO PRODUCTS" not in workbook.sheetnames:
        raise ValueError(f"{path.name}: nėra lapo 'ODOO PRODUCTS'")
    ws = workbook["ODOO PRODUCTS"]
    products: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = clean(row[1] if len(row) > 1 else None)
        if sku:
            products[sku] = {
                "odoo_id": row[0] if row else "",
                "active": row[3] if len(row) > 3 else "",
                "odoo_category": clean(row[7] if len(row) > 7 else None),
            }
    return products


def style_header(cells) -> None:
    fill = PatternFill("solid", fgColor=NAVY)
    for cell in cells:
        cell.fill = fill
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def autofit(ws, maximum: int = 45) -> None:
    for column in ws.columns:
        width = min(max(len(clean(cell.value)) for cell in column) + 2, maximum)
        ws.column_dimensions[get_column_letter(column[0].column)].width = max(width, 10)


def add_detail_sheet(wb: Workbook, name: str, skus: set[str], versions, odoo) -> None:
    ws = wb.create_sheet(name)
    headers = [
        "SKU", "Reform kategorija", "Šaltinio tipas", "Visos kategorijos",
        "Kategorijų konfliktas", "V6", "V8", "V10", "Yra produkcijoje",
        "Odoo kategorija", "Odoo ID", "Aktyvus", "REF (iki 10)",
    ]
    ws.append(headers)
    style_header(ws[1])
    for sku in sorted(skus, key=lambda value: (
        (
            versions["V10"].get(value)
            or versions["V8"].get(value)
            or versions["V6"][value]
        ).category,
        value,
    )):
        product = versions["V10"].get(sku) or versions["V8"].get(sku) or versions["V6"][sku]
        prod = odoo.get(sku, {})
        ws.append([
            sku, product.category, product.source_type, product.all_categories,
            product.category_conflict, sku in versions["V6"], sku in versions["V8"],
            sku in versions["V10"], sku in odoo if odoo else "",
            prod.get("odoo_category", ""), prod.get("odoo_id", ""),
            prod.get("active", ""), product.references,
        ])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autofit(ws)


def build_workbook(versions, sources, odoo, snapshot: Path | None, output: Path) -> None:
    sets = {version: set(products) for version, products in versions.items()}
    all_skus = set().union(*sets.values())
    categories = sorted({product.category for products in versions.values() for product in products.values()})

    wb = Workbook()
    ws = wb.active
    ws.title = "SUVESTINĖ"
    headers = [
        "Reform kategorija", "V6", "V8", "V10", "Nauji V8 vs V6",
        "Nauji V10 vs V6", "Nauji V10 vs V8", "Nebelikę V8 vs V6",
        "Nebelikę V10 vs V8", "V8 produkcijoje", "V10 produkcijoje",
    ]
    ws.append(headers)
    style_header(ws[1])
    for cat in categories:
        category_sets = {
            version: {sku for sku, product in products.items() if product.category == cat}
            for version, products in versions.items()
        }
        ws.append([
            cat, len(category_sets["V6"]), len(category_sets["V8"]),
            len(category_sets["V10"]), len(category_sets["V8"] - sets["V6"]),
            len(category_sets["V10"] - sets["V6"]),
            len(category_sets["V10"] - sets["V8"]),
            len(category_sets["V6"] - sets["V8"]),
            len(category_sets["V8"] - sets["V10"]),
            len(category_sets["V8"] & set(odoo)) if odoo else "",
            len(category_sets["V10"] & set(odoo)) if odoo else "",
        ])
    ws.append([
        "IŠ VISO", len(sets["V6"]), len(sets["V8"]), len(sets["V10"]),
        len(sets["V8"] - sets["V6"]), len(sets["V10"] - sets["V6"]),
        len(sets["V10"] - sets["V8"]), len(sets["V6"] - sets["V8"]),
        len(sets["V8"] - sets["V10"]),
        len(sets["V8"] & set(odoo)) if odoo else "",
        len(sets["V10"] & set(odoo)) if odoo else "",
    ])
    total_row = ws.max_row
    for cell in ws[total_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=7):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=9):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=LIGHT_RED)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{ws.max_row}"
    autofit(ws)

    add_detail_sheet(wb, "NAUJI V8 vs V6", sets["V8"] - sets["V6"], versions, odoo)
    add_detail_sheet(wb, "NAUJI V10 vs V6", sets["V10"] - sets["V6"], versions, odoo)
    add_detail_sheet(wb, "V10 POKYČIAI vs V8", sets["V10"] ^ sets["V8"], versions, odoo)
    add_detail_sheet(wb, "NEBELIKĘ", (sets["V6"] - sets["V8"]) | (sets["V6"] - sets["V10"]), versions, odoo)
    add_detail_sheet(wb, "VISOS KORTELĖS", all_skus, versions, odoo)

    meta = wb.create_sheet("ŠALTINIAI")
    meta.append(["Versija / taisyklė", "Reikšmė"])
    style_header(meta[1])
    for version in ("V6", "V8", "V10"):
        meta.append([version, str(sources[version].resolve())])
    meta.append(["Produkcija", str(snapshot.resolve()) if snapshot else "Netikrinta"])
    meta.append(["Įtraukta", "BOM SKU Code + Part Code; kiekvienas SKU skaičiuojamas vieną kartą"])
    meta.append(["Neįtraukta", "Tik Product Engine sugeneruojamos išvestinės kortelės"])
    autofit(meta, maximum=100)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--v6", required=True, type=Path)
    parser.add_argument("--v8", required=True, type=Path)
    parser.add_argument("--v10", required=True, type=Path)
    parser.add_argument("--snapshot", type=Path, help="Nebūtinas Odoo Snapshot.xlsx")
    parser.add_argument("--output", type=Path, default=Path("Reform_SKU_palyginimas.xlsx"))
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    sources = {"V6": args.v6, "V8": args.v8, "V10": args.v10}
    for label, path in sources.items():
        if not path.is_file():
            raise FileNotFoundError(f"{label}: failas nerastas: {path}")
    if args.snapshot and not args.snapshot.is_file():
        raise FileNotFoundError(f"Snapshot failas nerastas: {args.snapshot}")

    versions = {version: read_reform(path) for version, path in sources.items()}
    odoo = read_odoo_snapshot(args.snapshot)
    build_workbook(versions, sources, odoo, args.snapshot, args.output)
    sets = {version: set(products) for version, products in versions.items()}
    print(f"V6: {len(sets['V6'])}")
    print(f"V8: {len(sets['V8'])}")
    print(f"V10: {len(sets['V10'])}")
    print(f"Nauji V8 vs V6: {len(sets['V8'] - sets['V6'])}")
    print(f"Nauji V10 vs V6: {len(sets['V10'] - sets['V6'])}")
    print(f"Nauji V10 vs V8: {len(sets['V10'] - sets['V8'])}")
    print(f"Rezultatas: {args.output.resolve()}")


if __name__ == "__main__":
    main()
