"""Paruošia pasikeitusių esamų BOM naujų versijų importą Stage aplinkai.

Importo eiga sąmoningai padalinta į dvi dalis:
1. Dabartinei pasirinktai BOM versijai Sequence pakeičiamas į 10.
2. Sukuriama pilna nauja BOM versija su Sequence 0 ir Reform struktūra.

Programa pati Odoo nekeičia – tik sukuria peržiūros ir importo Excel failą.
"""

from __future__ import annotations

from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from bom_import_pilot_v1 import calculate_levels, canon, load_odoo_product_ids, read_sheet
from bom_import_pilot_v2 import load_operation_templates
from bom_import_v1 import (
    IMPORT_HEADERS,
    load_stage_subcategories,
    load_stage_workcenter_ids,
    style_sheet,
)
from config import load_settings
from odoo_client import OdooClient
from output_paths import environment_output_dir, environment_slug


VERSION_HEADERS = IMPORT_HEADERS[:7] + ["sequence"] + IMPORT_HEADERS[7:]


def find_reform_map(base: Path, output_dir: Path) -> Path:
    for path in (
        output_dir / "Reform_MAP.xlsx",
        base / "output" / "Reform_MAP.xlsx",
        base / "Reform_MAP.xlsx",
    ):
        if path.exists():
            return path
    raise FileNotFoundError("Nerastas Reform_MAP.xlsx. Pirmiausia vykdykite 4 žingsnį.")


def load_changed_parents(comparison_path: Path) -> set[str]:
    """Paimame tik esamus BOM, kuriems MAP palyginimas rado pakeitimų."""
    return {
        canon(row.get("Parent SKU"))
        for row in read_sheet(comparison_path, "BOM CHANGE SUMMARY")
        if canon(row.get("Parent SKU"))
    }


def load_reform_lines(path: Path) -> dict[str, list[dict]]:
    """Naujai versijai naudojame visą Reform BOM, ne vien pasikeitusias eilutes."""
    result = defaultdict(list)
    for row in read_sheet(path, "REFORM EDGES"):
        parent = canon(row.get("Parent SKU"))
        component = canon(row.get("Component SKU"))
        if parent and component:
            result[parent].append({"component": component, "quantity": row.get("Quantity")})
    return dict(result)


def load_selected_stage_boms(client: OdooClient, wanted: set[str]):
    """Stage parenka aktyvų BOM: mažiausia sequence, tada naujausias įrašas."""
    products = client.search_read_all(
        "product.product", [["default_code", "!=", False]],
        ["id", "default_code", "product_tmpl_id"], context={"active_test": False},
    )
    templates = defaultdict(list)
    for product in products:
        sku = canon(product.get("default_code"))
        value = product.get("product_tmpl_id")
        if sku in wanted and value:
            templates[int(value[0])].append(sku)
    boms = client.search_read_all(
        "mrp.bom", [["active", "=", True], ["product_tmpl_id", "in", sorted(templates)]],
        ["id", "code", "sequence", "type", "product_tmpl_id", "write_date"],
        order="product_tmpl_id asc, sequence asc, write_date desc, id desc",
        context={"active_test": False},
    )
    grouped = defaultdict(list)
    for bom in boms:
        template = bom.get("product_tmpl_id")
        if not template:
            continue
        for sku in templates.get(int(template[0]), []):
            grouped[sku].append(bom)
    selected = {}
    for sku, rows in grouped.items():
        rows.sort(key=lambda row: (
            int(row.get("sequence") or 0),
            str(row.get("write_date") or ""),
            int(row["id"]),
        ))
        lowest = int(rows[0].get("sequence") or 0)
        same = [row for row in rows if int(row.get("sequence") or 0) == lowest]
        same.sort(key=lambda row: (str(row.get("write_date") or ""), int(row["id"])), reverse=True)
        selected[sku] = same[0]
    return selected


def exact_operations_by_sku(path: Path) -> dict[str, dict]:
    """Esamai BOM versijai operacijas kopijuojame iš to paties Production SKU."""
    templates = load_operation_templates(path)
    return {canon(value["sku"]): value for value in templates.values()}


def prepare_versions(changed, reform_lines, levels, products, old_boms,
                     operation_by_sku, workcenters):
    ready, review, diagnostics = [], [], []
    for sku in sorted(changed, key=lambda value: (levels.get(value, 1), value)):
        messages = []
        parent = products.get(sku)
        old = old_boms.get(sku)
        lines = reform_lines.get(sku, [])
        technical_type = str((old or {}).get("type") or "")
        display_type = "KIT" if technical_type == "phantom" else "Manufacture this product" if technical_type == "normal" else ""
        operation_template = operation_by_sku.get(sku) if technical_type == "normal" else None
        operations = operation_template["operations"] if operation_template else []

        if not parent:
            messages.append("Stage nerastas BOM produktas")
        if not old:
            messages.append("Stage nerastas aktyvus senas BOM")
        if not lines:
            messages.append("Reform MAP nerastas pilnas BOM")
        if not display_type:
            messages.append("Nenustatytas seno BOM tipas")
        for line in lines:
            if line["component"] not in products:
                messages.append(f"Stage nerastas komponentas: {line['component']}")
        for operation in operations:
            if operation["workcenter"] not in workcenters:
                messages.append(f"Stage nerastas darbo centras: {operation['workcenter']}")

        record = {
            "sku": sku, "level": levels.get(sku, 1), "type": display_type,
            "old": old, "lines": lines, "operations": operations,
            "operation_template": operation_template,
            "status": "NOT READY" if messages else "READY",
            "message": "; ".join(dict.fromkeys(messages)),
        }
        review.append(record)
        if messages:
            for message in dict.fromkeys(messages):
                diagnostics.append([sku, f"lv{record['level']}", message])
        else:
            ready.append(record)
    return ready, review, diagnostics


def version_rows(record, products, workcenters):
    """Sukuria pilną naują BOM versiją su Sequence 0."""
    sku = record["sku"]
    parent = products[sku]
    lines = record["lines"]
    operations = record["operations"]
    count = max(len(lines), len(operations), 1)
    reference = f"{date.today():%Y%m%d}_Reform_{sku}"
    rows = []
    for index in range(count):
        first = index == 0
        line = lines[index] if index < len(lines) else None
        operation = operations[index] if index < len(operations) else None
        component = products[line["component"]] if line else None
        manufacture = record["type"] == "Manufacture this product"
        rows.append([
            parent["display_sku"] if first else None,
            parent["template_id"] if first else None,
            1 if first else None,
            component["display_sku"] if component else None,
            component["product_id"] if component else None,
            line["quantity"] if line else None,
            record["type"] if first else None,
            0 if first else None,
            reference if first else None,
            True if first and manufacture else None,
            True if first and operations else None,
            operation["name"] if operation else None,
            workcenters[operation["workcenter"]] if operation else None,
            operation["time_mode"] if operation else None,
            operation["time"] if operation else None,
            operation["sequence"] if operation else None,
        ])
    return rows


def write_output(path, ready, review, diagnostics, products, workcenters):
    wb = Workbook()
    wb.remove(wb.active)

    # Šį lapą importuojame PIRMĄ: dabartinis BOM lieka aktyvus, bet jo
    # prioritetas sumažinamas iš 0 į 10.
    ws = wb.create_sheet("1_OLD_BOM_SEQUENCE")
    ws.append([".id", "sequence"])
    for record in ready:
        ws.append([int(record["old"]["id"]), 10])
    style_sheet(ws)

    # Naujas pilnas BOM importuojamas tik po seno Sequence pakeitimo.
    for level in sorted({record["level"] for record in ready}):
        ws = wb.create_sheet(f"2_NEW_BOM_lv{level}")
        ws.append(VERSION_HEADERS)
        for record in (row for row in ready if row["level"] == level):
            for row in version_rows(record, products, workcenters):
                ws.append(row)
        style_sheet(ws)

    ws = wb.create_sheet("REVIEW", 0)
    ws.append(["Status", "Level", "Parent SKU", "BOM Type", "Old BOM ID", "Old Sequence", "Components", "Operations", "Message"])
    for record in review:
        old = record.get("old") or {}
        ws.append([
            record["status"], f"lv{record['level']}", record["sku"], record["type"],
            old.get("id", ""), old.get("sequence", ""), len(record["lines"]),
            len(record["operations"]), record["message"],
        ])
    style_sheet(ws)

    ws = wb.create_sheet("DIAGNOSTICS")
    ws.append(["Parent SKU", "Level", "Message"])
    for row in diagnostics:
        ws.append(row)
    style_sheet(ws)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main():
    base = Path(__file__).resolve().parent
    if environment_slug() != "stage":
        raise PermissionError("BOM versijų generatorius leidžiamas tik Stage.")
    output_dir = environment_output_dir(base)
    comparison_path = output_dir / "MAP_Comparison.xlsx"
    reform_path = find_reform_map(base, output_dir)
    operations_path = base / "output" / "production" / "BOM_Operations_Reference.xlsx"
    output_path = output_dir / "BOM_Version_Import.xlsx"

    # 1 ŽINGSNIS: nustatome pasikeitusius tėvus ir visą naujos versijos sudėtį.
    changed = load_changed_parents(comparison_path)
    reform_lines = load_reform_lines(reform_path)
    all_parents = set(reform_lines)
    levels = calculate_levels(all_parents, reform_lines)

    # 2 ŽINGSNIS: Stage pateikia dabartinių produktų ir senų BOM Database ID.
    settings = load_settings()
    client = OdooClient(settings)
    uid = client.authenticate()
    wanted = changed | {line["component"] for sku in changed for line in reform_lines.get(sku, [])}
    products = load_odoo_product_ids(client, wanted)
    old_boms = load_selected_stage_boms(client, changed)
    workcenters = load_stage_workcenter_ids(client)

    # 3 ŽINGSNIS: Manufacture versijoms kopijuojame to paties Production SKU
    # operacijas. KIT versijos operacijų neturi.
    operation_by_sku = exact_operations_by_sku(operations_path)
    ready, review, diagnostics = prepare_versions(
        changed, reform_lines, levels, products, old_boms,
        operation_by_sku, workcenters,
    )
    write_output(output_path, ready, review, diagnostics, products, workcenters)

    print("Prisijungta prie Stage Odoo. UID=", uid)
    print("\nBOM VERSIJŲ IMPORTO FAILAS SUKURTAS")
    print("Failas:", output_path)
    print("Esami BOM su pakeitimais:", len(review))
    print("Paruošti naujai versijai:", len(ready))
    print("Dar neparuošti:", len(review) - len(ready))
    print("Seno BOM Sequence: 10")
    print("Naujo BOM Sequence: 0")
    print("Odoo pakeitimų neatlikta.")


if __name__ == "__main__":
    main()
