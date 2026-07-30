from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from kit_reordering_rules_audit import (
    AuditError,
    analyze,
    build_report,
    load_settings_from_file,
    read_kit_skus,
)


class KitReorderingRulesAuditTests(unittest.TestCase):
    def test_reads_unique_parent_skus_and_closes_workbook(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "kit.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Product/Internal Reference", "BoM Lines/Quantity"])
            sheet.append(["SKU-1", 1])
            sheet.append(["SKU-1", 2])
            sheet.append(["SKU-2", 1])
            workbook.save(path)
            workbook.close()
            self.assertEqual(read_kit_skus(path), ["SKU-1", "SKU-2"])
            path.unlink()

    def test_rejects_stage_environment(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / ".env.stage"
            path.write_text(
                "ODOO_URL=https://stageodoo.furnibox.lt\n"
                "ODOO_DB=db\nODOO_LOGIN=user\nODOO_API_KEY=key\n",
                encoding="utf-8",
            )
            with self.assertRaises(AuditError):
                load_settings_from_file(path)

    def test_active_rule_is_blocker_but_archived_rule_is_not(self):
        products = [
            {
                "id": 10,
                "default_code": "SKU-1",
                "name": "One",
                "active": True,
                "product_tmpl_id": [100, "One"],
                "categ_id": [1, "Cat"],
            },
            {
                "id": 20,
                "default_code": "SKU-2",
                "name": "Two",
                "active": True,
                "product_tmpl_id": [200, "Two"],
                "categ_id": [1, "Cat"],
            },
        ]
        rules = [
            {"id": 1, "active": True, "product_id": [10, "SKU-1 One"]},
            {"id": 2, "active": False, "product_id": [20, "SKU-2 Two"]},
        ]
        result = analyze(["SKU-1", "SKU-2"], products, rules)
        self.assertEqual(result["blocker_skus"], ["SKU-1"])
        self.assertEqual(len(result["active_rules"]), 1)

    def test_report_contains_only_active_rules_as_import_blockers(self):
        products = [
            {
                "id": 10,
                "default_code": "SKU-1",
                "name": "One",
                "active": True,
                "product_tmpl_id": [100, "One"],
                "categ_id": [1, "Cat"],
            }
        ]
        base = {
            "product_id": [10, "SKU-1 One"],
            "warehouse_id": [1, "WH"],
            "location_id": [2, "Stock"],
            "company_id": [1, "Company"],
            "product_min_qty": 0,
            "product_max_qty": 0,
            "qty_multiple": 1,
            "trigger": "auto",
            "route_id": False,
        }
        rules = [
            {**base, "id": 1, "active": True},
            {**base, "id": 2, "active": False},
        ]
        with tempfile.TemporaryDirectory() as folder:
            output = Path(folder) / "audit.xlsx"
            result = build_report(
                output,
                Path("kit.xlsx"),
                "https://odoo.furnibox.lt",
                ["SKU-1"],
                products,
                rules,
            )
            workbook = load_workbook(output, read_only=True)
            try:
                self.assertEqual(workbook["IMPORT BLOCKERS"].max_row, 2)
                self.assertEqual(workbook["ALL RULES"].max_row, 3)
                self.assertEqual(result["status"], "BLOCKED")
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
