from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SOURCE_SKU_COLUMN = "Invoice lines/Product/Internal Reference"
SOURCE_REAL_PRICE_COLUMN = "Reali pirkimo kaina"
SOURCE_ADJUSTED_PRICE_COLUMN = "Adjustint kaina"

GENERATED_SHEET = "TAMARA ADJUSTMENTS"
GENERATED_SKU_COLUMN = "Internal Reference"
GENERATED_REAL_PRICE_COLUMN = "Real Purchase Price (reference)"
GENERATED_ADJUSTED_PRICE_COLUMN = "Adjusted Purchase Price"


@dataclass(frozen=True)
class PurchasePriceAdjustmentCandidate:
    sku: str
    excel_real_price: float | None
    current_adjustment: float | None
    new_adjustment: float
    status: str


def _headers(sheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value not in (None, "")
    }


def _to_float(
    value: Any,
    *,
    field_name: str,
    sku: str,
    allow_empty: bool = False,
) -> float | None:
    if value in (None, ""):
        if allow_empty:
            return None
        raise ValueError(
            f"SKU {sku}: laukas '{field_name}' negali būti tuščias."
        )

    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"SKU {sku}: laukas '{field_name}' nėra skaičius: {value!r}"
        ) from exc


def load_purchase_price_excel_adjustments(
    path: Path,
) -> dict[str, dict[str, float | None]]:
    if not path.exists():
        raise FileNotFoundError(
            f"Nerastas Pirkimo kain? Excel failas: {path}"
        )

    workbook = load_workbook(
        path,
        data_only=True,
        read_only=True,
    )

    if GENERATED_SHEET in workbook.sheetnames:
        sheet = workbook[GENERATED_SHEET]
    else:
        sheet = workbook.active
    columns = _headers(sheet)

    if {
        GENERATED_SKU_COLUMN,
        GENERATED_ADJUSTED_PRICE_COLUMN,
    }.issubset(columns):
        sku_column = GENERATED_SKU_COLUMN
        real_price_column = GENERATED_REAL_PRICE_COLUMN
        adjusted_price_column = GENERATED_ADJUSTED_PRICE_COLUMN
    else:
        sku_column = SOURCE_SKU_COLUMN
        real_price_column = SOURCE_REAL_PRICE_COLUMN
        adjusted_price_column = SOURCE_ADJUSTED_PRICE_COLUMN

    required_columns = (sku_column, adjusted_price_column)

    for required in required_columns:
        if required not in columns:
            workbook.close()
            raise ValueError(
                f"Pirkimo kain? Excel faile nerastas stulpelis: {required}"
            )

    result: dict[str, dict[str, float | None]] = {}
    duplicates: set[str] = set()

    for row_number in range(2, sheet.max_row + 1):
        raw_sku = sheet.cell(
            row_number,
            columns[sku_column],
        ).value

        if raw_sku in (None, ""):
            continue

        sku = str(raw_sku).strip()

        raw_adjusted_price = sheet.cell(
            row_number,
            columns[adjusted_price_column],
        ).value

        if raw_adjusted_price in (None, ""):
            continue

        if sku in result:
            duplicates.add(sku)
            continue

        excel_real_price = None
        if real_price_column in columns:
            excel_real_price = _to_float(
                sheet.cell(
                    row_number,
                    columns[real_price_column],
                ).value,
                field_name=real_price_column,
                sku=sku,
                allow_empty=True,
            )

        adjusted_price = _to_float(
            raw_adjusted_price,
            field_name=adjusted_price_column,
            sku=sku,
        )

        assert adjusted_price is not None

        # Zero cannot produce a releasable component price and must not
        # overwrite a newer positive Odoo purchase price. Treat it as an
        # unprovided Tamara price; the pricing engine will use its normal
        # fallback and report BLOCKED if that fallback is also non-positive.
        if adjusted_price <= 0:
            continue

        result[sku] = {
            "excel_real_price": excel_real_price,
            "new_adjustment": adjusted_price,
        }

    workbook.close()

    if duplicates:
        preview = ", ".join(sorted(duplicates)[:20])
        raise ValueError(
            "Pirkimo kain? Excel faile kartojasi SKU: "
            f"{preview}"
        )

    return result


def build_adjustment_preview(
    excel_adjustments: dict[str, dict[str, float | None]],
    current_adjustments: dict[str, dict[str, Any]],
) -> list[PurchasePriceAdjustmentCandidate]:
    rows: list[PurchasePriceAdjustmentCandidate] = []

    for sku in sorted(
        excel_adjustments,
        key=str.casefold,
    ):
        excel_row = excel_adjustments[sku]

        current_document = current_adjustments.get(sku)
        current_adjustment = None

        if current_document is not None:
            current_adjustment = float(
                current_document["adjusted_purchase_price"]
            )

        new_adjustment = float(
            excel_row["new_adjustment"]
        )

        if current_adjustment is None:
            status = "NEW"
        elif current_adjustment == new_adjustment:
            status = "SAME"
        else:
            status = "CHANGED"

        rows.append(
            PurchasePriceAdjustmentCandidate(
                sku=sku,
                excel_real_price=excel_row["excel_real_price"],
                current_adjustment=current_adjustment,
                new_adjustment=new_adjustment,
                status=status,
            )
        )

    return rows


def summarize_preview(
    rows: list[PurchasePriceAdjustmentCandidate],
) -> dict[str, int]:
    summary = {
        "TOTAL": len(rows),
        "NEW": 0,
        "CHANGED": 0,
        "SAME": 0,
    }

    for row in rows:
        summary[row.status] += 1

    return summary
