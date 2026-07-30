from __future__ import annotations

import argparse
import json
from collections import defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import load_settings
from odoo_client import OdooClient
from bom_release import load_latest_dataset_record
from output_paths import environment_output_dir
from operation_contract import (
    manufacture_operations_required,
    recognized_operation_names,
    required_operation_names,
)
QTY_TOLERANCE = 0.000001


def canon(value: Any) -> str:
    return str(value or "").strip().upper()


def number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def m2o_id(value: Any) -> int | None:
    if isinstance(value, (list, tuple)) and value:
        try:
            return int(value[0])
        except (TypeError, ValueError):
            return None
    return value if isinstance(value, int) else None


def load_dataset(path: Path | None) -> tuple[dict[str, Any], Path]:
    """Naudoja tą patį Validated Dataset šaltinį kaip BOM Release Analysis."""
    return load_latest_dataset_record(path)


def category(product: dict[str, Any]) -> str:
    return canon(
        product.get("reform_category")
        or product.get("product_type")
        or product.get("part_group")
        or product.get("category")
    )


def is_cabinet(product: dict[str, Any]) -> bool:
    return category(product) == "CABINETS"


def is_cabinet_part(product: dict[str, Any]) -> bool:
    return "CABINET PART" in category(product)


def is_kit(value: str) -> bool:
    return canon(value) in {"KIT", "PHANTOM"}


def component_map(rows: list[dict[str, Any]]) -> dict[str, float]:
    result: dict[str, float] = defaultdict(float)
    for row in rows:
        sku = canon(row.get("sku"))
        if sku:
            result[sku] += number(row.get("quantity"))
    return dict(result)


def same_qty(left: float, right: float) -> bool:
    return abs(left - right) <= QTY_TOLERANCE


def normalized_bom_type(value: str) -> str:
    value = canon(value)
    if value in {"KIT", "PHANTOM"}:
        return "KIT"
    if value in {"MANUFACTURE", "MANUFACTURE THIS PRODUCT", "NORMAL"}:
        return "MANUFACTURE"
    return value


@dataclass(frozen=True)
class Issue:
    test_code: str
    severity: str
    parent_sku: str
    related_sku: str
    message: str
    expected: str = ""
    actual: str = ""


class Acceptance:
    def __init__(self, dataset: dict[str, Any], client: OdooClient | None = None) -> None:
        self.dataset = dataset
        self.client = client
        self.issues: list[Issue] = []
        self.metrics: dict[str, int] = {}
        self.odoo_release_mismatches = 0
        sku_counts: dict[str, int] = defaultdict(int)
        for row in dataset["products"]:
            sku = canon(row.get("sku"))
            if sku:
                sku_counts[sku] += 1
        self.products = {
            canon(row.get("sku")): row
            for row in dataset["products"]
            if canon(row.get("sku"))
        }
        duplicates = {sku: count for sku, count in sku_counts.items() if count > 1}
        for sku, count in sorted(duplicates.items()):
            self.issues.append(Issue(
                "DUPLICATE_PRODUCT_SKU", "ERROR", sku, "",
                "Validated Dataset turi kelis produktus su tuo pačiu SKU.", "1", str(count),
            ))
        self.metrics["duplicate_product_skus"] = len(duplicates)

    def dataset_structures(self) -> dict[str, dict[str, Any]]:
        structures: dict[str, dict[str, Any]] = {}
        for sku, product in self.products.items():
            rows = product.get("components") or []
            # Product rows and BOM parent rows live in the same dataset. A plain
            # component product must not be treated as an expected empty BOM.
            if not rows and not str(product.get("bom_type") or "").strip():
                continue
            structures[sku] = {
                "bom_type": str(product.get("bom_type") or ""),
                "components": component_map(rows),
                "component_rows": rows,
                "reference": str(product.get("reference") or product.get("bom_reference") or ""),
                "sequence": product.get("sequence"),
            }
        return structures

    def odoo_structures(self, reference: str, sequence: int) -> dict[str, dict[str, Any]]:
        if self.client is None:
            raise RuntimeError("Nepateiktas OdooClient.")

        products = self.client.search_read_all(
            "product.product",
            [],
            ["id", "default_code", "product_tmpl_id", "active"],
            context={"active_test": False},
        )
        by_id = {int(row["id"]): row for row in products}
        by_template: dict[int, list[dict[str, Any]]] = defaultdict(list)
        for row in products:
            template_id = m2o_id(row.get("product_tmpl_id"))
            if template_id is not None:
                by_template[template_id].append(row)

        boms = self.client.search_read_all(
            "mrp.bom",
            [["active", "=", True]],
            ["id", "code", "type", "sequence", "product_id", "product_tmpl_id", "product_qty"],
        )

        parent_by_bom: dict[int, str] = {}
        structures: dict[str, dict[str, Any]] = {}
        expected = set(self.dataset_structures())
        selected: list[tuple[dict[str, Any], str]] = []
        candidates_by_parent: dict[str, list[dict[str, Any]]] = defaultdict(list)

        for bom in boms:
            parent = self._resolve_parent(bom, by_id, by_template)
            if not parent:
                continue
            bom_reference = str(bom.get("code") or "")
            bom_sequence = int(bom.get("sequence") or 0)
            if parent in expected:
                candidates_by_parent[parent].append(bom)

            # Release BOMs plus every active BOM of an expected parent are needed.
            # Otherwise a wrong reference/sequence would simply disappear from the
            # result and could incorrectly produce PASS.
            if parent not in expected and (not reference or bom_reference != reference):
                continue
            if parent in expected:
                if reference and bom_reference != reference:
                    continue
                if bom_sequence != sequence:
                    continue
            elif bom_reference != reference or bom_sequence != sequence:
                continue
            selected.append((bom, parent))

        for parent, candidates in sorted(candidates_by_parent.items()):
            sequence_zero = [bom for bom in candidates if int(bom.get("sequence") or 0) == 0]
            if len(sequence_zero) > 1:
                self.issues.append(Issue(
                    "MULTIPLE_SEQUENCE_ZERO", "ERROR", parent, "",
                    "Parent SKU turi kelis aktyvius BOM su sequence 0.",
                    "0 or 1", str(len(sequence_zero)),
                ))
            if any(
                str(bom.get("code") or "") == reference
                and int(bom.get("sequence") or 0) == sequence
                for bom in candidates
            ):
                continue
            self.odoo_release_mismatches += 1
            actual = sorted({
                f"{str(bom.get('code') or '')} @ {int(bom.get('sequence') or 0)}"
                for bom in candidates
            })
            self.issues.append(Issue(
                "WRONG_REFERENCE_OR_SEQUENCE", "ERROR", parent, "",
                "Parent turi aktyvių BOM, bet nė vienas neatitinka tikrinamo release reference ir sequence.",
                f"{reference} @ {sequence}", "; ".join(actual),
            ))

        duplicate_count: dict[str, int] = defaultdict(int)
        for bom, parent in selected:
            duplicate_count[parent] += 1
            parent_by_bom[int(bom["id"])] = parent
            structures[parent] = {
                "bom_type": str(bom.get("type") or ""),
                "components": defaultdict(float),
                "component_rows": [],
                "base_qty": number(bom.get("product_qty")) or 1.0,
                "reference": str(bom.get("code") or ""),
                "sequence": int(bom.get("sequence") or 0),
            }

        for parent, count in duplicate_count.items():
            if count > 1:
                self.issues.append(Issue(
                    "DUPLICATE_ACTIVE_BOM", "ERROR", parent, "",
                    "Tam pačiam Parent SKU rasti keli tikrinamo release/sequence BOM.",
                    "1", str(count),
                ))

        if parent_by_bom:
            lines = self.client.search_read_all(
                "mrp.bom.line",
                [["bom_id", "in", sorted(parent_by_bom)]],
                ["bom_id", "product_id", "product_qty"],
            )
            for line in lines:
                bom_id = m2o_id(line.get("bom_id"))
                product_id = m2o_id(line.get("product_id"))
                if bom_id not in parent_by_bom or product_id not in by_id:
                    continue
                parent = parent_by_bom[bom_id]
                component = canon(by_id[product_id].get("default_code"))
                if not component:
                    continue
                amount = number(line.get("product_qty")) / structures[parent]["base_qty"]
                structures[parent]["components"][component] += amount
                structures[parent]["component_rows"].append({"sku": component, "quantity": amount})

        for structure in structures.values():
            structure["components"] = dict(structure["components"])
        return structures

    @staticmethod
    def _resolve_parent(
        bom: dict[str, Any],
        by_id: dict[int, dict[str, Any]],
        by_template: dict[int, list[dict[str, Any]]],
    ) -> str:
        product_id = m2o_id(bom.get("product_id"))
        if product_id is not None:
            return canon(by_id.get(product_id, {}).get("default_code"))
        template_id = m2o_id(bom.get("product_tmpl_id"))
        if template_id is None:
            return ""
        candidates = [row for row in by_template.get(template_id, []) if canon(row.get("default_code"))]
        active = [row for row in candidates if row.get("active", True)]
        candidates = active or candidates
        return canon(candidates[0].get("default_code")) if len(candidates) == 1 else ""

    def run(self, structures: dict[str, dict[str, Any]]) -> None:
        self._operations()
        self._cabinet_pairs(structures)
        self._kit_structure(structures)
        self._fpack_apack_parts(structures)
        self._hrd_subset(structures)
        self._orphans(structures)
        self._basic_integrity(structures)

    def _operations(self) -> None:
        errors = 0
        for sku, product in sorted(self.products.items()):
            bom_type = normalized_bom_type(
                str(product.get("bom_type") or "")
            )
            operations = list(product.get("operations") or [])
            product_category = category(product)
            if (
                bom_type == "MANUFACTURE"
                and not operations
                and manufacture_operations_required(
                    sku=sku,
                    category=product_category,
                )
            ):
                errors += 1
                self.issues.append(Issue(
                    "MISSING_OPERATIONS", "ERROR", sku, "",
                    "Manufacture BOM neturi nė vienos operacijos.",
                    ">=1", "0",
                ))
            if bom_type == "KIT" and operations:
                errors += 1
                self.issues.append(Issue(
                    "KIT_HAS_OPERATIONS", "ERROR", sku, "",
                    "KIT BOM negali turėti gamybos operacijų.",
                    "0", str(len(operations)),
                ))

            names = recognized_operation_names(operations)
            required = required_operation_names(
                sku=sku,
                category=product_category,
            )
            missing = sorted(required - names)
            if missing:
                errors += 1
                self.issues.append(Issue(
                    "MISSING_REQUIRED_OPERATION", "ERROR", sku, "",
                    "Trūksta privalomų operacijų: "
                    + ", ".join(missing),
                    ", ".join(sorted(required)),
                    ", ".join(sorted(names)),
                ))
        self.metrics["operation_contract_errors"] = errors

    def compare_dataset_to_odoo(self, actual: dict[str, dict[str, Any]], reference: str, sequence: int) -> None:
        expected = self.dataset_structures()
        missing = sorted(set(expected) - set(actual))
        extra = sorted(set(actual) - set(expected))
        qty_errors = 0
        type_errors = 0
        reference_errors = self.odoo_release_mismatches

        for sku in missing:
            self.issues.append(Issue(
                "DATASET_ODOO_MISSING", "ERROR", sku, "",
                "Validated Dataset BOM nerastas Odoo release.", "BOM exists", "Missing",
            ))
        for sku in extra:
            self.issues.append(Issue(
                "DATASET_ODOO_EXTRA", "ERROR", sku, "",
                "Odoo release turi BOM, kurio nėra Validated Dataset.", "Missing", "BOM exists",
            ))

        for sku in sorted(set(expected) & set(actual)):
            left = expected[sku]
            right = actual[sku]
            if normalized_bom_type(left["bom_type"]) != normalized_bom_type(right["bom_type"]):
                type_errors += 1
                self.issues.append(Issue(
                    "DATASET_ODOO_TYPE", "ERROR", sku, "",
                    "BOM tipas Odoo nesutampa su Validated Dataset.",
                    normalized_bom_type(left["bom_type"]), normalized_bom_type(right["bom_type"]),
                ))
            if right.get("reference") != reference or right.get("sequence") != sequence:
                reference_errors += 1
                self.issues.append(Issue(
                    "DATASET_ODOO_RELEASE", "ERROR", sku, "",
                    "Odoo BOM reference arba sequence neatitinka tikrinamo release.",
                    f"reference={reference}; sequence={sequence}",
                    f"reference={right.get('reference')}; sequence={right.get('sequence')}",
                ))

            left_components = left["components"]
            right_components = right["components"]
            component_keys = set(left_components) | set(right_components)
            differences = [
                component for component in sorted(component_keys)
                if not same_qty(left_components.get(component, 0), right_components.get(component, 0))
            ]
            if differences:
                qty_errors += 1
                self.issues.append(Issue(
                    "DATASET_ODOO_COMPONENTS", "ERROR", sku, "",
                    "Odoo BOM komponentai arba kiekiai nesutampa su Validated Dataset.",
                    json.dumps(left_components, ensure_ascii=False, sort_keys=True),
                    json.dumps(right_components, ensure_ascii=False, sort_keys=True),
                ))

        self.metrics["dataset_odoo_missing"] = len(missing)
        self.metrics["dataset_odoo_extra"] = len(extra)
        self.metrics["dataset_odoo_qty_or_component_mismatch"] = qty_errors
        self.metrics["dataset_odoo_type_mismatch"] = type_errors
        self.metrics["dataset_odoo_reference_or_sequence_mismatch"] = reference_errors

    def _cabinet_pairs(self, structures: dict[str, dict[str, Any]]) -> None:
        errors = 0
        for sku, product in sorted(self.products.items()):
            if not is_cabinet(product):
                continue
            if sku.endswith("-A"):
                base = sku[:-2]
                if base not in self.products or not is_cabinet(self.products[base]):
                    errors += 1
                    self.issues.append(Issue(
                        "CABINET_PAIR", "ERROR", sku, base,
                        "Surinktas -A kabinetas neturi bazinio CABINET produkto poros.",
                    ))
                continue
            assembled = f"{sku}-A"
            if assembled not in self.products or not is_cabinet(self.products[assembled]):
                errors += 1
                self.issues.append(Issue(
                    "CABINET_PAIR", "ERROR", sku, assembled,
                    "Bazinis kabinetas neturi surinkto -A produkto poros.",
                ))
            if sku not in structures:
                errors += 1
                self.issues.append(Issue(
                    "CABINET_BOM_EXISTS", "ERROR", sku, "",
                    "Bazinis kabinetas neturi tikrinamo BOM.",
                ))
            if assembled not in structures:
                errors += 1
                self.issues.append(Issue(
                    "CABINET_BOM_EXISTS", "ERROR", assembled, "",
                    "Surinktas kabinetas neturi tikrinamo BOM.",
                ))
        self.metrics["cabinet_pair_errors"] = errors

    def _kit_structure(self, structures: dict[str, dict[str, Any]]) -> None:
        errors = 0
        for sku, product in sorted(self.products.items()):
            if not is_cabinet(product):
                continue
            structure = structures.get(sku)
            if not structure:
                continue
            components = sorted(structure["components"])
            if not is_kit(structure["bom_type"]):
                errors += 1
                self.issues.append(Issue(
                    "CABINET_KIT_TYPE", "ERROR", sku, "",
                    "Cabinet BOM tipas nėra KIT.", "KIT", structure["bom_type"],
                ))
            if len(components) != 2:
                errors += 1
                self.issues.append(Issue(
                    "CABINET_KIT_COUNT", "ERROR", sku, "",
                    "Cabinet KIT turi turėti tiksliai 2 tiesioginius komponentus.",
                    "2", str(len(components)),
                ))
                continue
            assembled = sku.endswith("-A")
            pack_prefix = "APACK-" if assembled else "FPACK-"
            packs = [value for value in components if value.startswith(pack_prefix)]
            hrds = [value for value in components if "HRD" in value]
            hrd_ok = len(hrds) == 1 and hrds[0].endswith("-A") == assembled
            if len(packs) != 1:
                errors += 1
                self.issues.append(Issue(
                    "CABINET_PACK_COMPONENT", "ERROR", sku, "",
                    "Nerastas vienas teisingas FPACK/APACK komponentas.",
                    pack_prefix, "; ".join(components),
                ))
            if not hrd_ok:
                errors += 1
                self.issues.append(Issue(
                    "CABINET_HRD_COMPONENT", "ERROR", sku, "",
                    "Nerastas vienas teisingas HRD/HRD-A komponentas.",
                    "HRD-A" if assembled else "HRD", "; ".join(components),
                ))
        self.metrics["cabinet_kit_structure_errors"] = errors

    def _part_map(self, structure: dict[str, Any]) -> dict[str, float]:
        return {
            sku: amount
            for sku, amount in structure["components"].items()
            if sku in self.products and is_cabinet_part(self.products[sku])
        }

    def _fpack_apack_parts(self, structures: dict[str, dict[str, Any]]) -> None:
        errors = 0
        for cabinet, product in sorted(self.products.items()):
            if not is_cabinet(product) or cabinet.endswith("-A"):
                continue
            base = structures.get(cabinet)
            assembled = structures.get(f"{cabinet}-A")
            if not base or not assembled:
                continue
            fpack = next((sku for sku in base["components"] if sku.startswith("FPACK-")), None)
            apack = next((sku for sku in assembled["components"] if sku.startswith("APACK-")), None)
            if not fpack or not apack:
                continue
            fpack_structure = structures.get(fpack)
            apack_structure = structures.get(apack)
            if not fpack_structure or not apack_structure:
                errors += 1
                self.issues.append(Issue(
                    "PACK_BOM_EXISTS", "ERROR", cabinet, f"{fpack} | {apack}",
                    "Nerastas FPACK arba APACK BOM detalių palyginimui.",
                ))
                continue
            left = self._part_map(fpack_structure)
            right = self._part_map(apack_structure)
            if left != right:
                errors += 1
                missing = sorted(set(left) - set(right))
                extra = sorted(set(right) - set(left))
                qty_diff = sorted(
                    sku for sku in set(left) & set(right)
                    if abs(left[sku] - right[sku]) > QTY_TOLERANCE
                )
                self.issues.append(Issue(
                    "FPACK_APACK_PARTS", "ERROR", cabinet, f"{fpack} | {apack}",
                    "FPACK ir APACK Cabinet Parts struktūra nesutampa.",
                    "Missing=0; Extra=0; Qty=0",
                    f"Missing={missing}; Extra={extra}; Qty={qty_diff}",
                ))
        self.metrics["fpack_apack_part_mismatch"] = errors

    def _hrd_subset(self, structures: dict[str, dict[str, Any]]) -> None:
        errors = 0
        for cabinet, product in sorted(self.products.items()):
            if not is_cabinet(product) or cabinet.endswith("-A"):
                continue
            base = structures.get(cabinet)
            assembled = structures.get(f"{cabinet}-A")
            if not base or not assembled:
                continue
            hrd = next((sku for sku in base["components"] if "HRD" in sku and not sku.endswith("-A")), None)
            hrd_a = next((sku for sku in assembled["components"] if "HRD" in sku and sku.endswith("-A")), None)
            if not hrd or not hrd_a:
                continue
            hrd_structure = structures.get(hrd)
            hrd_a_structure = structures.get(hrd_a)
            if not hrd_structure or not hrd_a_structure:
                errors += 1
                self.issues.append(Issue(
                    "HRD_BOM_EXISTS", "ERROR", cabinet, f"{hrd} | {hrd_a}",
                    "Nerastas HRD arba HRD-A BOM subset patikrai.",
                ))
                continue
            invalid = sorted(
                sku for sku, amount in hrd_a_structure["components"].items()
                if sku not in hrd_structure["components"]
                or amount - hrd_structure["components"].get(sku, 0) > QTY_TOLERANCE
            )
            if invalid:
                errors += 1
                self.issues.append(Issue(
                    "HRD_A_SUBSET", "ERROR", cabinet, f"{hrd} | {hrd_a}",
                    "HRD-A turi komponentų ar kiekių, kurių nėra baziniame HRD.",
                    "HRD-A subset of HRD", "; ".join(invalid),
                ))
        self.metrics["invalid_hrd_a_subset"] = errors

    def _orphans(self, structures: dict[str, dict[str, Any]]) -> None:
        used = {
            component
            for structure in structures.values()
            for component in structure["components"]
        }
        orphan = sorted(
            sku for sku, product in self.products.items()
            if is_cabinet_part(product) and sku not in used
        )
        for sku in orphan:
            self.issues.append(Issue(
                "ORPHAN_CABINET_PART", "ERROR", "", sku,
                "Aktuali Dataset Cabinet Part nenaudojama nė viename tikrinamame BOM.",
            ))
        self.metrics["orphan_cabinet_parts"] = len(orphan)

    def _basic_integrity(self, structures: dict[str, dict[str, Any]]) -> None:
        empty = 0
        invalid = 0
        duplicate = 0
        for parent, structure in structures.items():
            if not structure["components"]:
                empty += 1
                self.issues.append(Issue("EMPTY_BOM", "ERROR", parent, "", "BOM neturi komponentų."))
            for component, amount in structure["components"].items():
                if amount <= QTY_TOLERANCE:
                    invalid += 1
                    self.issues.append(Issue(
                        "INVALID_QTY", "ERROR", parent, component,
                        "BOM komponento kiekis nulinis arba neigiamas.", "> 0", str(amount),
                    ))
            occurrences: dict[str, int] = defaultdict(int)
            for row in structure.get("component_rows") or []:
                component = canon(row.get("sku"))
                if component:
                    occurrences[component] += 1
            for component, count in occurrences.items():
                if count > 1:
                    duplicate += 1
                    self.issues.append(Issue(
                        "DUPLICATE_COMPONENT", "ERROR", parent, component,
                        "Tas pats komponentas BOM įrašytas keliose eilutėse.", "1", str(count),
                    ))
        self.metrics["empty_boms"] = empty
        self.metrics["invalid_component_qty"] = invalid
        self.metrics["duplicate_components"] = duplicate
        # Validated Dataset contains BOM parents, not the complete product
        # master. Leaf components therefore do not have to occur in
        # dataset["products"]. For Odoo-sourced structures the component SKU
        # has already been resolved through product.product.
        self.metrics["missing_components"] = 0


def style_sheet(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in range(1, ws.max_column + 1):
        width = max(
            len(str(ws.cell(row=row, column=column).value or ""))
            for row in range(1, min(ws.max_row, 500) + 1)
        )
        ws.column_dimensions[get_column_letter(column)].width = min(width + 2, 70)


def write_report(
    *,
    source: str,
    reference: str,
    sequence: int | None,
    dataset: dict[str, Any],
    issues: list[Issue],
    metrics: dict[str, int],
    output: Path,
) -> tuple[Path, Path]:
    output = output.resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    status = "PASS" if not any(issue.severity == "ERROR" for issue in issues) else "FAIL"

    wb = Workbook()
    summary = wb.active
    summary.title = "SUMMARY"
    summary.append(["Metric", "Value"])
    summary.append(["Status", status])
    summary.append(["Source", source])
    summary.append(["Release Reference", reference])
    summary.append(["Sequence", sequence])
    summary.append(["Dataset ID", dataset.get("dataset_id", "")])
    summary.append(["Dataset Batch", dataset.get("batch_reference", "")])
    summary.append(["Errors", sum(issue.severity == "ERROR" for issue in issues)])
    summary.append(["Warnings", sum(issue.severity == "WARNING" for issue in issues)])
    for key, value in sorted(metrics.items()):
        summary.append([key, value])

    issues_ws = wb.create_sheet("ISSUES")
    issues_ws.append(["Severity", "Test Code", "Parent SKU", "Related SKU", "Message", "Expected", "Actual"])
    for issue in issues:
        issues_ws.append([
            issue.severity,
            issue.test_code,
            issue.parent_sku,
            issue.related_sku,
            issue.message,
            issue.expected,
            issue.actual,
        ])

    for ws in (summary, issues_ws):
        style_sheet(ws)

    wb.save(output)
    json_path = output.with_suffix(".json")
    json_path.write_text(
        json.dumps({
            "status": status,
            "source": source,
            "release_reference": reference,
            "sequence": sequence,
            "dataset_id": dataset.get("dataset_id", ""),
            "dataset_batch_reference": dataset.get("batch_reference", ""),
            "created_at_utc": datetime.now(timezone.utc).isoformat(),
            "metrics": metrics,
            "issues": [asdict(issue) for issue in issues],
        }, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return output, json_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Furnibox BOM release acceptance testai.")
    parser.add_argument("--source", choices=["dataset", "odoo"], default="dataset")
    parser.add_argument("--dataset", type=Path)
    parser.add_argument("--release-reference", default="")
    parser.add_argument("--sequence", type=int, default=0)
    parser.add_argument("--output-dir", type=Path)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.source == "odoo" and not args.release_reference.strip():
        raise SystemExit("--source odoo privalo turėti --release-reference.")
    dataset, dataset_path = load_dataset(args.dataset)
    client = None

    if args.source == "odoo":
        settings = load_settings()
        client = OdooClient(settings)
        uid = client.authenticate()
        print(f"Prisijungta prie Odoo. UID={uid}")

    acceptance = Acceptance(dataset, client)
    structures = (
        acceptance.dataset_structures()
        if args.source == "dataset"
        else acceptance.odoo_structures(args.release_reference, args.sequence)
    )
    acceptance.run(structures)
    if args.source == "odoo":
        acceptance.compare_dataset_to_odoo(structures, args.release_reference, args.sequence)

    base = Path(__file__).resolve().parent
    output_dir = args.output_dir.resolve() if args.output_dir else environment_output_dir(base)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    ref = (args.release_reference or "NO_REFERENCE").replace("/", "_")
    output = output_dir / f"BOM_Release_Acceptance_{args.source.upper()}_{ref}_{stamp}.xlsx"

    xlsx_path, json_path = write_report(
        source=args.source.upper(),
        reference=args.release_reference,
        sequence=args.sequence if args.source == "odoo" else None,
        dataset=dataset,
        issues=acceptance.issues,
        metrics=acceptance.metrics,
        output=output,
    )

    errors = sum(issue.severity == "ERROR" for issue in acceptance.issues)
    status = "PASS" if errors == 0 else "FAIL"
    print()
    print("=" * 80)
    print("BOM RELEASE ACCEPTANCE")
    print("=" * 80)
    print("Dataset:", dataset_path)
    print("Source:", args.source.upper())
    print("Status:", status)
    print("Errors:", errors)
    for key, value in sorted(acceptance.metrics.items()):
        print(f"{key:<38} {value}")
    print("Excel:", xlsx_path)
    print("JSON:", json_path)

    if status != "PASS":
        raise SystemExit(2)


if __name__ == "__main__":
    main()
