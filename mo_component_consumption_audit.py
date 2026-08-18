"""Read-only audit of completed MOs with under-consumed components."""
from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


EPSILON = 1e-6


def m2o_id(value: Any) -> int | None:
    return int(value[0]) if isinstance(value, (list, tuple)) and value else None


def m2o_name(value: Any) -> str:
    return str(value[1]) if isinstance(value, (list, tuple)) and len(value) > 1 else ""


def qty(row: dict[str, Any], *names: str) -> float:
    for name in names:
        if name in row:
            return float(row.get(name) or 0.0)
    return 0.0


@dataclass(frozen=True)
class ConsumptionGap:
    mo_id: int
    mo: str
    finished_product_sku: str
    finished_product: str
    produced_qty: float
    bom: str
    completion_date: str
    company: str
    component_sku: str
    component: str
    planned_qty: float
    consumed_qty: float
    missing_qty: float
    uom: str
    move_state: str


def detect_gaps(
    productions: Iterable[dict[str, Any]],
    moves: Iterable[dict[str, Any]],
    products_by_id: dict[int, dict[str, Any]],
) -> list[ConsumptionGap]:
    productions_by_move: dict[int, dict[str, Any]] = {}
    for production in productions:
        for move_id in production.get("move_raw_ids", []):
            productions_by_move[int(move_id)] = production

    gaps: list[ConsumptionGap] = []
    for move in moves:
        production = productions_by_move.get(int(move["id"]))
        if not production:
            continue
        planned = qty(move, "product_uom_qty")
        consumed = qty(move, "quantity", "quantity_done", "qty_done")
        missing = planned - consumed
        if planned <= EPSILON or missing <= EPSILON:
            continue

        component_id = m2o_id(move.get("product_id"))
        finished_id = m2o_id(production.get("product_id"))
        component = products_by_id.get(component_id or -1, {})
        finished = products_by_id.get(finished_id or -1, {})
        gaps.append(ConsumptionGap(
            mo_id=int(production["id"]),
            mo=str(production.get("name") or production["id"]),
            finished_product_sku=str(finished.get("default_code") or finished_id or ""),
            finished_product=str(finished.get("display_name") or m2o_name(production.get("product_id"))),
            produced_qty=qty(production, "qty_produced", "product_qty"),
            bom=m2o_name(production.get("bom_id")),
            completion_date=str(production.get("date_finished") or ""),
            company=m2o_name(production.get("company_id")),
            component_sku=str(component.get("default_code") or component_id or ""),
            component=str(component.get("display_name") or m2o_name(move.get("product_id"))),
            planned_qty=planned,
            consumed_qty=consumed,
            missing_qty=missing,
            uom=m2o_name(move.get("product_uom")) or m2o_name(move.get("product_uom_id")),
            move_state=str(move.get("state") or ""),
        ))
    return sorted(gaps, key=lambda row: (row.completion_date, row.mo, row.component_sku))


def _fields(client, model: str) -> set[str]:
    if client.uid is None:
        client.authenticate()
    return set(client.models.execute_kw(
        client.settings.db, client.uid, client.settings.api_key,
        model, "fields_get", [], {"attributes": ["type"]},
    ))


def collect_data(client, days: int = 550) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[int, dict[str, Any]], str]:
    mo_fields = _fields(client, "mrp.production")
    move_fields = _fields(client, "stock.move")
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")
    domain: list[list[Any]] = [["state", "=", "done"]]
    if "date_finished" in mo_fields:
        domain.append(["date_finished", ">=", cutoff])
    productions = client.search_read_all(
        "mrp.production", domain,
        [field for field in [
            "name", "state", "product_id", "product_qty", "qty_produced",
            "bom_id", "date_finished", "company_id", "move_raw_ids",
        ] if field in mo_fields],
        order="date_finished asc, id asc" if "date_finished" in mo_fields else "id asc",
    )
    move_ids = sorted({int(move_id) for mo in productions for move_id in mo.get("move_raw_ids", [])})
    query_fields = [field for field in [
        "product_id", "product_uom_qty", "quantity", "quantity_done",
        "qty_done", "product_uom", "product_uom_id", "state",
    ] if field in move_fields]
    moves: list[dict[str, Any]] = []
    for offset in range(0, len(move_ids), 1000):
        moves.extend(client.search_read_all(
            "stock.move", [["id", "in", move_ids[offset:offset + 1000]]],
            query_fields,
        ))
    product_ids = sorted({
        product_id for product_id in
        [m2o_id(mo.get("product_id")) for mo in productions]
        + [m2o_id(move.get("product_id")) for move in moves]
        if product_id
    })
    products = client.search_read_all(
        "product.product", [["id", "in", product_ids]],
        ["default_code", "display_name"], context={"active_test": False},
    ) if product_ids else []
    return productions, moves, {int(row["id"]): row for row in products}, cutoff


def _write_xlsx(path: Path, rows: list[dict[str, Any]], summary: dict[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Santrauka"
    summary_sheet.sheet_view.showGridLines = False
    summary_sheet.append(["Užbaigtų MO komponentų sunaudojimo auditas", ""])
    summary_sheet.merge_cells("A1:B1")
    summary_sheet["A1"].font = Font(size=16, bold=True, color="14532D")
    summary_sheet["A1"].fill = PatternFill("solid", fgColor="DCFCE7")
    summary_sheet["A1"].alignment = Alignment(vertical="center")
    summary_sheet.row_dimensions[1].height = 28
    summary_rows = [
        ("Sugeneruota (UTC)", summary["generated_at"]),
        ("Tikrintas laikotarpis, dienomis", summary["days_checked"]),
        ("Tikrinta nuo (UTC)", summary["completion_date_from"]),
        ("Patikrinta užbaigtų MO", summary["completed_mos_checked"]),
        ("Patikrinta komponentų judėjimų", summary["raw_component_moves_checked"]),
        ("MO su per mažu sunaudojimu", summary["mos_with_short_consumption"]),
        ("Komponentų neatitikimų", summary["component_gaps"]),
    ]
    for item in summary_rows:
        summary_sheet.append(item)
    for cell in summary_sheet["A"]:
        cell.font = Font(bold=cell.row > 1)
    summary_sheet.column_dimensions["A"].width = 38
    summary_sheet.column_dimensions["B"].width = 34

    sheet = workbook.create_sheet("Neatitikimai")
    sheet.sheet_view.showGridLines = False
    headers = [
        "MO", "Gatavo produkto kodas", "Gatavas produktas", "Pagaminta",
        "BOM", "Užbaigimo data", "Įmonė", "Komponento kodas", "Komponentas",
        "Planuota", "Sunaudota", "Trūksta", "Matavimo vnt.", "Judėjimo būsena",
    ]
    keys = [
        "mo", "finished_product_sku", "finished_product", "produced_qty",
        "bom", "completion_date", "company", "component_sku", "component",
        "planned_qty", "consumed_qty", "missing_qty", "uom", "move_state",
    ]
    sheet.append(headers)
    for row in rows:
        values = [row.get(key, "") for key in keys]
        if values[5]:
            try:
                values[5] = datetime.fromisoformat(str(values[5]))
            except ValueError:
                pass
        sheet.append(values)

    header_fill = PatternFill("solid", fgColor="14532D")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 32
    sheet.freeze_panes = "A2"
    widths = [16, 22, 42, 12, 42, 20, 22, 20, 42, 12, 12, 12, 16, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for column in (4, 10, 11, 12):
        for cell in sheet.iter_cols(min_col=column, max_col=column, min_row=2, max_row=sheet.max_row):
            cell[0].number_format = "#,##0.00"
    for cell in sheet.iter_cols(min_col=6, max_col=6, min_row=2, max_row=sheet.max_row):
        cell[0].number_format = "yyyy-mm-dd hh:mm"
    for cell in sheet.iter_cols(min_col=12, max_col=12, min_row=2, max_row=sheet.max_row):
        cell[0].fill = PatternFill("solid", fgColor="FEE2E2")
        cell[0].font = Font(color="991B1B", bold=True)
    if rows:
        table = Table(displayName="MOComponentConsumptionGaps", ref=f"A1:N{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        sheet.add_table(table)
    else:
        sheet.append(["Neatitikimų nerasta"] + [""] * (len(headers) - 1))
        sheet.auto_filter.ref = f"A1:N{sheet.max_row}"
    workbook.save(path)


def run_mo_component_consumption_audit(client, output_dir: Path, days: int = 550) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    productions, moves, products_by_id, cutoff = collect_data(client, days=days)
    gaps = detect_gaps(productions, moves, products_by_id)
    rows = [asdict(row) for row in gaps]
    summary = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "read_only": True,
        "days_checked": days,
        "completion_date_from": cutoff,
        "completed_mos_checked": len(productions),
        "raw_component_moves_checked": len(moves),
        "mos_with_short_consumption": len({row.mo_id for row in gaps}),
        "component_gaps": len(gaps),
        "total_missing_qty": sum(row.missing_qty for row in gaps),
    }
    payload = {"summary": summary, "gaps": rows}
    (output_dir / "mo_component_consumption_audit.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8",
    )
    _write_xlsx(output_dir / "mo_component_consumption_gaps.xlsx", rows, summary)
    lines = [
        "# Užbaigtų MO komponentų sunaudojimo auditas", "",
        "**Tik skaitymas:** ataskaita Odoo duomenų nekeičia.", "",
        "## Santrauka", "",
        f"- Tikrintas laikotarpis: paskutinės {days} dienos (nuo {cutoff} UTC)",
        f"- Patikrinta užbaigtų MO: {len(productions)}",
        f"- MO su per mažu sunaudojimu: {summary['mos_with_short_consumption']}",
        f"- Trūkstamos komponentų eilutės: {len(gaps)}", "",
        "## Neatitikimai", "",
        "| MO | Gatavas produktas | Pagaminta | BOM | Komponentas | Planuota | Sunaudota | Trūksta | Data |",
        "|---|---|---:|---|---|---:|---:|---:|---|",
    ]
    for row in gaps:
        lines.append(
            f"| {row.mo} | {row.finished_product_sku} {row.finished_product} | {row.produced_qty:g} | "
            f"{row.bom} | {row.component_sku} {row.component} | {row.planned_qty:g} | "
            f"{row.consumed_qty:g} | {row.missing_qty:g} | {row.completion_date} |"
        )
    if not gaps:
        lines.append("| – | – | 0 | – | – | 0 | 0 | 0 | Neatitikimų nerasta |")
    (output_dir / "mo_component_consumption_audit.md").write_text(
        "\n".join(lines) + "\n", encoding="utf-8",
    )
    return summary
