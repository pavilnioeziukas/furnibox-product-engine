from product_lifecycle_audit import (
    STATUS_ARCHIVED,
    STATUS_CANDIDATE,
    STATUS_IN_USE,
    STATUS_KEEP,
    STATUS_PRICING,
    build_audit,
    write_workbook,
)
from openpyxl import load_workbook


def production_product(sku, product_id, active=True):
    return {"sku": sku, "id": product_id, "active": active, "name": sku}


def test_lifecycle_statuses_are_conservative():
    dataset = {"product_catalog": [{"sku": "CURRENT"}]}
    production = {
        "products": [
            production_product("CURRENT", 1),
            production_product("ARCHIVED", 2, False),
            production_product("IN-STOCK", 3),
            production_product("PRICED", 4),
            production_product("OLD", 5),
        ],
        "boms": [],
    }
    config = {"bom_products": [{"sku": "PRICED"}]}
    usage = {
        "stock": {3: 2.0}, "sales": set(), "purchases": set(),
        "manufacturing": set(), "raw_material": set(),
    }
    rows = {row["sku"]: row for row in build_audit(dataset, production, config, usage)}
    assert rows["CURRENT"]["status"] == STATUS_KEEP
    assert rows["ARCHIVED"]["status"] == STATUS_ARCHIVED
    assert rows["IN-STOCK"]["status"] == STATUS_IN_USE
    assert rows["PRICED"]["status"] == STATUS_PRICING
    assert rows["OLD"]["status"] == STATUS_CANDIDATE


def test_active_bom_dependency_blocks_archival():
    dataset = {"product_catalog": []}
    production = {
        "products": [
            production_product("PARENT", 1),
            production_product("COMPONENT", 2),
        ],
        "boms": [{
            "id": 7, "sku": "PARENT", "active": True,
            "components": [{"component_sku": "COMPONENT", "quantity": 1}],
        }],
    }
    usage = {
        "stock": {}, "sales": set(), "purchases": set(),
        "manufacturing": set(), "raw_material": set(),
    }
    rows = {row["sku"]: row for row in build_audit(dataset, production, {}, usage)}
    assert rows["PARENT"]["status"] == STATUS_CANDIDATE
    assert rows["PARENT"]["active_bom_ids"] == [7]
    assert rows["COMPONENT"]["status"] == STATUS_IN_USE
    assert rows["COMPONENT"]["used_by_active_boms"] == ["PARENT"]


def test_audit_workbook_has_control_sheets_and_filters(tmp_path):
    row = {
        "sku": "OLD", "name": "Old", "status": STATUS_CANDIDATE,
        "decision": "Review", "active": True, "in_target_dataset": False,
        "in_pricing_config": False, "active_bom_ids": [7],
        "used_by_active_boms": [], "internal_stock": 0.0,
        "open_sales": False, "open_purchases": False,
        "open_manufacturing": False, "required_by_open_manufacturing": False,
        "evidence": "",
    }
    output = tmp_path / "audit.xlsx"
    write_workbook(output, [row])
    workbook = load_workbook(output)
    assert workbook.sheetnames == [
        "SUMMARY", "ARCHIVE CANDIDATES", "IN USE",
        "PRICING CONFIG REVIEW", "EVIDENCE",
    ]
    sheet = workbook["ARCHIVE CANDIDATES"]
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:K2"
    assert sheet["A2"].value == "OLD"
