from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from cabinet_parts_price_v1 import (
    DEFAULT_PARAMETERS,
    PriceParameters,
    build_workbook,
    calculate_unit_price,
    furnix_transfer_price,
)


class CabinetPartPriceCalculationTests(unittest.TestCase):
    def test_standard_ww_part_matches_reference_formula(self):
        result = calculate_unit_price("EU-SIDE-SREW-800x590-WW", (800, 590))
        expected = 0.472 * (
            DEFAULT_PARAMETERS.processing_rate_per_m2
            + DEFAULT_PARAMETERS.ww_material_rate_per_m2
        ) + 1
        self.assertAlmostEqual(result.unit_price, expected, places=10)
        self.assertEqual(result.part_type, "STANDARD")

    def test_back_part_uses_only_back_rate_and_small_part_surcharge(self):
        result = calculate_unit_price("EU-BACK-SREW-787x179-WW", (787, 179))
        expected = 0.140873 * DEFAULT_PARAMETERS.back_rate_per_m2 + 1
        self.assertAlmostEqual(result.unit_price, expected, places=10)
        self.assertEqual(result.material_rate_per_m2, 0)

    def test_exact_threshold_does_not_receive_surcharge(self):
        result = calculate_unit_price("TEST-1000x500-BB", (1000, 500))
        self.assertEqual(result.small_part_surcharge, 0)

    def test_unknown_standard_color_is_not_silently_priced(self):
        with self.assertRaisesRegex(ValueError, "nežinomas spalvos kodas"):
            calculate_unit_price("TEST-100x100-XX", (100, 100))

    def test_furnix_markup_is_applied_to_unit_cost(self):
        parameters = PriceParameters(furnix_markup_percent=15)
        markup_eur, transfer_price = furnix_transfer_price(10, parameters)
        self.assertEqual(markup_eur, 1.5)
        self.assertEqual(transfer_price, 11.5)


class CabinetPartPriceWorkbookTests(unittest.TestCase):
    def test_workbook_contains_calculated_unique_prices_and_diagnostics(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            source = Path(temporary_directory) / "MAP_Comparison.xlsx"
            output = Path(temporary_directory) / "prices.xlsx"
            wb = Workbook()
            products = wb.active
            products.title = "NEW PRODUCTS"
            products.append(["SKU", "Category", "Product Name 1", "Product Name 2", "Required Action"])
            products.append(["EU-SIDE-SREW-800x590-WW", "All / CABINET PART", "", "", "CREATE PRODUCT"])
            products.append(["EU-BACK-SREW-787x179-WW", "All / CABINET PART", "", "", "CREATE PRODUCT"])
            products.append(["EU-PART-100x100-XX", "All / CABINET PART", "", "", "CREATE PRODUCT"])
            products.append(["NOT-A-PART-100x100-WW", "All / CABINET", "", "", "CREATE PRODUCT"])
            lines = wb.create_sheet("NEW BOM LINES")
            lines.append(["Parent SKU", "Component SKU", "Quantity", "Required Action"])
            lines.append(["FPACK-TEST", "EU-SIDE-SREW-800x590-WW", 2, "ADD BOM LINE"])
            lines.append(["FPACK-TEST", "EU-BACK-SREW-787x179-WW", 1, "ADD BOM LINE"])
            lines.append(["FPACK-TEST", "EU-PART-100x100-XX", 1, "ADD BOM LINE"])
            wb.save(source)

            parameters = PriceParameters(furnix_markup_percent=15)
            rows, unique_parts, fpack_count, diagnostics = build_workbook(
                source, output, parameters=parameters
            )

            self.assertEqual((rows, unique_parts, fpack_count, diagnostics), (3, 3, 1, 1))
            result = load_workbook(output, data_only=True)
            self.assertEqual(
                result.sheetnames,
                [
                    "CABINET PART PRICES",
                    "FPACK PRICE BREAKDOWN",
                    "CALCULATION DETAILS",
                    "DIAGNOSTICS",
                    "INFO",
                    "PARAMETERS",
                ],
            )
            prices = result["CABINET PART PRICES"]
            self.assertEqual(prices.max_row, 3)

            price_headers = {
                cell.value: cell.column
                for cell in prices[1]
            }
            self.assertIn("Internal Reference", price_headers)
            self.assertIn("Furnix Unit Cost", price_headers)
            self.assertIn("Furnix Sales Price to Furnibox", price_headers)
            self.assertIn("Product Status", price_headers)

            price_rows = {
                prices.cell(row=row, column=price_headers["Internal Reference"]).value: row
                for row in range(2, prices.max_row + 1)
            }
            self.assertEqual(
                set(price_rows),
                {
                    "EU-SIDE-SREW-800x590-WW",
                    "EU-BACK-SREW-787x179-WW",
                },
            )

            for row in price_rows.values():
                self.assertIsNotNone(
                    prices.cell(row=row, column=price_headers["Furnix Unit Cost"]).value
                )
                self.assertAlmostEqual(
                    prices.cell(
                        row=row,
                        column=price_headers["Furnix Sales Price to Furnibox"],
                    ).value,
                    prices.cell(row=row, column=price_headers["Furnix Unit Cost"]).value
                    * 1.15,
                    places=4,
                )
                self.assertEqual(
                    prices.cell(row=row, column=price_headers["Product Status"]).value,
                    "ODOO NOT CHECKED",
                )

            errors = result["DIAGNOSTICS"]
            self.assertEqual(errors["A2"].value, "PRICE CALCULATION ERROR")

            breakdown = result["FPACK PRICE BREAKDOWN"]
            breakdown_headers = {cell.value: cell.column for cell in breakdown[1]}
            component_total = sum(
                breakdown.cell(
                    row=row,
                    column=breakdown_headers["Component Total Purchase Price, EUR"],
                ).value
                for row in range(2, breakdown.max_row + 1)
            )
            self.assertAlmostEqual(
                component_total,
                breakdown.cell(
                    row=2,
                    column=breakdown_headers["FPACK Cabinet Parts Purchase Price, EUR"],
                ).value,
                places=4,
            )


if __name__ == "__main__":
    unittest.main()
