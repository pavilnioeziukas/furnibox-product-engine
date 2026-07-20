"""Paruošia vieno naujo BOM bandomąjį Odoo importo failą Stage aplinkai.

Programa TIK nuskaito Odoo ir sukuria Excel. Ji nekuria ir nekeičia BOM Odoo.
"""

from __future__ import annotations

import argparse
from collections import defaultdict, deque
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import load_settings
from odoo_client import OdooClient
from output_paths import environment_output_dir, environment_slug


IMPORT_HEADERS = [
    "default_code",
    "Product_tmpl_id",
    "qty",
    "BoM Lines/Component/Internal Reference",
    "Product_id",
    "product_qty",
    "BoM Type",
    "Reference",
]


def canon(value: object) -> str:
    """SKU palyginimui pašalina tarpus ir suvienodina raidžių dydį."""
    return str(value or "").strip().upper()


def read_sheet(path: Path, sheet_name: str) -> list[dict]:
    """Nuskaito Excel lapą ir kiekvieną eilutę grąžina kaip žodyną."""
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Faile {path.name} nėra lapo {sheet_name!r}.")
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        wb.close()
        raise ValueError(f"Lapas {sheet_name!r} yra tuščias.")
    result = [dict(zip(headers, row)) for row in rows]
    wb.close()
    return result


def load_new_bom_graph(comparison_path: Path):
    """Paimame tik naujų BOM tėvus ir jų eilutes iš MAP palyginimo."""
    bom_rows = read_sheet(comparison_path, "NEW BOMS")
    line_rows = read_sheet(comparison_path, "NEW BOM LINES")

    parents = {
        canon(row.get("Parent SKU")): row
        for row in bom_rows
        if canon(row.get("Parent SKU"))
    }
    lines: dict[str, list[dict]] = defaultdict(list)
    for row in line_rows:
        parent = canon(row.get("Parent SKU"))
        component = canon(row.get("Component SKU"))
        if parent and component:
            lines[parent].append({
                "component": component,
                "quantity": row.get("Quantity"),
            })
    return parents, dict(lines)


def calculate_levels(parents: set[str], lines: dict[str, list[dict]]) -> dict[str, int]:
    """Priskiria lygius kaip Edgaro sename faile: šaknys lv1, jų BOM vaikai lv2.

    Komponentas daro įtaką lygiui tik jeigu jis pats taip pat yra naujo BOM
    tėvas. Paprastos detalės nėra atskiri BOM ir lygio negauna.
    """
    children = {
        parent: {
            row["component"] for row in lines.get(parent, [])
            if row["component"] in parents
        }
        for parent in parents
    }
    has_parent = {child for values in children.values() for child in values}
    roots = sorted(parents - has_parent)
    levels: dict[str, int] = {}
    queue = deque((root, 1) for root in roots)

    while queue:
        sku, level = queue.popleft()
        # Jei tas pats BOM pasiekiamas keliais keliais, paliekame giliausią lygį.
        if level <= levels.get(sku, 0):
            continue
        levels[sku] = level
        for child in sorted(children.get(sku, set())):
            queue.append((child, level + 1))

    # Ciklas arba izoliuota klaidinga struktūra neturi sustabdyti diagnostikos.
    for sku in parents:
        levels.setdefault(sku, 1)
    return levels


def load_bom_types(review_path: Path) -> dict[str, str]:
    """Konvertuoja techninį BOM tipą į Odoo importo ekrano reikšmę."""
    result = {}
    for row in read_sheet(review_path, "BOM TYPE REVIEW"):
        sku = canon(row.get("Parent SKU"))
        proposed = canon(row.get("Proposed BOM Type"))
        if proposed == "PHANTOM":
            result[sku] = "KIT"
        elif proposed == "NORMAL":
            result[sku] = "Manufacture this product"
    return result


def choose_pilot(parents: set[str], lines: dict[str, list[dict]],
                 bom_types: dict[str, str]) -> str:
    """Parenka mažiausią saugų KIT pilotą iš tos pačios rezultatų poros."""
    candidates = [
        sku for sku in parents
        if bom_types.get(sku) == "KIT" and lines.get(sku)
    ]
    if not candidates:
        raise ValueError("Nerastas nė vienas naujas KIT BOM su komponentais.")
    return min(candidates, key=lambda sku: (len(lines[sku]), sku))


def load_odoo_product_ids(client: OdooClient, wanted_skus: set[str]) -> dict[str, dict]:
    """Iš Stage pasiima template, variantą ir abiejų modelių External ID."""
    rows = client.search_read_all(
        "product.product",
        [["default_code", "!=", False]],
        ["id", "default_code", "product_tmpl_id"],
    )
    found = {}
    product_ids = set()
    template_ids = set()
    for row in rows:
        sku = canon(row.get("default_code"))
        if sku not in wanted_skus:
            continue
        template = row.get("product_tmpl_id")
        if not template:
            continue
        product_id = int(row["id"])
        template_id = int(template[0])
        found[sku] = {
            "display_sku": str(row["default_code"]).strip(),
            "product_id": product_id,
            "template_id": template_id,
        }
        product_ids.add(product_id)
        template_ids.add(template_id)

    def xmlids(model: str, ids: set[int]) -> dict[int, str]:
        if not ids:
            return {}
        records = client.search_read_all(
            "ir.model.data",
            [["model", "=", model], ["res_id", "in", sorted(ids)]],
            ["module", "name", "res_id"],
        )
        grouped: dict[int, list[str]] = defaultdict(list)
        for record in records:
            grouped[int(record["res_id"])].append(
                f"{record['module']}.{record['name']}"
            )
        result = {}
        for record_id, values in grouped.items():
            # Jei yra modulio XML ID, jis stabilesnis už __export__ ID.
            values.sort(key=lambda value: (value.startswith("__export__."), value))
            result[record_id] = values[0]
        return result

    product_xmlids = xmlids("product.product", product_ids)
    template_xmlids = xmlids("product.template", template_ids)
    for value in found.values():
        value["product_xmlid"] = product_xmlids.get(value["product_id"], "")
        value["template_xmlid"] = template_xmlids.get(value["template_id"], "")
    return found


def write_pilot(path: Path, pilot: str, level: int, lines: list[dict], bom_type: str,
                products: dict[str, dict]) -> None:
    """Sukuria vieną importo lapą tiksliai pagal seno failo 8 stulpelius."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"BOM_import(lv{level})"
    ws.append(IMPORT_HEADERS)

    parent = products[pilot]
    reference = f"{date.today():%Y%m%d}_Pilot_{pilot}"
    for index, line in enumerate(lines):
        component = products[line["component"]]
        first = index == 0
        ws.append([
            parent["display_sku"] if first else None,
            parent["template_xmlid"] if first else None,
            1 if first else None,
            component["display_sku"],
            component["product_xmlid"],
            line["quantity"],
            bom_type if first else None,
            reference if first else None,
        ])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = [32, 42, 9, 43, 42, 13, 28, 38]
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width

    # Diagnostikos lape paliekame žmogui aiškų įrodymą, ką failas ruošia.
    diag = wb.create_sheet("CHECK")
    diag.append(["Patikra", "Reikšmė"])
    diag.append(["Aplinka", "Stage"])
    diag.append(["Pilotinis BOM", parent["display_sku"]])
    diag.append(["Apskaičiuotas lygis", f"lv{level}"])
    diag.append(["BOM tipas", bom_type])
    diag.append(["Komponentų skaičius", len(lines)])
    diag.append(["Svarbu", "Failas tik paruoštas peržiūrai; Odoo pakeitimų neatlikta"])
    for cell in diag[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    diag.column_dimensions["A"].width = 28
    diag.column_dimensions["B"].width = 75

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Paruošia vieno BOM Stage importo pilotą.")
    parser.add_argument(
        "--pilot-sku",
        help="Pasirenkamas konkretus SKU; nenurodžius parenkamas mažiausias KIT.",
    )
    parser.add_argument("--comparison", type=Path)
    parser.add_argument("--types", type=Path)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()

    base = Path(__file__).resolve().parent
    output_dir = environment_output_dir(base)

    # 1 ŽINGSNIS: saugos patikra. Šis pilotas leidžiamas tik Stage.
    if environment_slug() != "stage":
        raise PermissionError(
            "BOM importo pilotas leidžiamas tik Stage. GUI pasirinkite Stage."
        )

    comparison_path = args.comparison or output_dir / "MAP_Comparison.xlsx"
    types_path = args.types or output_dir / "BOM_Type_Review.xlsx"
    output_path = args.output or output_dir / "BOM_Import_Pilot.xlsx"
    # 2 ŽINGSNIS: nuskaitome naujų BOM struktūrą ir apskaičiuojame lv1/lv2/...
    parents, all_lines = load_new_bom_graph(comparison_path)
    bom_types = load_bom_types(types_path)
    pilot = canon(args.pilot_sku) if args.pilot_sku else choose_pilot(
        set(parents), all_lines, bom_types
    )
    if pilot not in parents:
        raise ValueError(f"SKU {pilot} nerastas lape NEW BOMS.")
    lines = all_lines.get(pilot, [])
    if not lines:
        raise ValueError(f"SKU {pilot} neturi eilučių lape NEW BOM LINES.")
    levels = calculate_levels(set(parents), all_lines)

    # 3 ŽINGSNIS: pasiimame tik peržiūroje patvirtintą BOM tipą.
    bom_type = bom_types.get(pilot)
    if not bom_type:
        raise ValueError(f"SKU {pilot} neturi patvirtinto BOM tipo.")

    # 4 ŽINGSNIS: iš Stage API pasiimame aktualius External ID.
    wanted = {pilot} | {row["component"] for row in lines}
    settings = load_settings()
    client = OdooClient(settings)
    uid = client.authenticate()
    products = load_odoo_product_ids(client, wanted)

    diagnostics = []
    for sku in sorted(wanted):
        value = products.get(sku)
        if not value:
            diagnostics.append(f"Stage nerastas produktas: {sku}")
            continue
        if sku == pilot and not value["template_xmlid"]:
            diagnostics.append(f"BOM produktas neturi product.template External ID: {sku}")
        if sku != pilot and not value["product_xmlid"]:
            diagnostics.append(f"Komponentas neturi product.product External ID: {sku}")
    if diagnostics:
        raise ValueError("\n".join(diagnostics))

    # 5 ŽINGSNIS: sukuriame tik vieno BOM Excel; Odoo rašymo operacijų nėra.
    write_pilot(output_path, pilot, levels[pilot], lines, bom_type, products)
    print("Prisijungta prie Stage Odoo. UID=", uid)
    print("\nBOM IMPORTO PILOTAS SUKURTAS")
    print("Failas:", output_path)
    print("Pilotinis SKU:", products[pilot]["display_sku"])
    print("Lygis:", f"lv{levels[pilot]}")
    print("BOM tipas:", bom_type)
    print("Komponentų:", len(lines))
    print("Odoo pakeitimų neatlikta.")


if __name__ == "__main__":
    main()
