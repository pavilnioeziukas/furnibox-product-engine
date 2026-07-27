"""Generate Odoo imports for every missing Reform and paired assembled SKU."""

from __future__ import annotations

import argparse
import re
from collections import Counter, defaultdict
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import load_settings
from odoo_client import OdooClient
from product_detection_v2 import canon, find_bom_input, load_reform_universe, read_odoo_products, role
from product_import_v3 import (
    apack_sku,
    external_ids,
    infer_category_map,
    infer_route_profile,
    packaging_name,
    read_all,
)


# Šios parent kategorijos buvo rankiniu būdu patikrintos per pirmąjį Stage
# importą. Todėl naujuose paleidimuose jų nebereikia grąžinti į REVIEW lapą.
REVIEW_PARENT_CATEGORIES = set()
CLEAR_COMPONENT_GROUPS = {"CABINET PART", "SHELF PART", "ACCESSORIES", "CABINET ACCESSORIES"}

# 2026-07-16 Stage pilote patvirtintos trūkstamų komponentų taisyklės.
# Jos paimtos iš sėkmingai importuoto Odoo_Remaining_46_Products_Import.xlsx.
# External ID naudojami todėl, kad po Stage atnaujinimo jie išlieka tokie patys
# kaip dabartinėje Production bazėje.
APPROVED_COMPONENT_RULES = {
    "INTERIOR STORAGE": {
        "category": "__export__.product_category_12_8d605b38",
        "routes": "purchase_stock.route_warehouse0_buy",
        "vendor": "__export__.res_partner_9_b2e3c075",
    },
    "FRONT HARDWARE": {
        "category": "__export__.product_category_10_74eb6563",
        "routes": "purchase_stock.route_warehouse0_buy",
        "vendor": "__export__.res_partner_9_b2e3c075",
    },
    "BLANK": {
        "category": "__export__.product_category_16_9cb79bdc",
        "routes": "stock.route_warehouse0_mto",
        "vendor": "__export__.res_partner_86_3ef81eee",
    },
    "FASTENER": {
        "category": "__export__.product_category_17_842598e0",
        "routes": "purchase_stock.route_warehouse0_buy",
        "vendor": "__export__.res_partner_9_b2e3c075",
    },
    "FASTENERS": {
        "category": "__export__.product_category_17_842598e0",
        "routes": "purchase_stock.route_warehouse0_buy",
        "vendor": "__export__.res_partner_9_b2e3c075",
    },
}

# Reform šaltinyje šiam vienam produktui pavadinimas yra #N/A. Edgaras
# patvirtino laikiną aiškiai matomą pavadinimą, kad kortelę būtų galima rasti
# ir vėliau pervadinti Odoo.
APPROVED_PRODUCT_NAMES = {
    "EUB-P-ACC02-MIS012": "Rename",
}
EXPECTED_PARENT_CATEGORY_NAMES = {
    # SKU be -A yra flatpack produkto kortelė. Surinkta kortelė generuojama
    # atskirai su -A ir gauna CABINETS (Assembled) kategoriją.
    "CABINETS": "All / CABINETS",
}
IMPORT_HEADERS = [
    "Internal reference", "name", "route_ids/id", "type", "categ_id",
    "invoice_policy", "packaging_name", "variant_seller_ids/partner_id/id",
]


def hrd_assembled_sku(sku: str) -> str:
    """Grąžina HRD surinkto produkto SKU."""
    code = canon(sku)
    if code.endswith("-A"):
        return code
    return f"{code}-A"


def cabinet_assembled_sku(sku: str) -> str:
    """Grąžina surinktos spintelės SKU."""
    code = canon(sku)
    if code.endswith("-A"):
        return code
    return f"{code}-A"


def is_cabinet_bom_parent(sku: str, product: dict) -> bool:
    """Reform CABINETS tėvas, kuriam reikia atskiros surinktos -A kortelės."""
    code = canon(sku)
    return bool(
        product.get("is_parent")
        and canon(product.get("category")) == "CABINETS"
        and not code.endswith("-A")
    )


def is_hrd_bom_parent(sku: str, product: dict) -> bool:
    """Tikras Reform HRD BOM tėvas, kuriam reikia atskiros -A kortelės."""
    code = canon(sku)
    return bool(
        product.get("is_parent")
        and not code.endswith("-A")
        and "HRD" in code
    )


def find_reference_export(base: Path) -> Path:
    candidates = []
    for folder in (base / "data", base):
        candidates.extend(folder.glob("Product (product.template)*.xlsx"))
    candidates = [path for path in candidates if not path.name.startswith("~$")]
    if not candidates:
        raise FileNotFoundError(
            "Nerastas Stage produktų eksportas 'Product (product.template)*.xlsx'. "
            "Nukopijuokite jį į data aplanką arba naudokite --reference."
        )
    return max(candidates, key=lambda path: path.stat().st_mtime)


def load_reference_export(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    index = {str(value).strip(): i for i, value in enumerate(headers) if value is not None}
    required = {"default_code", "categ_id/id", "route_ids/id", "variant_seller_ids/partner_id/id"}
    missing = required - set(index)
    if missing:
        raise ValueError("Stage eksporte trūksta stulpelių: " + ", ".join(sorted(missing)))
    result = {}
    for row in rows:
        sku = canon(row[index["default_code"]])
        if not sku:
            continue
        result[sku] = {
            "category": str(row[index["categ_id/id"]] or "").strip(),
            "routes": str(row[index["route_ids/id"]] or "").strip(),
            "vendor": str(row[index["variant_seller_ids/partner_id/id"]] or "").strip(),
        }
    wb.close()
    return result


def infer_component_rules(reform_products, reference):
    votes = defaultdict(Counter)
    for sku, existing in reference.items():
        product = reform_products.get(sku)
        if not product or not product["is_component"]:
            continue
        group = product.get("part_group") or "BLANK"
        combo = (existing["category"], existing["routes"], existing["vendor"])
        votes[group][combo] += 1

    diagnostics = []
    rules = {}
    analysis_rows = []
    for group, combinations in sorted(votes.items()):
        for rank, (combo, count) in enumerate(combinations.most_common(), start=1):
            analysis_rows.append({
                "Reform Part Group": group,
                "Rank": rank,
                "Existing Product Count": count,
                "Category External ID": combo[0],
                "Route External IDs": combo[1],
                "Vendor External ID": combo[2],
            })

    # Strong, highly consistent detail rules.
    for group in ("CABINET PART", "SHELF PART"):
        if votes[group]:
            combo, count = votes[group].most_common(1)[0]
            rules[group] = {"category": combo[0], "routes": combo[1], "vendor": combo[2], "evidence": count}

    # Business rule: both accessory groups use ACCESSORIES MTS+Manufacture route
    # and the standard unknown supplier, while retaining their own categories.
    accessory_combos = votes["ACCESSORIES"]
    accessory_category = Counter()
    accessory_vendor = Counter()
    accessory_routes = Counter()
    for (category, routes, vendor), count in accessory_combos.items():
        if category:
            accessory_category[category] += count
        if vendor:
            accessory_vendor[vendor] += count
        if "mrp.route_warehouse0_manufacture" in routes and "purchase_stock.route_warehouse0_buy" not in routes:
            accessory_routes[routes] += count
    selected_routes = accessory_routes.most_common(1)[0][0] if accessory_routes else ""
    selected_vendor = accessory_vendor.most_common(1)[0][0] if accessory_vendor else ""
    for group in ("ACCESSORIES", "CABINET ACCESSORIES"):
        categories = Counter()
        for (category, routes, vendor), count in votes[group].items():
            if category:
                categories[category] += count
        category = categories.most_common(1)[0][0] if categories else ""
        rules[group] = {
            "category": category,
            "routes": selected_routes,
            "vendor": selected_vendor,
            "evidence": sum(votes[group].values()),
        }

    # Rankiniu būdu patvirtintos taisyklės turi pirmenybę prieš statistinį
    # spėjimą iš reference eksporto. Taip tas pats produktas po Stage bazės
    # atnaujinimo nebegrįžta į rankinę peržiūrą.
    for group, approved in APPROVED_COMPONENT_RULES.items():
        rules[group] = {**approved, "evidence": "approved Stage pilot"}

    for group in CLEAR_COMPONENT_GROUPS:
        rule = rules.get(group, {})
        missing_fields = [name for name in ("category", "routes", "vendor") if not rule.get(name)]
        if missing_fields:
            diagnostics.append(f"{group}: nenustatyta " + ", ".join(missing_fields))
    return rules, analysis_rows, diagnostics


def parent_metadata(reform_products):
    return {
        sku: {
            "category": product.get("category", ""),
            "sku": product.get("sku", sku),
            "name_1": product.get("name_1", ""),
            "name_2": product.get("name_2", ""),
        }
        for sku, product in reform_products.items() if product["is_parent"]
    }


def enforce_parent_category_names(client: OdooClient, category_map: dict, required_categories: set[str]):
    """Resolve business-critical categories by exact Odoo name, not SKU majority votes."""
    expected = {
        reform_category: complete_name
        for reform_category, complete_name in EXPECTED_PARENT_CATEGORY_NAMES.items()
        if reform_category in required_categories
    }
    if not expected:
        return []

    categories = read_all(client, "product.category", [], ["id", "name", "complete_name"])
    category_xmlids = external_ids(client, "product.category")
    by_complete_name = {
        str(row.get("complete_name") or row.get("name") or "").strip().casefold(): row
        for row in categories
    }
    messages = []
    for reform_category, expected_name in expected.items():
        row = by_complete_name.get(expected_name.casefold())
        if not row:
            raise RuntimeError(
                f"Saugos patikra: Odoo nerasta kategorija '{expected_name}' "
                f"Reform kategorijai '{reform_category}'. Importo failas nesukurtas."
            )
        category_id = int(row["id"])
        xmlid = category_xmlids.get(category_id)
        if not xmlid:
            raise RuntimeError(
                f"Saugos patikra: kategorija '{expected_name}' neturi External ID. "
                "Importo failas nesukurtas."
            )
        previous = category_map.get(reform_category, {})
        previous_name = previous.get("category_name", "nenustatyta")
        category_map[reform_category] = {
            "category_id": category_id,
            "category_name": row.get("complete_name") or row.get("name"),
            "external_id": xmlid,
            "matched_existing_products": previous.get("matched_existing_products", 0),
            "alternatives": previous.get("alternatives", ""),
        }
        messages.append(
            f"{reform_category}: priverstinai parinkta '{expected_name}' "
            f"({xmlid}); automatinis pasirinkimas buvo '{previous_name}'"
        )
    return messages


def resolve_exact_category(client: OdooClient, complete_name: str) -> dict:
    """Randa vieną verslo taisyklėje nurodytą Odoo kategoriją ir jos XML ID."""
    categories = read_all(
        client, "product.category", [], ["id", "name", "complete_name"]
    )
    matches = [
        row for row in categories
        if str(row.get("complete_name") or row.get("name") or "")
        .strip().casefold() == complete_name.casefold()
    ]
    if len(matches) != 1:
        raise RuntimeError(
            f"Saugos patikra: tikėtasi vienos kategorijos '{complete_name}', "
            f"rasta {len(matches)}. Importo failas nesukurtas."
        )
    row = matches[0]
    xmlid = external_ids(client, "product.category").get(int(row["id"]))
    if not xmlid:
        raise RuntimeError(
            f"Saugos patikra: kategorija '{complete_name}' neturi External ID."
        )
    return {
        "category_id": int(row["id"]),
        "category_name": row.get("complete_name") or row.get("name"),
        "external_id": xmlid,
    }


def infer_hrd_a_profiles(client: OdooClient):
    """Infer HRD -> HRD-A category and route rules from existing Odoo pairs."""
    templates = read_all(
        client,
        "product.template",
        [["default_code", "ilike", "HRD"]],
        ["id", "default_code", "categ_id", "route_ids"],
    )
    by_sku = {
        canon(row.get("default_code")): row
        for row in templates
        if canon(row.get("default_code"))
    }
    category_xmlids = external_ids(client, "product.category")
    route_xmlids = external_ids(client, "stock.route")
    votes = defaultdict(Counter)
    fallback_votes = Counter()

    for assembled_sku, assembled in by_sku.items():
        if not assembled_sku.endswith("-A"):
            continue
        assembled_category = assembled.get("categ_id")
        assembled_category_id = int(assembled_category[0]) if assembled_category else 0
        route_ids = tuple(sorted(int(route_id) for route_id in (assembled.get("route_ids") or [])))
        if assembled_category_id and route_ids:
            fallback_votes[(assembled_category_id, route_ids)] += 1
        source = by_sku.get(assembled_sku[:-2])
        if not source:
            continue
        source_category = source.get("categ_id")
        source_category_id = int(source_category[0]) if source_category else 0
        if source_category_id and assembled_category_id and route_ids:
            votes[source_category_id][(assembled_category_id, route_ids)] += 1

    profiles = {}
    diagnostics = []
    analysis_rows = []
    for source_category_id, combinations in sorted(votes.items()):
        ranked = combinations.most_common()
        for rank, ((assembled_category_id, route_ids), count) in enumerate(ranked, start=1):
            analysis_rows.append({
                "Source Category ID": source_category_id,
                "Rank": rank,
                "Existing HRD-A Pair Count": count,
                "HRD-A Category External ID": category_xmlids.get(assembled_category_id, ""),
                "HRD-A Route External IDs": ",".join(
                    route_xmlids.get(route_id, "") for route_id in route_ids
                ),
            })
        (assembled_category_id, route_ids), count = ranked[0]
        category_external_id = category_xmlids.get(assembled_category_id)
        route_external_ids = [route_xmlids.get(route_id) for route_id in route_ids]
        if not category_external_id or not all(route_external_ids):
            diagnostics.append(
                f"HRD-A etalonas pradinei kategorijai ID {source_category_id} "
                "neturi visų External ID"
            )
            continue
        profiles[source_category_id] = {
            "category": category_external_id,
            "routes": ",".join(route_external_ids),
            "evidence": count,
        }

    # Daugumos Production HRD-A kortelių pradinė HRD kortelė tuo pačiu SKU
    # nebeegzistuoja, todėl tikslios HRD -> HRD-A poros neužtenka. Tokiais
    # atvejais naudojame vyraujantį realių Production HRD-A profilį.
    if fallback_votes:
        (assembled_category_id, route_ids), count = fallback_votes.most_common(1)[0]
        category_external_id = category_xmlids.get(assembled_category_id)
        route_external_ids = [route_xmlids.get(route_id) for route_id in route_ids]
        if category_external_id and all(route_external_ids):
            profiles[0] = {
                "category": category_external_id,
                "routes": ",".join(route_external_ids),
                "evidence": count,
            }
            analysis_rows.append({
                "Source Category ID": "FALLBACK",
                "Rank": 1,
                "Existing HRD-A Pair Count": count,
                "HRD-A Category External ID": category_external_id,
                "HRD-A Route External IDs": ",".join(route_external_ids),
            })
        else:
            diagnostics.append(
                "Vyraujantis Production HRD-A profilis neturi visų External ID"
            )
    return profiles, by_sku, analysis_rows, diagnostics


def parent_row(sku, product, category_map, route_profile):
    category = product.get("category", "")
    mapping = category_map.get(category)
    if not mapping:
        return None
    return {
        "Internal reference": product["sku"],
        "name": (
            APPROVED_PRODUCT_NAMES.get(sku)
            or product.get("name_1")
            or product.get("name_2")
            or product["sku"]
        ),
        "route_ids/id": route_profile["manufacture"],
        "type": "Storable Product",
        "categ_id": mapping["external_id"],
        "invoice_policy": "Delivered quantities",
        "packaging_name": packaging_name(sku) if sku.startswith("FPACK-") else "",
        "variant_seller_ids/partner_id/id": "",
        "Reform Role": role(product),
        "Reform Category / Part Group": category,
        "Review reason": "",
    }


def component_row(sku, product, component_rules):
    group = product.get("part_group") or "BLANK"
    rule = component_rules.get(group)
    if not rule:
        return None
    return {
        "Internal reference": product["sku"],
        "name": product["sku"],
        "route_ids/id": rule["routes"],
        "type": "Storable Product",
        "categ_id": rule["category"],
        "invoice_policy": "Delivered quantities",
        "packaging_name": "",
        "variant_seller_ids/partner_id/id": rule["vendor"],
        "Reform Role": role(product),
        "Reform Category / Part Group": group,
        "Review reason": "",
    }


def build_rows(
    missing_skus, existing_skus, reform_products, category_map,
    cabinet_assembled_category, apack_category, route_profile, component_rules,
    hrd_a_profiles, odoo_hrd_templates,
):
    ready, review, diagnostics = [], [], []
    generated_apack = 0
    generated_cabinet_assembled = 0
    generated_hrd_assembled = 0
    for sku in sorted(missing_skus):
        product = reform_products[sku]
        if product["is_parent"]:
            row = parent_row(sku, product, category_map, route_profile)
            if row is None:
                diagnostics.append(f"{sku}: nenustatyta pagrindinio produkto Odoo kategorija")
                continue
            if product.get("category") in REVIEW_PARENT_CATEGORIES:
                row["Review reason"] = "Parent category requires manual review"
                review.append(row)
            else:
                ready.append(row)
            if sku.startswith("FPACK-"):
                generated_code = canon(apack_sku(sku))
                if generated_code in existing_skus:
                    continue
                if not apack_category:
                    diagnostics.append(f"{sku}: nenustatyta APACK kategorija")
                else:
                    duplicate = dict(row)
                    duplicate["Internal reference"] = apack_sku(sku)
                    duplicate["route_ids/id"] = route_profile["apack"]
                    duplicate["categ_id"] = apack_category["external_id"]
                    duplicate["packaging_name"] = packaging_name(sku)
                    duplicate["Reform Role"] = "GENERATED APACK"
                    duplicate["Reform Category / Part Group"] = "APACK"
                    duplicate["Review reason"] = ""
                    ready.append(duplicate)
                    generated_apack += 1
        else:
            row = component_row(sku, product, component_rules)
            if row is None:
                row = {
                    "Internal reference": product["sku"], "name": product["sku"],
                    "route_ids/id": "", "type": "Storable Product", "categ_id": "",
                    "invoice_policy": "Delivered quantities", "packaging_name": "",
                    "variant_seller_ids/partner_id/id": "", "Reform Role": role(product),
                    "Reform Category / Part Group": product.get("part_group") or "BLANK",
                    "Review reason": "No approved component rule for this Part Group",
                }
                review.append(row)
            else:
                ready.append(row)

    # CABINETS-A nėra atskiras Reform šaltinio SKU. Kiekvienai bazinei
    # CABINETS kortelei sukuriame surinktą -A porą, jeigu jos Odoo dar nėra.
    scheduled_skus = {canon(row["Internal reference"]) for row in ready + review}
    for source_sku, product in sorted(reform_products.items()):
        if not is_cabinet_bom_parent(source_sku, product):
            continue
        generated_code = cabinet_assembled_sku(source_sku)
        if generated_code in existing_skus or generated_code in scheduled_skus:
            continue
        duplicate = parent_row(source_sku, product, category_map, route_profile)
        if duplicate is None:
            diagnostics.append(
                f"{source_sku}: nenustatyta bazinio CABINETS produkto kategorija"
            )
            continue
        duplicate["Internal reference"] = generated_code
        duplicate["route_ids/id"] = route_profile["manufacture"]
        duplicate["categ_id"] = cabinet_assembled_category["external_id"]
        duplicate["packaging_name"] = ""
        duplicate["Reform Role"] = "GENERATED CABINET-A"
        duplicate["Reform Category / Part Group"] = "CABINETS (Assembled)"
        duplicate["Review reason"] = ""
        ready.append(duplicate)
        scheduled_skus.add(generated_code)
        generated_cabinet_assembled += 1

    # HRD-A nėra atskiras Reform šaltinio SKU, todėl jo negalima generuoti
    # vien tik pereinant per missing_skus. Kuriame -A visiems Reform HRD BOM
    # tėvams, jeigu tokios kortelės Odoo dar nėra.
    for source_sku, product in sorted(reform_products.items()):
        if not is_hrd_bom_parent(source_sku, product):
            continue
        generated_code = hrd_assembled_sku(source_sku)
        if generated_code in existing_skus or generated_code in scheduled_skus:
            continue
        source_template = odoo_hrd_templates.get(canon(source_sku))
        source_category = source_template.get("categ_id") if source_template else None
        source_category_id = int(source_category[0]) if source_category else 0
        hrd_a_profile = (
            hrd_a_profiles.get(source_category_id)
            or hrd_a_profiles.get(0)
        )
        if not hrd_a_profile:
            diagnostics.append(
                f"{source_sku}: nenustatytas HRD-A etalonas pradinei Odoo "
                f"kategorijai ID {source_category_id or 'NERASTA'}; "
                f"kortelė {generated_code} neįtraukta"
            )
            continue
        duplicate = parent_row(source_sku, product, category_map, route_profile)
        if duplicate is None:
            diagnostics.append(
                f"{source_sku}: nenustatyta pagrindinio HRD produkto Odoo kategorija"
            )
            continue
        duplicate["Internal reference"] = generated_code
        duplicate["route_ids/id"] = hrd_a_profile["routes"]
        duplicate["categ_id"] = hrd_a_profile["category"]
        duplicate["packaging_name"] = packaging_name(source_sku)
        duplicate["Reform Role"] = "GENERATED HRD-A"
        duplicate["Reform Category / Part Group"] = "HRD-A"
        duplicate["Review reason"] = ""
        ready.append(duplicate)
        scheduled_skus.add(generated_code)
        generated_hrd_assembled += 1

    return (
        ready, review, diagnostics, generated_apack,
        generated_cabinet_assembled, generated_hrd_assembled,
    )


def style_sheet(ws):
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        values = [str(ws.cell(row, col).value or "") for row in range(1, min(ws.max_row, 500) + 1)]
        ws.column_dimensions[get_column_letter(col)].width = min(max(map(len, values)) + 2, 55)


def add_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    style_sheet(ws)
    return ws


def write_workbook(path, main_rows, review_mode, analysis_rows, diagnostics, info):
    wb = Workbook()
    wb.remove(wb.active)
    headers = IMPORT_HEADERS + (["Reform Role", "Reform Category / Part Group", "Review reason"] if review_mode else [])
    add_sheet(wb, "REVIEW" if review_mode else "PRODUCT IMPORT", headers, main_rows)
    add_sheet(wb, "RULE ANALYSIS", [
        "Reform Part Group", "Rank", "Existing Product Count", "Category External ID",
        "Route External IDs", "Vendor External ID",
    ], analysis_rows)
    add_sheet(wb, "DIAGNOSTICS", ["Message"], ({"Message": item} for item in diagnostics))
    add_sheet(wb, "INFO", ["Parameter", "Value"],
              ({"Parameter": key, "Value": value} for key, value in info.items()))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="Generuoja visų trūkstamų Reform produktų Stage importą.")
    parser.add_argument(
        "--env-file",
        type=Path,
        help="Aplinkos failas, pvz. .env.stage. Jo reikšmės turi pirmenybę prieš .env.",
    )
    parser.add_argument("--reference", type=Path, help="Stage product.template eksportas")
    parser.add_argument("--bom-input", type=Path, help="Reform BOM Input")
    parser.add_argument("--output-dir", type=Path)
    args = parser.parse_args()
    base = Path(__file__).resolve().parent
    if args.env_file:
        env_path = args.env_file if args.env_file.is_absolute() else base / args.env_file
        if not env_path.exists():
            raise FileNotFoundError(f"Nerastas aplinkos failas: {env_path}")
        load_dotenv(env_path, override=True)
    reference_path = args.reference or find_reference_export(base)
    bom_input_path = args.bom_input or find_bom_input(base)
    output_dir = args.output_dir

    print("Nuskaitomi visi Reform produktai ir komponentai...")
    reform_products, _, _ = load_reform_universe(bom_input_path)
    reference = load_reference_export(reference_path)

    settings = load_settings()
    output_dir = output_dir or settings.output_dir
    print("Naudojama Odoo aplinka:", settings.url)
    client = OdooClient(settings)
    uid = client.authenticate()
    print(f"Prisijungta prie Odoo. UID={uid}")
    odoo_products, _ = read_odoo_products(client)
    missing_skus = set(reform_products) - set(odoo_products)
    print("Visi Reform SKU:", len(reform_products))
    print("Trūksta pasirinktoje Odoo aplinkoje:", len(missing_skus))

    parents = parent_metadata(reform_products)
    # HRD šaltiniai paprastai jau egzistuoja Odoo, todėl jų kategorijos
    # nepatektų į atranką, paremtą vien missing_skus. Vis dėlto jos reikalingos
    # generuojant naujas HRD-A korteles.
    hrd_parent_categories = {
        product.get("category", "")
        for sku, product in reform_products.items()
        if is_hrd_bom_parent(sku, product) and product.get("category")
    }
    required_categories = (
        {parents[sku]["category"] for sku in missing_skus if sku in parents}
        | hrd_parent_categories
    )
    category_map, apack_category, category_diagnostics = infer_category_map(
        client, parents, required_categories
    )
    category_override_messages = enforce_parent_category_names(
        client, category_map, required_categories
    )
    cabinet_assembled_category = resolve_exact_category(
        client, "All / CABINETS (Assembled)"
    )
    for message in category_override_messages:
        print("Kategorijos saugos patikra:", message)
    route_profile, route_rows, route_diagnostics = infer_route_profile(client)
    (
        hrd_a_profiles,
        odoo_hrd_templates,
        hrd_a_analysis_rows,
        hrd_a_diagnostics,
    ) = infer_hrd_a_profiles(client)
    component_rules, analysis_rows, component_diagnostics = infer_component_rules(reform_products, reference)
    (
        ready, review, row_diagnostics, generated_apack,
        generated_cabinet_assembled, generated_hrd_assembled,
    ) = build_rows(
        missing_skus, set(odoo_products), reform_products, category_map,
        cabinet_assembled_category, apack_category, route_profile,
        component_rules, hrd_a_profiles, odoo_hrd_templates,
    )
    diagnostics = (
        category_diagnostics + route_diagnostics + component_diagnostics
        + hrd_a_diagnostics + row_diagnostics
    )
    info = {
        "Odoo URL": settings.url,
        "Reform source": str(bom_input_path),
        "Stage rule source": str(reference_path),
        "All unique Reform SKU": len(reform_products),
        "Missing Reform SKU": len(missing_skus),
        "Ready import rows": len(ready),
        "Review rows": len(review),
        "Generated APACK": generated_apack,
        "Generated CABINET-A": generated_cabinet_assembled,
        "Generated HRD-A": generated_hrd_assembled,
        "Manufacture routes": route_profile["manufacture"],
        "APACK routes": route_profile["apack"],
        "HRD-A source-category profiles": len(hrd_a_profiles),
        "Category safety overrides": " | ".join(category_override_messages) or "None",
    }
    ready_path = output_dir / "Odoo_All_Missing_Products_Import.xlsx"
    review_path = output_dir / "Odoo_All_Missing_Products_Review.xlsx"
    write_workbook(ready_path, ready, False, analysis_rows, diagnostics, info)
    write_workbook(review_path, review, True, analysis_rows, diagnostics, info)

    print("\nVISŲ PRODUKTŲ IMPORTO FAILAI SUKURTI")
    print("Paruošta importui:", len(ready))
    print("Peržiūrai:", len(review))
    print("Sugeneruota APACK:", generated_apack)
    print("Sugeneruota CABINET-A:", generated_cabinet_assembled)
    print("Sugeneruota HRD-A:", generated_hrd_assembled)
    print("Diagnostikos įrašai:", len(diagnostics))
    print("Importas:", ready_path)
    print("Peržiūra:", review_path)


if __name__ == "__main__":
    main()
