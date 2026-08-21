"""Compile Tamara's BOM workbook into a versioned SKU/category reference."""

from __future__ import annotations

import argparse
import hashlib
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


def clean(value) -> str:
    return str(value or "").strip()


def tokens(expression: str) -> list[str]:
    return [token.strip() for token in clean(expression).split("+") if token.strip()]


def join_unique(*expressions: str) -> str:
    result = []
    for expression in expressions:
        for token in tokens(expression):
            if token not in result:
                result.append(token)
    return "+".join(result)


def compile_reference(source: Path) -> dict:
    workbook = load_workbook(source, data_only=True, read_only=True)
    assignments: dict[str, dict] = {}
    conflicts = []

    def assign(sku: str, expression: str, sheet: str) -> None:
        sku = clean(sku)
        expression = join_unique(expression)
        if not sku or not expression:
            return
        normalized = sku.casefold()
        existing = assignments.get(normalized)
        if existing and existing["expression"] != expression:
            conflicts.append({
                "sku": sku,
                "first": existing,
                "second": {"expression": expression, "sheet": sheet},
            })
            return
        assignments[normalized] = {
            "sku": sku,
            "expression": expression,
            "sheet": sheet,
        }

    for sheet in (
        "ASS CABINETS ",
        "SHELF",
        "SHELF PREPACK",
        "SHELF LED LEDROD",
        "SHELF LED LEDROD  PREPACK",
    ):
        for row in workbook[sheet].iter_rows(min_row=2, values_only=True):
            assign(row[0], row[2], sheet)

    # APACK has category 12 plus the package/component category combination
    # recorded in the final column.
    for row in workbook["ASS PREPACK CABINETS (2)"].iter_rows(
        min_row=2, values_only=True
    ):
        if clean(row[9]):
            assign(row[0], f"12+{clean(row[9])}", "ASS PREPACK CABINETS (2)")

    # FPACK category is stored in column D; BOM PAP package selection is in I.
    fpack_parts: dict[str, dict] = defaultdict(lambda: {"base": "", "packs": []})
    for row in workbook["PREPACK CABINETS"].iter_rows(min_row=2, values_only=True):
        sku = clean(row[0])
        if not sku:
            continue
        fpack_parts[sku]["base"] = clean(row[3]) or fpack_parts[sku]["base"]
        package = clean(row[8])
        if package:
            package = f"{package}.1" if "." not in package else package
            fpack_parts[sku]["packs"].append(package)
    for sku, values in fpack_parts.items():
        assign(
            sku,
            join_unique(values["base"], *values["packs"]),
            "PREPACK CABINETS",
        )

    # Hardware BOM totals combine the parent category with every distinct
    # component category used by that BOM.
    for sheet, base_code in (
        ("BOM HRD", "9"),
        ("ASS BOM HRD ", "9"),
        ("BOM SLF HRD", "7"),
    ):
        grouped: dict[str, list[str]] = defaultdict(list)
        for row in workbook[sheet].iter_rows(min_row=2, values_only=True):
            sku = clean(row[0])
            if not sku:
                continue
            grouped[sku].append(clean(row[8]) or clean(row[9]))
        for sku, category_codes in grouped.items():
            assign(sku, join_unique(base_code, *category_codes), sheet)

    workbook.close()
    digest = hashlib.sha256(source.read_bytes()).hexdigest()
    return {
        "schema_version": 1,
        "source_file": source.name,
        "source_sha256": digest,
        "assignment_count": len(assignments),
        "conflicts": conflicts,
        "sku_expressions": sorted(
            assignments.values(), key=lambda row: row["sku"].casefold()
        ),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    result = compile_reference(args.source)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(result, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print("Assignments:", result["assignment_count"])
    print("Conflicts:", len(result["conflicts"]))


if __name__ == "__main__":
    main()
