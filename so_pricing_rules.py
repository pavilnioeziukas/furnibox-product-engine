"""Persistent, application-owned configuration for Reform SO pricing."""
from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCHEMA_VERSION = 3
ADDON_FIELDS = ("assembly", "storage", "packaging", "put_on_pallet", "other", "markup")


# Tamara's CATEGORY sheet is the business source for transformed Furnibox
# products.  These are category rates, not SKU-specific prices.  They live in
# application configuration so a rate can be changed once without rebuilding
# the transformer or maintaining thousands of individual SKU overrides.
DEFAULT_BOM_CATEGORY_RATES = (
    ("1", "PREPACK CATEGORY 1", 0.0, 0.4, 0.4, 0.0, 0.04, 0.0),
    ("2", "PREPACK CATEGORY 2", 0.0, 0.3, 0.2, 0.0, 0.02, 0.0),
    ("3", "PREPACK CATEGORY 3", 0.0, 0.0, 0.0, 0.0, 0.0, 0.0),
    ("4", "PREPACK CATEGORY 4", 0.0, 0.0, 1.5, 0.0, 0.2, 0.0),
    ("5", "PREPACK CATEGORY 5", 0.0, 0.5, 0.5, 0.0, 0.0, 0.0),
    ("6", "INTERIOR STORAGE", 0.0, 4.0, 1.0, 0.0, 0.04, 0.0),
    ("7", "SHELF HARDWARE", 0.3, 0.05, 0.03, 0.02, 0.04, 0.0),
    ("8", "SLF PP", 0.0, 0.0, 1.0, 0.0, 0.0, 0.0),
    ("8.1", "SLF ROD PP", 5.0, 0.5, 1.0, 0.0, 0.0, 0.0),
    ("8.2", "SLF LED ROD PP", 7.16, 0.7, 1.0, 0.0, 0.0, 0.0),
    ("9", "CABINET HARDWARE", 1.5, 0.0, 0.0, 0.0, 0.1, 0.0),
    ("10", "PNL/PCL", 0.0, 0.0, 1.0, 0.0, 1.68, 0.0),
    ("11", "LED", 0.0, 0.1, 0.1, 0.0, 0.02, 0.0),
    ("12", "ASS PREPACK", 50.0, 0.0, 3.14, 1.0, 5.9, 3.0),
    ("20.1", "EU FPACK", 0.0, 0.1, 0.0, 0.0, 0.0, 0.0),
    ("21.1", "US FPACK", 0.0, 0.1, 0.0, 0.0, 0.0, 0.0),
    ("22.1", "EU ASS PACK", 0.0, 0.1, 0.0, 0.0, 0.0, 0.0),
    ("23.1", "US ASS PACK", 0.0, 0.1, 0.0, 0.0, 0.0, 0.0),
    ("24.1", "PAP HRD", 0.0, 0.0, 0.0, 0.0, 0.0, 0.13),
    ("25.1", "EU SHELF PREPACK", 0.0, 0.1, 0.0, 0.0, 0.05, 0.0),
    ("26.1", "US SHELF PREPACK", 0.0, 0.1, 0.0, 0.0, 0.05, 0.0),
    ("30", "COMPONENTS / FASTENERS", 0.0, 0.005, 0.005, 0.0, 0.0, 0.0),
    ("31", "COMPONENTS / FRONT HARDWARE", 0.0, 0.04, 0.01, 0.0, 0.0, 0.0),
    ("32", "COMPONENTS / INTERIOR STORAGE", 0.0, 1.0, 1.0, 0.0, 0.0, 0.0),
    ("33", "COMPONENTS / CABINET HARDWARE", 0.0, 0.02, 0.01, 0.0, 0.0, 0.0),
    ("34", "COMPONENTS / CABINET ACCESSORIES", 0.0, 0.02, 0.02, 0.0, 0.0, 0.05),
    ("35", "COMPONENTS / SHELF HARDWARE", 0.0, 0.15, 0.05, 0.02, 0.0, 0.0),
    ("36", "PAPER PRINT", 0.0, 0.0, 0.0, 0.0, 0.0, 0.0),
)


def default_bom_category_rates() -> list[dict[str, Any]]:
    return [
        {
            "code": code,
            "name": name,
            **dict(zip(ADDON_FIELDS, values)),
        }
        for code, name, *values in DEFAULT_BOM_CATEGORY_RATES
    ]


def _text(value: Any) -> str:
    return str(value or "").strip()


def _number(value: Any, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"Tikėtasi skaičiaus, gauta {value!r}")
    return float(value)


@dataclass(frozen=True)
class PricingRule:
    sku: str
    category_id: str
    category_name: str
    odoo_category: str
    assembly: float = 0.0
    storage: float = 0.0
    packaging: float = 0.0
    put_on_pallet: float = 0.0
    other: float = 0.0
    markup: float = 0.0

    @property
    def addons(self) -> tuple[float, float, float, float, float, float]:
        return tuple(getattr(self, field) for field in ADDON_FIELDS)


@dataclass(frozen=True)
class NonBomRule:
    sku: str
    name: str
    product_category: str
    pricing_category: str
    preparation: float = 0.0
    storage: float = 0.0
    bag: float = 0.0
    sticker: float = 0.0


def empty_config() -> dict[str, Any]:
    return {
        "schema_version": SCHEMA_VERSION,
        "adjustment_rate": -0.07,
        "bom_categories": [],
        "bom_category_rates": default_bom_category_rates(),
        "bom_skus": [],
        "bom_products": [],
        "non_bom_categories": [],
        "non_bom_skus": [],
        "migration_warnings": [],
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }


def _category_code(prefix: str, index: int) -> str:
    return f"{prefix}-{index:03d}"


def upgrade_config(document: dict[str, Any]) -> dict[str, Any]:
    """Convert the per-SKU v1 structure to categories plus SKU assignments."""
    if document.get("schema_version") == SCHEMA_VERSION:
        return document
    if document.get("schema_version") == 2:
        upgraded = dict(document)
        upgraded["schema_version"] = SCHEMA_VERSION
        upgraded.setdefault("bom_category_rates", default_bom_category_rates())
        return upgraded
    if document.get("schema_version") != 1:
        raise ValueError("Nepalaikoma SO kainodaros konfigūracijos versija.")
    upgraded = empty_config()
    upgraded.update({key: value for key, value in document.items() if key not in {
        "schema_version", "pricing_rules", "non_bom_rules",
    }})

    category_ids: dict[tuple[Any, ...], str] = {}
    for raw in document.get("pricing_rules", []):
        rule = PricingRule(**raw)
        signature = (rule.category_id, rule.category_name, rule.odoo_category, *rule.addons)
        category_id = category_ids.get(signature)
        if category_id is None:
            category_id = _category_code("BOM", len(category_ids) + 1)
            category_ids[signature] = category_id
            upgraded["bom_categories"].append({
                "id": category_id,
                "name": rule.category_name or rule.category_id or rule.odoo_category or "Be pavadinimo",
                "source_category_id": rule.category_id,
                "odoo_category": rule.odoo_category,
                **{field: getattr(rule, field) for field in ADDON_FIELDS},
            })
        upgraded["bom_skus"].append({"sku": rule.sku, "category_id": category_id})

    non_bom_ids: dict[tuple[Any, ...], str] = {}
    for raw in document.get("non_bom_rules", []):
        rule = NonBomRule(**raw)
        signature = (rule.pricing_category, rule.preparation, rule.storage, rule.bag, rule.sticker)
        category_id = non_bom_ids.get(signature)
        if category_id is None:
            category_id = _category_code("NONBOM", len(non_bom_ids) + 1)
            non_bom_ids[signature] = category_id
            upgraded["non_bom_categories"].append({
                "id": category_id,
                "name": rule.pricing_category or "Be pavadinimo",
                "preparation": rule.preparation, "storage": rule.storage,
                "bag": rule.bag, "sticker": rule.sticker,
            })
        upgraded["non_bom_skus"].append({
            "sku": rule.sku, "name": rule.name, "product_category": rule.product_category,
            "category_id": category_id,
        })
    upgraded["schema_version"] = 2
    return upgrade_config(upgraded)


def validate_config(document: dict[str, Any]) -> dict[str, Any]:
    document = upgrade_config(document)
    adjustment = document.get("adjustment_rate")
    if isinstance(adjustment, bool) or not isinstance(adjustment, (int, float)) or not -1 < adjustment <= 0:
        raise ValueError("Korekcija turi būti didesnė nei -100 % ir ne didesnė nei 0 %.")
    seen: set[str] = set()
    category_ids = {row.get("id") for row in document.get("bom_categories", [])}
    if None in category_ids or len(category_ids) != len(document.get("bom_categories", [])):
        raise ValueError("BOM kainodaros kategorijų ID turi būti unikalūs.")
    rate_codes: set[str] = set()
    for raw in document.get("bom_category_rates", []):
        code = _text(raw.get("code"))
        if not code or code.casefold() in rate_codes:
            raise ValueError("Verslo BOM kategorijų kodai turi būti unikalūs.")
        rate_codes.add(code.casefold())
        for field in ADDON_FIELDS:
            _number(raw.get(field))
    for raw in document.get("bom_skus", []):
        normalized = _text(raw.get("sku")).casefold()
        if not normalized:
            raise ValueError("BOM SKU priskyrime trūksta SKU.")
        if normalized in seen:
            raise ValueError(f"Pasikartojantis BOM SKU: {raw.get('sku')}")
        if raw.get("category_id") not in category_ids:
            raise ValueError(f"BOM SKU {raw.get('sku')} priskirta neegzistuojanti kategorija.")
        seen.add(normalized)
    seen.clear()
    for raw in document.get("bom_products", []):
        sku = _text(raw.get("sku"))
        if not sku:
            raise ValueError("BOM produktų sąraše trūksta SKU.")
        normalized = sku.casefold()
        if normalized in seen:
            raise ValueError(f"Pasikartojantis BOM produktas: {sku}")
        seen.add(normalized)
    seen.clear()
    non_bom_category_ids = {row.get("id") for row in document.get("non_bom_categories", [])}
    if None in non_bom_category_ids or len(non_bom_category_ids) != len(document.get("non_bom_categories", [])):
        raise ValueError("Ne BOM kainodaros kategorijų ID turi būti unikalūs.")
    for raw in document.get("non_bom_skus", []):
        normalized = _text(raw.get("sku")).casefold()
        if not normalized:
            raise ValueError("Ne BOM taisyklėje trūksta SKU.")
        if normalized in seen:
            raise ValueError(f"Pasikartojanti ne BOM taisyklė: {raw.get('sku')}")
        if raw.get("category_id") not in non_bom_category_ids:
            raise ValueError(f"Ne BOM SKU {raw.get('sku')} priskirta neegzistuojanti kategorija.")
        seen.add(normalized)
    return document


def load_config(path: Path) -> dict[str, Any]:
    if not path.exists():
        return empty_config()
    return validate_config(json.loads(path.read_text(encoding="utf-8")))


def save_config(path: Path, document: dict[str, Any]) -> None:
    document = dict(document)
    document["updated_at"] = datetime.now(timezone.utc).isoformat()
    document = validate_config(document)
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(document, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(path)


def migrate_legacy_workbook(source: Path) -> dict[str, Any]:
    """Read the legacy workbook once and return application-owned rules."""
    workbook = load_workbook(source, data_only=True, read_only=True)
    required = {"bomai", "Kainodaros kategorijos", "Ne BOM pozicijos"}
    missing = required.difference(workbook.sheetnames)
    if missing:
        workbook.close()
        raise ValueError(f"Sename faile trūksta lapų: {', '.join(sorted(missing))}")

    pricing_rules: list[dict[str, Any]] = []
    migration_warnings: list[str] = []
    seen: set[str] = set()
    for row in workbook["Kainodaros kategorijos"].iter_rows(min_row=2, values_only=True):
        sku = _text(row[0])
        if not sku:
            continue
        normalized = sku.casefold()
        values = tuple(_number(value) for value in row[4:10])
        if normalized in seen:
            migration_warnings.append(f"Pasikartojanti taisyklė {sku}: perkelta pirmoji seno Excel eilutė.")
            continue
        seen.add(normalized)
        pricing_rules.append(asdict(PricingRule(
            sku=sku,
            category_id=_text(row[1]),
            category_name=_text(row[2]),
            odoo_category=_text(row[3]),
            **dict(zip(ADDON_FIELDS, values)),
        )))

    bom_products: list[dict[str, str]] = []
    seen.clear()
    for row in workbook["bomai"].iter_rows(min_row=3, values_only=True):
        sku = _text(row[1])
        if sku and sku.casefold() not in seen:
            seen.add(sku.casefold())
            bom_products.append({"sku": sku, "product_category": _text(row[2])})

    non_bom_rules: list[dict[str, Any]] = []
    seen.clear()
    for row in workbook["Ne BOM pozicijos"].iter_rows(min_row=2, values_only=True):
        sku = _text(row[0])
        if not sku:
            continue
        normalized = sku.casefold()
        if normalized in seen:
            workbook.close()
            raise ValueError(f"Pasikartojanti ne BOM taisyklė: {sku}")
        seen.add(normalized)
        non_bom_rules.append(asdict(NonBomRule(
            sku=sku,
            name=_text(row[1]),
            product_category=_text(row[2]),
            pricing_category=_text(row[3]),
            preparation=_number(row[6]),
            storage=_number(row[7]),
            bag=_number(row[8]),
            sticker=_number(row[9]),
        )))
    workbook.close()
    document = empty_config()
    document["pricing_rules"] = pricing_rules
    document["bom_products"] = bom_products
    document["non_bom_rules"] = non_bom_rules
    document["migration_warnings"] = migration_warnings
    document["schema_version"] = 1
    return validate_config(document)


def migrate_generated_audit(source: Path) -> dict[str, Any]:
    """Recover the already-verified rules from a generated pricing audit."""
    workbook = load_workbook(source, data_only=True, read_only=True)
    required = {"SO LINE PRICES", "BOM CATEGORY BREAKDOWN", "NON-BOM RULES"}
    missing = required.difference(workbook.sheetnames)
    if missing:
        workbook.close()
        raise ValueError(f"Audito faile trūksta lapų: {', '.join(sorted(missing))}")
    pricing: dict[str, dict[str, Any]] = {}
    for row in workbook["BOM CATEGORY BREAKDOWN"].iter_rows(min_row=2, values_only=True):
        sku = _text(row[2])
        if not sku:
            continue
        multiplier = _number(row[6], 1.0)
        if multiplier == 0:
            continue
        pricing.setdefault(sku.casefold(), asdict(PricingRule(
            sku, _text(row[3]), _text(row[4]), _text(row[5]),
            *(_number(value) / multiplier for value in row[7:13]),
        )))
    products: list[dict[str, str]] = []
    for row in workbook["SO LINE PRICES"].iter_rows(min_row=2, values_only=True):
        if _text(row[2]) == "BOM":
            products.append({"sku": _text(row[0]), "product_category": _text(row[3])})
    non_bom: list[dict[str, Any]] = []
    for row in workbook["NON-BOM RULES"].iter_rows(min_row=2, values_only=True):
        if _text(row[0]):
            non_bom.append(asdict(NonBomRule(
                sku=_text(row[0]), name=_text(row[1]), product_category=_text(row[2]),
                pricing_category=_text(row[3]), preparation=_number(row[5]), storage=_number(row[6]),
                bag=_number(row[7]), sticker=_number(row[8]),
            )))
    workbook.close()
    document = empty_config()
    document["pricing_rules"] = list(pricing.values())
    document["bom_products"] = products
    document["non_bom_rules"] = non_bom
    document["schema_version"] = 1
    return validate_config(document)


def pricing_rules_from_config(document: dict[str, Any]) -> list[PricingRule]:
    categories = {row["id"]: row for row in document["bom_categories"]}
    result = []
    for assignment in document["bom_skus"]:
        category = categories[assignment["category_id"]]
        result.append(PricingRule(
            sku=assignment["sku"], category_id=category.get("source_category_id", ""),
            category_name=category["name"], odoo_category=category.get("odoo_category", ""),
            **{field: float(category.get(field, 0)) for field in ADDON_FIELDS},
        ))
    return result


def compose_bom_category_rule(
    sku: str,
    expression: str,
    document: dict[str, Any],
) -> PricingRule:
    """Compose one SKU rule from Tamara business category codes."""
    rates = {
        _text(row.get("code")).casefold(): row
        for row in validate_config(document).get("bom_category_rates", [])
    }
    codes = [code.strip() for code in _text(expression).split("+") if code.strip()]
    missing = [code for code in codes if code.casefold() not in rates]
    if missing:
        raise ValueError(
            f"BOM kainodaros kombinacijoje {expression!r} nėra kategorijų: "
            + ", ".join(missing)
        )
    values = tuple(
        sum(_number(rates[code.casefold()].get(field)) for code in codes)
        for field in ADDON_FIELDS
    )
    names = " + ".join(rates[code.casefold()]["name"] for code in codes)
    return PricingRule(sku, expression, names, "", *values)


def non_bom_rules_from_config(document: dict[str, Any]) -> list[NonBomRule]:
    categories = {row["id"]: row for row in document["non_bom_categories"]}
    result = []
    for assignment in document["non_bom_skus"]:
        category = categories[assignment["category_id"]]
        result.append(NonBomRule(
            sku=assignment["sku"], name=assignment.get("name", ""),
            product_category=assignment.get("product_category", ""), pricing_category=category["name"],
            preparation=float(category.get("preparation", 0)), storage=float(category.get("storage", 0)),
            bag=float(category.get("bag", 0)), sticker=float(category.get("sticker", 0)),
        ))
    return result
