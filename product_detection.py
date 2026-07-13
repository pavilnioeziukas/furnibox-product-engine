from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


BASE_DIR = Path(__file__).resolve().parent

REFORM_FILE = BASE_DIR / "data" / "Reform BOM Transformer.xlsm"
ODOO_FILE = BASE_DIR / "output" / "Odoo_Snapshot.xlsx"
OUTPUT_FILE = BASE_DIR / "output" / "Product_Detection.xlsx"

REFORM_SHEET = "BOM VERTICAL"
ODOO_SHEET = "ODOO PRODUCTS"

REFORM_HEADER = "BOM SKU Code"
ODOO_HEADER = "Internal Reference"


def normalize_sku(value: object) -> str:
    if value is None:
        return ""

    return str(value).strip().upper()


def find_header_column(
    worksheet,
    header_name: str,
    max_rows: int = 20,
) -> tuple[int, int]:
    expected = normalize_sku(header_name)

    for row in worksheet.iter_rows(
        min_row=1,
        max_row=min(max_rows, worksheet.max_row),
        values_only=False,
    ):
        for cell in row:
            if normalize_sku(cell.value) == expected:
                return cell.row, cell.column

    raise ValueError(
        f"Lape '{worksheet.title}' nerasta antraštė '{header_name}'."
    )


def read_reform_products() -> dict[str, str]:
    if not REFORM_FILE.exists():
        raise FileNotFoundError(
            f"Nerastas Reform failas:\n{REFORM_FILE}"
        )

    workbook = load_workbook(
        REFORM_FILE,
        read_only=True,
        data_only=True,
        keep_vba=True,
    )

    try:
        if REFORM_SHEET not in workbook.sheetnames:
            raise ValueError(
                f"Reform faile nerastas lapas '{REFORM_SHEET}'."
            )

        worksheet = workbook[REFORM_SHEET]

        header_row, sku_column = find_header_column(
            worksheet,
            REFORM_HEADER,
        )

        products: dict[str, str] = {}

        for row in worksheet.iter_rows(
            min_row=header_row + 1,
            min_col=sku_column,
            max_col=sku_column,
            values_only=True,
        ):
            raw_sku = row[0]
            normalized_sku = normalize_sku(raw_sku)

            if not normalized_sku:
                continue

            if normalized_sku not in products:
                products[normalized_sku] = str(raw_sku).strip()

        return products

    finally:
        workbook.close()


def read_odoo_products() -> dict[str, dict]:
    if not ODOO_FILE.exists():
        raise FileNotFoundError(
            f"Nerastas Odoo Snapshot failas:\n{ODOO_FILE}"
        )

    workbook = load_workbook(
        ODOO_FILE,
        read_only=True,
        data_only=True,
    )

    try:
        if ODOO_SHEET not in workbook.sheetnames:
            raise ValueError(
                f"Odoo faile nerastas lapas '{ODOO_SHEET}'."
            )

        worksheet = workbook[ODOO_SHEET]

        header_row, sku_column = find_header_column(
            worksheet,
            ODOO_HEADER,
        )

        headers: dict[str, int] = {}

        header_values = next(
            worksheet.iter_rows(
                min_row=header_row,
                max_row=header_row,
                values_only=True,
            )
        )

        for column_number, value in enumerate(
            header_values,
            start=1,
        ):
            if value is not None:
                headers[str(value).strip()] = column_number

        required_columns = {
            "ID": headers.get("ID"),
            "Internal Reference": headers.get("Internal Reference"),
            "Name": headers.get("Name"),
            "Active": headers.get("Active"),
            "Category": headers.get("Category"),
        }

        missing_columns = [
            name
            for name, column in required_columns.items()
            if column is None
        ]

        if missing_columns:
            raise ValueError(
                "ODOO PRODUCTS lape trūksta stulpelių: "
                + ", ".join(missing_columns)
            )

        max_required_column = max(required_columns.values())

        products: dict[str, dict] = {}

        for row in worksheet.iter_rows(
            min_row=header_row + 1,
            max_col=max_required_column,
            values_only=True,
        ):
            raw_sku = row[sku_column - 1]
            normalized_sku = normalize_sku(raw_sku)

            if not normalized_sku:
                continue

            products[normalized_sku] = {
                "sku": str(raw_sku).strip(),
                "id": row[required_columns["ID"] - 1],
                "name": row[required_columns["Name"] - 1],
                "active": row[required_columns["Active"] - 1],
                "category": row[required_columns["Category"] - 1],
            }

        return products

    finally:
        workbook.close()


def build_detection_rows(
    reform_products: dict[str, str],
    odoo_products: dict[str, dict],
) -> list[dict]:
    rows = []

    for normalized_sku, reform_sku in sorted(
        reform_products.items(),
        key=lambda item: item[1],
    ):
        odoo_product = odoo_products.get(normalized_sku)

        if odoo_product:
            rows.append(
                {
                    "Reform SKU": reform_sku,
                    "Odoo SKU": odoo_product["sku"],
                    "Odoo Product ID": odoo_product["id"],
                    "Odoo Product Name": odoo_product["name"],
                    "Odoo Category": odoo_product["category"],
                    "Odoo Active": odoo_product["active"],
                    "Exists in Odoo": "YES",
                    "Next Step": "CHECK BOM",
                }
            )
        else:
            rows.append(
                {
                    "Reform SKU": reform_sku,
                    "Odoo SKU": "",
                    "Odoo Product ID": "",
                    "Odoo Product Name": "",
                    "Odoo Category": "",
                    "Odoo Active": "",
                    "Exists in Odoo": "NO",
                    "Next Step": "CREATE PRODUCT",
                }
            )

    return rows


def write_output(rows: list[dict]) -> None:
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PRODUCT DETECTION"

    headers = [
        "Reform SKU",
        "Odoo SKU",
        "Odoo Product ID",
        "Odoo Product Name",
        "Odoo Category",
        "Odoo Active",
        "Exists in Odoo",
        "Next Step",
    ]

    worksheet.append(headers)

    for row in rows:
        worksheet.append([row[header] for header in headers])

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    column_widths = {
        "A": 34,
        "B": 34,
        "C": 18,
        "D": 45,
        "E": 45,
        "F": 14,
        "G": 18,
        "H": 22,
    }

    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    new_fill = PatternFill(
        fill_type="solid",
        fgColor="FCE4D6",
    )

    existing_fill = PatternFill(
        fill_type="solid",
        fgColor="E2F0D9",
    )

    for row_number in range(2, worksheet.max_row + 1):
        exists = worksheet.cell(
            row=row_number,
            column=7,
        ).value

        row_fill = (
            existing_fill
            if exists == "YES"
            else new_fill
        )

        for column_number in range(1, len(headers) + 1):
            worksheet.cell(
                row=row_number,
                column=column_number,
            ).fill = row_fill

    summary = workbook.create_sheet("SUMMARY")

    total = len(rows)
    existing = sum(
        1
        for row in rows
        if row["Exists in Odoo"] == "YES"
    )
    new = total - existing

    summary.append(["Rodiklis", "Reikšmė"])
    summary.append(["Reform produktai", total])
    summary.append(["Yra Odoo", existing])
    summary.append(["Nėra Odoo", new])

    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font

    summary.column_dimensions["A"].width = 25
    summary.column_dimensions["B"].width = 15

    workbook.save(OUTPUT_FILE)


def main() -> None:
    print("Nuskaitomi Reform produktai...")
    reform_products = read_reform_products()
    print(f"Unikalių Reform produktų: {len(reform_products)}")

    print("Nuskaitomi Odoo produktai...")
    odoo_products = read_odoo_products()
    print(
        "Odoo produktų su Internal Reference: "
        f"{len(odoo_products)}"
    )

    print("Atliekamas produktų palyginimas...")
    rows = build_detection_rows(
        reform_products,
        odoo_products,
    )

    print("Kuriamas Product_Detection.xlsx...")
    write_output(rows)

    existing = sum(
        1
        for row in rows
        if row["Exists in Odoo"] == "YES"
    )
    new = len(rows) - existing

    print()
    print("=" * 60)
    print("PRODUCT DETECTION BAIGTAS")
    print(f"Reform produktai: {len(rows)}")
    print(f"Yra Odoo: {existing}")
    print(f"Nėra Odoo: {new}")
    print(f"Rezultatas: {OUTPUT_FILE}")
    print("=" * 60)


if __name__ == "__main__":
    main()