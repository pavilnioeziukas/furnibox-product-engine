"""Generate Odoo product import files from Reform BOM and product detection data."""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import load_settings
from odoo_client import OdooClient


REVIEW_CATEGORIES = {"CABINET HARDWARE", "FRONT HARDWARE", "INTERIOR STORAGE"}
IMPORT_HEADERS = [
    "Internal reference",
    "name",
    "route_ids/id",
    "type",
    "categ_id",
    "invoice_policy",
    "packaging_name",
]


def canon(value: object) -> str:
    return str(value or "").strip().upper()


def find_file(base: Path, filename: str) -> Path:
    for candidate in (base / filename, base / "output" / filename, base / "data" / filename):
        if candidate.exists():
            return candidate
    raise FileNotFoundError(f"Nerastas {filename}. Nurodykite kelią komandiniu parametru.")


def find_bom_input(base: Path) -> Path:
    candidates = []
    for folder in (base / "data", base):
        for pattern in ("Reform_BOM_Input*.xlsx", "BOM_for Furnibox*.xlsx"):
            candidates.extend(folder.glob(pattern))
    candidates = [path for path in candidates if not path.name.startswith("~$")]
    if not candidates:
        raise FileNotFoundError("Nerastas Reform BOM Input failas. Naudokite --bom-input.")
    return max(candidates, key=lambda path: path.stat().st_mtime)


def load_new_product_skus(path: Path) -> set[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["PRODUCT DETECTION"]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    index = {str(value).strip(): i for i, value in enumerate(headers) if value is not None}
    result = set()
    for row in rows:
        if canon(row[index["Exists in Odoo"]]) in {"NO", "NE", "FALSE", "0"}:
            result.add(canon(row[index["Reform SKU"]]))
    wb.close()
    return result


def load_reform_metadata(path: Path) -> dict[str, dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        header_row = None
        headers = None
        for row_no, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True), 1):
            values = [str(value).strip() if value is not None else "" for value in row]
            if "BOM SKU Code" in values:
                header_row, headers = row_no, values
                break
        if not header_row:
            continue
        index = {value: i for i, value in enumerate(headers) if value}
        metadata = {}
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            sku_raw = row[index["BOM SKU Code"]]
            sku = canon(sku_raw)
            if not sku:
                continue
            metadata[sku] = {
                "sku": str(sku_raw).strip(),
                "category": str(row[index["Product Catagory"]] or "UNCLASSIFIED").strip(),
                "name_1": str(row[index["SKU name 1"]] or "").strip(),
                "name_2": str(row[index["SKU name 2"]] or "").strip(),
            }
        wb.close()
        return metadata
    wb.close()
    raise ValueError(f"Faile {path.name} nerasta 'BOM SKU Code' antraštė.")


def load_bom_parents(path: Path) -> set[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["REFORM EDGES"]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    index = {str(value).strip(): i for i, value in enumerate(headers) if value is not None}
    result = {canon(row[index["Parent SKU"]]) for row in rows if row[index["Parent SKU"]]}
    wb.close()
    return result


def read_all(client: OdooClient, model: str, domain, fields, batch_size=1000):
    return client.search_read_all(model, domain, fields, batch_size=batch_size)


def external_ids(client: OdooClient, model: str) -> dict[int, str]:
    rows = read_all(client, "ir.model.data", [["model", "=", model]], ["module", "name", "res_id"])
    grouped = defaultdict(list)
    for row in rows:
        if row.get("res_id"):
            grouped[int(row["res_id"])].append(f"{row['module']}.{row['name']}")
    result = {}
    for record_id, values in grouped.items():
        # Prefer a stable module XML ID; otherwise use the Odoo export ID.
        values.sort(key=lambda value: (value.startswith("__export__."), value))
        result[record_id] = values[0]
    return result


def infer_category_map(client: OdooClient, metadata: dict[str, dict], required_categories: set[str]):
    categories = read_all(client, "product.category", [], ["id", "name", "complete_name"])
    category_by_id = {int(row["id"]): row for row in categories}
    category_xmlids = external_ids(client, "product.category")
    products = read_all(client, "product.product", [["default_code", "!=", False]], ["default_code", "categ_id"])

    votes = defaultdict(Counter)
    apack_votes = Counter()
    for product in products:
        sku = canon(product.get("default_code"))
        category_value = product.get("categ_id")
        if not category_value:
            continue
        category_id = int(category_value[0])
        if sku.startswith("APACK-"):
            apack_votes[category_id] += 1
        if sku in metadata:
            votes[metadata[sku]["category"]][category_id] += 1

    mapping = {}
    diagnostics = []
    # Only categories required by the products in this import belong in diagnostics.
    # The full BOM input may contain unrelated categories (for example SINK) that
    # have no new products in the current import batch.
    for reform_category in sorted(required_categories):
        category_votes = votes[reform_category]
        if not category_votes:
            diagnostics.append(f"Nerastas Odoo kategorijos pavyzdys Reform kategorijai: {reform_category}")
            continue
        category_id, count = category_votes.most_common(1)[0]
        xmlid = category_xmlids.get(category_id)
        if not xmlid:
            diagnostics.append(f"Odoo kategorija ID={category_id} neturi External ID: {reform_category}")
            continue
        mapping[reform_category] = {
            "category_id": category_id,
            "category_name": category_by_id[category_id].get("complete_name") or category_by_id[category_id]["name"],
            "external_id": xmlid,
            "matched_existing_products": count,
            "alternatives": "; ".join(
                f"{category_by_id[cid].get('complete_name') or category_by_id[cid]['name']} ({qty})"
                for cid, qty in category_votes.most_common()
            ),
        }

    apack = None
    if apack_votes:
        category_id, count = apack_votes.most_common(1)[0]
        xmlid = category_xmlids.get(category_id)
        if xmlid:
            apack = {
                "category_id": category_id,
                "category_name": category_by_id[category_id].get("complete_name") or category_by_id[category_id]["name"],
                "external_id": xmlid,
                "matched_existing_products": count,
                "alternatives": "; ".join(
                    f"{category_by_id[cid].get('complete_name') or category_by_id[cid]['name']} ({qty})"
                    for cid, qty in apack_votes.most_common()
                ),
            }
    if not apack:
        diagnostics.append("Nepavyko nustatyti APACK kategorijos arba jos External ID.")
    return mapping, apack, diagnostics


def infer_route_profile(client: OdooClient):
    """Infer exact route External IDs from existing products in selected Odoo."""
    route_xmlids = external_ids(client, "stock.route")
    route_records = read_all(client, "stock.route", [], ["id", "name"])
    route_names = {int(row["id"]): row["name"] for row in route_records}

    def profile_for_prefix(prefix: str):
        products = read_all(
            client, "product.product",
            [["default_code", "ilike", f"{prefix}-%"]],
            ["default_code", "product_tmpl_id"],
        )
        template_ids = sorted({
            int(row["product_tmpl_id"][0])
            for row in products if row.get("product_tmpl_id")
        })
        if not template_ids:
            return (), Counter()
        templates = read_all(
            client, "product.template", [["id", "in", template_ids]], ["route_ids"]
        )
        votes = Counter(
            tuple(sorted(int(value) for value in row.get("route_ids", [])))
            for row in templates
        )
        votes.pop((), None)
        return (votes.most_common(1)[0][0] if votes else ()), votes

    fpack_ids, fpack_votes = profile_for_prefix("FPACK")
    apack_ids, apack_votes = profile_for_prefix("APACK")
    diagnostics = []
    if not fpack_ids:
        diagnostics.append("Nepavyko nustatyti FPACK maršrutų pasirinktoje Odoo aplinkoje.")
    if not apack_ids:
        diagnostics.append("Nepavyko nustatyti APACK maršrutų pasirinktoje Odoo aplinkoje.")

    common_ids = set(fpack_ids) & set(apack_ids)
    mto_id = None
    for route_id in common_ids:
        xmlid = route_xmlids.get(route_id, "")
        name = canon(route_names.get(route_id, ""))
        if xmlid.endswith("route_warehouse0_mto") or "MTO" in name or "REPLENISH ON ORDER" in name:
            mto_id = route_id
            break
    if mto_id is None and len(common_ids) == 1:
        mto_id = next(iter(common_ids))
    if mto_id is None:
        diagnostics.append("Nepavyko vienareikšmiškai nustatyti bendro MTO maršruto.")

    def to_xmlids(route_ids, label):
        values = []
        for route_id in route_ids:
            xmlid = route_xmlids.get(route_id)
            if xmlid:
                values.append(xmlid)
            else:
                diagnostics.append(f"{label} maršrutas ID={route_id} neturi External ID.")
        return values

    fpack_xmlids = to_xmlids(fpack_ids, "FPACK")
    apack_xmlids = to_xmlids(apack_ids, "APACK")
    mto_xmlids = to_xmlids([mto_id], "MTO") if mto_id is not None else []
    profile = {
        "manufacture": ",".join(fpack_xmlids),
        "apack": ",".join(apack_xmlids),
        "mto": ",".join(mto_xmlids),
    }
    route_rows = []
    for usage, route_ids, votes in (
        ("BOM product / FPACK", fpack_ids, fpack_votes),
        ("APACK", apack_ids, apack_votes),
        ("Component without BOM", (mto_id,) if mto_id is not None else (), Counter()),
    ):
        route_rows.append({
            "Usage": usage,
            "Route Names": ", ".join(route_names.get(route_id, str(route_id)) for route_id in route_ids),
            "Route External IDs": ",".join(route_xmlids.get(route_id, "") for route_id in route_ids),
            "Matched Existing Products": votes.most_common(1)[0][1] if votes else "",
            "Observed Route Sets": "; ".join(
                f"{','.join(route_xmlids.get(route_id, str(route_id)) for route_id in route_set)} ({count})"
                for route_set, count in votes.most_common()
            ),
        })
    return profile, route_rows, diagnostics


def packaging_name(sku: str) -> str:
    core = sku[:-2] if sku.endswith("-A") else sku
    return core.rsplit("-", 1)[-1]


def apack_sku(fpack_sku: str) -> str:
    parts = fpack_sku.split("-", 2)
    if len(parts) != 3 or parts[0] != "FPACK":
        raise ValueError(f"Netinkamas FPACK SKU: {fpack_sku}")
    return f"APACK-{parts[1]}-C-{parts[2]}-A"


def routes_for(sku: str, bom_parents: set[str], route_profile: dict[str, str]) -> str:
    if sku.startswith("APACK-"):
        return route_profile["apack"]
    if sku in bom_parents:
        return route_profile["manufacture"]
    return route_profile["mto"]


def build_rows(new_skus, metadata, bom_parents, category_map, apack_category, route_profile):
    import_rows = []
    review_rows = []
    diagnostics = []
    for sku in sorted(new_skus):
        meta = metadata.get(sku)
        if not meta:
            diagnostics.append(f"SKU nėra Reform BOM Input faile: {sku}")
            continue
        category = meta["category"]
        mapping = category_map.get(category)
        if not mapping:
            diagnostics.append(f"SKU {sku}: kategorija {category} neturi Odoo External ID")
            continue
        row = {
            "Internal reference": meta["sku"],
            "name": meta["name_1"] or meta["name_2"] or meta["sku"],
            "route_ids/id": routes_for(sku, bom_parents, route_profile),
            "type": "Storable Product",
            "categ_id": mapping["external_id"],
            "invoice_policy": "Delivered quantities",
            "packaging_name": packaging_name(sku) if sku.startswith("FPACK-") else "",
            "Reform Category": category,
            "Product Name 2": meta["name_2"],
            "Review reason": "",
        }
        if category in REVIEW_CATEGORIES:
            row["Review reason"] = "Accessory category requires manual review"
            review_rows.append(row)
        else:
            import_rows.append(row)

        if sku.startswith("FPACK-"):
            if not apack_category:
                diagnostics.append(f"APACK nesukurtas, nes nenustatyta APACK kategorija: {sku}")
                continue
            duplicate = dict(row)
            duplicate["Internal reference"] = apack_sku(sku)
            duplicate["route_ids/id"] = route_profile["apack"]
            duplicate["categ_id"] = apack_category["external_id"]
            duplicate["packaging_name"] = packaging_name(sku)
            duplicate["Reform Category"] = "APACK (generated from FPACK)"
            duplicate["Review reason"] = ""
            # FPACK review is currently not expected, but preserve the same decision if it occurs.
            (review_rows if category in REVIEW_CATEGORIES else import_rows).append(duplicate)
    return import_rows, review_rows, diagnostics


def style_sheet(ws):
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx in range(1, ws.max_column + 1):
        values = [str(ws.cell(row, col_idx).value or "") for row in range(1, min(ws.max_row, 500) + 1)]
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(map(len, values)) + 2, 55)


def write_workbook(path: Path, rows, extra_columns=False, category_rows=None,
                   route_rows=None, diagnostics=None, info=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "PRODUCT IMPORT" if not extra_columns else "REVIEW"
    headers = IMPORT_HEADERS + (["Reform Category", "Product Name 2", "Review reason"] if extra_columns else [])
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    style_sheet(ws)

    if category_rows is not None:
        cat = wb.create_sheet("CATEGORY MAP")
        cat_headers = ["Reform Category", "Odoo Category", "Odoo Category External ID",
                       "Matched Existing Products", "Observed Alternatives"]
        cat.append(cat_headers)
        for row in category_rows:
            cat.append([row.get(header, "") for header in cat_headers])
        style_sheet(cat)
    if route_rows is not None:
        route = wb.create_sheet("ROUTE MAP")
        route_headers = ["Usage", "Route Names", "Route External IDs",
                         "Matched Existing Products", "Observed Route Sets"]
        route.append(route_headers)
        for row in route_rows:
            route.append([row.get(header, "") for header in route_headers])
        style_sheet(route)
    if diagnostics is not None:
        diag = wb.create_sheet("DIAGNOSTICS")
        diag.append(["Message"])
        for message in diagnostics:
            diag.append([message])
        style_sheet(diag)
    if info is not None:
        inf = wb.create_sheet("INFO")
        inf.append(["Parameter", "Value"])
        for key, value in info.items():
            inf.append([key, value])
        style_sheet(inf)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="Generuoja naujų produktų Odoo importo failus.")
    parser.add_argument("--products", type=Path, help="Product_Detection.xlsx")
    parser.add_argument("--reform-map", type=Path, help="Reform_MAP.xlsx")
    parser.add_argument("--bom-input", type=Path, help="Reform BOM Input .xlsx")
    parser.add_argument("--output-dir", type=Path, help="Rezultatų aplankas")
    args = parser.parse_args()

    base = Path(__file__).resolve().parent
    products_path = args.products or find_file(base, "Product_Detection.xlsx")
    reform_map_path = args.reform_map or find_file(base, "Reform_MAP.xlsx")
    bom_input_path = args.bom_input or find_bom_input(base)
    output_dir = args.output_dir or base / "output"

    new_skus = load_new_product_skus(products_path)
    metadata = load_reform_metadata(bom_input_path)
    bom_parents = load_bom_parents(reform_map_path)

    settings = load_settings()
    client = OdooClient(settings)
    uid = client.authenticate()
    print(f"Prisijungta prie Odoo. UID={uid}")
    required_categories = {
        metadata[sku]["category"] for sku in new_skus if sku in metadata
    }
    category_map, apack_category, category_diagnostics = infer_category_map(
        client, metadata, required_categories
    )
    route_profile, route_rows, route_diagnostics = infer_route_profile(client)
    import_rows, review_rows, row_diagnostics = build_rows(
        new_skus, metadata, bom_parents, category_map, apack_category, route_profile
    )
    diagnostics = category_diagnostics + route_diagnostics + row_diagnostics

    category_rows = []
    for reform_category, mapping in sorted(category_map.items()):
        category_rows.append({
            "Reform Category": reform_category,
            "Odoo Category": mapping["category_name"],
            "Odoo Category External ID": mapping["external_id"],
            "Matched Existing Products": mapping["matched_existing_products"],
            "Observed Alternatives": mapping["alternatives"],
        })
    if apack_category:
        category_rows.append({
            "Reform Category": "APACK (generated from FPACK)",
            "Odoo Category": apack_category["category_name"],
            "Odoo Category External ID": apack_category["external_id"],
            "Matched Existing Products": apack_category["matched_existing_products"],
            "Observed Alternatives": apack_category["alternatives"],
        })

    ready_path = output_dir / "Odoo_New_Products_Import.xlsx"
    review_path = output_dir / "Odoo_New_Products_Review.xlsx"
    info = {
        "New Reform products": len(new_skus),
        "Ready import rows (includes generated APACK)": len(import_rows),
        "Review rows": len(review_rows),
        "Generated APACK": sum(row["Internal reference"].startswith("APACK-") for row in import_rows + review_rows),
        "Product detection source": str(products_path),
        "Reform map source": str(reform_map_path),
        "BOM input source": str(bom_input_path),
        "Manufacture routes": route_profile["manufacture"],
        "APACK routes": route_profile["apack"],
        "MTO-only route": route_profile["mto"],
    }
    write_workbook(ready_path, import_rows, category_rows=category_rows,
                   route_rows=route_rows, diagnostics=diagnostics, info=info)
    write_workbook(review_path, review_rows, extra_columns=True,
                   category_rows=category_rows, route_rows=route_rows,
                   diagnostics=diagnostics, info=info)

    print("\nPRODUKTŲ IMPORTO FAILAI SUKURTI")
    print("Paruošta importui:", len(import_rows))
    print("Peržiūrai:", len(review_rows))
    print("Sugeneruota APACK:", info["Generated APACK"])
    print("Importas:", ready_path)
    print("Peržiūra:", review_path)
    print("Diagnostikos įrašai:", len(diagnostics))


if __name__ == "__main__":
    main()
