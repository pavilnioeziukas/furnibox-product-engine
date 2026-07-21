#!/usr/bin/env python3
"""
Iš Odoo produkcinės aplinkos pagal Sales Order numerius ištraukia
surenkamų spintelių kiekį.

Atrankos taisyklė:
- produkto Internal Reference (default_code) baigiasi "-A"

Rezultatas:
- assembled_cabinets_by_so.xlsx
  - SUMMARY: suvestinė pagal SO
  - DETAILS: atrinktos SO eilutės
  - NOT_FOUND: nerasti SO

Naudojimas:
1. Įdiek priklausomybę:
   pip install openpyxl

2. Nustatyk prisijungimo duomenis per aplinkos kintamuosius:
   set ODOO_URL=https://odoo.furnibox.lt
   set ODOO_DB=DUOMENU_BAZES_PAVADINIMAS
   set ODOO_USERNAME=tavo.elpastas@furnibox.lt
   set ODOO_API_KEY=TAVO_API_RAKTAS

3. Paleisk:
   python extract_assembled_cabinets.py orders.txt
"""

from __future__ import annotations

import argparse
import os
import sys
import xmlrpc.client
from collections import defaultdict
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


DEFAULT_OUTPUT = "assembled_cabinets_by_so.xlsx"


def load_orders(path: Path) -> list[str]:
    if not path.exists():
        raise FileNotFoundError(f"Nerastas failas: {path}")

    orders: list[str] = []
    seen: set[str] = set()

    for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
        order = raw_line.strip()
        if not order or order.startswith("#"):
            continue
        if order not in seen:
            seen.add(order)
            orders.append(order)

    if not orders:
        raise ValueError("Užsakymų sąrašas tuščias.")

    return orders


def get_required_env(name: str) -> str:
    value = os.getenv(name, "").strip()
    if not value:
        raise RuntimeError(f"Nenustatytas aplinkos kintamasis {name}.")
    return value


def connect_odoo() -> tuple[xmlrpc.client.ServerProxy, int]:
    url = get_required_env("ODOO_URL").rstrip("/")
    db = get_required_env("ODOO_DB")
    username = get_required_env("ODOO_USERNAME")
    api_key = get_required_env("ODOO_API_KEY")

    common = xmlrpc.client.ServerProxy(
        f"{url}/xmlrpc/2/common",
        allow_none=True,
    )
    uid = common.authenticate(db, username, api_key, {})
    if not uid:
        raise RuntimeError("Nepavyko prisijungti prie Odoo. Patikrink DB, vartotoją ir API raktą.")

    models = xmlrpc.client.ServerProxy(
        f"{url}/xmlrpc/2/object",
        allow_none=True,
    )

    return models, uid


def execute_kw(
    models: xmlrpc.client.ServerProxy,
    uid: int,
    model: str,
    method: str,
    args: list[Any],
    kwargs: dict[str, Any] | None = None,
) -> Any:
    db = get_required_env("ODOO_DB")
    api_key = get_required_env("ODOO_API_KEY")
    return models.execute_kw(db, uid, api_key, model, method, args, kwargs or {})


def fetch_data(
    models: xmlrpc.client.ServerProxy,
    uid: int,
    order_names: list[str],
) -> tuple[list[dict[str, Any]], list[str]]:
    orders = execute_kw(
        models,
        uid,
        "sale.order",
        "search_read",
        [[("name", "in", order_names)]],
        {
            "fields": ["id", "name", "state", "partner_id"],
            "limit": 0,
        },
    )

    orders_by_id = {order["id"]: order for order in orders}
    found_names = {order["name"] for order in orders}
    not_found = [name for name in order_names if name not in found_names]

    if not orders:
        return [], not_found

    lines = execute_kw(
        models,
        uid,
        "sale.order.line",
        "search_read",
        [[
            ("order_id", "in", list(orders_by_id)),
            ("display_type", "=", False),
        ]],
        {
            "fields": [
                "order_id",
                "product_id",
                "product_uom_qty",
                "product_uom",
                "name",
            ],
            "limit": 0,
        },
    )

    product_ids = sorted({
        line["product_id"][0]
        for line in lines
        if line.get("product_id")
    })

    products: list[dict[str, Any]] = []
    if product_ids:
        products = execute_kw(
            models,
            uid,
            "product.product",
            "read",
            [product_ids],
            {
                "fields": ["default_code", "name"],
            },
        )

    products_by_id = {product["id"]: product for product in products}

    details: list[dict[str, Any]] = []

    for line in lines:
        product_value = line.get("product_id")
        if not product_value:
            continue

        product_id = product_value[0]
        product = products_by_id.get(product_id, {})
        internal_reference = (product.get("default_code") or "").strip()

        if not internal_reference.upper().endswith("-A"):
            continue

        order_id = line["order_id"][0]
        order = orders_by_id[order_id]

        details.append({
            "so": order["name"],
            "state": order.get("state", ""),
            "customer": order["partner_id"][1] if order.get("partner_id") else "",
            "internal_reference": internal_reference,
            "product_name": product.get("name") or product_value[1],
            "quantity": line.get("product_uom_qty") or 0,
            "uom": line["product_uom"][1] if line.get("product_uom") else "",
            "line_description": line.get("name") or "",
        })

    return details, not_found


def autosize_worksheet(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)


def save_excel(
    output_path: Path,
    order_names: list[str],
    details: list[dict[str, Any]],
    not_found: list[str],
) -> None:
    totals: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    line_counts: dict[str, int] = defaultdict(int)
    states: dict[str, str] = {}
    customers: dict[str, str] = {}

    for row in details:
        so = row["so"]
        totals[so] += Decimal(str(row["quantity"]))
        line_counts[so] += 1
        states[so] = row["state"]
        customers[so] = row["customer"]

    wb = Workbook()

    summary = wb.active
    summary.title = "SUMMARY"
    summary.append([
        "Sales Order",
        "Status",
        "Customer",
        "Assembled cabinet lines",
        "Assembled cabinets qty",
        "Result",
    ])

    found_set = set(order_names) - set(not_found)

    for so in order_names:
        if so in not_found:
            summary.append([so, "", "", 0, 0, "SO NOT FOUND"])
        else:
            qty = float(totals[so])
            summary.append([
                so,
                states.get(so, ""),
                customers.get(so, ""),
                line_counts.get(so, 0),
                qty,
                "FOUND -A PRODUCTS" if qty else "NO -A PRODUCTS",
            ])

    details_ws = wb.create_sheet("DETAILS")
    details_ws.append([
        "Sales Order",
        "Status",
        "Customer",
        "Internal Reference",
        "Product",
        "Quantity",
        "UoM",
        "SO line description",
    ])

    order_position = {name: index for index, name in enumerate(order_names)}
    sorted_details = sorted(
        details,
        key=lambda row: (
            order_position.get(row["so"], 999999),
            row["internal_reference"],
        ),
    )

    for row in sorted_details:
        details_ws.append([
            row["so"],
            row["state"],
            row["customer"],
            row["internal_reference"],
            row["product_name"],
            float(Decimal(str(row["quantity"]))),
            row["uom"],
            row["line_description"],
        ])

    not_found_ws = wb.create_sheet("NOT_FOUND")
    not_found_ws.append(["Sales Order"])
    for so in not_found:
        not_found_ws.append([so])

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        autosize_worksheet(ws)

    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Suskaičiuoja Odoo SO eilutes, kurių produkto Internal Reference baigiasi '-A'."
    )
    parser.add_argument(
        "orders_file",
        nargs="?",
        default="orders.txt",
        help="TXT failas su vienu SO numeriu eilutėje. Numatyta: orders.txt",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=DEFAULT_OUTPUT,
        help=f"Rezultato XLSX failas. Numatyta: {DEFAULT_OUTPUT}",
    )
    args = parser.parse_args()

    try:
        order_names = load_orders(Path(args.orders_file))
        print(f"Nuskaityta SO: {len(order_names)}")

        models, uid = connect_odoo()
        print("Prisijungta prie Odoo.")

        details, not_found = fetch_data(models, uid, order_names)
        save_excel(Path(args.output), order_names, details, not_found)

        total_qty = sum(Decimal(str(row["quantity"])) for row in details)
        print(f"Rasta '-A' eilučių: {len(details)}")
        print(f"Bendras surenkamų spintelių kiekis: {total_qty}")
        print(f"Nerasta SO: {len(not_found)}")
        print(f"Rezultatas: {Path(args.output).resolve()}")
        return 0

    except Exception as exc:
        print(f"KLAIDA: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
