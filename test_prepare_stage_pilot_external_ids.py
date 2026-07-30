from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from bom_release.generator import IMPORT_HEADERS
from prepare_stage_pilot_external_ids import (
    StageExternalIdError,
    prepare_stage_pilot,
)


class FakeClient:
    def __init__(self, products, external_ids):
        self.products = products
        self.external_ids = external_ids

    def search_read_all(
        self, model, domain, fields, order="id asc", context=None
    ):
        if model == "product.product":
            return list(self.products)
        if model == "ir.model.data":
            wanted_model = domain[0][2]
            wanted_ids = set(domain[1][2])
            return [
                row
                for row in self.external_ids
                if row["model"] == wanted_model
                and row["res_id"] in wanted_ids
            ]
        raise AssertionError(model)


def make_source(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOM import"
    sheet.append(IMPORT_HEADERS)
    first = [None] * len(IMPORT_HEADERS)
    first[0] = "APACK-EU-C-CAB01-COS001-A"
    first[1] = "production.parent"
    first[3] = "CON7X50"
    first[4] = "production.component"
    first[5] = 2
    first[6] = "Manufacture this product"
    first[8] = "__export__.stock_picking_type_21_0e009783"
    sheet.append(first)
    second = [None] * len(IMPORT_HEADERS)
    second[3] = "SCREW1"
    second[4] = "production.screw"
    second[5] = 4
    sheet.append(second)
    workbook.save(path)
    workbook.close()


class PrepareStagePilotExternalIdsTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.base = Path(self.temp.name)
        self.source = self.base / "pilot.xlsx"
        self.output = self.base / "stage.xlsx"
        self.audit = self.base / "audit.json"
        make_source(self.source)
        self.products = [
            {
                "id": 11,
                "default_code": "APACK-EU-C-CAB01-COS001-A",
                "product_tmpl_id": [101, "APACK"],
                "active": True,
            },
            {
                "id": 12,
                "default_code": "CON7X50",
                "product_tmpl_id": [102, "CON"],
                "active": True,
            },
            {
                "id": 13,
                "default_code": "SCREW1",
                "product_tmpl_id": [103, "SCREW"],
                "active": True,
            },
        ]
        self.external_ids = [
            {
                "model": "product.template",
                "res_id": 101,
                "module": "__export__",
                "name": "stage_parent",
            },
            {
                "model": "product.product",
                "res_id": 12,
                "module": "__export__",
                "name": "stage_component",
            },
            {
                "model": "product.product",
                "res_id": 13,
                "module": "__export__",
                "name": "stage_screw",
            },
        ]

    def tearDown(self):
        self.temp.cleanup()

    def test_remaps_parent_and_components_but_not_operation_type(self):
        result = prepare_stage_pilot(
            FakeClient(self.products, self.external_ids),
            self.source,
            self.output,
            self.audit,
        )
        workbook = load_workbook(self.output, read_only=True)
        rows = list(workbook["BOM import"].iter_rows(values_only=True))
        workbook.close()
        self.assertEqual(rows[1][1], "__export__.stage_parent")
        self.assertEqual(rows[1][4], "__export__.stage_component")
        self.assertEqual(rows[2][4], "__export__.stage_screw")
        self.assertEqual(
            rows[1][8], "__export__.stock_picking_type_21_0e009783"
        )
        self.assertEqual(result["status"], "PASS")
        self.assertEqual(json.loads(self.audit.read_text())["odoo_changes"], 0)

    def test_missing_sku_is_blocked(self):
        with self.assertRaisesRegex(StageExternalIdError, "Stage nerasti SKU"):
            prepare_stage_pilot(
                FakeClient(self.products[:-1], self.external_ids),
                self.source,
                self.output,
            )

    def test_duplicate_sku_is_blocked(self):
        duplicate = dict(self.products[1], id=99)
        with self.assertRaisesRegex(StageExternalIdError, "neunikalūs SKU"):
            prepare_stage_pilot(
                FakeClient(self.products + [duplicate], self.external_ids),
                self.source,
                self.output,
            )

    def test_missing_external_id_is_blocked(self):
        with self.assertRaisesRegex(
            StageExternalIdError, "neturi product.product External ID"
        ):
            prepare_stage_pilot(
                FakeClient(self.products, self.external_ids[:-1]),
                self.source,
                self.output,
            )

    def test_export_external_id_has_priority(self):
        external_ids = self.external_ids + [
            {
                "model": "product.product",
                "res_id": 12,
                "module": "custom",
                "name": "component",
            }
        ]
        prepare_stage_pilot(
            FakeClient(self.products, external_ids),
            self.source,
            self.output,
        )
        workbook = load_workbook(self.output, read_only=True)
        value = workbook["BOM import"].cell(2, 5).value
        workbook.close()
        self.assertEqual(value, "__export__.stage_component")


if __name__ == "__main__":
    unittest.main()
