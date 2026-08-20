"""Paruošia naujų BOM tipų pasiūlymus peržiūrai.

Programa nieko neimportuoja ir nekeičia Odoo. Ji perskaito:
1) output/Odoo_MAP.xlsx – aktyvius esamus Odoo BOM;
2) output/MAP_Comparison.xlsx – produktus, kuriems reikia naujo BOM;
ir sukuria output/BOM_Type_Review.xlsx.
"""

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from output_paths import environment_output_dir


def text(value: object) -> str:
    """Suvienodina tekstą palyginimui, bet nekeičia šaltinio failų."""
    return str(value or "").strip()


def canon(value: object) -> str:
    return text(value).upper()


def sequence_number(value: object) -> float:
    """Tuščią ar netinkamą Sequence laikome blogesniu už bet kurį skaičių."""
    try:
        return float(value)
    except (TypeError, ValueError):
        return float("inf")


def write_date(value: object) -> datetime:
    """Paverčia Odoo Write Date į datą, kad galėtume pasirinkti naujausią."""
    if isinstance(value, datetime):
        return value
    raw = text(value)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(raw[:19], fmt)
        except ValueError:
            pass
    return datetime.min


def family_key(sku: str) -> tuple[str, str]:
    """Sukuria produkto struktūros raktą analogų paieškai.

    EU ir US bei CAB01/CAB02/CAB03 suvienodinami, nes tai lygiagrečios
    tos pačios konstrukcijos šeimos. Paskutinio kodo skaitinė dalis nuimama:
    HIG013 ir HIG020 patenka į HIG šeimą.
    """
    parts = canon(sku).split("-")
    if parts and parts[-1] == "A":
        parts = parts[:-1]
    last = parts[-1] if parts else ""
    match = re.match(r"([A-Z]+)", last)
    code_family = match.group(1) if match else last

    lane = []
    for part in parts[:-1]:
        part = re.sub(r"^CAB\d+$", "CAB#", part)
        if part in {"EU", "US", "EUB", "USB"}:
            part = "REGION"
        lane.append(part)
    return "-".join(lane), code_family


def choose_active_reference(rows: list[dict]) -> dict:
    """Pasirenka vieną etaloninį BOM pagal sutartą prioritetą.

    Odoo_MAP.xlsx turi tik aktyvius BOM. Iš jų:
    1. imame mažiausią Sequence (0, jei toks yra);
    2. jei Sequence vienoda, imame naujausią Write Date;
    3. jei ir data vienoda, imame didžiausią BOM ID.
    """
    lowest_sequence = min(sequence_number(row["Sequence"]) for row in rows)
    candidates = [
        row for row in rows
        if sequence_number(row["Sequence"]) == lowest_sequence
    ]
    return max(
        candidates,
        key=lambda row: (write_date(row["Write Date"]), int(row["BOM ID"] or 0)),
    )


def load_reference_boms(path: Path) -> list[dict]:
    """Nuskaito aktyvius BOM ir kiekvienam produktui palieka vieną etaloną."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["BOM SELECTION"]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    index = {name: i for i, name in enumerate(headers)}

    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        sku = canon(row[index["Parent SKU"]])
        if not sku:
            continue
        record = {name: row[pos] for name, pos in index.items()}
        grouped[sku].append(record)
    wb.close()

    references = []
    for sku, records in grouped.items():
        chosen = choose_active_reference(records)
        chosen["Canonical SKU"] = sku
        references.append(chosen)
    return references


def load_new_boms(path: Path) -> list[dict]:
    """Nuskaito tik MAP palyginimo lape NEW BOMS esančius produktus."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["NEW BOMS"]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    result = [dict(zip(headers, row)) for row in rows]
    wb.close()
    return result


def business_rule(category: str, sku: str) -> tuple[str, str] | None:
    """Taiko Edgaro patvirtintas taisykles prieš analogų paiešką."""
    category = canon(category)
    sku = canon(sku)
    if category == "CABINETS":
        return "phantom", "PATVIRTINTA TAISYKLĖ: CABINETS = KIT"
    if category == "PREPACK CABINETS" or sku.startswith(("FPACK-", "APACK-")):
        return "normal", "PATVIRTINTA TAISYKLĖ: PREPACK/FPACK/APACK = MANUFACTURE"
    if category == "CABINET SHELF":
        return "phantom", "PATVIRTINTA TAISYKLĖ: CABINET SHELF = KIT; GAMYBA VYKSTA SHELF PP"
    if category == "SINK":
        # Plautuvės ir sifono dalys gaunamos bei laikomos atskirai. Pardavimo
        # metu Odoo turi išskleisti komplektą į komponentus, o ne kurti MO.
        return "phantom", "PATVIRTINTA TAISYKLĖ: SINK KOMPONENTAI ATEINA ATSKIRAI = KIT"
    if category == "LED HARDWARE":
        # Linijinio LED apšvietimo rinkiniai išduodami kaip komponentų
        # komplektai. Jiems neturi būti kuriamas MO ar priskiriamos operacijos.
        return "phantom", "PATVIRTINTA TAISYKLĖ: LED HARDWARE = KIT"
    if category == "INTERIOR STORAGE":
        # Šie trys komplektai turi kelis komponentus / tvirtinimo detales ir
        # atitinka esamus Production gamybinius analogus.
        if sku.endswith(("-MIS010", "-MIS050", "-MIS051")):
            return "normal", "PATVIRTINTA TAISYKLĖ: KELIŲ DALIŲ INTERIOR STORAGE = MANUFACTURE"
        # Patvirtinti vieno realaus komponento apvalkalai: ištraukiamos
        # lentynos, turn shelf, plate holder ir LED spot light.
        if (
            re.search(r"-(?:EUB|USB)-P-ACC02-SLF20[0-8]$", "-" + sku)
            or sku.endswith(("-MIS015", "-MIS016", "-MIS017"))
            or sku.endswith(("-MIS950", "-MIS951", "-MIS952"))
        ):
            return "phantom", "PATVIRTINTA TAISYKLĖ: VIENO KOMPONENTO INTERIOR STORAGE = KIT"
    if category in {"FRONT HARDWARE", "CABINET HARDWARE"}:
        # Šie produktai yra realiai komplektuojami iš kelių furnitūros,
        # tvirtinimo ir dokumentacijos komponentų. Production analogai taip
        # pat naudoja gamybinį BOM tipą.
        return "normal", "PATVIRTINTA TAISYKLĖ: HARDWARE KOMPLEKTAI = MANUFACTURE"
    return None


def infer_from_analogs(sku: str, references: list[dict]) -> tuple[str, str, str, str]:
    """Parenka artimiausią tos pačios struktūrinės šeimos etaloną.

    Jei visa šeima turi vieną tipą, rezultatas yra HIGH.
    Jei šeimoje yra abu tipai, siūlomas artimiausio SKU tipas, bet rezultatas
    pažymimas REVIEW, kad žmogus jį patikrintų prieš importą.
    """
    wanted_family = family_key(sku)
    candidates = [
        row for row in references
        if family_key(row["Canonical SKU"]) == wanted_family
    ]
    if not candidates:
        return "", "REVIEW", "", "NERASTA TOS PAČIOS STRUKTŪROS ETALONO"

    candidates.sort(
        key=lambda row: SequenceMatcher(
            None, canon(sku), row["Canonical SKU"]
        ).ratio(),
        reverse=True,
    )
    analog = candidates[0]
    types = {canon(row["BOM Type"]).lower() for row in candidates}
    proposed = canon(analog["BOM Type"]).lower()
    if len(types) == 1:
        confidence = "HIGH"
        reason = "VIENODAS VISŲ TOS PAČIOS STRUKTŪROS ETALONŲ TIPAS"
    else:
        confidence = "REVIEW"
        reason = "ŠEIMOJE YRA ABU TIPAI – PARINKTAS ARTIMIAUSIAS SKU"
    return proposed, confidence, text(analog["Parent SKU"]), reason


def add_review_sheet(wb: Workbook, rows: list[dict]) -> None:
    """Sukuria lengvai filtruojamą peržiūros lentelę."""
    ws = wb.active
    ws.title = "BOM TYPE REVIEW"
    headers = [
        "Parent SKU", "Category", "Product Name 1", "Product Name 2",
        "BOM Line Count", "Proposed BOM Type", "Business Meaning",
        "Confidence", "Reference SKU", "Assignment Reason",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column, header in enumerate(headers, start=1):
        values = [
            len(text(ws.cell(row=row, column=column).value))
            for row in range(1, min(ws.max_row, 400) + 1)
        ]
        ws.column_dimensions[get_column_letter(column)].width = min(
            max(max(values, default=len(header)) + 2, 14), 55
        )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Paruošia naujų BOM tipų pasiūlymus peržiūrai."
    )
    parser.add_argument("--odoo-map", type=Path, help="Odoo_MAP.xlsx kelias")
    parser.add_argument("--comparison", type=Path, help="MAP_Comparison.xlsx kelias")
    parser.add_argument("--output", type=Path, help="Rezultato BOM_Type_Review.xlsx kelias")
    args = parser.parse_args()

    base = Path(__file__).resolve().parent
    # Tikslinės aplinkos palyginimas ir rezultatas laikomi jos aplanke.
    # Tačiau BOM tipo etalonas VISADA imamas iš Production, nes Stage DB
    # yra senesnė ir gali turėti pasenusius arba prieštaringus BOM tipus.
    target_output_dir = environment_output_dir(base)
    production_output_dir = base / "output" / "production"
    odoo_path = args.odoo_map or production_output_dir / "Odoo_MAP.xlsx"
    comparison_path = args.comparison or target_output_dir / "MAP_Comparison.xlsx"
    output_path = args.output or target_output_dir / "BOM_Type_Review.xlsx"

    # 1 ŽINGSNIS: pasirenkame po vieną aktyvų PRODUCTION etaloninį BOM
    # kiekvienam produktui, nepriklausomai nuo testuojamos aplinkos.
    print("BOM tipo etalonas:", odoo_path)
    references = load_reference_boms(odoo_path)

    # 2 ŽINGSNIS: paimame produktus, kuriems MAP palyginimas siūlo kurti BOM.
    new_boms = load_new_boms(comparison_path)

    # 3 ŽINGSNIS: pirmiausia taikome patvirtintas verslo taisykles.
    # Jei taisyklė netinka, ieškome panašios Internal Reference struktūros.
    review_rows = []
    for product in new_boms:
        sku = text(product.get("Parent SKU"))
        category = text(product.get("Category"))
        fixed = business_rule(category, sku)
        if fixed:
            proposed, reason = fixed
            confidence, analog = "HIGH", ""
        else:
            proposed, confidence, analog, reason = infer_from_analogs(
                sku, references
            )
        review_rows.append({
            **product,
            "Proposed BOM Type": proposed,
            "Business Meaning": (
                "KIT" if proposed == "phantom"
                else "MANUFACTURE" if proposed == "normal"
                else ""
            ),
            "Confidence": confidence,
            "Reference SKU": analog,
            "Assignment Reason": reason,
        })

    # 4 ŽINGSNIS: sukuriame tik peržiūros failą. Odoo ši programa nekeičia.
    wb = Workbook()
    add_review_sheet(wb, review_rows)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    high = sum(row["Confidence"] == "HIGH" for row in review_rows)
    review = len(review_rows) - high
    print("\nBOM TYPE PERŽIŪRA SUKURTA")
    print("Failas:", output_path)
    print("Visi nauji BOM:", len(review_rows))
    print("HIGH confidence:", high)
    print("Reikia peržiūrėti:", review)
    print("Odoo pakeitimų neatlikta.")


if __name__ == "__main__":
    main()
