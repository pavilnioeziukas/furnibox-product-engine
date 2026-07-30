from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from bom_release.generator import IMPORT_HEADERS, canon
from stage_product_id_map import (
    StageProductIdMapError,
    load_map,
    read_bom_skus,
)


def remap_bom(
    source: Path,
    output: Path,
    id_map: dict[str, Any],
    audit_path: Path | None = None,
) -> dict[str, Any]:
    parent_skus, component_skus = read_bom_skus([source])
    records = id_map["records"]
    missing_parents = sorted(
        sku
        for sku in parent_skus
        if not records.get(sku, {}).get("product_template_external_id")
    )
    missing_components = sorted(
        sku
        for sku in component_skus
        if not records.get(sku, {}).get("product_product_external_id")
    )
    if missing_parents or missing_components:
        details: list[str] = []
        if missing_parents:
            details.append(
                "Žodyne nėra parent product.template External ID: "
                + ", ".join(missing_parents)
            )
        if missing_components:
            details.append(
                "Žodyne nėra komponentų product.product External ID: "
                + ", ".join(missing_components)
            )
        raise StageProductIdMapError("; ".join(details))

    workbook = load_workbook(source, data_only=False)
    try:
        sheet = workbook["BOM import"]
        if list(next(sheet.iter_rows(values_only=True))) != IMPORT_HEADERS:
            raise StageProductIdMapError(
                f"{source.name}: neteisingi importo stulpeliai."
            )
        parent_changes = 0
        component_changes = 0
        operation_types: set[str] = set()
        for row_number in range(2, sheet.max_row + 1):
            parent = canon(sheet.cell(row_number, 1).value)
            component = canon(sheet.cell(row_number, 4).value)
            if parent:
                value = records[parent]["product_template_external_id"]
                cell = sheet.cell(row_number, 2)
                parent_changes += int(cell.value != value)
                cell.value = value
            if component:
                value = records[component]["product_product_external_id"]
                cell = sheet.cell(row_number, 5)
                component_changes += int(cell.value != value)
                cell.value = value
            operation_type = str(sheet.cell(row_number, 9).value or "").strip()
            if operation_type:
                operation_types.add(operation_type)
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output)
    finally:
        workbook.close()

    audit = {
        "status": "PASS",
        "source": str(source.resolve()),
        "output": str(output.resolve()),
        "map_stage_url": id_map.get("stage_url"),
        "map_exported_at_utc": id_map.get("exported_at_utc"),
        "parent_skus": parent_skus,
        "component_skus": component_skus,
        "parent_external_ids_changed": parent_changes,
        "component_external_ids_changed": component_changes,
        "operation_type_external_ids_unchanged": sorted(operation_types),
        "odoo_changes": 0,
    }
    if audit_path:
        audit_path.parent.mkdir(parents=True, exist_ok=True)
        audit_path.write_text(
            json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    return audit


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Pagal išsaugotą Stage ID žodyną persieja BOM produkto ir "
            "komponentų External ID. Prie Odoo nesijungia."
        )
    )
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--id-map", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--audit", type=Path)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    audit = remap_bom(
        args.source, args.output, load_map(args.id_map), args.audit
    )
    print("=" * 80)
    print("BOM STAGE EXTERNAL ID REMAP")
    print("=" * 80)
    print("Statusas:", audit["status"])
    print("BOM produktai:", len(audit["parent_skus"]))
    print("Komponentai:", len(audit["component_skus"]))
    print("Pakeisti parent External ID:", audit["parent_external_ids_changed"])
    print(
        "Pakeisti komponentų External ID:",
        audit["component_external_ids_changed"],
    )
    print("Failas:", args.output.resolve())
    if args.audit:
        print("Auditas:", args.audit.resolve())
    print("Odoo pakeitimai: 0")


if __name__ == "__main__":
    main()
