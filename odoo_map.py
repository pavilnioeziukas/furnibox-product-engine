"""Generate a canonical Odoo BOM map from Odoo API.

Selection rule for multiple active BOMs of the same product:
1. Lowest sequence (0 is preferred when present).
2. Newest write_date when sequence is equal.

Outputs:
- ODOO EDGES: canonical Parent SKU -> Component SKU relationships.
- ODOO MAP: human-readable lv1/lv2/lv3 layout.
- BOM SELECTION: selected and rejected active BOM versions.
- DIAGNOSTICS: missing SKUs, ambiguous templates and cycles.
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

@dataclass(frozen=True)
class SelectedBom:
    bom_id: int
    parent_sku: str
    reference: str
    sequence: int
    bom_type: str
    write_date: str


def m2o_id(value: Any) -> int | None:
    if isinstance(value, (list, tuple)) and value:
        return int(value[0])
    if isinstance(value, int) and not isinstance(value, bool):
        return value
    return None


def parse_date(value: Any) -> datetime:
    if not value:
        return datetime(1970, 1, 1)
    text = str(value)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return datetime(1970, 1, 1)


def normalize_sku(value: Any) -> str:
    return str(value or "").strip()


def fetch_data(client: Any) -> tuple[list[dict], list[dict], list[dict]]:
    products = client.products()
    boms = client.search_read_all(
        "mrp.bom",
        [["active", "=", True]],
        [
            "id", "code", "active", "sequence", "product_tmpl_id",
            "product_id", "product_qty", "product_uom_id", "type",
            "company_id", "write_date",
        ],
        order="product_tmpl_id asc, sequence asc, write_date desc, id desc",
        context={"active_test": False},
    )
    bom_lines = client.search_read_all(
        "mrp.bom.line",
        [["bom_id.active", "=", True]],
        [
            "id", "bom_id", "product_id", "product_qty",
            "product_uom_id", "sequence", "company_id", "write_date",
        ],
        order="bom_id asc, sequence asc, id asc",
    )
    return products, boms, bom_lines


def build_product_indexes(products: list[dict]) -> tuple[dict[int, str], dict[int, list[str]]]:
    by_variant: dict[int, str] = {}
    by_template: dict[int, list[str]] = defaultdict(list)
    for product in products:
        sku = normalize_sku(product.get("default_code"))
        if not sku:
            continue
        product_id = int(product["id"])
        template_id = m2o_id(product.get("product_tmpl_id"))
        by_variant[product_id] = sku
        if template_id is not None and sku not in by_template[template_id]:
            by_template[template_id].append(sku)
    return by_variant, by_template


def resolve_parent_skus(
    bom: dict,
    by_variant: dict[int, str],
    by_template: dict[int, list[str]],
) -> tuple[list[str], str | None]:
    variant_id = m2o_id(bom.get("product_id"))
    if variant_id is not None:
        sku = by_variant.get(variant_id)
        return ([sku] if sku else []), (None if sku else "BOM variant has no Internal Reference")

    template_id = m2o_id(bom.get("product_tmpl_id"))
    skus = by_template.get(template_id or -1, [])
    if not skus:
        return [], "BOM template has no product with Internal Reference"
    if len(skus) > 1:
        return skus, "Template has multiple variants; BOM applied to every variant"
    return skus, None


def select_boms(
    boms: list[dict],
    by_variant: dict[int, str],
    by_template: dict[int, list[str]],
) -> tuple[dict[str, SelectedBom], list[list[Any]], list[list[Any]]]:
    candidates: dict[str, list[dict]] = defaultdict(list)
    diagnostics: list[list[Any]] = []

    for bom in boms:
        skus, note = resolve_parent_skus(bom, by_variant, by_template)
        if note:
            diagnostics.append(["BOM TARGET", bom.get("id"), "", note])
        for sku in skus:
            candidates[sku].append(bom)

    selected: dict[str, SelectedBom] = {}
    selection_rows: list[list[Any]] = []
    for sku, rows in sorted(candidates.items()):
        ordered = sorted(
            rows,
            key=lambda r: (
                int(r.get("sequence") or 0),
                -parse_date(r.get("write_date")).timestamp(),
                -int(r["id"]),
            ),
        )
        winner = ordered[0]
        selected[sku] = SelectedBom(
            bom_id=int(winner["id"]),
            parent_sku=sku,
            reference=normalize_sku(winner.get("code")),
            sequence=int(winner.get("sequence") or 0),
            bom_type=normalize_sku(winner.get("type")),
            write_date=normalize_sku(winner.get("write_date")),
        )
        for position, bom in enumerate(ordered):
            selection_rows.append([
                sku,
                int(bom["id"]),
                normalize_sku(bom.get("code")),
                int(bom.get("sequence") or 0),
                normalize_sku(bom.get("type")),
                normalize_sku(bom.get("write_date")),
                "SELECTED" if position == 0 else "REJECTED",
            ])
    return selected, selection_rows, diagnostics


def build_edges(
    selected: dict[str, SelectedBom],
    bom_lines: list[dict],
    by_variant: dict[int, str],
) -> tuple[list[list[Any]], dict[str, list[tuple[str, float]]], list[list[Any]]]:
    parents_by_bom: dict[int, list[str]] = defaultdict(list)
    for sku, bom in selected.items():
        parents_by_bom[bom.bom_id].append(sku)

    edges: list[list[Any]] = []
    graph: dict[str, list[tuple[str, float]]] = defaultdict(list)
    diagnostics: list[list[Any]] = []
    seen: set[tuple[str, str, float, int]] = set()

    for line in bom_lines:
        bom_id = m2o_id(line.get("bom_id"))
        if bom_id not in parents_by_bom:
            continue
        component_id = m2o_id(line.get("product_id"))
        component_sku = by_variant.get(component_id or -1, "")
        if not component_sku:
            diagnostics.append(["COMPONENT", line.get("id"), bom_id, "Component has no Internal Reference"])
            continue
        qty = float(line.get("product_qty") or 0)
        for parent_sku in parents_by_bom[bom_id]:
            dedupe_key = (parent_sku, component_sku, qty, int(line["id"]))
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            bom = selected[parent_sku]
            edges.append([
                parent_sku, component_sku, qty,
                bom.bom_id, bom.reference, bom.sequence,
                bom.bom_type, bom.write_date, int(line["id"]),
            ])
            graph[parent_sku].append((component_sku, qty))
    edges.sort(key=lambda r: (r[0], r[8], r[1]))
    return edges, graph, diagnostics


def build_display_map(
    edges: list[list[Any]],
    graph: dict[str, list[tuple[str, float]]],
) -> tuple[list[list[Any]], list[list[Any]]]:
    direct_by_parent: dict[str, list[tuple[str, float]]] = defaultdict(list)
    for row in edges:
        direct_by_parent[row[0]].append((row[1], row[2]))

    output: list[list[Any]] = []
    diagnostics: list[list[Any]] = []
    for lv1 in sorted(direct_by_parent):
        for lv2, qty_lv2 in direct_by_parent[lv1]:
            children = graph.get(lv2, [])
            if not children:
                output.append([lv1, lv2, qty_lv2, "", ""])
                continue
            for index, (lv3, qty_lv3) in enumerate(children):
                if lv3 in {lv1, lv2}:
                    diagnostics.append(["CYCLE", "", "", f"{lv1} -> {lv2} -> {lv3}"])
                output.append([
                    lv1 if index == 0 else "",
                    lv2 if index == 0 else "",
                    qty_lv2 if index == 0 else "",
                    lv3,
                    qty_lv3,
                ])
    return output, diagnostics


def add_table(ws, headers: list[str], rows: list[list[Any]], table_name: str) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    if rows:
        table = Table(displayName=table_name, ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    for column in ws.columns:
        values = [len(str(c.value or "")) for c in column[:200]]
        ws.column_dimensions[column[0].column_letter].width = min(max(values, default=10) + 2, 45)


def write_workbook(
    output_path: Path,
    edges: list[list[Any]],
    display_map: list[list[Any]],
    selection_rows: list[list[Any]],
    diagnostics: list[list[Any]],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    add_table(
        wb.create_sheet("ODOO EDGES"),
        ["Parent SKU", "Component SKU", "Quantity", "BOM ID", "BOM Reference", "BOM Sequence", "BOM Type", "Write Date", "BOM Line ID"],
        edges,
        "OdooEdges",
    )
    add_table(
        wb.create_sheet("ODOO MAP"),
        ["lv1", "lv2", "qty_lv2", "lv3", "qty_lv3"],
        display_map,
        "OdooMap",
    )
    add_table(
        wb.create_sheet("BOM SELECTION"),
        ["Parent SKU", "BOM ID", "BOM Reference", "Sequence", "BOM Type", "Write Date", "Status"],
        selection_rows,
        "BomSelection",
    )
    add_table(
        wb.create_sheet("DIAGNOSTICS"),
        ["Type", "Record ID", "Related ID", "Message"],
        diagnostics,
        "MapDiagnostics",
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    from config import load_settings
    from odoo_client import OdooClient

    settings = load_settings()
    client = OdooClient(settings)
    uid = client.authenticate()
    print(f"Prisijungta prie Odoo. UID={uid}")

    products, boms, bom_lines = fetch_data(client)
    print(f"Produktai: {len(products)}")
    print(f"Aktyvūs BOM: {len(boms)}")
    print(f"Aktyvių BOM eilutės: {len(bom_lines)}")

    by_variant, by_template = build_product_indexes(products)
    selected, selection_rows, diagnostics = select_boms(boms, by_variant, by_template)
    edges, graph, edge_diagnostics = build_edges(selected, bom_lines, by_variant)
    display_map, map_diagnostics = build_display_map(edges, graph)
    diagnostics.extend(edge_diagnostics)
    diagnostics.extend(map_diagnostics)

    output_path = settings.output_dir / "Odoo_MAP.xlsx"
    write_workbook(output_path, edges, display_map, selection_rows, diagnostics)

    print("\nODOO MAP SUKURTAS")
    print("Failas:", output_path)
    print("Pasirinkti BOM:", len(selected))
    print("Tiesioginiai ryšiai:", len(edges))
    print("MAP eilutės:", len(display_map))
    print("Diagnostikos įrašai:", len(diagnostics))


if __name__ == "__main__":
    main()
