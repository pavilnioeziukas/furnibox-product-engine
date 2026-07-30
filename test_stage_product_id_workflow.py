from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from bom_release.generator import IMPORT_HEADERS
from remap_bom_external_ids import remap_bom
from stage_product_id_map import (
    StageProductIdMapError,
    export_stage_product_id_map,
    read_bom_skus,
)


class FakeClient:
    def __init__(self, products, external_ids, templates=None):
        self.products = products
        self.external_ids = external_ids
        self.templates = templates or []
        self.calls = []

    def search_read_all(
        self, model, domain, fields, order="id asc", context=None,
        batch_size=1000,
    ):
        self.calls.append((model, domain, context))
        if model == "product.product":
            conditions = [item for item in domain if isinstance(item, list)]
            field = conditions[0][0]
            if field == "default_code":
                operator = conditions[0][1]
                if operator == "!=":
                    return [r for r in self.products if r["default_code"]]
                wanted = {str(item[2]).strip().upper() for item in conditions}
                return [
                    r for r in self.products
                    if str(r["default_code"]).strip().upper() in wanted
                ]
            wanted = set(conditions[0][2])
            return [
                r
                for r in self.products
                if r.get("product_tmpl_id")
                and r["product_tmpl_id"][0] in wanted
            ]
        if model == "product.template":
            conditions = [item for item in domain if isinstance(item, list)]
            operator = conditions[0][1]
            if operator == "!=":
                return [r for r in self.templates if r["default_code"]]
            wanted = {str(item[2]).strip().upper() for item in conditions}
            return [
                r for r in self.templates
                if str(r["default_code"]).strip().upper() in wanted
            ]
        if model == "ir.model.data":
            wanted_model = domain[0][2]
            wanted_ids = set(domain[1][2])
            return [
                r
                for r in self.external_ids
                if r["model"] == wanted_model and r["res_id"] in wanted_ids
            ]
        raise AssertionError(model)


def make_bom(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOM import"
    sheet.append(IMPORT_HEADERS)
    row = [None] * len(IMPORT_HEADERS)
    row[0] = "APACK-1"
    row[1] = "production.parent"
    row[3] = "PART-1"
    row[4] = "production.component"
    row[5] = 2
    row[8] = "operation.type"
    sheet.append(row)
    workbook.save(path)
    workbook.close()


class StageProductIdWorkflowTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.base = Path(self.temp.name)
        self.source = self.base / "source.xlsx"
        self.output = self.base / "output.xlsx"
        make_bom(self.source)
        self.products = [
            {
                "id": 11,
                "default_code": "APACK-1",
                "product_tmpl_id": [101, "APACK"],
                "active": True,
            },
            {
                "id": 12,
                "default_code": "PART-1",
                "product_tmpl_id": [102, "PART"],
                "active": False,
            },
        ]
        self.external_ids = [
            {
                "model": "product.template",
                "res_id": 101,
                "module": "__export__",
                "name": "stage_parent",
            },
            {
                "model": "product.product",
                "res_id": 12,
                "module": "__export__",
                "name": "stage_component",
            },
        ]
        self.templates = [
            {"id": 101, "default_code": "APACK-1", "active": True},
            {"id": 102, "default_code": "PART-1", "active": False},
        ]

    def tearDown(self):
        self.temp.cleanup()

    def test_exports_template_and_product_ids_by_internal_reference(self):
        parents, components = read_bom_skus([self.source])
        client = FakeClient(self.products, self.external_ids, self.templates)
        result = export_stage_product_id_map(
            client, parents, components, "https://stage.example"
        )
        self.assertEqual(
            result["records"]["APACK-1"]["product_template_external_id"],
            "__export__.stage_parent",
        )
        self.assertEqual(
            result["records"]["PART-1"]["product_product_external_id"],
            "__export__.stage_component",
        )
        self.assertEqual(result["odoo_changes"], 0)
        product_calls = [
            call for call in client.calls if call[0] == "product.product"
        ]
        self.assertTrue(product_calls)
        self.assertEqual(product_calls[0][2], {"active_test": False})

    def test_catalog_query_is_bounded_and_normalized(self):
        products = [
            dict(self.products[0], default_code=" apack-1 "),
            dict(self.products[1], default_code="part-1"),
        ]
        templates = [
            dict(self.templates[0], default_code=" APACK-1 "),
            dict(self.templates[1], default_code="PART-1"),
        ]
        client = FakeClient(products, self.external_ids, templates)
        result = export_stage_product_id_map(
            client, ["APACK-1"], ["PART-1"], "https://stage.example"
        )
        self.assertIn("APACK-1", result["records"])
        self.assertFalse(
            any(
                call[1] == [["default_code", "!=", False]]
                for call in client.calls
            )
        )

    def test_remap_changes_only_two_external_id_columns(self):
        parents, components = read_bom_skus([self.source])
        id_map = export_stage_product_id_map(
            FakeClient(self.products, self.external_ids, self.templates),
            parents,
            components,
            "https://stage.example",
        )
        remap_bom(self.source, self.output, id_map)
        before_book = load_workbook(self.source, read_only=True)
        before = before_book["BOM import"]
        after_book = load_workbook(self.output, read_only=True)
        after = after_book["BOM import"]
        before_values = list(before.iter_rows(values_only=True))[1]
        after_values = list(after.iter_rows(values_only=True))[1]
        before_book.close()
        after_book.close()
        self.assertEqual(after_values[1], "__export__.stage_parent")
        self.assertEqual(after_values[4], "__export__.stage_component")
        for index in range(len(IMPORT_HEADERS)):
            if index not in {1, 4}:
                self.assertEqual(before_values[index], after_values[index])

    def test_missing_internal_reference_is_blocked(self):
        with self.assertRaisesRegex(
            StageProductIdMapError, "Stage nerasti Internal Reference"
        ):
            export_stage_product_id_map(
                FakeClient(
                    self.products[:-1], self.external_ids, self.templates[:-1]
                ),
                ["APACK-1"],
                ["PART-1"],
                "https://stage.example",
            )

    def test_duplicate_internal_reference_is_blocked(self):
        duplicate = dict(self.products[1], id=99)
        with self.assertRaisesRegex(
            StageProductIdMapError, "neunikalūs Internal Reference"
        ):
            export_stage_product_id_map(
                FakeClient(
                    self.products + [duplicate], self.external_ids, self.templates
                ),
                ["APACK-1"],
                ["PART-1"],
                "https://stage.example",
            )

    def test_missing_external_id_is_blocked(self):
        with self.assertRaisesRegex(
            StageProductIdMapError, "neturi External ID"
        ):
            export_stage_product_id_map(
                FakeClient(
                    self.products, self.external_ids[:-1], self.templates
                ),
                ["APACK-1"],
                ["PART-1"],
                "https://stage.example",
            )

    def test_stale_map_is_blocked_for_unknown_sku(self):
        id_map = {
            "schema_version": 1,
            "environment": "stage",
            "records": {
                "APACK-1": {
                    "product_template_external_id": "__export__.parent"
                }
            },
        }
        with self.assertRaisesRegex(
            StageProductIdMapError, "komponentų product.product"
        ):
            remap_bom(self.source, self.output, id_map)

    def test_component_can_be_resolved_from_template_to_single_variant(self):
        products = [self.products[0], dict(self.products[1], default_code="")]
        result = export_stage_product_id_map(
            FakeClient(products, self.external_ids, self.templates),
            ["APACK-1"],
            ["PART-1"],
            "https://stage.example",
        )
        self.assertEqual(
            result["records"]["PART-1"]["product_product_external_id"],
            "__export__.stage_component",
        )


if __name__ == "__main__":
    unittest.main()
